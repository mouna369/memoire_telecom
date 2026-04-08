"""
DIAGNOSTIC — Trouver le bon champ ID entre Excel et MongoDB
Usage : python3 diagnostic_id.py
"""

from pymongo  import MongoClient
from openpyxl import load_workbook
from bson     import ObjectId
import glob
import sys

MONGO_URI  = "mongodb://yousrahadjabderrahmane_db_user:C8wjIvWqOBUjK66u@ac-1ksfahb-shard-00-00.gejzu4a.mongodb.net:27017,ac-1ksfahb-shard-00-01.gejzu4a.mongodb.net:27017,ac-1ksfahb-shard-00-02.gejzu4a.mongodb.net:27017/?ssl=true&replicaSet=atlas-mdnqx7-shard-0&authSource=admin&appName=Cluster0"
DB_NAME    = "telecom_algerie_new"
COLLECTION = "dataset_unifie"

print("\n" + "="*60)
print("  DIAGNOSTIC — TROUVER LE BON CHAMP ID")
print("="*60)

# Connexion MongoDB
try:
    client     = MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)
    client.server_info()
    collection = client[DB_NAME][COLLECTION]
    print(f"\n  Connecté à MongoDB Atlas")
    print(f"  Total documents : {collection.count_documents({})}")
except Exception as e:
    print(f"  Connexion échouée : {e}")
    sys.exit(1)

# Lire les 5 premiers IDs depuis le fichier Excel
fichiers = glob.glob("validation_*.xlsx")
if not fichiers:
    print("  Aucun fichier Excel trouvé")
    sys.exit(1)

fichier = fichiers[0]
print(f"\n  Fichier Excel utilisé : {fichier}")

wb = load_workbook(fichier, read_only=True)
ws = wb["Validation"]

# Trouver la colonne MONGO_ID
headers = {}
for cell in ws[1]:
    if cell.value:
        headers[str(cell.value).strip().upper()] = cell.column - 1

idx_id = headers.get("MONGO_ID")
if idx_id is None:
    print("  Colonne MONGO_ID introuvable")
    sys.exit(1)

# Lire les 5 premiers IDs du Excel
ids_excel = []
for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
    val = row[idx_id] if idx_id < len(row) else None
    if val:
        ids_excel.append(str(val).strip())

wb.close()

print(f"\n  5 premiers MONGO_ID dans le Excel :")
for id_val in ids_excel:
    print(f"   → '{id_val}' (longueur: {len(id_val)})")

# Tester les différentes façons de chercher dans MongoDB
print(f"\n  {'─'*50}")
print(f"  TEST 1 : chercher par champ 'mongo_id' (string)")
print(f"  {'─'*50}")
for id_val in ids_excel[:3]:
    doc = collection.find_one({"mongo_id": id_val})
    if doc:
        print(f"   TROUVE  : mongo_id='{id_val}'")
        print(f"            _id = {doc['_id']}")
    else:
        print(f"   ABSENT  : mongo_id='{id_val}'")

print(f"\n  {'─'*50}")
print(f"  TEST 2 : chercher par champ '_id' (ObjectId)")
print(f"  {'─'*50}")
for id_val in ids_excel[:3]:
    try:
        oid = ObjectId(id_val)
        doc = collection.find_one({"_id": oid})
        if doc:
            print(f"   TROUVE  : _id=ObjectId('{id_val}')")
            print(f"            mongo_id = {doc.get('mongo_id', 'N/A')}")
        else:
            print(f"   ABSENT  : _id=ObjectId('{id_val}')")
    except Exception as e:
        print(f"   ERREUR  : '{id_val}' → {e}")

print(f"\n  {'─'*50}")
print(f"  TEST 3 : chercher par champ '_id' (string direct)")
print(f"  {'─'*50}")
for id_val in ids_excel[:3]:
    doc = collection.find_one({"_id": id_val})
    if doc:
        print(f"   TROUVE  : _id='{id_val}' (string)")
    else:
        print(f"   ABSENT  : _id='{id_val}' (string)")

# Afficher un vrai document MongoDB pour comparaison
print(f"\n  {'─'*50}")
print(f"  EXEMPLE d'un vrai document MongoDB :")
print(f"  {'─'*50}")
exemple = collection.find_one({})
if exemple:
    print(f"   _id       : {exemple.get('_id')} (type: {type(exemple.get('_id')).__name__})")
    print(f"   mongo_id  : {exemple.get('mongo_id', 'ABSENT')} (type: {type(exemple.get('mongo_id')).__name__ if exemple.get('mongo_id') else 'N/A'})")
    print(f"   label     : {exemple.get('label', 'N/A')}")
    print(f"   source    : {exemple.get('source', 'N/A')}")

print(f"\n{'='*60}")
print(f"  Regardez quel TEST a dit TROUVE → c'est le bon champ")
print(f"{'='*60}\n")

client.close()
