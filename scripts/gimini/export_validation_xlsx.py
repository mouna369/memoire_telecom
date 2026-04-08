# ╔══════════════════════════════════════════════════════════════════╗
# ║  EXPORT MONGODB → EXCEL pour validation manuelle (VERSION SPLIT)║
# ║                                                                  ║
# ║  Génère DEUX fichiers Excel : un pour toi, un pour ton binôme   ║
# ╚══════════════════════════════════════════════════════════════════╝

from pymongo  import MongoClient
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, GradientFill)
from openpyxl.utils  import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import random
import sys
import os

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CONFIG
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
MONGO_URI   = "mongodb://yousrahadjabderrahmane_db_user:C8wjIvWqOBUjK66u@ac-1ksfahb-shard-00-00.gejzu4a.mongodb.net:27017,ac-1ksfahb-shard-00-01.gejzu4a.mongodb.net:27017,ac-1ksfahb-shard-00-02.gejzu4a.mongodb.net:27017/?ssl=true&replicaSet=atlas-mdnqx7-shard-0&authSource=admin&appName=Cluster0"
DB_NAME          = "telecom_algerie_new"
COLLECTION       = "dataset_unifie"
CONF_SEUIL_BAS   = 0.75
SAMPLE_PAR_LABEL = 150
OUTPUT_DIR       = "."

# Demander les deux annotateurs
print("\n  📝 PARTAGE DES TWEETS ENTRE LES DEUX ANNOTATEURS")
print("  " + "="*50)
ANNOTATEUR_1 = input("  Prénom du 1er annotateur (ex: Sara) : ").strip()
ANNOTATEUR_2 = input("  Prénom du 2nd annotateur (ex: Amina) : ").strip()

if not ANNOTATEUR_1 or not ANNOTATEUR_2:
    print("  ❌ Les deux prénoms sont requis.")
    sys.exit(1)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CONNEXION
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
try:
    client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)
    client.server_info()
    col = client[DB_NAME][COLLECTION]
    print(f"  ✅ Connecté à MongoDB Atlas")
except Exception as e:
    print(f"  ❌ Connexion échouée : {e}")
    sys.exit(1)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CONVERSION CONFIDENCE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CONVERT_CONF = {
    "$convert": {
        "input"  : "$confidence",
        "to"     : "double",
        "onError": None,
        "onNull" : None
    }
}

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CHARGER LES TWEETS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def charger_tweets():
    print(f"\n  Chargement des tweets...")

    # Partie 1 : douteux conf < 0.75
    pipeline_douteux = [
        {"$addFields": {"conf_num": CONVERT_CONF}},
        {"$match": {
            "conf_num": {"$lt": CONF_SEUIL_BAS},
            f"validation.{ANNOTATEUR_1}": {"$exists": False},
            f"validation.{ANNOTATEUR_2}": {"$exists": False}
        }},
        {"$sort":    {"conf_num": 1}},
        {"$project": {"conf_num": 0}}
    ]
    douteux = list(col.aggregate(pipeline_douteux))
    print(f"   Douteux (conf < {CONF_SEUIL_BAS}) : {len(douteux):>4} tweets")

    # Partie 2 : échantillon 150 par label
    labels = ["negatif", "neutre", "positif"]
    sample = []
    for label in labels:
        pipeline_sample = [
            {"$addFields": {"conf_num": CONVERT_CONF}},
            {"$match": {
                "label"   : label,
                "conf_num": {"$gte": CONF_SEUIL_BAS},
                f"validation.{ANNOTATEUR_1}": {"$exists": False},
                f"validation.{ANNOTATEUR_2}": {"$exists": False}
            }},
            {"$sample":  {"size": SAMPLE_PAR_LABEL}},
            {"$project": {"conf_num": 0}}
        ]
        s = list(col.aggregate(pipeline_sample))
        sample.extend(s)
        print(f"   Échantillon '{label}'         : {len(s):>4} tweets")

    # Fusion + dédup
    seen, tous = set(), []
    for doc in douteux + sample:
        sid = str(doc["_id"])
        if sid not in seen:
            seen.add(sid)
            try:
                cf = float(doc.get("confidence", 1))
                doc["__type__"] = "DOUTEUX" if cf < CONF_SEUIL_BAS else "SAMPLE"
            except:
                doc["__type__"] = "CONF_INVALIDE"
            tous.append(doc)

    random.shuffle(tous)
    print(f"\n  Total à valider : {len(tous):>4} tweets")
    return tous

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# STYLES
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

LABEL_FILLS = {
    "negatif" : make_fill("FFD7D7"),
    "neutre"  : make_fill("FFF9D7"),
    "positif" : make_fill("D7FFD7"),
    "ambigu"  : make_fill("D7E8FF"),
}

TYPE_FILLS = {
    "DOUTEUX"       : make_fill("FFE0B2"),
    "SAMPLE"        : make_fill("E8F5E9"),
    "CONF_INVALIDE" : make_fill("F3E5F5"),
}

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CRÉER UN FICHIER EXCEL POUR UN ANNOTATEUR
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def creer_excel(tweets, annotateur, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation"

    # En-têtes
    headers = [
        ("A", "MONGO_ID",         16, "B0BEC5"),
        ("B", "TYPE",              12, "B0BEC5"),
        ("C", "SOURCE",            12, "B0BEC5"),
        ("D", "TEXTE_TWEET",       60, "90CAF9"),
        ("E", "LABEL_GEMINI",      16, "A5D6A7"),
        ("F", "CONFIDENCE",        14, "A5D6A7"),
        ("G", "SCORE",             10, "A5D6A7"),
        ("H", "RAISON_GEMINI",     35, "A5D6A7"),
        ("I", "MON_LABEL",         18, "FFCC02"),
        ("J", "COMMENTAIRE",       30, "FFE082"),
    ]

    header_font  = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_letter, title, width, color in headers:
        cell = ws[f"{col_letter}1"]
        cell.value     = title
        cell.font      = header_font
        cell.fill      = make_fill(color)
        cell.alignment = header_align
        cell.border    = make_border()
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Validation dropdown
    dv = DataValidation(
        type="list",
        formula1='"negatif,neutre,positif,ambigu,SUPPRIMER"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Valeur invalide",
        error='Choisir : negatif / neutre / positif / ambigu / SUPPRIMER'
    )
    ws.add_data_validation(dv)

    # Remplir les données
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    for i, doc in enumerate(tweets, start=2):
        texte = (
            doc.get("Commentaire_Client_Original") or
            doc.get("Commentaire_Client") or
            doc.get("normalized_full") or ""
        )
        label      = doc.get("label", "")
        confidence = doc.get("confidence", "")
        score      = doc.get("score", "")
        reason     = doc.get("reason", "")
        source     = doc.get("source", "")
        typ        = doc.get("__type__", "SAMPLE")
        mongo_id   = str(doc.get("_id", ""))

        row_data = [
            mongo_id,
            typ,
            source,
            texte,
            label,
            confidence,
            score,
            reason,
            "",
            "",
        ]

        for j, val in enumerate(row_data, start=1):
            cell           = ws.cell(row=i, column=j, value=str(val) if val else "")
            cell.border    = make_border()
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = left if j in (4, 8, 10) else center

        # Couleur ligne selon label Gemini
        label_fill = LABEL_FILLS.get(label, make_fill("FFFFFF"))
        for j in range(1, 11):
            if j not in (9, 10):
                ws.cell(row=i, column=j).fill = label_fill

        ws.cell(row=i, column=2).fill = TYPE_FILLS.get(typ, make_fill("FFFFFF"))
        ws.cell(row=i, column=9).fill  = make_fill("FFFDE7")
        ws.cell(row=i, column=10).fill = make_fill("FFF8E1")

        nb_lignes = max(1, len(str(texte)) // 60)
        ws.row_dimensions[i].height = min(15 * nb_lignes, 80)
        dv.sqref = f"I2:I{len(tweets)+1}"

    # Feuille Instructions
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        (f"VALIDATION POUR {annotateur.upper()}", "TITRE"),
        ("", ""),
        ("COMMENT REMPLIR CE FICHIER", "SECTION"),
        (f"1. Tu as {len(tweets)} tweets à valider", "TEXTE"),
        ("2. Colonne MON_LABEL : choisir parmi negatif/neutre/positif/ambigu/SUPPRIMER", "TEXTE"),
        ("3. Si d'accord avec Gemini → recopier le label", "TEXTE"),
        ("4. Si pas d'accord → mettre le bon label", "TEXTE"),
        ("", ""),
        ("APRÈS VALIDATION", "SECTION"),
        (f"Sauvegarde puis utilise le script d'import", "TEXTE"),
    ]

    ws2.column_dimensions["A"].width = 70
    for row_i, (text, style) in enumerate(instructions, start=1):
        cell = ws2.cell(row=row_i, column=1, value=text)
        if style == "TITRE":
            cell.font = Font(bold=True, size=14, color="FFFFFF", name="Arial")
            cell.fill = make_fill("1565C0")
        elif style == "SECTION":
            cell.font = Font(bold=True, size=11, color="FFFFFF", name="Arial")
            cell.fill = make_fill("1976D2")
        else:
            cell.font = Font(size=10, name="Arial")
        cell.alignment = Alignment(vertical="center")
        ws2.row_dimensions[row_i].height = 20

    wb.save(filepath)
    print(f"  ✅ Fichier créé : {filepath} ({len(tweets)} tweets)")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# MAIN
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def main():
    print(f"\n  {'='*55}")
    print(f"  EXPORT VALIDATION SPLIT")
    print(f"  {'='*55}")

    tweets = charger_tweets()

    if not tweets:
        print(f"\n  ✅ Aucun tweet à valider !")
        return

    # Diviser les tweets en deux moitiés
    moitie = len(tweets) // 2
    tweets_1 = tweets[:moitie]
    tweets_2 = tweets[moitie:]

    print(f"\n  📊 RÉPARTITION :")
    print(f"     {ANNOTATEUR_1} → {len(tweets_1)} tweets")
    print(f"     {ANNOTATEUR_2} → {len(tweets_2)} tweets")

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    
    # Créer les deux fichiers
    filepath_1 = os.path.join(OUTPUT_DIR, f"validation_{ANNOTATEUR_1}_{ts}.xlsx")
    filepath_2 = os.path.join(OUTPUT_DIR, f"validation_{ANNOTATEUR_2}_{ts}.xlsx")
    
    creer_excel(tweets_1, ANNOTATEUR_1, filepath_1)
    creer_excel(tweets_2, ANNOTATEUR_2, filepath_2)

    print(f"""
  {'='*55}
  ✅ DEUX FICHIERS GÉNÉRÉS :
  
  1. validation_{ANNOTATEUR_1}_{ts}.xlsx ({len(tweets_1)} tweets)
  2. validation_{ANNOTATEUR_2}_{ts}.xlsx ({len(tweets_2)} tweets)
  
  {'─'*55}
  📝 WORKFLOW :
  
  1. Chacun prend son fichier Excel
  2. Remplir la colonne MON_LABEL pour chaque tweet
  3. Sauvegarder
  4. Lancer import_validation_xlsx.py (chacun son tour)
  
  {'='*55}
    """)

if __name__ == "__main__":
    main()