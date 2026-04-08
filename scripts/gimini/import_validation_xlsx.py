"""
╔══════════════════════════════════════════════════════════════════╗
║  IMPORT VALIDATION XLSX → MONGODB  (VERSION FINALE)             ║
║  Cherche par _id = ObjectId(mongo_id_excel)                      ║
║  Usage : python3 import_validation_xlsx.py                       ║
╚══════════════════════════════════════════════════════════════════╝
"""

from pymongo  import MongoClient, UpdateOne
from datetime import datetime
from openpyxl import load_workbook
from bson     import ObjectId
import glob
import sys
import os

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CONFIG
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
MONGO_URI  = "mongodb://yousrahadjabderrahmane_db_user:C8wjIvWqOBUjK66u@ac-1ksfahb-shard-00-00.gejzu4a.mongodb.net:27017,ac-1ksfahb-shard-00-01.gejzu4a.mongodb.net:27017,ac-1ksfahb-shard-00-02.gejzu4a.mongodb.net:27017/?ssl=true&replicaSet=atlas-mdnqx7-shard-0&authSource=admin&appName=Cluster0"
DB_NAME    = "telecom_algerie_new"
COLLECTION = "dataset_unifie"

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CONNEXION MONGODB
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
print("\n" + "="*60)
print("  IMPORT VALIDATION XLSX → MONGODB (VERSION FINALE)")
print("="*60)

try:
    client     = MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)
    client.server_info()
    collection = client[DB_NAME][COLLECTION]
    print(f"\n  Connecté à MongoDB Atlas")
    print(f"  Total documents : {collection.count_documents({})}")
except Exception as e:
    print(f"\n  Connexion échouée : {e}")
    sys.exit(1)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TROUVER LES FICHIERS EXCEL
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
fichiers = glob.glob("validation_*.xlsx")

if not fichiers:
    print("\n  Aucun fichier validation_*.xlsx trouvé")
    sys.exit(1)

print(f"\n  Fichiers trouvés :")
for i, f in enumerate(fichiers):
    print(f"   [{i}] {f}")

print("\n  Entrez les numéros à importer (ex: 0 1) ou 'all' pour tous :")
choix = input("  > ").strip()

if choix.lower() == "all":
    fichiers_choisis = fichiers
else:
    try:
        indices          = [int(x) for x in choix.split()]
        fichiers_choisis = [fichiers[i] for i in indices]
    except:
        print("  Choix invalide")
        sys.exit(1)

print(f"\n  Fichiers à importer : {fichiers_choisis}")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TRAITEMENT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
labels_valides  = {"negatif", "neutre", "positif", "ambigu"}
ts              = datetime.now()
total_maj       = 0
total_supprimes = 0
total_ignores   = 0
total_erreurs   = 0

for fichier in fichiers_choisis:
    print(f"\n  {'─'*50}")
    print(f"  Traitement : {fichier}")
    print(f"  {'─'*50}")

    parties    = os.path.basename(fichier).replace(".xlsx","").split("_")
    annotateur = parties[1] if len(parties) > 1 else "inconnu"
    print(f"  Annotateur : {annotateur}")

    try:
        wb = load_workbook(fichier, read_only=True)
        ws = wb["Validation"]
    except Exception as e:
        print(f"  Erreur ouverture fichier : {e}")
        continue

    # Lire l'en-tête
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip().upper()] = cell.column - 1

    manquantes = [c for c in ["MONGO_ID","LABEL_GEMINI","MON_LABEL"] if c not in headers]
    if manquantes:
        print(f"  Colonnes manquantes : {manquantes} — fichier ignoré")
        continue

    idx_id      = headers["MONGO_ID"]
    idx_gemini  = headers["LABEL_GEMINI"]
    idx_human   = headers["MON_LABEL"]
    idx_comment = headers.get("COMMENTAIRE", None)

    operations  = []
    a_supprimer = []
    ignores     = 0
    nb_lignes   = 0
    oid_erreurs = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        mongo_id_val = row[idx_id]     if idx_id     < len(row) else None
        gemini_lbl   = row[idx_gemini] if idx_gemini  < len(row) else None
        human_lbl    = row[idx_human]  if idx_human   < len(row) else None
        commentaire  = row[idx_comment] if idx_comment and idx_comment < len(row) else ""

        # Ignorer lignes vides
        if not mongo_id_val or not human_lbl:
            ignores += 1
            continue

        human_lbl    = str(human_lbl).strip().lower()
        gemini_lbl   = str(gemini_lbl).strip().lower() if gemini_lbl else ""
        mongo_id_str = str(mongo_id_val).strip()
        nb_lignes   += 1

        # Convertir en ObjectId — diagnostic a confirmé que c'est le bon type
        try:
            oid = ObjectId(mongo_id_str)
        except Exception:
            oid_erreurs += 1
            ignores     += 1
            continue

        # Cas SUPPRIMER
        if human_lbl == "supprimer":
            a_supprimer.append(oid)
            continue

        # Label invalide
        if human_lbl not in labels_valides:
            ignores += 1
            continue

        # Filtre par _id = ObjectId  ← confirmé par le diagnostic TEST 2
        filtre = {"_id": oid}

        update_doc = {
            "$set": {
                f"validation.{annotateur}": {
                    "label"       : human_lbl,
                    "date"        : ts,
                    "gemini_label": gemini_lbl,
                    "corrige"     : human_lbl != gemini_lbl,
                    "commentaire" : str(commentaire) if commentaire else ""
                }
            }
        }

        # Si label corrigé → mettre à jour le label principal
        if human_lbl != gemini_lbl and human_lbl != "ambigu":
            update_doc["$set"]["label"]          = human_lbl
            update_doc["$set"]["label_source"]   = "validation_manuelle"
            update_doc["$set"]["label_original"] = gemini_lbl

        operations.append(UpdateOne(filtre, update_doc))

    print(f"  Lignes lues          : {nb_lignes}")
    print(f"  Mises à jour prévues : {len(operations)}")
    print(f"  À supprimer          : {len(a_supprimer)}")
    print(f"  Ignorées             : {ignores}")
    if oid_erreurs:
        print(f"  IDs invalides        : {oid_erreurs}")

    # Bulk write
    if operations:
        try:
            result    = collection.bulk_write(operations, ordered=False)
            maj       = result.modified_count
            total_maj += maj
            print(f"  Documents mis à jour : {maj}")
        except Exception as e:
            print(f"  Erreur bulk_write : {e}")
            total_erreurs += 1

    # Supprimer
    if a_supprimer:
        try:
            result_del      = collection.delete_many({"_id": {"$in": a_supprimer}})
            sup             = result_del.deleted_count
            total_supprimes += sup
            print(f"  Documents supprimés  : {sup}")
        except Exception as e:
            print(f"  Erreur suppression : {e}")

    total_ignores += ignores
    wb.close()

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# RÉSUMÉ FINAL
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
print(f"\n{'='*60}")
print(f"  RESUME FINAL")
print(f"{'='*60}")
print(f"  Documents mis à jour : {total_maj}")
print(f"  Documents supprimés  : {total_supprimes}")
print(f"  Lignes ignorées      : {total_ignores}")
if total_erreurs:
    print(f"  Erreurs              : {total_erreurs}")
print(f"{'='*60}")
print(f"  DONE ✅ — MongoDB mis à jour")
print(f"{'='*60}\n")

client.close()