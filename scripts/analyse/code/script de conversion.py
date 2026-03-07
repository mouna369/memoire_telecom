#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# convert_csv_to_xlsx.py
# Convertit le CSV exporté depuis MongoDB en XLSX propre et formaté

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re, json, os

# ============================================================
# CONFIGURATION — change ces chemins selon ton besoin
# ============================================================
CSV_PATH  = "/home/mouna/projet_telecom/donnees/brutes/telecom_algerie.commentaires_sans_chiffres_certains.csv"
XLSX_PATH = "/home/mouna/projet_telecom/donnees/transformees/commentaires_propres_avec_certain_chiffre.xlsx"

# ============================================================
print("="*60)
print("📊 CONVERSION CSV → XLSX")
print("="*60)

# 1. LIRE LE CSV
print(f"\n📂 Lecture : {CSV_PATH}")
df = pd.read_csv(CSV_PATH, encoding="utf-8")
print(f"✅ {len(df)} lignes chargées")

# 2. NETTOYER METADATA
def extraire_metadata(meta_str):
    try:
        if pd.isna(meta_str):
            return None, None, None
        s = re.sub(
            r'datetime\.datetime\((\d+),\s*(\d+),\s*(\d+),\s*(\d+),\s*(\d+),\s*(\d+).*?\)',
            r'"\1-\2-\3 \4:\5:\6"', str(meta_str)
        ).replace("'", '"')
        meta = json.loads(s)
        return meta.get("fichier",""), meta.get("ligne",""), meta.get("date_import","")
    except:
        return None, None, None

if "metadata" in df.columns:
    print("🔄 Nettoyage metadata...")
    df[["meta_fichier","meta_ligne","meta_date_import"]] = df["metadata"].apply(
        lambda x: pd.Series(extraire_metadata(x))
    )
    df = df.drop(columns=["metadata"])
    print("✅ Metadata → 3 colonnes")

# 3. ORDRE DES COLONNES
cols = ["_id","Commentaire_Client","commentaire_moderateur","date",
        "source","moderateur","statut","meta_fichier","meta_ligne","meta_date_import"]
df = df[[c for c in cols if c in df.columns]]

# 4. CRÉER XLSX
print("\n📝 Création XLSX...")
wb = Workbook()
ws = wb.active
ws.title = "Commentaires"

BLEU   = "1F4E79"
CLAIR  = "DEEAF1"
BLANC  = "FFFFFF"

f_head = Font(name="Arial", bold=True, color="FFFFFF", size=11)
f_data = Font(name="Arial", size=10)
a_c    = Alignment(horizontal="center", vertical="center", wrap_text=True)
a_g    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
brd    = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"),  bottom=Side(style="thin"))

NOMS = {
    "_id":"ID", "Commentaire_Client":"Commentaire Client",
    "commentaire_moderateur":"Réponse Modérateur", "date":"Date",
    "source":"Source", "moderateur":"Modérateur", "statut":"Statut",
    "meta_fichier":"Fichier Source", "meta_ligne":"Ligne",
    "meta_date_import":"Date Import"
}

LARG = {
    "_id":25, "Commentaire_Client":60, "commentaire_moderateur":50,
    "date":18, "source":12, "moderateur":25, "statut":12,
    "meta_fichier":30, "meta_ligne":10, "meta_date_import":20
}

# En-têtes
for ci, col in enumerate(df.columns, 1):
    c = ws.cell(row=1, column=ci)
    c.value = NOMS.get(col, col)
    c.font  = f_head
    c.fill  = PatternFill("solid", start_color=BLEU)
    c.alignment = a_c
    c.border = brd
ws.row_dimensions[1].height = 30

# Données
print("   📥 Écriture des données...")
for ri, row in enumerate(df.itertuples(index=False), 2):
    bg = CLAIR if ri % 2 == 0 else BLANC
    for ci, val in enumerate(row, 1):
        c = ws.cell(row=ri, column=ci)
        c.value     = "" if pd.isna(val) else str(val)
        c.font      = f_data
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = a_g
        c.border    = brd
    if ri % 5000 == 0:
        print(f"   ... {ri} lignes écrites")

# Largeurs + figer
for ci, col in enumerate(df.columns, 1):
    ws.column_dimensions[get_column_letter(ci)].width = LARG.get(col, 20)
ws.freeze_panes = "A2"

# Feuille Statistiques
ws2 = wb.create_sheet("Statistiques")
stats = [
    ["📊 STATISTIQUES", ""], ["",""],
    ["Total documents",     len(df)],
    ["Sources uniques",     df["source"].nunique() if "source" in df.columns else "N/A"],
    ["Modérateurs uniques", df["moderateur"].nunique() if "moderateur" in df.columns else "N/A"],
    ["Commentaires vides",  df["Commentaire_Client"].isna().sum() if "Commentaire_Client" in df.columns else "N/A"],
    ["",""], ["📱 PAR SOURCE",""],
]
if "source" in df.columns:
    for s, n in df["source"].value_counts().items():
        stats.append([f"  {s}", n])
stats += [["",""], ["📋 PAR STATUT",""]]
if "statut" in df.columns:
    for s, n in df["statut"].value_counts().items():
        stats.append([f"  {s}", n])
if "date" in df.columns:
    stats += [["",""], ["📅 PÉRIODE",""],
              ["Date la plus ancienne", df["date"].min()],
              ["Date la plus récente",  df["date"].max()]]

for ri, (label, val) in enumerate(stats, 1):
    ws2.cell(row=ri, column=1).value = label
    ws2.cell(row=ri, column=2).value = val
    if label and not label.startswith("  ") and val == "":
        ws2.cell(row=ri, column=1).font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    else:
        ws2.cell(row=ri, column=1).font = Font(name="Arial", bold=True, size=10)
        ws2.cell(row=ri, column=2).font = Font(name="Arial", size=10)

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 25

# Sauvegarder
wb.save(XLSX_PATH)

print(f"\n{'='*60}")
print(f"✅ XLSX CRÉÉ : {XLSX_PATH}")
print(f"{'='*60}")
print(f"   📊 Lignes        : {len(df)}")
print(f"   📋 Colonnes      : {len(df.columns)}")
print(f"   📌 En-tête figée : Oui")
print(f"   📈 Feuille Stats : Oui")
print(f"   📁 Chemin        : {os.path.abspath(XLSX_PATH)}")
print(f"{'='*60}")