# #!/usr/bin/env python3
# # -*- coding: utf-8 -*-

# # scripts/analyse/code/analyse_doublons_detail.py
# # Analyse détaillée des 3 types de doublons
# # ✅ Pure Python + MongoDB

# from pymongo import MongoClient
# from openpyxl import Workbook
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.utils import get_column_letter
# import os

# # ============================================================
# # CONFIGURATION
# # ============================================================
# MONGO_URI    = "mongodb://localhost:27018/"
# DB_NAME      = "telecom_algerie"
# COLLECTION   = "commentaires_sans_urls_arobase"
# RAPPORT_PATH = "/home/mouna/projet_telecom/scripts/analyse/code/Rapports/rapport_doublons_detail.xlsx"

# # ============================================================
# # CONNEXION MONGODB
# # ============================================================
# print("="*65)
# print("🔍 ANALYSE DÉTAILLÉE DES DOUBLONS")
# print("="*65)

# client     = MongoClient(MONGO_URI)
# db         = client[DB_NAME]
# collection = db[COLLECTION]
# total      = collection.count_documents({})
# print(f"✅ {total} documents trouvés")

# # ============================================================
# # AGRÉGATION MONGODB
# # ============================================================
# print("\n🔎 Analyse en cours...")

# # Grouper par commentaire et collecter sources, dates, modérateurs
# pipeline = [
#     {
#         "$group": {
#             "_id"          : "$Commentaire_Client",
#             "nb_total"     : {"$sum": 1},
#             "sources"      : {"$addToSet": "$source"},
#             "dates"        : {"$addToSet": "$date"},
#             "moderateurs"  : {"$addToSet": "$moderateur"},
#         }
#     },
#     {"$match": {"nb_total": {"$gt": 1}}},  # Garder seulement les doublons
#     {"$sort" : {"nb_total": -1}}
# ]

# doublons = list(collection.aggregate(pipeline))
# print(f"✅ {len(doublons)} groupes de doublons trouvés")

# # ============================================================
# # CLASSIFIER LES 3 TYPES
# # ============================================================
# type_sources    = []  # Même commentaire, sources différentes
# type_dates      = []  # Même commentaire, dates différentes
# type_moderateurs = [] # Même commentaire, modérateurs différents
# type_tous       = []  # Présent dans plusieurs types

# for d in doublons:
#     texte        = d["_id"] or ""
#     nb_total     = d["nb_total"]
#     nb_sources   = len(d["sources"])
#     nb_dates     = len(d["dates"])
#     nb_mods      = len(d["moderateurs"])

#     texte_court  = texte[:80] + "..." if len(texte) > 80 else texte
#     sources_str  = " / ".join([s for s in d["sources"] if s])
#     dates_str    = " / ".join(sorted([str(dt) for dt in d["dates"] if dt])[:3])
#     mods_str     = " / ".join([m for m in d["moderateurs"] if m])

#     row = [texte_court, nb_total, nb_sources, nb_dates, nb_mods,
#            sources_str, dates_str, mods_str]

#     if nb_sources > 1:
#         type_sources.append(row)
#     if nb_dates > 1:
#         type_dates.append(row)
#     if nb_mods > 1:
#         type_moderateurs.append(row)

# print(f"   • Doublons sources différentes    : {len(type_sources)}")
# print(f"   • Doublons dates différentes      : {len(type_dates)}")
# print(f"   • Doublons modérateurs différents : {len(type_moderateurs)}")

# # ============================================================
# # STYLES EXCEL
# # ============================================================
# ROUGE_F  = "8B0000"
# ROUGE_C  = "F4CCCC"
# ORANGE_F = "7F4000"
# ORANGE_C = "FCE5CD"
# BLEU_F   = "1F4E79"
# BLEU_C   = "DEEAF1"
# VIOLET_F = "4A148C"
# VIOLET_C = "E1BEE7"
# BLANC    = "FFFFFF"

# def style_entete(cell, couleur):
#     cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
#     cell.fill      = PatternFill("solid", start_color=couleur)
#     cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#     cell.border    = Border(left=Side(style="thin"), right=Side(style="thin"),
#                             top=Side(style="thin"),  bottom=Side(style="thin"))

# def style_data(cell, bg):
#     cell.font      = Font(name="Arial", size=10)
#     cell.fill      = PatternFill("solid", start_color=bg)
#     cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
#     cell.border    = Border(left=Side(style="thin"), right=Side(style="thin"),
#                             top=Side(style="thin"),  bottom=Side(style="thin"))

# def creer_feuille(wb, titre, couleur, donnees):
#     ws = wb.create_sheet(titre)

#     entetes  = ["#", "COMMENTAIRE", "NB COPIES", "NB SOURCES",
#                 "NB DATES", "NB MODÉRATEURS", "SOURCES", "DATES", "MODÉRATEURS"]
#     largeurs = [5, 55, 10, 10, 10, 15, 25, 35, 30]

#     for ci, (e, l) in enumerate(zip(entetes, largeurs), 1):
#         cell = ws.cell(row=1, column=ci, value=e)
#         style_entete(cell, couleur)
#         ws.column_dimensions[get_column_letter(ci)].width = l
#     ws.row_dimensions[1].height = 35
#     ws.freeze_panes = "A2"

#     for ri, row in enumerate(donnees, 2):
#         bg = ROUGE_C if ri % 2 == 0 else BLANC
#         ws.cell(row=ri, column=1, value=ri-1)
#         style_data(ws.cell(row=ri, column=1), bg)

#         for ci, val in enumerate(row, 2):
#             cell = ws.cell(row=ri, column=ci, value=str(val) if val else "")
#             style_data(cell, bg)
#             # NB COPIES en gras rouge
#             if ci == 3:
#                 cell.font      = Font(name="Arial", bold=True, size=11, color="8B0000")
#                 cell.alignment = Alignment(horizontal="center", vertical="center")

#     return ws

# # ============================================================
# # CRÉER EXCEL
# # ============================================================
# print("\n📊 Création du fichier Excel...")
# wb = Workbook()
# wb.remove(wb.active)

# # Feuille 1 — Résumé
# ws_resume = wb.create_sheet("📊 Résumé")
# resume = [
#     ["📊 ANALYSE DÉTAILLÉE DES DOUBLONS", ""],
#     ["", ""],
#     ["Collection", f"{DB_NAME}.{COLLECTION}"],
#     ["Total documents", total],
#     ["", ""],
#     ["TYPE 1 — Même commentaire, sources différentes",     len(type_sources)],
#     ["TYPE 2 — Même commentaire, dates différentes",       len(type_dates)],
#     ["TYPE 3 — Même commentaire, modérateurs différents",  len(type_moderateurs)],
#     ["", ""],
#     ["Total groupes de doublons", len(doublons)],
#     ["Total documents en double", sum(d["nb_total"] - 1 for d in doublons)],
#     ["Taux de doublons",          f"{sum(d['nb_total']-1 for d in doublons)/total*100:.2f}%"],
# ]

# COULEURS_RESUME = {
#     "📊 ANALYSE DÉTAILLÉE DES DOUBLONS": ("1F4E79", 12),
#     "TYPE 1 — Même commentaire, sources différentes":    ("7F4000", 11),
#     "TYPE 2 — Même commentaire, dates différentes":      ("8B0000", 11),
#     "TYPE 3 — Même commentaire, modérateurs différents": ("4A148C", 11),
# }

# for ri, (label, val) in enumerate(resume, 1):
#     c1 = ws_resume.cell(row=ri, column=1, value=label)
#     c2 = ws_resume.cell(row=ri, column=2, value=val)
#     if label in COULEURS_RESUME:
#         couleur, size = COULEURS_RESUME[label]
#         c1.font = Font(name="Arial", bold=True, size=size, color=couleur)
#         c2.font = Font(name="Arial", bold=True, size=size, color=couleur)
#     else:
#         c1.font = Font(name="Arial", bold=True, size=10)
#         c2.font = Font(name="Arial", size=10)

# ws_resume.column_dimensions["A"].width = 50
# ws_resume.column_dimensions["B"].width = 20

# # Feuilles 3 types
# creer_feuille(wb, "🔵 Sources Différentes",    BLEU_F,   type_sources)
# creer_feuille(wb, "🔴 Dates Différentes",      ROUGE_F,  type_dates)
# creer_feuille(wb, "🟣 Modérateurs Différents", VIOLET_F, type_moderateurs)

# # ============================================================
# # SAUVEGARDER
# # ============================================================
# os.makedirs(os.path.dirname(RAPPORT_PATH), exist_ok=True)
# wb.save(RAPPORT_PATH)
# client.close()

# print(f"\n✅ Rapport Excel sauvegardé : {RAPPORT_PATH}")
# print("\n" + "="*65)
# print("📊 RÉSUMÉ FINAL")
# print("="*65)
# print(f"   📥 Total documents                        : {total}")
# print(f"   🔵 Doublons sources différentes           : {len(type_sources)}")
# print(f"   🔴 Doublons dates différentes             : {len(type_dates)}")
# print(f"   🟣 Doublons modérateurs différents        : {len(type_moderateurs)}")
# print(f"   ❌ Total groupes de doublons              : {len(doublons)}")
# print(f"   ❌ Total documents en double              : {sum(d['nb_total']-1 for d in doublons)}")
# print(f"   📊 Taux de doublons                       : {sum(d['nb_total']-1 for d in doublons)/total*100:.2f}%")
# print("="*65)




#genration des fichier 
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# scripts/analyse/code/analyse_doublons_detail.py
# Analyse détaillée des 3 types de doublons avec exemples concrets
# ✅ Pure Python + MongoDB

from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime
import os

# ============================================================
# CONFIGURATION
# ============================================================
MONGO_URI    = "mongodb://localhost:27018/"
DB_NAME      = "telecom_algerie"
COLLECTION   = "commentaires_sans_urls_arobase"
RAPPORT_PATH = "/home/mouna/projet_telecom/scripts/analyse/Rapports/rapport_doublons_detail.xlsx"
RAPPORT_TXT  = "/home/mouna/projet_telecom/scripts/analyse/Rapports/rapport_doublons_detail.txt"

# ============================================================
# CONNEXION MONGODB
# ============================================================
print("="*65)
print("🔍 ANALYSE DÉTAILLÉE DES DOUBLONS AVEC EXEMPLES")
print("="*65)

client     = MongoClient(MONGO_URI)
db         = client[DB_NAME]
collection = db[COLLECTION]
total      = collection.count_documents({})
print(f"✅ {total} documents trouvés")

# ============================================================
# RÉCUPÉRER TOUS LES DOCUMENTS
# ============================================================
print("\n🔎 Chargement des données...")

groupes = defaultdict(list)

for doc in collection.find({}, {
    "Commentaire_Client": 1,
    "date"              : 1,
    "source"            : 1,
    "moderateur"        : 1,
}):
    texte = doc.get("Commentaire_Client", "")
    if not texte:
        continue
    groupes[texte.strip()].append({
        "source"    : str(doc.get("source",     "") or ""),
        "date"      : str(doc.get("date",       "") or ""),
        "moderateur": str(doc.get("moderateur", "") or ""),
    })

print(f"✅ {len(groupes)} commentaires uniques trouvés")

# ============================================================
# CLASSIFIER LES 3 TYPES
# ============================================================
type_sources     = []
type_dates       = []
type_moderateurs = []

for texte, copies in groupes.items():
    if len(copies) < 2:
        continue

    sources     = set(c["source"]      for c in copies)
    dates       = set(c["date"]        for c in copies)
    moderateurs = set(c["moderateur"]  for c in copies)

    texte_court = texte[:80] + "..." if len(texte) > 80 else texte
    exemples    = copies[:4]

    row = {
        "texte"    : texte_court,
        "nb_copies": len(copies),
        "nb_sources": len(sources),
        "nb_dates" : len(dates),
        "nb_mods"  : len(moderateurs),
        "exemples" : exemples,
        "sources"  : sources,
        "dates"    : dates,
        "moderateurs": moderateurs,
    }

    if len(sources) > 1:
        type_sources.append(row)
    if len(dates) > 1:
        type_dates.append(row)
    if len(moderateurs) > 1:
        type_moderateurs.append(row)

type_sources.sort(    key=lambda x: x["nb_copies"], reverse=True)
type_dates.sort(      key=lambda x: x["nb_copies"], reverse=True)
type_moderateurs.sort(key=lambda x: x["nb_copies"], reverse=True)

total_doublons  = sum(len(g)-1 for g in groupes.values() if len(g) > 1)
total_groupes   = len([g for g in groupes.values() if len(g) > 1])
taux_doublons   = total_doublons / total * 100

print(f"   🔵 Doublons sources différentes    : {len(type_sources)}")
print(f"   🔴 Doublons dates différentes      : {len(type_dates)}")
print(f"   🟣 Doublons modérateurs différents : {len(type_moderateurs)}")

# ============================================================
# GÉNÉRER LE RAPPORT TXT
# ============================================================
print("\n📝 Génération du rapport TXT...")

os.makedirs(os.path.dirname(RAPPORT_TXT), exist_ok=True)

with open(RAPPORT_TXT, "w", encoding="utf-8") as f:

    f.write("=" * 70 + "\n")
    f.write("RAPPORT — ANALYSE DÉTAILLÉE DES DOUBLONS\n")
    f.write("=" * 70 + "\n")
    f.write(f"Date       : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    f.write(f"Collection : {DB_NAME}.{COLLECTION}\n")
    f.write("\n")

    # ── RÉSUMÉ GLOBAL ──
    f.write("=" * 70 + "\n")
    f.write("📊 RÉSUMÉ GLOBAL\n")
    f.write("=" * 70 + "\n")
    f.write(f"   Total documents                        : {total}\n")
    f.write(f"   Commentaires uniques                   : {len(groupes)}\n")
    f.write(f"   Total groupes de doublons              : {total_groupes}\n")
    f.write(f"   Total documents en double              : {total_doublons}\n")
    f.write(f"   Taux de doublons                       : {taux_doublons:.2f}%\n")
    f.write("\n")
    f.write(f"   🔵 TYPE 1 — Sources différentes        : {len(type_sources)} groupes\n")
    f.write(f"   🔴 TYPE 2 — Dates différentes          : {len(type_dates)} groupes\n")
    f.write(f"   🟣 TYPE 3 — Modérateurs différents     : {len(type_moderateurs)} groupes\n")
    f.write("\n")

    # ── TYPE 1 — SOURCES DIFFÉRENTES ──
    f.write("=" * 70 + "\n")
    f.write("🔵 TYPE 1 — MÊME COMMENTAIRE, SOURCES DIFFÉRENTES\n")
    f.write("   → Même client qui a posté sur plusieurs plateformes\n")
    f.write("   → Garder 1 copie par source\n")
    f.write("=" * 70 + "\n\n")

    for i, row in enumerate(type_sources[:20], 1):
        f.write(f"  [{i}] NB COPIES : {row['nb_copies']} | "
                f"SOURCES : {', '.join(row['sources'])}\n")
        f.write(f"       COMMENTAIRE : {row['texte']}\n")
        f.write(f"       DÉTAIL :\n")
        for ex in row["exemples"]:
            f.write(f"         • Source={ex['source']:<15} "
                    f"Date={ex['date']:<25} "
                    f"Modérateur={ex['moderateur']}\n")
        f.write("\n")

    if len(type_sources) > 20:
        f.write(f"  ... et {len(type_sources)-20} autres groupes\n\n")

    # ── TYPE 2 — DATES DIFFÉRENTES ──
    f.write("=" * 70 + "\n")
    f.write("🔴 TYPE 2 — MÊME COMMENTAIRE, DATES DIFFÉRENTES\n")
    f.write("   → Client qui répète sa plainte à des dates différentes\n")
    f.write("   → Garder toutes les copies (utile pour analyse temporelle)\n")
    f.write("=" * 70 + "\n\n")

    for i, row in enumerate(type_dates[:20], 1):
        f.write(f"  [{i}] NB COPIES : {row['nb_copies']} | "
                f"NB DATES : {row['nb_dates']}\n")
        f.write(f"       COMMENTAIRE : {row['texte']}\n")
        f.write(f"       DÉTAIL :\n")
        for ex in row["exemples"]:
            f.write(f"         • Source={ex['source']:<15} "
                    f"Date={ex['date']:<25} "
                    f"Modérateur={ex['moderateur']}\n")
        f.write("\n")

    if len(type_dates) > 20:
        f.write(f"  ... et {len(type_dates)-20} autres groupes\n\n")

    # ── TYPE 3 — MODÉRATEURS DIFFÉRENTS ──
    f.write("=" * 70 + "\n")
    f.write("🟣 TYPE 3 — MÊME COMMENTAIRE, MODÉRATEURS DIFFÉRENTS\n")
    f.write("   → Commentaire traité par plusieurs modérateurs\n")
    f.write("   → Garder toutes les copies (utile pour analyse performance)\n")
    f.write("=" * 70 + "\n\n")

    for i, row in enumerate(type_moderateurs[:20], 1):
        f.write(f"  [{i}] NB COPIES : {row['nb_copies']} | "
                f"MODÉRATEURS : {', '.join(row['moderateurs'])}\n")
        f.write(f"       COMMENTAIRE : {row['texte']}\n")
        f.write(f"       DÉTAIL :\n")
        for ex in row["exemples"]:
            f.write(f"         • Source={ex['source']:<15} "
                    f"Date={ex['date']:<25} "
                    f"Modérateur={ex['moderateur']}\n")
        f.write("\n")

    if len(type_moderateurs) > 20:
        f.write(f"  ... et {len(type_moderateurs)-20} autres groupes\n\n")

    # ── DÉCISIONS ──
    f.write("=" * 70 + "\n")
    f.write("✅ DÉCISIONS DE TRAITEMENT\n")
    f.write("=" * 70 + "\n")
    f.write("   🔵 Sources différentes    → Garder 1 copie par source\n")
    f.write("   🔴 Dates différentes      → Garder toutes (analyse temporelle)\n")
    f.write("   🟣 Modérateurs différents → Garder toutes (analyse performance)\n")
    f.write("   ❌ Même texte+jour+source → Supprimer toutes sauf la plus ancienne\n")
    f.write("\n")
    f.write("=" * 70 + "\n")
    f.write("FIN DU RAPPORT\n")
    f.write("=" * 70 + "\n")

print(f"✅ Rapport TXT sauvegardé : {RAPPORT_TXT}")

# ============================================================
# STYLES EXCEL
# ============================================================
ROUGE_F  = "8B0000"
ROUGE_C  = "F4CCCC"
BLEU_F   = "1F4E79"
BLEU_C   = "DEEAF1"
VIOLET_F = "4A148C"
VIOLET_C = "E1BEE7"
BLANC    = "FFFFFF"

def style_entete(cell, couleur):
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill      = PatternFill("solid", start_color=couleur)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = Border(left=Side(style="thin"), right=Side(style="thin"),
                            top=Side(style="thin"),  bottom=Side(style="thin"))

def style_data(cell, bg, bold=False, couleur_texte="000000", centre=False):
    cell.font      = Font(name="Arial", bold=bold, size=10, color=couleur_texte)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(
        horizontal="center" if centre else "left",
        vertical="center", wrap_text=True
    )
    cell.border    = Border(left=Side(style="thin"), right=Side(style="thin"),
                            top=Side(style="thin"),  bottom=Side(style="thin"))

def style_titre_groupe(cell, couleur_bg, couleur_texte):
    cell.font      = Font(name="Arial", bold=True, size=10, color=couleur_texte)
    cell.fill      = PatternFill("solid", start_color=couleur_bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border    = Border(left=Side(style="medium"), right=Side(style="medium"),
                            top=Side(style="medium"),  bottom=Side(style="thin"))

def creer_feuille_avec_exemples(wb, titre, couleur_header, couleur_groupe,
                                 couleur_texte_groupe, donnees, couleur_exemple):
    ws = wb.create_sheet(titre)
    entetes  = ["#", "COMMENTAIRE", "NB COPIES", "COPIE #",
                "SOURCE", "DATE", "MODÉRATEUR"]
    largeurs = [5, 55, 10, 8, 20, 25, 30]

    for ci, (e, l) in enumerate(zip(entetes, largeurs), 1):
        cell = ws.cell(row=1, column=ci, value=e)
        style_entete(cell, couleur_header)
        ws.column_dimensions[get_column_letter(ci)].width = l
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"

    ri = 2
    for idx, row in enumerate(donnees, 1):
        texte     = row["texte"]
        nb_copies = row["nb_copies"]
        exemples  = row["exemples"]
        nb_rows   = len(exemples)

        cell = ws.cell(row=ri, column=1, value=idx)
        style_titre_groupe(cell, couleur_groupe, couleur_texte_groupe)

        cell = ws.cell(row=ri, column=2, value=texte)
        style_titre_groupe(cell, couleur_groupe, couleur_texte_groupe)
        if nb_rows > 1:
            ws.merge_cells(start_row=ri, start_column=2,
                           end_row=ri+nb_rows-1, end_column=2)

        cell           = ws.cell(row=ri, column=3, value=nb_copies)
        cell.font      = Font(name="Arial", bold=True, size=12, color="8B0000")
        cell.fill      = PatternFill("solid", start_color=couleur_groupe)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = Border(left=Side(style="thin"), right=Side(style="thin"),
                                top=Side(style="medium"), bottom=Side(style="thin"))
        if nb_rows > 1:
            ws.merge_cells(start_row=ri, start_column=3,
                           end_row=ri+nb_rows-1, end_column=3)

        for ex_idx, exemple in enumerate(exemples):
            bg = couleur_exemple if ex_idx % 2 == 0 else BLANC

            cell = ws.cell(row=ri, column=4, value=f"Copie {ex_idx+1}")
            style_data(cell, bg, bold=True, centre=True)

            cell = ws.cell(row=ri, column=5, value=exemple["source"])
            style_data(cell, bg)

            cell = ws.cell(row=ri, column=6, value=exemple["date"])
            style_data(cell, bg)

            cell = ws.cell(row=ri, column=7, value=exemple["moderateur"])
            style_data(cell, bg)

            ws.row_dimensions[ri].height = 25
            ri += 1

        for ci in range(1, 8):
            cell = ws.cell(row=ri, column=ci, value="")
            cell.fill = PatternFill("solid", start_color="EEEEEE")
        ri += 1

    return ws

# ============================================================
# CRÉER EXCEL
# ============================================================
print("\n📊 Création du fichier Excel...")
wb = Workbook()
wb.remove(wb.active)

ws_resume = wb.create_sheet("📊 Résumé")
resume = [
    ["📊 ANALYSE DÉTAILLÉE DES DOUBLONS", ""],
    ["", ""],
    ["Collection",                                           f"{DB_NAME}.{COLLECTION}"],
    ["Date analyse",                                         datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
    ["Total documents",                                      total],
    ["", ""],
    ["🔵 TYPE 1 — Même commentaire, sources différentes",    len(type_sources)],
    ["🔴 TYPE 2 — Même commentaire, dates différentes",      len(type_dates)],
    ["🟣 TYPE 3 — Même commentaire, modérateurs différents", len(type_moderateurs)],
    ["", ""],
    ["Total groupes de doublons",  total_groupes],
    ["Total documents en double",  total_doublons],
    ["Taux de doublons",           f"{taux_doublons:.2f}%"],
    ["", ""],
    ["✅ DÉCISIONS", ""],
    ["Sources différentes",    "Garder 1 copie par source"],
    ["Dates différentes",      "Garder toutes (analyse temporelle)"],
    ["Modérateurs différents", "Garder toutes (analyse performance)"],
    ["Même texte+jour+source", "Supprimer sauf la plus ancienne"],
]

STYLES_RESUME = {
    "📊 ANALYSE DÉTAILLÉE DES DOUBLONS"                       : ("1F4E79", 13),
    "🔵 TYPE 1 — Même commentaire, sources différentes"       : ("1F4E79", 11),
    "🔴 TYPE 2 — Même commentaire, dates différentes"         : ("8B0000", 11),
    "🟣 TYPE 3 — Même commentaire, modérateurs différents"    : ("4A148C", 11),
    "✅ DÉCISIONS"                                            : ("1E6B3C", 11),
}

for ri, (label, val) in enumerate(resume, 1):
    c1 = ws_resume.cell(row=ri, column=1, value=label)
    c2 = ws_resume.cell(row=ri, column=2, value=val)
    if label in STYLES_RESUME:
        couleur, size = STYLES_RESUME[label]
        c1.font = Font(name="Arial", bold=True, size=size, color=couleur)
        c2.font = Font(name="Arial", bold=True, size=size, color=couleur)
    else:
        c1.font = Font(name="Arial", bold=True, size=10)
        c2.font = Font(name="Arial", size=10)

ws_resume.column_dimensions["A"].width = 55
ws_resume.column_dimensions["B"].width = 35

creer_feuille_avec_exemples(wb, "🔵 Sources Différentes",
    BLEU_F, BLEU_C, "1F4E79", type_sources, "EBF5FB")

creer_feuille_avec_exemples(wb, "🔴 Dates Différentes",
    ROUGE_F, ROUGE_C, "8B0000", type_dates, "FDEDEC")

creer_feuille_avec_exemples(wb, "🟣 Modérateurs Différents",
    VIOLET_F, VIOLET_C, "4A148C", type_moderateurs, "F3E5F5")

os.makedirs(os.path.dirname(RAPPORT_PATH), exist_ok=True)
wb.save(RAPPORT_PATH)
client.close()

print(f"✅ Rapport Excel sauvegardé : {RAPPORT_PATH}")
print("\n" + "="*65)
print("📊 RÉSUMÉ FINAL")
print("="*65)
print(f"   📥 Total documents                        : {total}")
print(f"   🔵 Doublons sources différentes           : {len(type_sources)}")
print(f"   🔴 Doublons dates différentes             : {len(type_dates)}")
print(f"   🟣 Doublons modérateurs différents        : {len(type_moderateurs)}")
print(f"   ❌ Total documents en double              : {total_doublons}")
print(f"   📊 Taux de doublons                       : {taux_doublons:.2f}%")
print("="*65)
print(f"📄 Rapport TXT  : {RAPPORT_TXT}")
print(f"📊 Rapport Excel: {RAPPORT_PATH}")
print("="*65)