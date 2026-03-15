# #!/usr/bin/env python3
# # -*- coding: utf-8 -*-

# # scripts/analyse/code/analyse_doublons_detail.py
# # Analyse détaillée des 3 types de doublons avec exemples concrets
# # ✅ Pure Python + MongoDB

# from pymongo import MongoClient
# from openpyxl import Workbook
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.utils import get_column_letter
# from collections import defaultdict
# from datetime import datetime
# import os

# # ============================================================
# # CONFIGURATION
# # ============================================================
# MONGO_URI    = "mongodb://localhost:27018/"
# DB_NAME      = "telecom_algerie"
# COLLECTION   = "commentaires_sans_urls_arobase"
# RAPPORT_PATH = "/home/mouna/projet_telecom/scripts/analyse/Rapports/rapport_doublons_detail.xlsx"
# RAPPORT_TXT  = "/home/mouna/projet_telecom/scripts/analyse/Rapports/rapport_doublons_detail.txt"

# # ============================================================
# # CONNEXION MONGODB
# # ============================================================
# print("="*65)
# print("🔍 ANALYSE DÉTAILLÉE DES DOUBLONS AVEC EXEMPLES")
# print("="*65)

# client     = MongoClient(MONGO_URI)
# db         = client[DB_NAME]
# collection = db[COLLECTION]
# total      = collection.count_documents({})
# print(f"✅ {total} documents trouvés")

# # ============================================================
# # RÉCUPÉRER TOUS LES DOCUMENTS
# # ============================================================
# print("\n🔎 Chargement des données...")

# groupes = defaultdict(list)

# for doc in collection.find({}, {
#     "Commentaire_Client": 1,
#     "date"              : 1,
#     "source"            : 1,
#     "moderateur"        : 1,
# }):
#     texte = doc.get("Commentaire_Client", "")
#     if not texte:
#         continue
#     groupes[texte.strip()].append({
#         "source"    : str(doc.get("source",     "") or ""),
#         "date"      : str(doc.get("date",       "") or ""),
#         "moderateur": str(doc.get("moderateur", "") or ""),
#     })

# print(f"✅ {len(groupes)} commentaires uniques trouvés")

# # ============================================================
# # CLASSIFIER LES 3 TYPES
# # ============================================================
# type_sources     = []
# type_dates       = []
# type_moderateurs = []

# for texte, copies in groupes.items():
#     if len(copies) < 2:
#         continue

#     sources     = set(c["source"]      for c in copies)
#     dates       = set(c["date"]        for c in copies)
#     moderateurs = set(c["moderateur"]  for c in copies)

#     texte_court = texte[:80] + "..." if len(texte) > 80 else texte
#     exemples    = copies[:4]

#     row = {
#         "texte"    : texte_court,
#         "nb_copies": len(copies),
#         "nb_sources": len(sources),
#         "nb_dates" : len(dates),
#         "nb_mods"  : len(moderateurs),
#         "exemples" : exemples,
#         "sources"  : sources,
#         "dates"    : dates,
#         "moderateurs": moderateurs,
#     }

#     if len(sources) > 1:
#         type_sources.append(row)
#     if len(dates) > 1:
#         type_dates.append(row)
#     if len(moderateurs) > 1:
#         type_moderateurs.append(row)

# type_sources.sort(    key=lambda x: x["nb_copies"], reverse=True)
# type_dates.sort(      key=lambda x: x["nb_copies"], reverse=True)
# type_moderateurs.sort(key=lambda x: x["nb_copies"], reverse=True)

# total_doublons  = sum(len(g)-1 for g in groupes.values() if len(g) > 1)
# total_groupes   = len([g for g in groupes.values() if len(g) > 1])
# taux_doublons   = total_doublons / total * 100

# print(f"   🔵 Doublons sources différentes    : {len(type_sources)}")
# print(f"   🔴 Doublons dates différentes      : {len(type_dates)}")
# print(f"   🟣 Doublons modérateurs différents : {len(type_moderateurs)}")

# # ============================================================
# # GÉNÉRER LE RAPPORT TXT
# # ============================================================
# print("\n📝 Génération du rapport TXT...")

# os.makedirs(os.path.dirname(RAPPORT_TXT), exist_ok=True)

# with open(RAPPORT_TXT, "w", encoding="utf-8") as f:

#     f.write("=" * 70 + "\n")
#     f.write("RAPPORT — ANALYSE DÉTAILLÉE DES DOUBLONS\n")
#     f.write("=" * 70 + "\n")
#     f.write(f"Date       : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
#     f.write(f"Collection : {DB_NAME}.{COLLECTION}\n")
#     f.write("\n")

#     # ── RÉSUMÉ GLOBAL ──
#     f.write("=" * 70 + "\n")
#     f.write("📊 RÉSUMÉ GLOBAL\n")
#     f.write("=" * 70 + "\n")
#     f.write(f"   Total documents                        : {total}\n")
#     f.write(f"   Commentaires uniques                   : {len(groupes)}\n")
#     f.write(f"   Total groupes de doublons              : {total_groupes}\n")
#     f.write(f"   Total documents en double              : {total_doublons}\n")
#     f.write(f"   Taux de doublons                       : {taux_doublons:.2f}%\n")
#     f.write("\n")
#     f.write(f"   🔵 TYPE 1 — Sources différentes        : {len(type_sources)} groupes\n")
#     f.write(f"   🔴 TYPE 2 — Dates différentes          : {len(type_dates)} groupes\n")
#     f.write(f"   🟣 TYPE 3 — Modérateurs différents     : {len(type_moderateurs)} groupes\n")
#     f.write("\n")

#     # ── TYPE 1 — SOURCES DIFFÉRENTES ──
#     f.write("=" * 70 + "\n")
#     f.write("🔵 TYPE 1 — MÊME COMMENTAIRE, SOURCES DIFFÉRENTES\n")
#     f.write("   → Même client qui a posté sur plusieurs plateformes\n")
#     f.write("   → Garder 1 copie par source\n")
#     f.write("=" * 70 + "\n\n")

#     for i, row in enumerate(type_sources[:20], 1):
#         f.write(f"  [{i}] NB COPIES : {row['nb_copies']} | "
#                 f"SOURCES : {', '.join(row['sources'])}\n")
#         f.write(f"       COMMENTAIRE : {row['texte']}\n")
#         f.write(f"       DÉTAIL :\n")
#         for ex in row["exemples"]:
#             f.write(f"         • Source={ex['source']:<15} "
#                     f"Date={ex['date']:<25} "
#                     f"Modérateur={ex['moderateur']}\n")
#         f.write("\n")

#     if len(type_sources) > 20:
#         f.write(f"  ... et {len(type_sources)-20} autres groupes\n\n")

#     # ── TYPE 2 — DATES DIFFÉRENTES ──
#     f.write("=" * 70 + "\n")
#     f.write("🔴 TYPE 2 — MÊME COMMENTAIRE, DATES DIFFÉRENTES\n")
#     f.write("   → Client qui répète sa plainte à des dates différentes\n")
#     f.write("   → Garder toutes les copies (utile pour analyse temporelle)\n")
#     f.write("=" * 70 + "\n\n")

#     for i, row in enumerate(type_dates[:20], 1):
#         f.write(f"  [{i}] NB COPIES : {row['nb_copies']} | "
#                 f"NB DATES : {row['nb_dates']}\n")
#         f.write(f"       COMMENTAIRE : {row['texte']}\n")
#         f.write(f"       DÉTAIL :\n")
#         for ex in row["exemples"]:
#             f.write(f"         • Source={ex['source']:<15} "
#                     f"Date={ex['date']:<25} "
#                     f"Modérateur={ex['moderateur']}\n")
#         f.write("\n")

#     if len(type_dates) > 20:
#         f.write(f"  ... et {len(type_dates)-20} autres groupes\n\n")

#     # ── TYPE 3 — MODÉRATEURS DIFFÉRENTS ──
#     f.write("=" * 70 + "\n")
#     f.write("🟣 TYPE 3 — MÊME COMMENTAIRE, MODÉRATEURS DIFFÉRENTS\n")
#     f.write("   → Commentaire traité par plusieurs modérateurs\n")
#     f.write("   → Garder toutes les copies (utile pour analyse performance)\n")
#     f.write("=" * 70 + "\n\n")

#     for i, row in enumerate(type_moderateurs[:20], 1):
#         f.write(f"  [{i}] NB COPIES : {row['nb_copies']} | "
#                 f"MODÉRATEURS : {', '.join(row['moderateurs'])}\n")
#         f.write(f"       COMMENTAIRE : {row['texte']}\n")
#         f.write(f"       DÉTAIL :\n")
#         for ex in row["exemples"]:
#             f.write(f"         • Source={ex['source']:<15} "
#                     f"Date={ex['date']:<25} "
#                     f"Modérateur={ex['moderateur']}\n")
#         f.write("\n")

#     if len(type_moderateurs) > 20:
#         f.write(f"  ... et {len(type_moderateurs)-20} autres groupes\n\n")

#     # ── DÉCISIONS ──
#     f.write("=" * 70 + "\n")
#     f.write("✅ DÉCISIONS DE TRAITEMENT\n")
#     f.write("=" * 70 + "\n")
#     f.write("   🔵 Sources différentes    → Garder 1 copie par source\n")
#     f.write("   🔴 Dates différentes      → Garder toutes (analyse temporelle)\n")
#     f.write("   🟣 Modérateurs différents → Garder toutes (analyse performance)\n")
#     f.write("   ❌ Même texte+jour+source → Supprimer toutes sauf la plus ancienne\n")
#     f.write("\n")
#     f.write("=" * 70 + "\n")
#     f.write("FIN DU RAPPORT\n")
#     f.write("=" * 70 + "\n")

# print(f"✅ Rapport TXT sauvegardé : {RAPPORT_TXT}")

# # ============================================================
# # STYLES EXCEL
# # ============================================================
# ROUGE_F  = "8B0000"
# ROUGE_C  = "F4CCCC"
# BLEU_F   = "1F4E79"
# BLEU_C   = "DEEAF1"
# VIOLET_F = "4A148C"
# VIOLET_C = "E1BEE7"
# BLANC    = "FFFFFF"

# def style_entete(cell, couleur):
#     cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
#     cell.fill      = PatternFill("solid", start_color=couleur)
#     cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#     cell.border    = Border(left=Side(style="thin"), right=Side(style="thin"),
#                             top=Side(style="thin"),  bottom=Side(style="thin"))

# def style_data(cell, bg, bold=False, couleur_texte="000000", centre=False):
#     cell.font      = Font(name="Arial", bold=bold, size=10, color=couleur_texte)
#     cell.fill      = PatternFill("solid", start_color=bg)
#     cell.alignment = Alignment(
#         horizontal="center" if centre else "left",
#         vertical="center", wrap_text=True
#     )
#     cell.border    = Border(left=Side(style="thin"), right=Side(style="thin"),
#                             top=Side(style="thin"),  bottom=Side(style="thin"))

# def style_titre_groupe(cell, couleur_bg, couleur_texte):
#     cell.font      = Font(name="Arial", bold=True, size=10, color=couleur_texte)
#     cell.fill      = PatternFill("solid", start_color=couleur_bg)
#     cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
#     cell.border    = Border(left=Side(style="medium"), right=Side(style="medium"),
#                             top=Side(style="medium"),  bottom=Side(style="thin"))

# def creer_feuille_avec_exemples(wb, titre, couleur_header, couleur_groupe,
#                                  couleur_texte_groupe, donnees, couleur_exemple):
#     ws = wb.create_sheet(titre)
#     entetes  = ["#", "COMMENTAIRE", "NB COPIES", "COPIE #",
#                 "SOURCE", "DATE", "MODÉRATEUR"]
#     largeurs = [5, 55, 10, 8, 20, 25, 30]

#     for ci, (e, l) in enumerate(zip(entetes, largeurs), 1):
#         cell = ws.cell(row=1, column=ci, value=e)
#         style_entete(cell, couleur_header)
#         ws.column_dimensions[get_column_letter(ci)].width = l
#     ws.row_dimensions[1].height = 35
#     ws.freeze_panes = "A2"

#     ri = 2
#     for idx, row in enumerate(donnees, 1):
#         texte     = row["texte"]
#         nb_copies = row["nb_copies"]
#         exemples  = row["exemples"]
#         nb_rows   = len(exemples)

#         cell = ws.cell(row=ri, column=1, value=idx)
#         style_titre_groupe(cell, couleur_groupe, couleur_texte_groupe)

#         cell = ws.cell(row=ri, column=2, value=texte)
#         style_titre_groupe(cell, couleur_groupe, couleur_texte_groupe)
#         if nb_rows > 1:
#             ws.merge_cells(start_row=ri, start_column=2,
#                            end_row=ri+nb_rows-1, end_column=2)

#         cell           = ws.cell(row=ri, column=3, value=nb_copies)
#         cell.font      = Font(name="Arial", bold=True, size=12, color="8B0000")
#         cell.fill      = PatternFill("solid", start_color=couleur_groupe)
#         cell.alignment = Alignment(horizontal="center", vertical="center")
#         cell.border    = Border(left=Side(style="thin"), right=Side(style="thin"),
#                                 top=Side(style="medium"), bottom=Side(style="thin"))
#         if nb_rows > 1:
#             ws.merge_cells(start_row=ri, start_column=3,
#                            end_row=ri+nb_rows-1, end_column=3)

#         for ex_idx, exemple in enumerate(exemples):
#             bg = couleur_exemple if ex_idx % 2 == 0 else BLANC

#             cell = ws.cell(row=ri, column=4, value=f"Copie {ex_idx+1}")
#             style_data(cell, bg, bold=True, centre=True)

#             cell = ws.cell(row=ri, column=5, value=exemple["source"])
#             style_data(cell, bg)

#             cell = ws.cell(row=ri, column=6, value=exemple["date"])
#             style_data(cell, bg)

#             cell = ws.cell(row=ri, column=7, value=exemple["moderateur"])
#             style_data(cell, bg)

#             ws.row_dimensions[ri].height = 25
#             ri += 1

#         for ci in range(1, 8):
#             cell = ws.cell(row=ri, column=ci, value="")
#             cell.fill = PatternFill("solid", start_color="EEEEEE")
#         ri += 1

#     return ws

# # ============================================================
# # CRÉER EXCEL
# # ============================================================
# print("\n📊 Création du fichier Excel...")
# wb = Workbook()
# wb.remove(wb.active)

# ws_resume = wb.create_sheet("📊 Résumé")
# resume = [
#     ["📊 ANALYSE DÉTAILLÉE DES DOUBLONS", ""],
#     ["", ""],
#     ["Collection",                                           f"{DB_NAME}.{COLLECTION}"],
#     ["Date analyse",                                         datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
#     ["Total documents",                                      total],
#     ["", ""],
#     ["🔵 TYPE 1 — Même commentaire, sources différentes",    len(type_sources)],
#     ["🔴 TYPE 2 — Même commentaire, dates différentes",      len(type_dates)],
#     ["🟣 TYPE 3 — Même commentaire, modérateurs différents", len(type_moderateurs)],
#     ["", ""],
#     ["Total groupes de doublons",  total_groupes],
#     ["Total documents en double",  total_doublons],
#     ["Taux de doublons",           f"{taux_doublons:.2f}%"],
#     ["", ""],
#     ["✅ DÉCISIONS", ""],
#     ["Sources différentes",    "Garder 1 copie par source"],
#     ["Dates différentes",      "Garder toutes (analyse temporelle)"],
#     ["Modérateurs différents", "Garder toutes (analyse performance)"],
#     ["Même texte+jour+source", "Supprimer sauf la plus ancienne"],
# ]

# STYLES_RESUME = {
#     "📊 ANALYSE DÉTAILLÉE DES DOUBLONS"                       : ("1F4E79", 13),
#     "🔵 TYPE 1 — Même commentaire, sources différentes"       : ("1F4E79", 11),
#     "🔴 TYPE 2 — Même commentaire, dates différentes"         : ("8B0000", 11),
#     "🟣 TYPE 3 — Même commentaire, modérateurs différents"    : ("4A148C", 11),
#     "✅ DÉCISIONS"                                            : ("1E6B3C", 11),
# }

# for ri, (label, val) in enumerate(resume, 1):
#     c1 = ws_resume.cell(row=ri, column=1, value=label)
#     c2 = ws_resume.cell(row=ri, column=2, value=val)
#     if label in STYLES_RESUME:
#         couleur, size = STYLES_RESUME[label]
#         c1.font = Font(name="Arial", bold=True, size=size, color=couleur)
#         c2.font = Font(name="Arial", bold=True, size=size, color=couleur)
#     else:
#         c1.font = Font(name="Arial", bold=True, size=10)
#         c2.font = Font(name="Arial", size=10)

# ws_resume.column_dimensions["A"].width = 55
# ws_resume.column_dimensions["B"].width = 35

# creer_feuille_avec_exemples(wb, "🔵 Sources Différentes",
#     BLEU_F, BLEU_C, "1F4E79", type_sources, "EBF5FB")

# creer_feuille_avec_exemples(wb, "🔴 Dates Différentes",
#     ROUGE_F, ROUGE_C, "8B0000", type_dates, "FDEDEC")

# creer_feuille_avec_exemples(wb, "🟣 Modérateurs Différents",
#     VIOLET_F, VIOLET_C, "4A148C", type_moderateurs, "F3E5F5")

# os.makedirs(os.path.dirname(RAPPORT_PATH), exist_ok=True)
# wb.save(RAPPORT_PATH)
# client.close()

# print(f"✅ Rapport Excel sauvegardé : {RAPPORT_PATH}")
# print("\n" + "="*65)
# print("📊 RÉSUMÉ FINAL")
# print("="*65)
# print(f"   📥 Total documents                        : {total}")
# print(f"   🔵 Doublons sources différentes           : {len(type_sources)}")
# print(f"   🔴 Doublons dates différentes             : {len(type_dates)}")
# print(f"   🟣 Doublons modérateurs différents        : {len(type_moderateurs)}")
# print(f"   ❌ Total documents en double              : {total_doublons}")
# print(f"   📊 Taux de doublons                       : {taux_doublons:.2f}%")
# print("="*65)
# print(f"📄 Rapport TXT  : {RAPPORT_TXT}")
# print(f"📊 Rapport Excel: {RAPPORT_PATH}")
# print("="*65)
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# scripts/analyse/code/analyse_doublons_detail.py
# Analyse détaillée des 4 types de doublons avec exemples concrets
# ✅ Version alignée sur la logique MongoDB (normalisation 60 caractères, jour seulement)

from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime
import os
import re

# ============================================================
# CONFIGURATION
# ============================================================
MONGO_URI    = "mongodb://localhost:27018/"
DB_NAME      = "telecom_algerie"
COLLECTION   = "commentaires_sans_urls_arobase"
RAPPORT_PATH = "/home/mouna/projet_telecom/scripts/analyse/Rapports/rapport_doublons_detail.xlsx"
RAPPORT_TXT  = "/home/mouna/projet_telecom/scripts/analyse/Rapports/rapport_doublons_detail.txt"

# Nombre de caractères pour normalisation (comme MongoDB)
NB_CHARS_COMPARAISON = 60

# ============================================================
# FONCTIONS DE NORMALISATION (comme MongoDB)
# ============================================================
def normaliser_texte(texte, nb_chars=NB_CHARS_COMPARAISON):
    """
    Normalise le texte pour comparaison :
    - Supprime 'En voir plus' et variantes
    - Supprime espaces en trop
    - Prend les 60 premiers caractères (comme MongoDB)
    """
    if not texte:
        return ""
    texte = str(texte).strip()
    texte = re.sub(r'[Ee]n voir plus\.?', '', texte)
    texte = re.sub(r'See more\.?',        '', texte)
    texte = re.sub(r'أكثر\.?',           '', texte)
    texte = re.sub(r'\s+', ' ', texte).strip()
    return texte[:nb_chars]

def extraire_jour(date_str):
    """
    Extrait uniquement la date (AAAA-MM-JJ) en ignorant l'heure
    Comme MongoDB avec $dateToString
    """
    if not date_str:
        return "inconnu"
    date_str = str(date_str).strip()
    for fmt in ["%d/%m/%Y %H:%M", "%d/%m/%Y",
                "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"]:
        try:
            return datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
        except:
            continue
    return date_str[:10]

# ============================================================
# CONNEXION MONGODB
# ============================================================
print("="*65)
print("🔍 ANALYSE DÉTAILLÉE DES DOUBLONS (ALIGNÉE MONGODB)")
print("="*65)

client     = MongoClient(MONGO_URI)
db         = client[DB_NAME]
collection = db[COLLECTION]
total      = collection.count_documents({})
print(f"✅ {total} documents trouvés")

# ============================================================
# RÉCUPÉRER TOUS LES DOCUMENTS
# ============================================================
print("\n🔎 Chargement des données...")

groupes = defaultdict(list)

for doc in collection.find({}, {
    "Commentaire_Client": 1,
    "date"              : 1,
    "source"            : 1,
    "moderateur"        : 1,
}):
    texte = doc.get("Commentaire_Client", "")
    if not texte:
        continue
    
    # Normaliser le texte (60 premiers caractères) - comme MongoDB
    texte_norm = normaliser_texte(texte)
    
    # Extraire uniquement le jour - comme MongoDB
    jour = extraire_jour(doc.get("date", ""))
    
    groupes[texte_norm].append({
        "texte_original": texte,
        "texte_norm"    : texte_norm,
        "source"        : str(doc.get("source",     "") or ""),
        "date"          : str(doc.get("date",       "") or ""),
        "jour"          : jour,
        "moderateur"    : str(doc.get("moderateur", "") or ""),
    })

print(f"✅ {len(groupes)} textes normalisés uniques trouvés")

# ============================================================
# CLASSIFIER LES 4 TYPES (comme MongoDB)
# ============================================================
type_sources     = []
type_dates       = []
type_moderateurs = []
type_parfaits    = []

total_groupes_doublons = 0
total_documents_doublons = 0

for texte_norm, copies in groupes.items():
    if len(copies) < 2:
        continue
    
    total_groupes_doublons += 1
    total_documents_doublons += len(copies)
    
    sources     = set(c["source"]      for c in copies)
    jours       = set(c["jour"]        for c in copies)      # jours seulement (pas heures)
    moderateurs = set(c["moderateur"]  for c in copies)

    # Prendre un exemple de texte original pour l'affichage
    texte_exemple = copies[0]["texte_original"]
    texte_court = texte_exemple[:80] + "..." if len(texte_exemple) > 80 else texte_exemple
    exemples    = copies[:4]

    row = {
        "texte"       : texte_court,
        "texte_norm"  : texte_norm,
        "nb_copies"   : len(copies),
        "nb_sources"  : len(sources),
        "nb_jours"    : len(jours),
        "nb_mods"     : len(moderateurs),
        "exemples"    : exemples,
        "sources"     : sources,
        "jours"       : jours,
        "moderateurs" : moderateurs,
    }

    # 🔵 Sources différentes
    if len(sources) > 1:
        type_sources.append(row)
    
    # 🔴 Jours différents (pas dates complètes)
    if len(jours) > 1:
        type_dates.append(row)
    
    # 🟣 Modérateurs différents
    if len(moderateurs) > 1:
        type_moderateurs.append(row)
    
    # ⚪ Doublons parfaits (même jour, même source, même modérateur)
    if len(jours) == 1 and len(sources) == 1 and len(moderateurs) == 1:
        type_parfaits.append(row)

# Trier par nombre de copies
type_sources.sort(     key=lambda x: x["nb_copies"], reverse=True)
type_dates.sort(       key=lambda x: x["nb_copies"], reverse=True)
type_moderateurs.sort( key=lambda x: x["nb_copies"], reverse=True)
type_parfaits.sort(    key=lambda x: x["nb_copies"], reverse=True)

total_doublons = total_documents_doublons - total_groupes_doublons
taux_doublons  = total_doublons / total * 100

print(f"   📊 Groupes de doublons                : {total_groupes_doublons}")
print(f"   📥 Documents dans doublons            : {total_documents_doublons}")
print(f"   🔵 Doublons sources différentes       : {len(type_sources)}")
print(f"   🔴 Doublons jours différentes         : {len(type_dates)}")
print(f"   🟣 Doublons modérateurs différents    : {len(type_moderateurs)}")
print(f"   ⚪ Doublons parfaits                  : {len(type_parfaits)}")

# ============================================================
# GÉNÉRER LE RAPPORT TXT
# ============================================================
print("\n📝 Génération du rapport TXT...")

os.makedirs(os.path.dirname(RAPPORT_TXT), exist_ok=True)

with open(RAPPORT_TXT, "w", encoding="utf-8") as f:

    f.write("=" * 70 + "\n")
    f.write("RAPPORT — ANALYSE DÉTAILLÉE DES DOUBLONS (ALIGNÉE MONGODB)\n")
    f.write("=" * 70 + "\n")
    f.write(f"Date       : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    f.write(f"Collection : {DB_NAME}.{COLLECTION}\n")
    f.write(f"Méthode    : Normalisation {NB_CHARS_COMPARAISON} caractères, comparaison par jour\n")
    f.write("\n")

    # ── RÉSUMÉ GLOBAL ──
    f.write("=" * 70 + "\n")
    f.write("📊 RÉSUMÉ GLOBAL\n")
    f.write("=" * 70 + "\n")
    f.write(f"   Total documents                        : {total}\n")
    f.write(f"   Textes normalisés uniques              : {len(groupes)}\n")
    f.write(f"   Total groupes de doublons              : {total_groupes_doublons}\n")
    f.write(f"   Total documents en double              : {total_doublons}\n")
    f.write(f"   Taux de doublons                       : {taux_doublons:.2f}%\n")
    f.write("\n")
    f.write(f"   🔵 TYPE 1 — Sources différentes        : {len(type_sources)} groupes\n")
    f.write(f"   🔴 TYPE 2 — Jours différents           : {len(type_dates)} groupes\n")
    f.write(f"   🟣 TYPE 3 — Modérateurs différents     : {len(type_moderateurs)} groupes\n")
    f.write(f"   ⚪ TYPE 4 — Doublons parfaits          : {len(type_parfaits)} groupes\n")
    f.write("\n")

    # ── TYPE 1 — SOURCES DIFFÉRENTES ──
    if type_sources:
        f.write("=" * 70 + "\n")
        f.write("🔵 TYPE 1 — MÊME TEXTE NORMALISÉ, SOURCES DIFFÉRENTES\n")
        f.write("   → Même client qui a posté sur plusieurs plateformes\n")
        f.write("   → Garder 1 copie par source\n")
        f.write("=" * 70 + "\n\n")

        for i, row in enumerate(type_sources[:20], 1):
            f.write(f"  [{i}] NB COPIES : {row['nb_copies']} | "
                    f"SOURCES : {', '.join(row['sources'])}\n")
            f.write(f"       COMMENTAIRE : {row['texte']}\n")
            f.write(f"       DÉTAIL :\n")
            for ex in row["exemples"]:
                f.write(f"         • Source={ex['source']:<15} "
                        f"Jour={ex['jour']:<10} "
                        f"Modérateur={ex['moderateur']}\n")
            f.write("\n")

        if len(type_sources) > 20:
            f.write(f"  ... et {len(type_sources)-20} autres groupes\n\n")

    # ── TYPE 2 — JOURS DIFFÉRENTS ──
    if type_dates:
        f.write("=" * 70 + "\n")
        f.write("🔴 TYPE 2 — MÊME TEXTE NORMALISÉ, JOURS DIFFÉRENTS\n")
        f.write("   → Client qui répète sa plainte à des jours différents\n")
        f.write("   → Garder toutes les copies (utile pour analyse temporelle)\n")
        f.write("=" * 70 + "\n\n")

        for i, row in enumerate(type_dates[:20], 1):
            f.write(f"  [{i}] NB COPIES : {row['nb_copies']} | "
                    f"NB JOURS : {row['nb_jours']}\n")
            f.write(f"       COMMENTAIRE : {row['texte']}\n")
            f.write(f"       DÉTAIL :\n")
            for ex in row["exemples"]:
                f.write(f"         • Source={ex['source']:<15} "
                        f"Jour={ex['jour']:<10} "
                        f"Modérateur={ex['moderateur']}\n")
            f.write("\n")

        if len(type_dates) > 20:
            f.write(f"  ... et {len(type_dates)-20} autres groupes\n\n")

    # ── TYPE 3 — MODÉRATEURS DIFFÉRENTS ──
    if type_moderateurs:
        f.write("=" * 70 + "\n")
        f.write("🟣 TYPE 3 — MÊME TEXTE NORMALISÉ, MODÉRATEURS DIFFÉRENTS\n")
        f.write("   → Commentaire traité par plusieurs modérateurs\n")
        f.write("   → Garder toutes les copies (utile pour analyse performance)\n")
        f.write("=" * 70 + "\n\n")

        for i, row in enumerate(type_moderateurs[:20], 1):
            f.write(f"  [{i}] NB COPIES : {row['nb_copies']} | "
                    f"MODÉRATEURS : {', '.join(row['moderateurs'])}\n")
            f.write(f"       COMMENTAIRE : {row['texte']}\n")
            f.write(f"       DÉTAIL :\n")
            for ex in row["exemples"]:
                f.write(f"         • Source={ex['source']:<15} "
                        f"Jour={ex['jour']:<10} "
                        f"Modérateur={ex['moderateur']}\n")
            f.write("\n")

        if len(type_moderateurs) > 20:
            f.write(f"  ... et {len(type_moderateurs)-20} autres groupes\n\n")

    # ── TYPE 4 — DOUBLONS PARFAITS ──
    if type_parfaits:
        f.write("=" * 70 + "\n")
        f.write("⚪ TYPE 4 — DOUBLONS PARFAITS\n")
        f.write("   → Même texte normalisé, même jour, même source, même modérateur\n")
        f.write("   → À SUPPRIMER (garder seulement 1 copie)\n")
        f.write("=" * 70 + "\n\n")

        for i, row in enumerate(type_parfaits[:20], 1):
            f.write(f"  [{i}] NB COPIES : {row['nb_copies']} | "
                    f"MÊME SOURCE/MODÉRATEUR\n")
            f.write(f"       COMMENTAIRE : {row['texte']}\n")
            f.write(f"       DÉTAIL :\n")
            for ex in row["exemples"]:
                f.write(f"         • Source={ex['source']:<15} "
                        f"Jour={ex['jour']:<10} "
                        f"Modérateur={ex['moderateur']}\n")
            f.write("\n")

        if len(type_parfaits) > 20:
            f.write(f"  ... et {len(type_parfaits)-20} autres groupes\n\n")

    # ── STATISTIQUES FINALES ──
    f.write("=" * 70 + "\n")
    f.write("📊 STATISTIQUES FINALES\n")
    f.write("=" * 70 + "\n")
    f.write(f"   Total groupes de doublons              : {total_groupes_doublons}\n")
    f.write(f"   Total documents dans doublons          : {total_documents_doublons}\n")
    f.write(f"   Documents en double (à supprimer)      : {total_doublons}\n")
    f.write(f"   Taux de doublons                       : {taux_doublons:.2f}%\n")
    f.write("\n")
    f.write("=" * 70 + "\n")
    f.write("FIN DU RAPPORT\n")
    f.write("=" * 70 + "\n")

print(f"✅ Rapport TXT sauvegardé : {RAPPORT_TXT}")

# ============================================================
# STYLES EXCEL (inchangés)
# ============================================================
ROUGE_F  = "8B0000"
ROUGE_C  = "F4CCCC"
BLEU_F   = "1F4E79"
BLEU_C   = "DEEAF1"
VIOLET_F = "4A148C"
VIOLET_C = "E1BEE7"
VERT_F   = "1E6B3C"
VERT_C   = "D9EAD3"
BLANC    = "FFFFFF"

def style_entete(cell, couleur):
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill      = PatternFill("solid", start_color=couleur)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = Border(left=Side(style="thin"), right=Side(style="thin"),
                            top=Side(style="thin"),  bottom=Side(style="thin"))

def style_data(cell, bg, bold=False, couleur_texte="000000", centre=False):
    cell.font      = Font(name="Arial", bold=bold, size=10, color=couleur_texte)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(
        horizontal="center" if centre else "left",
        vertical="center", wrap_text=True
    )
    cell.border    = Border(left=Side(style="thin"), right=Side(style="thin"),
                            top=Side(style="thin"),  bottom=Side(style="thin"))

def style_titre_groupe(cell, couleur_bg, couleur_texte):
    cell.font      = Font(name="Arial", bold=True, size=10, color=couleur_texte)
    cell.fill      = PatternFill("solid", start_color=couleur_bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border    = Border(left=Side(style="medium"), right=Side(style="medium"),
                            top=Side(style="medium"),  bottom=Side(style="thin"))

def creer_feuille_avec_exemples(wb, titre, couleur_header, couleur_groupe,
                                 couleur_texte_groupe, donnees, couleur_exemple):
    if not donnees:
        return
        
    ws = wb.create_sheet(titre)
    entetes  = ["#", "COMMENTAIRE", "NB COPIES", "COPIE #",
                "SOURCE", "JOUR", "MODÉRATEUR"]  # "DATE" → "JOUR"
    largeurs = [5, 55, 10, 8, 20, 15, 30]

    for ci, (e, l) in enumerate(zip(entetes, largeurs), 1):
        cell = ws.cell(row=1, column=ci, value=e)
        style_entete(cell, couleur_header)
        ws.column_dimensions[get_column_letter(ci)].width = l
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"

    ri = 2
    for idx, row in enumerate(donnees, 1):
        texte     = row["texte"]
        nb_copies = row["nb_copies"]
        exemples  = row["exemples"]
        nb_rows   = len(exemples)

        cell = ws.cell(row=ri, column=1, value=idx)
        style_titre_groupe(cell, couleur_groupe, couleur_texte_groupe)

        cell = ws.cell(row=ri, column=2, value=texte)
        style_titre_groupe(cell, couleur_groupe, couleur_texte_groupe)
        if nb_rows > 1:
            ws.merge_cells(start_row=ri, start_column=2,
                           end_row=ri+nb_rows-1, end_column=2)

        cell           = ws.cell(row=ri, column=3, value=nb_copies)
        cell.font      = Font(name="Arial", bold=True, size=12, color="8B0000")
        cell.fill      = PatternFill("solid", start_color=couleur_groupe)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = Border(left=Side(style="thin"), right=Side(style="thin"),
                                top=Side(style="medium"), bottom=Side(style="thin"))
        if nb_rows > 1:
            ws.merge_cells(start_row=ri, start_column=3,
                           end_row=ri+nb_rows-1, end_column=3)

        for ex_idx, exemple in enumerate(exemples):
            bg = couleur_exemple if ex_idx % 2 == 0 else BLANC

            cell = ws.cell(row=ri, column=4, value=f"Copie {ex_idx+1}")
            style_data(cell, bg, bold=True, centre=True)

            cell = ws.cell(row=ri, column=5, value=exemple["source"])
            style_data(cell, bg)

            cell = ws.cell(row=ri, column=6, value=exemple["jour"])  # Maintenant "jour"
            style_data(cell, bg)

            cell = ws.cell(row=ri, column=7, value=exemple["moderateur"])
            style_data(cell, bg)

            ws.row_dimensions[ri].height = 25
            ri += 1

        for ci in range(1, 8):
            cell = ws.cell(row=ri, column=ci, value="")
            cell.fill = PatternFill("solid", start_color="EEEEEE")
        ri += 1

    return ws

# ============================================================
# CRÉER EXCEL
# ============================================================
print("\n📊 Création du fichier Excel...")
wb = Workbook()
wb.remove(wb.active)

ws_resume = wb.create_sheet("📊 Résumé")
resume = [
    ["📊 ANALYSE DÉTAILLÉE DES DOUBLONS (ALIGNÉE MONGODB)", ""],
    ["", ""],
    ["Collection",                                           f"{DB_NAME}.{COLLECTION}"],
    ["Date analyse",                                         datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
    ["Méthode",                                              f"Normalisation {NB_CHARS_COMPARAISON} caractères, comparaison par jour"],
    ["Total documents",                                      total],
    ["", ""],
    ["📊 Groupes de doublons",                               total_groupes_doublons],
    ["📥 Documents dans doublons",                           total_documents_doublons],
    ["❌ Documents en double",                               total_doublons],
    ["", ""],
    ["🔵 TYPE 1 — Sources différentes",                      len(type_sources)],
    ["🔴 TYPE 2 — Jours différentes",                        len(type_dates)],
    ["🟣 TYPE 3 — Modérateurs différents",                   len(type_moderateurs)],
    ["⚪ TYPE 4 — Doublons parfaits",                        len(type_parfaits)],
    ["", ""],
    ["📈 Taux de doublons",                                  f"{taux_doublons:.2f}%"],
    ["", ""],
    ["✅ DÉCISIONS", ""],
    ["Sources différentes",    "Garder 1 copie par source"],
    ["Jours différentes",      "Garder toutes (analyse temporelle)"],
    ["Modérateurs différents", "Garder toutes (analyse performance)"],
    ["Doublons parfaits",      "Supprimer toutes sauf une"],
]

STYLES_RESUME = {
    "📊 ANALYSE DÉTAILLÉE DES DOUBLONS (ALIGNÉE MONGODB)"    : ("1F4E79", 13),
    "📊 Groupes de doublons"                                 : ("1F4E79", 11),
    "🔵 TYPE 1 — Sources différentes"                        : ("1F4E79", 11),
    "🔴 TYPE 2 — Jours différentes"                          : ("8B0000", 11),
    "🟣 TYPE 3 — Modérateurs différents"                     : ("4A148C", 11),
    "⚪ TYPE 4 — Doublons parfaits"                          : ("1E6B3C", 11),
    "✅ DÉCISIONS"                                           : ("1E6B3C", 11),
}

for ri, (label, val) in enumerate(resume, 1):
    c1 = ws_resume.cell(row=ri, column=1, value=label)
    c2 = ws_resume.cell(row=ri, column=2, value=val)
    if label in STYLES_RESUME:
        couleur, size = STYLES_RESUME[label]
        c1.font = Font(name="Arial", bold=True, size=size, color=couleur)
        c2.font = Font(name="Arial", bold=True, size=size, color=couleur)
    else:
        c1.font = Font(name="Arial", bold=True, size=10)
        c2.font = Font(name="Arial", size=10)

ws_resume.column_dimensions["A"].width = 60
ws_resume.column_dimensions["B"].width = 35

# Créer les feuilles avec exemples
creer_feuille_avec_exemples(wb, "🔵 Sources Différentes",
    BLEU_F, BLEU_C, "1F4E79", type_sources, "EBF5FB")

creer_feuille_avec_exemples(wb, "🔴 Jours Différents",  # "Dates" → "Jours"
    ROUGE_F, ROUGE_C, "8B0000", type_dates, "FDEDEC")

creer_feuille_avec_exemples(wb, "🟣 Modérateurs Différents",
    VIOLET_F, VIOLET_C, "4A148C", type_moderateurs, "F3E5F5")

creer_feuille_avec_exemples(wb, "⚪ Doublons Parfaits",
    VERT_F, VERT_C, "1E6B3C", type_parfaits, "D9EAD3")

os.makedirs(os.path.dirname(RAPPORT_PATH), exist_ok=True)
wb.save(RAPPORT_PATH)
client.close()

print(f"✅ Rapport Excel sauvegardé : {RAPPORT_PATH}")
print("\n" + "="*65)
print("📊 RÉSUMÉ FINAL")
print("="*65)
print(f"   📥 Total documents                        : {total}")
print(f"   📊 Groupes de doublons                    : {total_groupes_doublons}")
print(f"   📥 Documents dans doublons                : {total_documents_doublons}")
print(f"   🔵 Doublons sources différentes           : {len(type_sources)}")
print(f"   🔴 Doublons jours différentes             : {len(type_dates)}")
print(f"   🟣 Doublons modérateurs différents        : {len(type_moderateurs)}")
print(f"   ⚪ Doublons parfaits                      : {len(type_parfaits)}")
print(f"   ❌ Documents en double (à supprimer)      : {total_doublons}")
print(f"   📊 Taux de doublons                       : {taux_doublons:.2f}%")
print("="*65)
print(f"📄 Rapport TXT  : {RAPPORT_TXT}")
print(f"📊 Rapport Excel: {RAPPORT_PATH}")
print("="*65)