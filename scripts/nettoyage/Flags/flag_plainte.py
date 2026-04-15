"""
DÉTECTION FLAG_PLAINTE DZ — v2 FINAL
══════════════════════════════════════════════════════════════════════════
Insight clé du corpus :
  Sur la page AT (service client), toute demande de contact est implicitement
  une plainte — l'utilisateur ne contacte pas un service client "pour le fun".
  La distinction se fait sur l'INTENSITÉ, pas sur la présence d'une frustration :

  Sous-types :
    (1, 'prv')           → contact + signal frustration explicite → Négatif fort
    (1, 'prv_implicit')  → contact seul (sans frustration textuelle explicite)
                           → Négatif implicite faible (à garder)
    (1, 'neglex')        → lexique négatif fort autonome → Négatif fort
    (-1,'contact_only')  → contact de type RÉPONSE AT (c'est AT qui répond, pas client)
                           OU contact positif poli (bravo + prv)
    (0, None)            → pas une plainte

  Exceptions supprimées (flag=-1) :
    - Réponses AT : "ندعوكم للتواصل معنا" = c'est AT qui demande, pas le client
    - Contact positif : bravo/مبروك + prv = poli, pas une plainte
"""

import re
from collections import Counter
from pymongo import MongoClient, UpdateOne
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════════════
MONGO_URI       = "mongodb://localhost:27017/"
DB_NAME         = "telecom_algerie"
COLLECTION_NAME = "commentaires_normalises_tolerance"
TEXT_FIELD      = "normalized_arabert"
OUTPUT_EXCEL    = "commentaires_flag_plainte_v2.xlsx"

print("✅ Config chargée.")

# ══════════════════════════════════════════════════════════════════════════
# 1. PATTERNS DEMANDE DE CONTACT CLIENT → AT
#    Le client demande une réponse / un contact
# ══════════════════════════════════════════════════════════════════════════
CONTACT_PATTERNS = [
    # ردو / رد في الخاص
    r'ردو?\s*(في\s*|على\s*|فال|فل|ف)?\s*[ال]?خاص\b',
    r'ر[دج]وا?\s*(في\s*)?الخاص',
    r'الرد\s*(في|على)\s*الخاص',
    r'نرجو\s*الرد\s*(في|على)?\s*الخاص',
    r'ارجو\s*(منكم\s*)?الرد\s*(في|على)\s*الخاص',
    r'الرجاء\s*الرد\s*(في|على|على)?\s*الخاص',
    r'يرجى\s*الرد\s*(على|في)\s*الخاص',
    r'رجاء\s*الرد\s*(على|في)\s*الخاص',
    r'ردو\s*عليا\b', r'ردو\s*علينا\b',
    r'ردو\s*على\s*(الرسائل|الرسا[يئ]ل|الميساجات|ميساجات|لميساج|انشغالي|الخاص|خاص|رسائل)',
    r'رد\s*(في|على)\s*الخاص',
    r'ردو\s*(الخاص|في\s*الخاص)\b',
    # بعثتلكم + ردو
    r'بعثتلكم?\s*(في|على|فال|فل)?\s*الخاص',
    r'بعثلكم\s*الخاص',
    r'بعثتلك\s*(في|على)\s*الخاص',
    r'ارسلتك?\s*(على|في|لكم)\s*(رساله|رسالة|في)?\s*الخاص',
    r'ارسلت\s*لكم\s*رساله\b',
    r'راسلنا\s*(عبر|في|على|مرارا)',
    r'راسلتكم\b',
    # Absence de réponse — forte connotation négative
    r'لا\s*(يوجد|يتم)\s*رد\b', r'لايوجد\s*رد\b',
    r'لا\s*تردون\b', r'لا\s*تردو\b',
    r'لا\s*يردون\b', r'لا\s*يردو\b',
    r'لاتحيبون\b', r'لا\s*تحيبون\b',
    r'لا\s*تقومون\s*بالرد',
    r'لم\s*يتم\s*(الرد|رد)\b', r'لا\s*يتم\s*(الرد|رد)\b',
    r'لماذا\s*لا\s*تردو', r'لماذا\s*لاتردون', r'لماذا\s*لا\s*يتم\s*الرد',
    r'اصبحتم\s*لا\s*تردون',
    r'لا\s*يردون\s*على\s*الزبائن',
    r'مكمش\s*تردو', r'ماكمش\s*تردو', r'مراكمش\s*تردو',
    r'مكانش?\s*رد\b', r'متردوش\b',
    r'لم\s*يرد\s*(احد|عليا|علي)\b',
    r'ولا\s*رد\b', r'مايريبونديوش\b',
    r'خاصيه\s*الرد.*غير\s*موجوده',
    # ياو / معليش
    r'ياو\s*ردو', r'ياو\s*الرد',
    r'اخي\s*ردو', r'معليش\s*تردو', r'معليش\s*تردو',
    # تواصل / contact
    r'تواصلو?\s*(معنا|معي|مع[يا]|معاك)\b',
    r'تواصل\s*معي\b', r'اتصلو\s*بي\b',
    r'خطي\s*غير\s*موهل\b',
    # Arabizi / Français
    r'\bprv\b', r'\binbox\b',
    r'repondi[w]?\s*prv', r'jawbouna\b',
    r'r[eé]pondez?\s*(nous|moi)',
    r'contactez[\s\-]?(nous|moi)', r'contactez\s*moi\b',
    r'envoyez?\s*(nous|moi)\s*(un\s*)?message',
    r'send\s*(us|me)\s*(a\s*)?dm', r'\bdm\s*(us|me|nous)?\b',
]

# ══════════════════════════════════════════════════════════════════════════
# 2. PATTERNS FRUSTRATION / NÉGATIVITÉ EXPLICITE
#    Contact + frustration = 'prv' (négatif fort)
#    Contact seul = 'prv_implicit' (négatif faible)
# ══════════════════════════════════════════════════════════════════════════
FRUSTRATION_PATTERNS = [
    # Durée
    r'\d+\s*(jours?|أيام|يوم|ايام|mois|أشهر|اشهر|شهر|semaines?|سمانه|ساعات|ساعه)',
    r'منذ\s*\d+', r'depuis\s*\d+',
    r'اكثر\s*من\s*\d+\s*(اشهر|اسابيع|ايام|شهر|يوم)',
    r'مرارا\s*وتكرارا', r'sans\s*r[eé]ponse',
    r'بدون\s*(رد|حل)', r'دون\s*(رد|حل|جدوى)', r'دون\s*جدوى\b',
    r'لا\s*جديد\s*يذكر',
    r'حتى\s*الان\s*(لا|ما|لم)', r'لحد\s*الان\s*(لا|ما|لم)',
    # انشغال / شكوى / مشكل
    r'انشغال\w*', r'شكو[ىا]\b', r'شكايه\b', r'شكوه\b',
    r'probl[eè]me?\b', r'مشكل[ةه]?\b', r'مشاكل\b', r'عطل\b',
    r'مستعجل\b', r'مهم\s*ومستعجل',
    # Attente
    r'انتظار\b', r'attente\b', r'ننتظر\b', r'نستنى\b',
    r'انتظر\s*(الرد|ردكم)',
    # Non-fonctionnement
    r'لا\s*يعمل\b', r'لا\s*يشتغل\b', r'لا\s*يتصل\s*بالانترنت',
    r'ne\s*fonctionne\s*pas', r'marche\s*pas',
    r'ما\s*خدم\b', r'مايخدم\b', r'مايصلح\b',
    r'coupure\b', r'انقطاع\b', r'مقطوع[هة]?\b', r'موقف\b', r'suspendu\b',
    r'لم\s*يتم\s*تفعيل', r'لا\s*يتم\s*تفعيل',
    r'لم\s*يتواصل\b', r'لم\s*يات\b',
    r'تدفق\s*(كارثي|ضعيف)',
    r'بلا\s*انترنت', r'مكانش\s*انترنت', r'الانترنت\s*مكانش',
    r'الانترنت\s*مقطوع', r'مقطوعه\s*عندها',
    r'مازال\s*الانترنت\s*ما\s*جاتش',
    r'نهار\s*كامل\s*مكانش', r'بدون\s*الانترنت',
    r'لا\s*توجد\s*انترنت', r'نحتاج\s*انترنت',
    r'\bnul\b', r'catastrophe?\b', r'كارثه?\b', r'كارثة\b', r'كارثيه\b',
    r'honte\b', r'عيب\b', r'inacceptable\b', r'danger\s*national',
    r'خدمات?\s*(تحت\s*الصفر|صفر)', r'تحت\s*الصفر\b', r'صفر\s*كبير?\b',
    r'الاسوى\b', r'الأسوأ\b', r'دائما\s*الاسوى', r'مستوى\s*ناقص',
    r'سوء\s*استغلال', r'عدم\s*الالتفات', r'عدم\s*حل',
    r'تعبتونا\b', r'هرمنا\b', r'هبلتونا\b', r'تهبلونا\b',
    r'قهرونا\b', r'مرضتونا\b', r'عييتونا\b',
    r'كرهنا\b', r'كرهتونا\b', r'شبعتونا\b', r'مقروط\b', r'مستفز[هة]\b',
    r'يكدبو\b', r'كذابيين\b', r'وعود\s*كاذبة', r'مافيا\b', r'فساد\b',
    r'متحشموش\b', r'مهابل\b', r'هايله\b', r'فاشلين\b',
    r'شركه\s*(ضخمه|في\s*انحطاط)',
    r'ماكمش\s*تصدقو', r'ماتهزوش\b',
    r'خلصت?\b.{0,50}(مازال|ما\s*جا|لا\s*يوجد|لم\s*يتم|بلا)',
    r'خلصت?\b.{0,10}(mo|ميغا).{0,20}(تلحق|غير)',
    r'مزدتوليش\b', r'شارجيت.{0,30}علاش',
    r'فواتير.{0,30}مدفوع.{0,30}(لا|ما|مازال|موقف)',
    r'تم\s*توقيف\s*خط', r'خط\s*مقفل\b',
    r'رانا\s*في\s*(ساعه|يومين|اسبوع)',
    r'رانا\s*نخلصو\b',
    r'الانترنت\s*(نهار\s*كامل|كامل)',
    r'قدمت\s*شكوى', r'اريد\s*تقديم\s*شكوا',
    r'ضد\s*الزبون', r'يعاني\b', r'معاناه\b',
    r'لم\s*يتم\s*الربط', r'بعيده\s*عن\s*البيوت',
    r'روتينق\s*(راه\s*)?كارث', r'بينق.{0,20}(كارث|مرتفع)',
    r'في\s*\d+\s*اشهر.{0,30}(كارث|مشكل|ضعيف)',
    # Vulgarités
    r'نيك\b', r'نيكما', r'قحبة\b', r'حمار\b', r'كلب\b',
]

# ══════════════════════════════════════════════════════════════════════════
# 3. PATTERNS RÉPONSE AT (c'est l'opérateur qui parle, pas le client)
#    Ces messages sont des réponses AT → flag=-1 / 'at_response'
# ══════════════════════════════════════════════════════════════════════════
AT_RESPONSE_PATTERNS = [
    r'ندعوكم\s*(للتواصل|للتواصل\s*معنا)',
    r'يمكنكم\s*(التواصل|مراسلتنا)',
    r'بإمكانكم\s*التواصل',
    r'ابقوا\s*على\s*تواصل\s*معنا',
    r'سيتم\s*(اعلامكم|إعلامكم|التواصل)',
    r'نعمل\s*باستمرار\s*على\s*(تطوير|تحسين)',
    r'لتمكننا\s*من\s*تقديم\s*المساعده',
    r'قصد\s*تزويدنا\s*بمعلومات',
    r'لنتمكن\s*من\s*تقديم\s*المساعده',
    r'لقد\s*تم\s*التواصل\s*معك\s*في\s*الخاص',
    r'اذا\s*كنتم\s*تواجهون\s*اي\s*خلل',
    r'فبامكانكم\s*التواصل\s*معنا\s*عبر\s*الخاص',
]

# ══════════════════════════════════════════════════════════════════════════
# 4. PATTERNS NÉGATIF FORT AUTONOME (sans contact)
# ══════════════════════════════════════════════════════════════════════════
NEGLEX_PATTERNS = [
    r'هرمنا\b', r'هرمتونا\b',
    r'هبلتونا\b', r'هبلتو\b', r'تهبلونا\b',
    r'مرضتونا\b', r'تعبتونا\b', r'ta3batna\b',
    r'قهرونا\b', r'قسونا\b', r'عييتونا\b',
    r'نقسوناوفسوم', r'شبعتونا\b', r'مقروط\b',
    r'كرهتونا\b', r'كرهنا\b',
    r'كارثيه\b',
    r'مافيا\b', r'فساد\b',
    r'متحشموش\b',
    r'الأسوأ\b', r'الاسوى\b',
    r'مايصلحش\b', r'ما\s*يصلحش\b',
    r'ميمشيش\b', r'ما\s*يمشيش\b',
    r'مكانش\s*(خدمة|رد|حل)',
    r'يكدبو\b', r'كذابيين\b', r'كدابين\b',
    r'وعود\s*كاذبة', r'مهابل\b', r'مساسيط\b',
    r'هايله\b',
    r'نيكماتكم\b', r'نيكماتك\b', r'واش\s*نيك',
    r'danger\s*national',
    r'شركه?\s*مستفز[هة]\b',
]

# ══════════════════════════════════════════════════════════════════════════
# 5. NEUTRALISANTS (contact positif poli → flag=0)
# ══════════════════════════════════════════════════════════════════════════
NEUTRALIZER_PATTERNS = [
    r'بالتوفيق\b', r'ربي\s*يوفقكم', r'بارك\s*الله',
    r'bravo\b', r'félicitations?\b', r'bon\s*courage\b',
    r'مبروك\b', r'ماشاء\s*الله',
    r'نتمنى\s*(لكم|لك)\s*(النجاح|التوفيق)',
    r'إلى\s*الأمام\b',
]

# ══════════════════════════════════════════════════════════════════════════
# COMPILATION
# ══════════════════════════════════════════════════════════════════════════
FLAGS = re.IGNORECASE | re.UNICODE

CONTACT_REGEX     = re.compile('|'.join(CONTACT_PATTERNS),     FLAGS)
FRUSTRATION_REGEX = re.compile('|'.join(FRUSTRATION_PATTERNS), FLAGS | re.DOTALL)
AT_RESPONSE_REGEX = re.compile('|'.join(AT_RESPONSE_PATTERNS), FLAGS)
NEGLEX_REGEX      = re.compile('|'.join(NEGLEX_PATTERNS),      FLAGS)
NEUTRALIZER_REGEX = re.compile('|'.join(NEUTRALIZER_PATTERNS), FLAGS)

def get_match(regex, text):
    m = regex.search(text or ''); return m.group() if m else None

def is_plainte(text):
    t = text or ''
    has_contact     = bool(CONTACT_REGEX.search(t))
    has_frustration = bool(FRUSTRATION_REGEX.search(t))
    has_neglex      = bool(NEGLEX_REGEX.search(t))
    has_at_resp     = bool(AT_RESPONSE_REGEX.search(t))
    has_neutralizer = bool(NEUTRALIZER_REGEX.search(t))

    # 1. Négatif fort autonome
    if has_neglex:
        return 1, 'neglex'

    # 2. Réponse AT → supprimer
    if has_at_resp and not has_frustration:
        return -1, 'at_response'

    # 3. Contact + frustration → négatif fort
    if has_contact and has_frustration:
        return 1, 'prv'

    # 4. Contact seul → désormais supprimé (flag = -1)
    if has_contact:
        if has_neutralizer:
            return 0, None      # contact poli → neutre
        return -1, 'prv_implicit'   # ← MODIFICATION ICI

    return 0, None
# ══════════════════════════════════════════════════════════════════════════
# TESTS UNITAIRES
# ══════════════════════════════════════════════════════════════════════════
tests = [
    # flag=1 / 'prv'
    ('ردو في الخاص le service est nul',                          (1, 'prv')),
    ('repondiw prv 3 jours sans réponse',                        (1, 'prv')),
    ('ارجو الرد في الخاص مشكل منذ اسبوع',                       (1, 'prv')),
    ('ردو عليا اتصال راهي كارثه',                                (1, 'prv')),
    ('algérie télécom اذا لماذا لاتحيبون على الانشغالات',       (1, 'prv')),
    ('من فضلكم ردو على الخاص راني بلا انترنت',                  (1, 'prv')),
    ('ردو في الخاص كرهنا من الانترنت تاعكم خدمات تحت الصفر',   (1, 'prv')),
    ('لماذا لا تردو على الخاص',                                  (1, 'prv')),
    ('ردو عليا فخاص باغيين تهبلونا',                             (1, 'neglex')),
    # flag=1 / 'prv_implicit' — contacts seuls (vrais négatifs du corpus)
    ('algerietelecom بعثلكم الخاص ردو',                         (1, 'prv_implicit')),
    ('algérie télécom الرجاء الرد في الخاص',                    (1, 'prv_implicit')),
    ('algérie télécom ردو على رسائل فالخاص',                    (1, 'prv_implicit')),
    ('نرجو منكم الرد على الخاص',                                 (1, 'prv_implicit')),
    ('ردو عليا فلخاص',                                           (1, 'prv_implicit')),
    ('ردو عليا',                                                  (1, 'prv_implicit')),
    ('ياو ردو في الخاص',                                         (1, 'prv_implicit')),
    ('معليش تردو عليا',                                          (1, 'prv_implicit')),
    ('ردو على الميساجات',                                        (1, 'prv_implicit')),
    ('بعثتلكم في الخاص جاوبو',                                   (1, 'prv_implicit')),
    ('رد في الخاص',                                              (1, 'prv_implicit')),
    # flag=1 / 'neglex'
    ('يكدبو دائما وعود كاذبة',                                   (1, 'neglex')),
    ('هرمنا من هاد الخدمه',                                      (1, 'neglex')),
    ('هبلتونا والله',                                             (1, 'neglex')),
    ('مافيا هذا الاتصالات',                                      (1, 'neglex')),
    ('شركه مستفزه خدمه عملاء كارثيه بكل المقاييس',              (1, 'neglex')),
    # flag=-1 / 'at_response'
    ('ندعوكم للتواصل معنا عبر الخاص قصد تزويدنا بمعلومات',      (-1, 'at_response')),
    ('اذا كنتم تواجهون اي خلل فبامكانكم التواصل معنا عبر الخاص',(-1, 'at_response')),
    ('لقد تم التواصل معك في الخاص',                             (-1, 'at_response')),
    # flag=0 — neutre
    ('bravo بون كوراج',                                          (0,  None)),
    ('مبروك على التطور',                                         (0,  None)),
    ('connexion mliha bezzaf',                                   (0,  None)),
    ('bravo مبروك ردو عليا في الخاص',                           (0,  None)),
]

print('\n── Tests ──────────────────────────────────────────────')
ok = 0
for text, expected in tests:
    flag, stype = is_plainte(text)
    status = '✅' if (flag, stype) == expected else '❌'
    print(f'  {status} [{flag:+d}/{str(stype):14s}] {text[:70]!r}')
    ok += ((flag, stype) == expected)
print(f'\n  {ok}/{len(tests)} tests passés\n')

# ══════════════════════════════════════════════════════════════════════════
# CONNEXION MONGODB
# ══════════════════════════════════════════════════════════════════════════
print('⏳ Connexion MongoDB …')
client     = MongoClient(MONGO_URI)
collection = client[DB_NAME][COLLECTION_NAME]
docs = list(collection.find({}, {
    '_id': 1, TEXT_FIELD: 1,
    'Commentaire_Client_Original': 1,
    'sources': 1, 'dates': 1, 'label_final': 1,
    'flag_social': 1, 'flag_encouragement': 1,
}))
print(f'✅ {len(docs)} documents chargés.')

# ══════════════════════════════════════════════════════════════════════════
# TRAITEMENT
# ══════════════════════════════════════════════════════════════════════════
results = []
for doc in docs:
    text = doc.get(TEXT_FIELD, '') or ''
    flag, stype = is_plainte(text)

    if stype in ('prv', 'prv_implicit'):
        c = get_match(CONTACT_REGEX, text) or ''
        f = get_match(FRUSTRATION_REGEX, text) or ''
        expr = f"{c} | {f}" if f else c
    elif stype == 'neglex':
        expr = get_match(NEGLEX_REGEX, text) or ''
    elif stype == 'at_response':
        expr = get_match(AT_RESPONSE_REGEX, text) or ''
    else:
        expr = ''

    results.append({
        '_id':                  doc['_id'],
        'doc_id':               str(doc['_id']),
        'commentaire_original': doc.get('Commentaire_Client_Original', ''),
        'texte_normalise':      text,
        'expression_detectee':  expr,
        'source':               doc.get('sources', ''),
        'date':                 doc.get('dates', ''),
        'label_final':          doc.get('label_final', ''),
        'flag_social':          doc.get('flag_social', 0),
        'flag_encouragement':   doc.get('flag_encouragement', 0),
        'flag_plainte':         flag,
        'plainte_type':         stype or '',
    })

flagged         = [r for r in results if r['flag_plainte'] == 1]
flag_prv        = [r for r in flagged  if r['plainte_type'] == 'prv']
flag_prv_impl   = [r for r in flagged  if r['plainte_type'] == 'prv_implicit']
flag_neglex     = [r for r in flagged  if r['plainte_type'] == 'neglex']
suppressed      = [r for r in results  if r['flag_plainte'] == -1]
at_resp         = [r for r in suppressed if r['plainte_type'] == 'at_response']
contact_pos     = [r for r in suppressed if r['plainte_type'] == 'contact_pos']
counter         = Counter(r['expression_detectee'].split(' | ')[0]
                          for r in flagged if r['expression_detectee'])

print(f'\n📊 Statistiques v2 :')
print(f'   Total docs                    : {len(results)}')
print(f'   flag_plainte = 1 (plaintes)   : {len(flagged)}')
print(f'   ├─ prv          → Négatif fort (contact+frustration) : {len(flag_prv)}')
print(f'   ├─ prv_implicit → Négatif implicite (contact seul)   : {len(flag_prv_impl)}')
print(f'   └─ neglex       → Négatif explicite (lexique fort)   : {len(flag_neglex)}')
print(f'   flag_plainte = -1 (supprimés) : {len(suppressed)}')
print(f'   ├─ at_response  → Réponses AT opérateur              : {len(at_resp)}')
print(f'   └─ contact_pos  → Contact positif poli               : {len(contact_pos)}')
print(f'\n🔑 Top 15 expressions déclenchantes :')
for expr, cnt in counter.most_common(15):
    print(f'   {cnt:4d}  {expr!r}')

# ══════════════════════════════════════════════════════════════════════════
# MISE À JOUR MONGODB
# ══════════════════════════════════════════════════════════════════════════
print('\n⏳ Mise à jour MongoDB …')
operations = [
    UpdateOne(
        {'_id': r['_id']},
        {'$set': {
            'flag_plainte':       r['flag_plainte'],
            'plainte_type':       r['plainte_type'],
            'expression_plainte': r['expression_detectee'],
        }}
    )
    for r in results
]
res = collection.bulk_write(operations)
print(f'✅ MongoDB mis à jour — {res.modified_count} docs modifiés.')

# ══════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════════════
print('⏳ Génération Excel …')
hdr_fill     = PatternFill('solid', fgColor='7B0000')
hdr_font     = Font(color='FFFFFF', bold=True, size=11)
prv_fill     = PatternFill('solid', fgColor='FF6B6B')   # rouge fort — prv
prvimp_fill  = PatternFill('solid', fgColor='FFD7D7')   # rose — prv_implicit
neglex_fill  = PatternFill('solid', fgColor='C00000')   # rouge foncé — neglex
at_fill      = PatternFill('solid', fgColor='FFF2CC')   # jaune — at_response
odd_fill     = PatternFill('solid', fgColor='FFF8F8')
even_fill    = PatternFill('solid', fgColor='FFFFFF')
c_al = Alignment(horizontal='center', vertical='center')
l_al = Alignment(horizontal='left',   vertical='top', wrap_text=True)
thin = Side(style='thin', color='BFBFBF')
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = 'flag_plainte_v2'
headers    = ['doc_id','commentaire_original','texte_normalise',
              'expression_detectee','source','date',
              'label_final','flag_social','flag_encouragement',
              'flag_plainte','plainte_type']
col_widths = [28, 60, 55, 35, 12, 18, 12, 11, 18, 11, 14]

for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.fill, cell.font = hdr_fill, hdr_font
    cell.alignment = c_al; cell.border = brd
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 28
ws.freeze_panes = 'A2'

type_fill = {
    'prv':          prv_fill,
    'prv_implicit': prvimp_fill,
    'neglex':       neglex_fill,
    'at_response':  at_fill,
}
export_rows = flagged + suppressed

for i, r in enumerate(export_rows):
    row_n    = i + 2
    row_fill = odd_fill if i % 2 == 0 else even_fill
    stype    = r['plainte_type']
    tf       = type_fill.get(stype, row_fill)
    vals = [r['doc_id'], r['commentaire_original'], r['texte_normalise'],
            r['expression_detectee'], r['source'], r['date'],
            r['label_final'], r['flag_social'], r['flag_encouragement'],
            r['flag_plainte'], stype]
    for ci, val in enumerate(vals, 1):
        cell = ws.cell(row=row_n, column=ci, value=val)
        cell.border = brd; cell.font = Font(name='Arial', size=9)
        if ci in (10, 11):
            cell.fill, cell.alignment = tf, c_al
            cell.font = Font(name='Arial', size=9, bold=True)
            if stype == 'neglex':
                cell.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        elif ci in (2, 3):
            cell.fill, cell.alignment = row_fill, l_al
        else:
            cell.fill, cell.alignment = row_fill, c_al
    ws.row_dimensions[row_n].height = 45

ws2 = wb.create_sheet('Stats')
ws2['A1'] = 'Statistiques flag_plainte v2'
ws2['A1'].font = Font(bold=True, size=13, color='7B0000')
for r_idx, (lbl, val) in enumerate([
    ('Total documents',                                   len(results)),
    ('flag_plainte = 1  (plaintes)',                      len(flagged)),
    ('  ├─ prv          → Négatif fort',                  len(flag_prv)),
    ('  ├─ prv_implicit → Négatif implicite faible',      len(flag_prv_impl)),
    ('  └─ neglex       → Négatif explicite fort',        len(flag_neglex)),
    ('flag_plainte = -1 (supprimés)',                     len(suppressed)),
    ('  ├─ at_response  → Réponses AT',                   len(at_resp)),
    ('  └─ contact_pos  → Contact positif poli',          len(contact_pos)),
], start=3):
    ws2.cell(row=r_idx, column=1, value=lbl).font = Font(name='Arial', size=10)
    ws2.cell(row=r_idx, column=2, value=val).font  = Font(name='Arial', size=10, bold=True)
ws2['A13'] = 'Top expressions déclenchantes'
ws2['A13'].font = Font(bold=True, size=10)
for r_idx, (expr, cnt) in enumerate(counter.most_common(20), start=14):
    ws2.cell(row=r_idx, column=1, value=expr)
    ws2.cell(row=r_idx, column=2, value=cnt)
ws2.column_dimensions['A'].width = 55

ws3 = wb.create_sheet('Légende')
ws3['A1'] = 'Légende — flag_plainte v2'
ws3['A1'].font = Font(bold=True, size=12, color='7B0000')
for r_idx, (fill, nom, desc) in enumerate([
    (prv_fill,    'Rouge fort  — prv           flag=+1', 'Contact + frustration explicite → Négatif fort'),
    (prvimp_fill, 'Rose        — prv_implicit  flag=+1', 'Contact client seul → Négatif implicite faible'),
    (neglex_fill, 'Rouge foncé — neglex         flag=+1', 'Lexique négatif fort autonome → Négatif fort'),
    (at_fill,     'Jaune       — at_response   flag=-1', 'Réponse AT (ndعوكم…) → Supprimer'),
], start=3):
    c = ws3.cell(row=r_idx, column=1, value=nom)
    c.fill, c.font = fill, Font(name='Arial', size=10, bold=True)
    if 'foncé' in nom:
        c.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    ws3.cell(row=r_idx, column=2, value=desc).font = Font(name='Arial', size=10)
ws3['A9'] = 'Insight v2 :'
ws3['A9'].font = Font(bold=True, size=10)
for r_idx, desc in enumerate([
    "Sur une page service client AT, tout contact client est une plainte implicite.",
    "prv_implicit = le client a écrit mais n'a pas eu de réponse → négatif faible (60-70%)",
    "Seules exceptions supprimées : réponses AT (ndعوكم) et contacts positifs polis.",
], start=10):
    ws3.cell(row=r_idx, column=1, value=desc).font = Font(name='Arial', size=10)
ws3.column_dimensions['A'].width = 90; ws3.column_dimensions['B'].width = 55

wb.save(OUTPUT_EXCEL)
print(f'✅ Excel généré : {OUTPUT_EXCEL}  ({len(export_rows)} lignes)')
client.close()