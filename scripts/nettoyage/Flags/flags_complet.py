"""
DÉTECTION FLAG_SOCIAL DZ — version REGEX uniquement
Détecte les commentaires de vœux/encouragements sociaux
Export Excel : uniquement flag_social = 1
"""

import re
from collections import Counter
from pymongo import MongoClient, UpdateOne
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════════════

# "mongodb+srv://yousrahadjabderrahmane_db_user:C8wjIvWqOBUjK66u@ac-1ksfahb-shard-00-00.gejzu4a.mongodb.net/?retryWrites=true&w=majority&appName=Cluster0"
MONGO_URI       = "mongodb://localhost:27018/"
DB_NAME         = "telecom_algerie"
COLLECTION_NAME = "commentaires_normalises"
TEXT_FIELD      = "normalized_arabert"
OUTPUT_EXCEL    = "commentaires_flag_social_1.xlsx"

print("✅ Config chargée.")

# ══════════════════════════════════════════════════════════════════════════
# PATTERNS POSITIFS — vœux / encouragements sociaux
# ══════════════════════════════════════════════════════════════════════════
SOCIAL_PATTERNS = [

    # ── Arabe / darija arabe ──────────────────────────────────────────────
    r'بارك\s*الله',
    r'الله\s*يبارك',
    r'الله\s*يعاون',
    r'ربي\s*يعاون',
    r'ربي\s*يجازي',
    r'جزاك\s*الله\s*خير',
    r'جزاكم\s*الله\s*خير',
    r'الله\s*يوفق',
    r'ربي\s*يحفظ',
    r'ربي\s*يوفق',
    r'بالتوفيق',
    r'رمضان\s*كريم',
    r'رمضان\s*مبارك',
    r'عيد\s*مبارك',
    r'عيد\s*سعيد',
    r'كل\s*عام\s*و[أا]نتم',
    r'كل\s*عام\s*وأنتم',
    r'مبروك',
    r'تبارك\s*الله',
    r'ماشاء\s*الله',
    r'ما\s*شاء\s*الله',
    r'الله\s*معاكم',
    r'الله\s*معاك\b',
    r'ربي\s*يوفقكم',
    r'ربي\s*يوفقك\b',
    r'الله\s*يوفقكم',
    r'شكرا\s*على\s*جهودكم',
    r'شكرا\s*على\s*المجهودات',
    r'ربي\s*يجازيكم',
    r'ربي\s*يعاونكم',
    r'إن\s*شاء\s*الله\s*ينجح',
    r'نشالله\s*ينجح',
    r'يسلموا?\b',
    r'الله\s*يسلمك',
    r'صح\s*فطورك',
    r'صح\s*فطوركم',
    r'صح\s*عيدك',
    r'صح\s*عيدكم',
    r'حفظكم\s*الله',
    r'وفقكم\s*الله',
    r'الله\s*يعطيكم\s*الصحة',
    r'ربي\s*يبارك\s*فيكم',
    r'ربي\s*يبارك\s*فيك\b',

    # ── Darija latin / arabizi ────────────────────────────────────────────
    r'rbi\s*y[3e]awkom',
    r'rabi\s*y[3e]awkom',
    r'barak\s*allah',
    r'baraka\s*llah',
    r'allah\s*ybarek',
    r'allah\s*yewfek',
    r'allah\s*yewfeqkom',
    r'tbarakallah',
    r'tbarkallah',
    r'nchallah\s*ynajh',
    r'inch[a]?allah\s*ynajh',
    r'saha\s*ftourkoum',
    r'saha\s*[ae]idkom',
    r'sa7a\s*[ae]idkom',
    r'kol\s*[3e]am\s*w',
    r'[ae]id\s*mobarak',
    r'[ae]id\s*sa[3i]id',
    r'mabrouk',
    r'mbr[ou]+k',
    r'bon\s*courage',
    r'bonne?\s*continuation',
    r'merci\s*pour\s*vos\s*efforts',
    r'merci\s*pour\s*le\s*travail',
    r'courage\s*[àa]\s*tou[st]',
    r'bonne?\s*ann[eé]e',
    r'assegwas\s*amegas',
    r'vive\s*alg[eé]rie',
    r'allah\s*ma[3e]akom',
    r'rabi\s*ma[3e]akom',
    r'barra\s*barra\s*likom',
    r'[ae]ychine\b',
    r'3aychine\b',
    r'rbi\s*yjazik',
    r'yeslamo\b',
    r'ya3tikom\s*essa7a',

    # ── Français ─────────────────────────────────────────────────────────
    r'joyeux\s*ramadan',
    r'bon\s*travail',
    r'bravo\s*(à\s*)?l[a-z]*\s*[eé]quipe',
    r'bonne?\s*f[eê]te',
    r'f[eé]licitations?',
    r'bonne?\s*ann[eé]e',
    r'bonne?\s*continuation',
    r'merci\s*pour\s*vos\s*efforts',
    r'bonne?\s*chance',
    r'bravo\s*pour',
    r'chapeau\s*pour',
    r'joyeuses?\s*f[eê]tes?',
    r'meilleurs?\s*v[oœ]ux',
    r'bonne?\s*journ[eé]e',
    r'bon\s*succ[eè]s',
    r'bonne?\s*r[eé]ussite',
    r'tous\s*mes\s*v[oœ]ux',
]

# ══════════════════════════════════════════════════════════════════════════
# PATTERNS NÉGATIFS — mots-clés qui ANNULENT le flag
# (pour éviter les faux positifs du type "جزاكم الله" dans une longue plainte)
# ══════════════════════════════════════════════════════════════════════════
NEGATIVE_PATTERNS = [
    r'مشكل',
    r'مشكلة',
    r'شكاية',
    r'شكوى', r'شكاوي',
    r'ما\s*خدم',
    r'مكانش',
    r'ما\s*عندي',
    r'ما\s*ركبوا',
    r'ما\s*تواصل',
    r'ما\s*جاو',
    r'ما\s*حلوا',
    r'ما\s*صلحوا',
    r'ما\s*ردوا',
    r'مشكل.*جزاك',
    r'انتظار',
    r'أنتظر',
    r'نستنى',
    r'وعود\s*كاذبة',
    r'سوء\s*الاستقبال',
    r'بدون\s*حل',
    r'دون\s*[أا]ي\s*حل',
    r'هل\s*يعقل',
    r'هل\s*منطق',
    r'نقطة\s*الصفر',
    r'معاناة',
    r'اضطهاد',
    r'ظلم',
    r'إرهاب',
    r'مافيا',
    r'فساد',
    r'سجن',
    r'ديكتاتور',
    r'تبون',
    r'بوتفليقة',
    r'فاضح',
    r'راكم\s*غير\s*تزيدوا\s*تبعدوا',
    r'خدمة.*شهر',
    r'شهر.*ما.*ركب',
    r'شهرين.*ولحد',
    r'ست\s*اشهر',
    r'شهر\s*ونص',
    r'يحرم\s*من\s*خدمة',
    r'لم\s*يتواصلوا',
    r'لم\s*نستطع',
    r'لم\s*يرد',
    r'خط\s*أخضر.*لا\s*يرد',
    r'تعبتونا',
    r'هبلونا',
    r'حيكم',
    r'اذلال',
]

SOCIAL_REGEX   = re.compile('|'.join(SOCIAL_PATTERNS),   flags=re.IGNORECASE | re.UNICODE)
NEGATIVE_REGEX = re.compile('|'.join(NEGATIVE_PATTERNS), flags=re.IGNORECASE | re.UNICODE)

def get_match_social(text):
    m = SOCIAL_REGEX.search(text or '')
    return m.group() if m else None

def is_social(text):
    t = text or ''
    # Doit matcher un pattern positif ET ne pas matcher un pattern négatif
    if not SOCIAL_REGEX.search(t):
        return 0
    if NEGATIVE_REGEX.search(t):
        return 0
    return 1

# ── Test rapide ───────────────────────────────────────────────────────────
tests_social = [
    # Vrais sociaux → 1
    ('ربي يوفقكم',                                              1),
    ('barak allah fik',                                         1),
    ('bon courage',                                             1),
    ('mabrouk',                                                 1),
    ('بالتوفيق والسداد',                                        1),
    # Faux positifs corrigés → 0
    ('والله يا ذيك راحه',                                       0),
    ('تعبتونا والله',                                           0),
    ('المودم مكانش',                                            0),
    # cas réels signalés → 0
    ('واش من اقرب راكم غير تزيدوا تبعدوا علينا بلاصيت فيبر',   0),
    ('قدمت عدت شكاوي جزاكم الله خيرا',                         0),  # plainte avec جزاكم
    ('ست اشهر يستنى انترنت هل يعقل',                           0),
    ('الله يوفق كل من يسعى لإصلاح هذا',                        1),  # court, positif pur
]
print('Test des patterns flag_social :')
ok = 0
for text, expected in tests_social:
    result  = is_social(text)
    match   = get_match_social(text)
    status  = '✅' if result == expected else '❌'
    print(f'  {status} [{result}] {text[:50]!r:55s} → match={match!r}')
    ok += (result == expected)
print(f'\n{ok}/{len(tests_social)} tests passés.')

# ══════════════════════════════════════════════════════════════════════════
# CHARGEMENT MONGODB — FLAG_SOCIAL
# ══════════════════════════════════════════════════════════════════════════
print('⏳ Connexion MongoDB …')
client     = MongoClient(MONGO_URI)
collection = client[DB_NAME][COLLECTION_NAME]

docs = list(collection.find({}, {
    '_id': 1,
    TEXT_FIELD: 1,
    'Commentaire_Client_Original': 1,
    'sources': 1,
    'dates': 1,
    'label_final': 1,
}))
print(f'✅ {len(docs)} documents chargés.')

# ══════════════════════════════════════════════════════════════════════════
# APPLICATION DES REGEX — FLAG_SOCIAL
# ══════════════════════════════════════════════════════════════════════════
print('⏳ Application des patterns …')

results_social = []
for doc in docs:
    text  = doc.get(TEXT_FIELD) or ''
    flag  = is_social(text)
    match = get_match_social(text) if flag else ''
    results_social.append({
        '_id':                  doc['_id'],
        'doc_id':               str(doc['_id']),
        'texte_normalise':      text,
        'commentaire_original': doc.get('Commentaire_Client_Original', ''),
        'source':               doc.get('sources', ''),
        'date':                 doc.get('dates', ''),
        'label_final':          doc.get('label_final', ''),
        'flag_social':          flag,
        'expression_detectee':  match or '',
    })

n_social = sum(1 for r in results_social if r['flag_social'] == 1)
total    = len(results_social)
print(f'✅ flag_social=1 : {n_social} / {total} ({100*n_social/total:.1f}%)')

# ══════════════════════════════════════════════════════════════════════════
# MISE À JOUR MONGODB — FLAG_SOCIAL
# ══════════════════════════════════════════════════════════════════════════
print('⏳ Mise à jour MongoDB …')

operations_social = [
    UpdateOne(
        {'_id': r['_id']},
        {'$set': {
            'flag_social':       r['flag_social'],
            'expression_sociale': r['expression_detectee'],
        }}
    )
    for r in results_social
]
res = collection.bulk_write(operations_social)
print(f'✅ MongoDB mis à jour — {res.modified_count} docs modifiés.')

# ══════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL — FLAG_SOCIAL uniquement flag_social = 1
# ══════════════════════════════════════════════════════════════════════════
print('⏳ Génération du fichier Excel …')

social_results = [r for r in results_social if r['flag_social'] == 1]

# ── Styles ────────────────────────────────────────────────────────────────
hdr_fill    = PatternFill('solid', fgColor='1F4E79')
hdr_font    = Font(color='FFFFFF', bold=True, size=11)
green_fill  = PatternFill('solid', fgColor='E2EFDA')
yellow_fill = PatternFill('solid', fgColor='FFF2CC')
blue_fill   = PatternFill('solid', fgColor='D9E1F2')
sec_fill    = PatternFill('solid', fgColor='D9E1F2')
c_al        = Alignment(horizontal='center', vertical='center')
l_al        = Alignment(horizontal='left',   vertical='center', wrap_text=True)
border      = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

def write_header(ws, cols, widths):
    ws.row_dimensions[1].height = 28
    for ci, (h, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = c_al
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w

wb_social  = Workbook()

# ── Feuille 1 : flag_social_1 ─────────────────────────────────────────────
ws1 = wb_social.active
ws1.title = 'flag_social_1'
write_header(ws1,
    ['doc_id', 'commentaire_original', 'texte_normalise',
     'expression_detectee', 'source', 'date', 'label_final', 'flag_social'],
    [28, 60, 60, 30, 14, 18, 14, 12]
)
for ri, r in enumerate(social_results, 2):
    ws1.row_dimensions[ri].height = 22
    vals   = [r['doc_id'], r['commentaire_original'], r['texte_normalise'],
              r['expression_detectee'], r['source'], r['date'],
              r['label_final'], r['flag_social']]
    fills  = [None, green_fill, green_fill, yellow_fill, None, None, None, blue_fill]
    aligns = [c_al, l_al, l_al, c_al, c_al, c_al, c_al, c_al]
    for ci, (val, fill, align) in enumerate(zip(vals, fills, aligns), 1):
        cell = ws1.cell(row=ri, column=ci, value=val)
        cell.border = border
        cell.alignment = align
        if fill:
            cell.fill = fill
        if ci == 8:
            cell.font = Font(bold=True, color='276221')

# ── Feuille 2 : Stats ─────────────────────────────────────────────────────
ws2 = wb_social.create_sheet('Stats')
ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 20

source_counts = Counter(r['source'] for r in social_results)
match_counts  = Counter(r['expression_detectee'].lower()
                        for r in social_results if r['expression_detectee'])

stats_rows = [
    ('RÉSULTATS — DÉTECTION flag_social = 1', ''),
    ('', ''),
    ('Total commentaires traités',   total),
    ('flag_social = 1',              n_social),
    ('flag_social = 0',              total - n_social),
    ('% commentaires sociaux',       f'{100*n_social/total:.2f}%'),
    ('Méthode',                      'Regex lexicale + filtre négatif'),
    ('', ''),
    ('RÉPARTITION PAR SOURCE', ''),
]
for src, cnt in source_counts.most_common():
    stats_rows.append((f'  {src}', cnt))

stats_rows += [('', ''), ('TOP 10 EXPRESSIONS DÉTECTÉES', '')]
for expr, cnt in match_counts.most_common(10):
    stats_rows.append((f'  {expr}', cnt))

for ri, (label, val) in enumerate(stats_rows, 1):
    ws2.row_dimensions[ri].height = 22
    ca = ws2.cell(row=ri, column=1, value=label)
    cb = ws2.cell(row=ri, column=2, value=val)
    is_sec = any(x in str(label) for x in ['RÉSULTATS', 'RÉPARTITION', 'TOP 10'])
    if is_sec:
        ca.font = Font(bold=True, size=11, color='1F4E79')
        ca.fill = sec_fill
        cb.fill = sec_fill
    ca.alignment = l_al
    cb.alignment = c_al
    if isinstance(val, int):
        cb.font = Font(bold=True)

wb_social.save(OUTPUT_EXCEL)
print(f'✅ Excel généré : {OUTPUT_EXCEL}')
print(f'   → {n_social} commentaires sociaux exportés')
print()
print('=' * 55)
print('  RÉSUMÉ FINAL — FLAG_SOCIAL')
print('=' * 55)
print(f'  Total docs traités    : {total}')
print(f'  flag_social = 1       : {n_social} ({100*n_social/total:.1f}%)')
print(f'  flag_social = 0       : {total - n_social} ({100*(total-n_social)/total:.1f}%)')
print(f'  Méthode               : Regex + filtre négatif')
print(f'  MongoDB mis à jour    : ✅')
print(f'  Excel généré          : {OUTPUT_EXCEL} ✅')
print('=' * 55)

client.close()


# ════════════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════════════
# DÉTECTION FLAG_SUGGESTION DZ — v1
# ════════════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════════════
"""
DÉTECTION FLAG_SUGGESTION DZ — v1
══════════════════════════════════════════════════════════════════════════
Insight clé du corpus :
  "Positif Conditionnel DZ" — satisfaction partielle très fréquente dans
  les avis algériens en ligne. L'utilisateur est globalement satisfait
  MAIS signale un manque ou un potentiel d'amélioration.

  Structure obligatoire : [Mot positif DZ] + [Connecteur adversatif DZ]
                          + [Verbe / mot d'amélioration DZ]

  Sous-types :
    (1, 'pur')        → 3 conditions réunies, aucun négatif fort
                        → Positif conditionnel (60–70%) — jamais Négatif
    (1, 'pos_adv')    → positif + adversatif SANS verbe d'amélioration
                        → Positif atténué faible (50–60%)
    (-1,'masked_neg') → les 3 conditions présentes MAIS négatif fort aussi
                        → flag_mixte l'emporte, flag_suggestion supprimé
    (0, None)         → pas un positif conditionnel

  Règle d'or :
    flag_suggestion = 1 (pur)     si A ∧ B ∧ C ∧ ¬NEG_FORT
    flag_suggestion = 1 (pos_adv) si A ∧ B ∧ ¬C ∧ ¬NEG_FORT
    flag_suggestion = -1          si (A ∧ B ∧ C) ∧ NEG_FORT  → flag_mixte domine
    flag_suggestion = 0           sinon
"""

# ══════════════════════════════════════════════════════════════════════════
# CONFIG — FLAG_SUGGESTION
# ══════════════════════════════════════════════════════════════════════════
MONGO_URI_SUG       = "mongodb://localhost:27018/"
DB_NAME_SUG         = "telecom_algerie"
COLLECTION_NAME_SUG = "commentaires_normalises"
TEXT_FIELD_SUG      = "normalized_arabert"
OUTPUT_EXCEL_SUG    = "commentaires_flag_suggestion_v1.xlsx"

print("✅ Config chargée.")

# ══════════════════════════════════════════════════════════════════════════
# 1. CONDITION A — MOTS POSITIFS DZ
#    Présence d'un terme évaluatif positif (qualité, satisfaction)
# ══════════════════════════════════════════════════════════════════════════
POS_PATTERNS_SUG = [
    # ── Darija latin / arabizi ─────────────────────────────────────────────
    r'\bmliha?\b',          # mliha / mlih
    r'\bmzian[ae]?\b',      # mzian / mziana
    r'\bzwin[ae]?\b',       # zwina / zwine
    r'\bhil[ae]\b',         # hila / hile
    r'\btamam\b',           # tamam (trop bien)
    r'\bmazyan[ae]?\b',     # mazyan / mazyana
    r'\bsari3\b',           # rapide (arabizi)
    r'\bb[ae]rak\s*allah\b',# baraka allah (positif pur, pas social ici)
    # ── Français ──────────────────────────────────────────────────────────
    r'\bbien\b',
    r'\bbon(ne)?\b',
    r'\bsuper\b',
    r'\bpas\s*mal\b',
    r'\bcorrect(e)?\b',
    r'\bsatisfais?ant(e)?\b',
    r'\brapide\b',
    r'\befficace\b',
    r"\bc'est\s*bien\b",        
    r'\bça\s*march[e]?\b',
    r'\bça\s*fonctionn\w+\b',
    r'\bbon\s*d[ée]but\b',
    r'\bencourageant(e)?\b',
    r'\bglobalement\s*(positif|bien|correct)\b',
    r'\bbon\s*travail\b',
    r'\bp[ae]s\s*mal\b',
    r'\bbonne\s*(qualité|vitesse|connexion)\b',
    r'\bla\s*connexion\s*(est\s*)?(bonne|correcte|rapide|stable)\b',
    # ── Arabe / Darija arabe ───────────────────────────────────────────────
    r'مليح[ةه]?',
    r'زوين[ةه]?',
    r'\bتمام\b',
    r'ممتاز[ةه]?',
    r'سريع[ةه]?',
    r'جيد[ةه]?',
    r'حسن[ةه]?',
    r'مقبول[ةه]?',
    r'هايل[ةه]?',
    r'رائع[ةه]?',
    r'مزيان[ةه]?',
    r'معقول[ةه]?',
    r'كويس[ةه]?',
    r'نظيف[ةه]?',
    r'مرتاح[ةه]?',
    r'راضي[ةه]?',
    r'خدمة\s*(مليحة|جيدة|حسنة|معقولة)',
    r'الخدمة\s*(مليحة|جيدة|ممتازة)',
    r'اتصال\s*(مليح|جيد|سريع)',
    r'الانترنت\s*(مليح|مليحة|سريع|تمام)',
    r'سرعة\s*(كويسة|مليحة|جيدة)',
]

# ══════════════════════════════════════════════════════════════════════════
# 2. CONDITION B — CONNECTEURS ADVERSATIFS DZ
#    Marqueur d'opposition / de restriction (mais, cependant, seulement)
# ══════════════════════════════════════════════════════════════════════════
ADV_PATTERNS = [
    # ── Darija latin ──────────────────────────────────────────────────────
    r'\bms\b',              # mais (darija lat.)
    r'\bbss\b',             # seulement / mais
    r'\blukan\b',           # si seulement
    r'\blou?kan\b',
    r'\bwalakin\b',         # mais (arabe dial.)
    # ── Français ──────────────────────────────────────────────────────────
    r'\bmais\b',
    r'\bcependant\b',
    r'\btoutefois\b',
    r'\bseulement\b',
    r'\bpar\s*contre\b',
    r'\bn[eé]anmoins\b',
    r'\bsi\s*possible\b',
    r'\bsi\s*seulement\b',
    r'\bsauf\s*que\b',
    r'\bexcept[eé]\s*que\b',
    r'\bà\s*condition\s*que\b',
    # ── Arabe / Darija arabe ───────────────────────────────────────────────
    r'\bبصح\b',             # mais (darija ar.) — PRIORITÉ HAUTE
    r'\bلكن\b',
    r'\bوالكن\b',
    r'\bغير\s*أن\b',
    r'\bإلا\s*أن\b',
    r'\bبس\b',
    r'\bلو\s*كان\b',
    r'لو\s*كانو',
    r'\bفقط\s*لو\b',
    r'\bولكن\b',
    r'\bمع\s*ذلك\b',
    r'\bرغم\s*ذلك\b',
    r'\bبالرغم\s*من\b',
]

# ══════════════════════════════════════════════════════════════════════════
# 3. CONDITION C — VERBES / MOTS D'AMÉLIORATION DZ
#    Demande explicite d'une action correctrice ou d'un ajout
# ══════════════════════════════════════════════════════════════════════════
AMEL_PATTERNS = [
    # ── Darija latin — verbes d'action ────────────────────────────────────
    r'\btehsen[ou]*\b',     # améliorez
    r'\btzid[ou]*\b',       # continuez / ajoutez
    r'\bbeddl[ou]*\b',      # changez
    r'\bcorrigi\b',         # corrigez
    r'\bdebugi\b',          # déboguez
    r'\brajou[ou]*\b',      # revenez sur
    r'\bml[ou]+\b',         # remplissez
    r'\bkamml[ou]*\b',      # complétez
    r'\bhasn[ou]*\b',       # améliorez (var.)
    # ── Français — demandes constructives ─────────────────────────────────
    r'amélior\w+',
    r'optim[iu]s\w+',
    r'augment\w+\s*(la\s*|le\s*)?(vitesse|débit|qualité)',
    r'r[eé]duir\w+\s*(le\s*|les\s*)?prix',
    r'baiss\w+\s*(le\s*|les\s*)?prix',
    r'peaufin\w+',
    r'ajust\w+',
    r'travaill\w+\s*sur',
    r'des?\s*am[eé]liorations?\b',
    r'des?\s*efforts?\b',
    r'un\s*peu\s*plus\b',
    r'encore\s*mieux\b',
    r'plus\s*de\s*(rapidité|vitesse|qualité|stabilité)',
    r'il\s*(faut|faudrait)\s*(encore|plus)',
    r'il\s*(reste|y\s*a)\s*(des\s*)?am[eé]liorations?\s*[àa]\s*faire',
    r'ce\s*serait\s*(bien|mieux|parfait)\s*(de|si)',
    r"j'aimerais?\s*(plus|voir|que)",
    r'on\s*attend\s*(mieux|plus)',
    r'si\s*possible\s*(am[eé]lior|r[eé]duir|augment|opt)',
    # ── Arabe / Darija arabe ───────────────────────────────────────────────
    r'تحسين',
    r'يتحسن',
    r'تحسنو',
    r'تزيدو',
    r'نتمنى\s*(يتحسن|تحسين|أكثر|أفضل)',
    r'يلزم\s*تحسن',
    r'خاصكم\s*تحسنو',
    r'لو\s*(زادو|تحسنو|رفعو)',
    r'نشاء\s*الله\s*يتحسن',
    r'إن\s*شاء\s*الله\s*يتحسن',
    r'المزيد\s*من\s*(الجهود|التحسين|التطوير)',
    r'نأمل\s*(في\s*)?(تحسين|تطوير)',
    r'نتمنى\s*(تطوير|تحسين)',
    r'يجب\s*(تحسين|تطوير|رفع)',
    r'أتمنى\s*(أن\s*)?(يتحسن|تحسين)',
    r'نقص\s*في\s*(السعر|الخدمة|التغطية)',    # manque = demande implicite
    r'لو\s*كان\s*السعر\s*(أقل|منخفض)',
    r'لو\s*تخفضو\s*السعر',
    r'ينقصه?\s*(فقط|بس)',                    # il manque juste
    r'ينقص\s*(شوية|قليل|قليلاً)',
]

# ══════════════════════════════════════════════════════════════════════════
# 4. NÉGATIFS FORTS — annulent flag_suggestion → flag_mixte domine
#    Si présents avec A+B+C → flag = -1 / 'masked_neg'
# ══════════════════════════════════════════════════════════════════════════
NEG_FORT_PATTERNS = [
    # ── Mots négatifs forts DZ ─────────────────────────────────────────────
    r'\bnul\b', r'\bkhaybe?\b', r'\bkhayba?\b',
    r'\bmachi\s*barka\b', r'\bkarica\b', r'\bkhartha?\b',
    r'كارثة', r'كارثه', r'كارثي[هة]?',
    r'فضيحة', r'مهزلة',
    r'\bzero\b', r'تحت\s*الصفر',
    # ── Non-fonctionnement ────────────────────────────────────────────────
    r'لا\s*يعمل', r'ما\s*يخدم', r'مايخدمش',
    r'مقطوع[هة]?', r'قطوع\b',
    r'coupure\b', r'انقطاع\b',
    # ── Plaintes explicites ───────────────────────────────────────────────
    r'شكوى', r'شكاية', r'شكاوي',
    r'مشكل[ةه]?', r'مشاكل\b',
    r'r[eé]clamation\b', r'plainte\b',
    # ── Durée d'attente + frustration ─────────────────────────────────────
    r'\d+\s*(أشهر|mois|شهر|jours?|أيام|ايام)\s*(بدون|sans|دون|ولم)',
    r'شهرين.*ما\s*ركب',
    r'ست\s*اشهر',
    r'مازال\s*(ما|لم)\s*(جا[ءو]?|ركب)',
    # ── Frustration forte ────────────────────────────────────────────────
    r'تعبتونا', r'هرمنا', r'هبلتونا',
    r'وعود\s*كاذبة',
    r'مافيا\b', r'فساد\b',
    r'هل\s*يعقل',
    # ── Sarcasme explicite ────────────────────────────────────────────────
    r'والله.*يا\s*(ذيك|راحه)',
    r'في\s*الأحلام\b', r'في\s*المنام\b',
]

# ══════════════════════════════════════════════════════════════════════════
# COMPILATION — FLAG_SUGGESTION
# ══════════════════════════════════════════════════════════════════════════
FLAGS_SUG = re.IGNORECASE | re.UNICODE

POS_REGEX_SUG      = re.compile('|'.join(POS_PATTERNS_SUG),  FLAGS_SUG)
ADV_REGEX          = re.compile('|'.join(ADV_PATTERNS),      FLAGS_SUG)
AMEL_REGEX         = re.compile('|'.join(AMEL_PATTERNS),     FLAGS_SUG)
NEG_FORT_REGEX     = re.compile('|'.join(NEG_FORT_PATTERNS), FLAGS_SUG | re.DOTALL)

def get_match_sug(regex, text):
    m = regex.search(text or ''); return m.group() if m else None

def is_suggestion(text):
    """
    Retourne (flag, subtype) :
      (1,  'pur')        → A ∧ B ∧ C ∧ ¬NEG  — Positif conditionnel fort
      (1,  'pos_adv')    → A ∧ B ∧ ¬C ∧ ¬NEG — Positif atténué faible
      (-1, 'masked_neg') → A ∧ B ∧ C ∧ NEG   — Masqué par négatif fort
      (0,  None)         → pas de positif conditionnel
    """
    t = text or ''
    has_pos  = bool(POS_REGEX_SUG.search(t))
    has_adv  = bool(ADV_REGEX.search(t))
    has_amel = bool(AMEL_REGEX.search(t))
    has_neg  = bool(NEG_FORT_REGEX.search(t))

    # Cas de base : au moins pos + adversatif requis
    if not (has_pos and has_adv):
        return 0, None

    # A ∧ B ∧ C ∧ NEG → masqué — flag_mixte domine
    if has_amel and has_neg:
        return -1, 'masked_neg'

    # A ∧ B ∧ C ∧ ¬NEG → positif conditionnel pur
    if has_amel and not has_neg:
        return 1, 'pur'

    # A ∧ B ∧ ¬C ∧ ¬NEG → positif atténué (adversatif sans verbe d'amélioration)
    if not has_amel and not has_neg:
        return 1, 'pos_adv'

    return 0, None

# ══════════════════════════════════════════════════════════════════════════
# TESTS UNITAIRES — FLAG_SUGGESTION
# ══════════════════════════════════════════════════════════════════════════
tests_sug = [
    # ── flag=1 / 'pur' — les 3 conditions ─────────────────────────────────
    ('connexion mliha ms tehseno fiha',                                    (1, 'pur')),
    ("l'appli zwina bss lukan tzidou fiha",                                (1, 'pur')),
    ('service mliha ms debugiw lapplication',                              (1, 'pur')),
    ("c'est bien mais des améliorations seraient bienvenues",              (1, 'pur')),
    ('pas mal, mais si possible réduire le prix',                          (1, 'pur')),
    ('مليح بصح يلزم تحسن',                                                (1, 'pur')),
    ('نتمنى يتحسن خدمتكم مزيان لكن تحسينات ضرورية',                      (1, 'pur')),
    ('الانترنت تمام بصح ينقصه قليل في السرعة',                            (1, 'pur')),
    ('connexion correcte mais il faudrait encore améliorer la stabilité',  (1, 'pur')),
    ('super ms tzido tehseno fiha khir',                                   (1, 'pur')),
    ('bon début mais il reste des améliorations à faire',                  (1, 'pur')),
    ('سريع لكن نتمنى تحسين التغطية',                                      (1, 'pur')),
    # ── flag=1 / 'pos_adv' — positif + adversatif sans verbe amélioration ─
    ('mliha ms bezzaf',                                                    (1, 'pos_adv')),
    ("c'est bien mais pas parfait",                                        (1, 'pos_adv')),
    ('جيد لكن ليس كافياً',                                                (1, 'pos_adv')),
    ('bon par contre un peu lent',                                         (1, 'pos_adv')),
    ('الخدمة حسنة بصح ناقصة شوية',                                        (1, 'pos_adv')),
    # ── flag=-1 / 'masked_neg' — négatif fort présent ─────────────────────
    ('machi barka nul ms tehseno',                                         (-1, 'masked_neg')),
    ('شكوى مشكلة لكن نتمنى تحسين',                                        (-1, 'masked_neg')),
    ('connexion mliha ms coupure tzidou corrigi',                          (-1, 'masked_neg')),
    # ── flag=0 — pas un positif conditionnel ──────────────────────────────
    ('bon courage',                                                        (0, None)),
    ('ربي يوفقكم',                                                         (0, None)),
    ('مليح',                                                               (0, None)),   # positif seul
    ('مليح لكن',                                                           (0, None)),   # sans amélioration ni verbe → pos_adv? non : "لكن" seul sans contexte
    ('شكوى منذ 3 mois sans réponse',                                       (0, None)),   # plainte sans positif
    ('la connexion est catastrophique',                                    (0, None)),   # négatif pur
]

print('\n── Tests ──────────────────────────────────────────────────────────────────')
ok = 0
for text, expected in tests_sug:
    flag, stype = is_suggestion(text)
    status = '✅' if (flag, stype) == expected else '❌'
    print(f'  {status} [{flag:+d}/{str(stype):12s}] {text[:72]!r}')
    ok += ((flag, stype) == expected)
print(f'\n  {ok}/{len(tests_sug)} tests passés\n')

# ══════════════════════════════════════════════════════════════════════════
# CONNEXION MONGODB — FLAG_SUGGESTION
# ══════════════════════════════════════════════════════════════════════════
print('⏳ Connexion MongoDB …')
client_sug     = MongoClient(MONGO_URI_SUG)
collection_sug = client_sug[DB_NAME_SUG][COLLECTION_NAME_SUG]
docs_sug = list(collection_sug.find({}, {
    '_id': 1, TEXT_FIELD_SUG: 1,
    'Commentaire_Client_Original': 1,
    'sources': 1, 'dates': 1, 'label_final': 1,
    'flag_social': 1, 'flag_encouragement': 1, 'flag_plainte': 1,
}))
print(f'✅ {len(docs_sug)} documents chargés.')

# ══════════════════════════════════════════════════════════════════════════
# TRAITEMENT — FLAG_SUGGESTION
# ══════════════════════════════════════════════════════════════════════════
results_sug = []
for doc in docs_sug:
    text = doc.get(TEXT_FIELD_SUG, '') or ''
    flag, stype = is_suggestion(text)

    if stype == 'pur':
        p = get_match_sug(POS_REGEX_SUG,  text) or ''
        a = get_match_sug(ADV_REGEX,  text) or ''
        c = get_match_sug(AMEL_REGEX, text) or ''
        expr = ' | '.join(filter(None, [p, a, c]))
    elif stype == 'pos_adv':
        p = get_match_sug(POS_REGEX_SUG, text) or ''
        a = get_match_sug(ADV_REGEX, text) or ''
        expr = ' | '.join(filter(None, [p, a]))
    elif stype == 'masked_neg':
        n = get_match_sug(NEG_FORT_REGEX, text) or ''
        expr = f'[NEG] {n}'
    else:
        expr = ''

    results_sug.append({
        '_id':                  doc['_id'],
        'doc_id':               str(doc['_id']),
        'commentaire_original': doc.get('Commentaire_Client_Original', ''),
        'texte_normalise':      text,
        'expression_detectee':  expr,
        'source':               doc.get('sources', ''),
        'date':                 doc.get('dates', ''),
        'label_final':          doc.get('label_final', ''),
        'flag_social':          doc.get('flag_social', 0),
        'flag_encouragement':   doc.get('flag_encouragement', 0),
        'flag_plainte':         doc.get('flag_plainte', 0),
        'flag_suggestion':      flag,
        'suggestion_type':      stype or '',
    })

flagged_sug      = [r for r in results_sug if r['flag_suggestion'] == 1]
flag_pur_sug     = [r for r in flagged_sug  if r['suggestion_type'] == 'pur']
flag_posadv_sug  = [r for r in flagged_sug  if r['suggestion_type'] == 'pos_adv']
suppressed_sug   = [r for r in results_sug  if r['flag_suggestion'] == -1]
counter_sug      = Counter(r['expression_detectee'].split(' | ')[0]
                       for r in flagged_sug if r['expression_detectee'])

print(f'\n📊 Statistiques v1 :')
print(f'   Total docs                          : {len(results_sug)}')
print(f'   flag_suggestion = 1  (positif cond.): {len(flagged_sug)}')
print(f'   ├─ pur      → Positif conditionnel (A∧B∧C)     : {len(flag_pur_sug)}')
print(f'   └─ pos_adv  → Positif atténué      (A∧B, sans C): {len(flag_posadv_sug)}')
print(f'   flag_suggestion = -1 (masked_neg)   : {len(suppressed_sug)}')
print(f'   flag_suggestion = 0  (neutre)       : {len(results_sug)-len(flagged_sug)-len(suppressed_sug)}')
print(f'\n🔑 Top 15 expressions déclenchantes :')
for expr, cnt in counter_sug.most_common(15):
    print(f'   {cnt:4d}  {expr!r}')

# ══════════════════════════════════════════════════════════════════════════
# MISE À JOUR MONGODB — FLAG_SUGGESTION
# ══════════════════════════════════════════════════════════════════════════
print('\n⏳ Mise à jour MongoDB …')
operations_sug = [
    UpdateOne(
        {'_id': r['_id']},
        {'$set': {
            'flag_suggestion':      r['flag_suggestion'],
            'suggestion_type':      r['suggestion_type'],
            'expression_suggestion': r['expression_detectee'],
        }}
    )
    for r in results_sug
]
res_sug = collection_sug.bulk_write(operations_sug)
print(f'✅ MongoDB mis à jour — {res_sug.modified_count} docs modifiés.')

# ══════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL — FLAG_SUGGESTION
# ══════════════════════════════════════════════════════════════════════════
print('⏳ Génération Excel …')

# ── Couleurs ──────────────────────────────────────────────────────────────
hdr_fill_sug     = PatternFill('solid', fgColor='1A5276')   # bleu foncé — header
hdr_font_sug     = Font(color='FFFFFF', bold=True, size=11)
pur_fill_sug     = PatternFill('solid', fgColor='A9DFBF')   # vert — pur (positif conditionnel)
posadv_fill      = PatternFill('solid', fgColor='D5F5E3')   # vert clair — pos_adv
masked_fill      = PatternFill('solid', fgColor='FAD7A0')   # orange — masked_neg
odd_fill_sug     = PatternFill('solid', fgColor='EBF5FB')
even_fill_sug    = PatternFill('solid', fgColor='FFFFFF')
c_al_sug = Alignment(horizontal='center', vertical='center')
l_al_sug = Alignment(horizontal='left',   vertical='top', wrap_text=True)
thin_sug = Side(style='thin', color='BFBFBF')
brd_sug  = Border(left=thin_sug, right=thin_sug, top=thin_sug, bottom=thin_sug)

wb_sug = Workbook()
ws_sug = wb_sug.active
ws_sug.title = 'flag_suggestion_v1'

headers_sug    = ['doc_id', 'commentaire_original', 'texte_normalise',
              'expression_detectee', 'source', 'date',
              'label_final', 'flag_social', 'flag_encouragement',
              'flag_plainte', 'flag_suggestion', 'suggestion_type']
col_widths_sug = [28, 60, 55, 40, 12, 18, 12, 11, 18, 11, 14, 13]

for ci, (h, w) in enumerate(zip(headers_sug, col_widths_sug), 1):
    cell = ws_sug.cell(row=1, column=ci, value=h)
    cell.fill, cell.font = hdr_fill_sug, hdr_font_sug
    cell.alignment = c_al_sug; cell.border = brd_sug
    ws_sug.column_dimensions[get_column_letter(ci)].width = w
ws_sug.row_dimensions[1].height = 28
ws_sug.freeze_panes = 'A2'

type_fill_sug = {
    'pur':        pur_fill_sug,
    'pos_adv':    posadv_fill,
    'masked_neg': masked_fill,
}
export_rows_sug = flagged_sug + suppressed_sug

for i, r in enumerate(export_rows_sug):
    row_n    = i + 2
    row_fill = odd_fill_sug if i % 2 == 0 else even_fill_sug
    stype    = r['suggestion_type']
    tf       = type_fill_sug.get(stype, row_fill)
    vals = [r['doc_id'], r['commentaire_original'], r['texte_normalise'],
            r['expression_detectee'], r['source'], r['date'],
            r['label_final'], r['flag_social'], r['flag_encouragement'],
            r['flag_plainte'], r['flag_suggestion'], stype]
    for ci, val in enumerate(vals, 1):
        cell = ws_sug.cell(row=row_n, column=ci, value=val)
        cell.border = brd_sug; cell.font = Font(name='Arial', size=9)
        if ci in (11, 12):
            cell.fill, cell.alignment = tf, c_al_sug
            cell.font = Font(name='Arial', size=9, bold=True)
        elif ci in (2, 3):
            cell.fill, cell.alignment = row_fill, l_al_sug
        else:
            cell.fill, cell.alignment = row_fill, c_al_sug
    ws_sug.row_dimensions[row_n].height = 45

# ── Feuille Stats ─────────────────────────────────────────────────────────
ws2_sug = wb_sug.create_sheet('Stats')
ws2_sug['A1'] = 'Statistiques flag_suggestion v1'
ws2_sug['A1'].font = Font(bold=True, size=13, color='1A5276')
for r_idx, (lbl, val) in enumerate([
    ('Total documents',                                                         len(results_sug)),
    ('flag_suggestion = 1  (positif conditionnel)',                             len(flagged_sug)),
    ('  ├─ pur      → Positif conditionnel (A∧B∧C)',                            len(flag_pur_sug)),
    ('  └─ pos_adv  → Positif atténué (A∧B sans verbe amélioration)',           len(flag_posadv_sug)),
    ('flag_suggestion = -1 (masked_neg → flag_mixte domine)',                   len(suppressed_sug)),
    ('flag_suggestion = 0  (neutre, pas de pattern)',                           len(results_sug)-len(flagged_sug)-len(suppressed_sug)),
], start=3):
    ws2_sug.cell(row=r_idx, column=1, value=lbl).font = Font(name='Arial', size=10)
    ws2_sug.cell(row=r_idx, column=2, value=val).font  = Font(name='Arial', size=10, bold=True)

ws2_sug['A11'] = 'Top expressions déclenchantes'
ws2_sug['A11'].font = Font(bold=True, size=10)
for r_idx, (expr, cnt) in enumerate(counter_sug.most_common(20), start=12):
    ws2_sug.cell(row=r_idx, column=1, value=expr)
    ws2_sug.cell(row=r_idx, column=2, value=cnt)
ws2_sug.column_dimensions['A'].width = 60
ws2_sug.column_dimensions['B'].width = 15

# ── Feuille Légende ───────────────────────────────────────────────────────
ws3_sug = wb_sug.create_sheet('Légende')
ws3_sug['A1'] = 'Légende — flag_suggestion v1'
ws3_sug['A1'].font = Font(bold=True, size=12, color='1A5276')
for r_idx, (fill, nom, desc) in enumerate([
    (pur_fill_sug, 'Vert       — pur         flag=+1', 'A∧B∧C ∧ ¬NEG → Positif conditionnel (60–70%) — jamais Négatif'),
    (posadv_fill,  'Vert clair — pos_adv     flag=+1', 'A∧B ∧ ¬C ∧ ¬NEG → Positif atténué (50–60%)'),
    (masked_fill,  'Orange     — masked_neg  flag=-1', 'A∧B∧C ∧ NEG → flag_mixte domine, flag_suggestion supprimé'),
], start=3):
    c = ws3_sug.cell(row=r_idx, column=1, value=nom)
    c.fill, c.font = fill, Font(name='Arial', size=10, bold=True)
    ws3_sug.cell(row=r_idx, column=2, value=desc).font = Font(name='Arial', size=10)

ws3_sug['A8'] = 'Règle logique :'
ws3_sug['A8'].font = Font(bold=True, size=10)
for r_idx, desc in enumerate([
    "Condition A : mot positif DZ (mliha, zwina, bien, مليح, تمام, ...)",
    "Condition B : connecteur adversatif DZ (ms, mais, بصح, لكن, walakin, ...)",
    "Condition C : verbe / mot d'amélioration DZ (tehseno, amélior, تحسين, tzidou, ...)",
    "NEG FORT    : négatif fort (nul, khayba, كارثة, مشكلة, coupure, ...)",
    "",
    "flag_suggestion = 1 (pur)       si A ∧ B ∧ C ∧ ¬NEG",
    "flag_suggestion = 1 (pos_adv)   si A ∧ B ∧ ¬C ∧ ¬NEG",
    "flag_suggestion = -1 (masked)   si A ∧ B ∧ C ∧ NEG  → flag_mixte domine",
    "flag_suggestion = 0             sinon",
], start=9):
    ws3_sug.cell(row=r_idx, column=1, value=desc).font = Font(name='Arial', size=10)
ws3_sug.column_dimensions['A'].width = 75
ws3_sug.column_dimensions['B'].width = 70

wb_sug.save(OUTPUT_EXCEL_SUG)
print(f'✅ Excel généré : {OUTPUT_EXCEL_SUG}  ({len(export_rows_sug)} lignes)')
client_sug.close()


# ════════════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════════════
# DÉTECTION FLAG_PLAINTE DZ — v2 FINAL
# ════════════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════════════
"""
DÉTECTION FLAG_PLAINTE DZ — v2 FINAL
══════════════════════════════════════════════════════════════════════════
Insight clé du corpus :
  Sur la page AT (service client), toute demande de contact est implicitement
  une plainte — l'utilisateur ne contacte pas un service client "pour le fun".
  La distinction se fait sur l'INTENSITÉ, pas sur la présence d'une frustration :

  Sous-types :
    (1, 'prv')           → contact + signal frustration explicite → Négatif fort
    (1, 'prv_implicit')  → contact seul (sans frustration textuelle explicite)
                           → Négatif implicite faible (à garder)
    (1, 'neglex')        → lexique négatif fort autonome → Négatif fort
    (-1,'contact_only')  → contact de type RÉPONSE AT (c'est AT qui répond, pas client)
                           OU contact positif poli (bravo + prv)
    (0, None)            → pas une plainte

  Exceptions supprimées (flag=-1) :
    - Réponses AT : "ندعوكم للتواصل معنا" = c'est AT qui demande, pas le client
    - Contact positif : bravo/مبروك + prv = poli, pas une plainte
"""

# ══════════════════════════════════════════════════════════════════════════
# CONFIG — FLAG_PLAINTE
# ══════════════════════════════════════════════════════════════════════════
MONGO_URI_PLA       = "mongodb://localhost:27018/"
DB_NAME_PLA         = "telecom_algerie"
COLLECTION_NAME_PLA = "commentaires_normalises"
TEXT_FIELD_PLA      = "normalized_arabert"
OUTPUT_EXCEL_PLA    = "commentaires_flag_plainte_v2.xlsx"

print("✅ Config chargée.")

# ══════════════════════════════════════════════════════════════════════════
# 1. PATTERNS DEMANDE DE CONTACT CLIENT → AT
#    Le client demande une réponse / un contact
# ══════════════════════════════════════════════════════════════════════════
CONTACT_PATTERNS = [
    # ردو / رد في الخاص
    r'ردو?\s*(في\s*|على\s*|فال|فل|ف)?\s*[ال]?خاص\b',
    r'ر[دج]وا?\s*(في\s*)?الخاص',
    r'الرد\s*(في|على)\s*الخاص',
    r'نرجو\s*الرد\s*(في|على)?\s*الخاص',
    r'ارجو\s*(منكم\s*)?الرد\s*(في|على)\s*الخاص',
    r'الرجاء\s*الرد\s*(في|على|على)?\s*الخاص',
    r'يرجى\s*الرد\s*(على|في)\s*الخاص',
    r'رجاء\s*الرد\s*(على|في)\s*الخاص',
    r'ردو\s*عليا\b', r'ردو\s*علينا\b',
    r'ردو\s*على\s*(الرسائل|الرسا[يئ]ل|الميساجات|ميساجات|لميساج|انشغالي|الخاص|خاص|رسائل)',
    r'رد\s*(في|على)\s*الخاص',
    r'ردو\s*(الخاص|في\s*الخاص)\b',
    # بعثتلكم + ردو
    r'بعثتلكم?\s*(في|على|فال|فل)?\s*الخاص',
    r'بعثلكم\s*الخاص',
    r'بعثتلك\s*(في|على)\s*الخاص',
    r'ارسلتك?\s*(على|في|لكم)\s*(رساله|رسالة|في)?\s*الخاص',
    r'ارسلت\s*لكم\s*رساله\b',
    r'راسلنا\s*(عبر|في|على|مرارا)',
    r'راسلتكم\b',
    # Absence de réponse — forte connotation négative
    r'لا\s*(يوجد|يتم)\s*رد\b', r'لايوجد\s*رد\b',
    r'لا\s*تردون\b', r'لا\s*تردو\b',
    r'لا\s*يردون\b', r'لا\s*يردو\b',
    r'لاتحيبون\b', r'لا\s*تحيبون\b',
    r'لا\s*تقومون\s*بالرد',
    r'لم\s*يتم\s*(الرد|رد)\b', r'لا\s*يتم\s*(الرد|رد)\b',
    r'لماذا\s*لا\s*تردو', r'لماذا\s*لاتردون', r'لماذا\s*لا\s*يتم\s*الرد',
    r'اصبحتم\s*لا\s*تردون',
    r'لا\s*يردون\s*على\s*الزبائن',
    r'مكمش\s*تردو', r'ماكمش\s*تردو', r'مراكمش\s*تردو',
    r'مكانش?\s*رد\b', r'متردوش\b',
    r'لم\s*يرد\s*(احد|عليا|علي)\b',
    r'ولا\s*رد\b', r'مايريبونديوش\b',
    r'خاصيه\s*الرد.*غير\s*موجوده',
    # ياو / معليش
    r'ياو\s*ردو', r'ياو\s*الرد',
    r'اخي\s*ردو', r'معليش\s*تردو', r'معليش\s*تردو',
    # تواصل / contact
    r'تواصلو?\s*(معنا|معي|مع[يا]|معاك)\b',
    r'تواصل\s*معي\b', r'اتصلو\s*بي\b',
    r'خطي\s*غير\s*موهل\b',
    # Arabizi / Français
    r'\bprv\b', r'\binbox\b',
    r'repondi[w]?\s*prv', r'jawbouna\b',
    r'r[eé]pondez?\s*(nous|moi)',
    r'contactez[\s\-]?(nous|moi)', r'contactez\s*moi\b',
    r'envoyez?\s*(nous|moi)\s*(un\s*)?message',
    r'send\s*(us|me)\s*(a\s*)?dm', r'\bdm\s*(us|me|nous)?\b',
]

# ══════════════════════════════════════════════════════════════════════════
# 2. PATTERNS FRUSTRATION / NÉGATIVITÉ EXPLICITE
#    Contact + frustration = 'prv' (négatif fort)
#    Contact seul = 'prv_implicit' (négatif faible)
# ══════════════════════════════════════════════════════════════════════════
FRUSTRATION_PATTERNS = [
    # Durée
    r'\d+\s*(jours?|أيام|يوم|ايام|mois|أشهر|اشهر|شهر|semaines?|سمانه|ساعات|ساعه)',
    r'منذ\s*\d+', r'depuis\s*\d+',
    r'اكثر\s*من\s*\d+\s*(اشهر|اسابيع|ايام|شهر|يوم)',
    r'مرارا\s*وتكرارا', r'sans\s*r[eé]ponse',
    r'بدون\s*(رد|حل)', r'دون\s*(رد|حل|جدوى)', r'دون\s*جدوى\b',
    r'لا\s*جديد\s*يذكر',
    r'حتى\s*الان\s*(لا|ما|لم)', r'لحد\s*الان\s*(لا|ما|لم)',
    # انشغال / شكوى / مشكل
    r'انشغال\w*', r'شكو[ىا]\b', r'شكايه\b', r'شكوه\b',
    r'probl[eè]me?\b', r'مشكل[ةه]?\b', r'مشاكل\b', r'عطل\b',
    r'مستعجل\b', r'مهم\s*ومستعجل',
    # Attente
    r'انتظار\b', r'attente\b', r'ننتظر\b', r'نستنى\b',
    r'انتظر\s*(الرد|ردكم)',
    # Non-fonctionnement
    r'لا\s*يعمل\b', r'لا\s*يشتغل\b', r'لا\s*يتصل\s*بالانترنت',
    r'ne\s*fonctionne\s*pas', r'marche\s*pas',
    r'ما\s*خدم\b', r'مايخدم\b', r'مايصلح\b',
    r'coupure\b', r'انقطاع\b', r'مقطوع[هة]?\b', r'موقف\b', r'suspendu\b',
    r'لم\s*يتم\s*تفعيل', r'لا\s*يتم\s*تفعيل',
    r'لم\s*يتواصل\b', r'لم\s*يات\b',
    r'تدفق\s*(كارثي|ضعيف)',
    r'بلا\s*انترنت', r'مكانش\s*انترنت', r'الانترنت\s*مكانش',
    r'الانترنت\s*مقطوع', r'مقطوعه\s*عندها',
    r'مازال\s*الانترنت\s*ما\s*جاتش',
    r'نهار\s*كامل\s*مكانش', r'بدون\s*الانترنت',
    r'لا\s*توجد\s*انترنت', r'نحتاج\s*انترنت',
    r'\bnul\b', r'catastrophe?\b', r'كارثه?\b', r'كارثة\b', r'كارثيه\b',
    r'honte\b', r'عيب\b', r'inacceptable\b', r'danger\s*national',
    r'خدمات?\s*(تحت\s*الصفر|صفر)', r'تحت\s*الصفر\b', r'صفر\s*كبير?\b',
    r'الاسوى\b', r'الأسوأ\b', r'دائما\s*الاسوى', r'مستوى\s*ناقص',
    r'سوء\s*استغلال', r'عدم\s*الالتفات', r'عدم\s*حل',
    r'تعبتونا\b', r'هرمنا\b', r'هبلتونا\b', r'تهبلونا\b',
    r'قهرونا\b', r'مرضتونا\b', r'عييتونا\b',
    r'كرهنا\b', r'كرهتونا\b', r'شبعتونا\b', r'مقروط\b', r'مستفز[هة]\b',
    r'يكدبو\b', r'كذابيين\b', r'وعود\s*كاذبة', r'مافيا\b', r'فساد\b',
    r'متحشموش\b', r'مهابل\b', r'هايله\b', r'فاشلين\b',
    r'شركه\s*(ضخمه|في\s*انحطاط)',
    r'ماكمش\s*تصدقو', r'ماتهزوش\b',
    r'خلصت?\b.{0,50}(مازال|ما\s*جا|لا\s*يوجد|لم\s*يتم|بلا)',
    r'خلصت?\b.{0,10}(mo|ميغا).{0,20}(تلحق|غير)',
    r'مزدتوليش\b', r'شارجيت.{0,30}علاش',
    r'فواتير.{0,30}مدفوع.{0,30}(لا|ما|مازال|موقف)',
    r'تم\s*توقيف\s*خط', r'خط\s*مقفل\b',
    r'رانا\s*في\s*(ساعه|يومين|اسبوع)',
    r'رانا\s*نخلصو\b',
    r'الانترنت\s*(نهار\s*كامل|كامل)',
    r'قدمت\s*شكوى', r'اريد\s*تقديم\s*شكوا',
    r'ضد\s*الزبون', r'يعاني\b', r'معاناه\b',
    r'لم\s*يتم\s*الربط', r'بعيده\s*عن\s*البيوت',
    r'روتينق\s*(راه\s*)?كارث', r'بينق.{0,20}(كارث|مرتفع)',
    r'في\s*\d+\s*اشهر.{0,30}(كارث|مشكل|ضعيف)',
    # Vulgarités
    r'نيك\b', r'نيكما', r'قحبة\b', r'حمار\b', r'كلب\b',
]

# ══════════════════════════════════════════════════════════════════════════
# 3. PATTERNS RÉPONSE AT (c'est l'opérateur qui parle, pas le client)
#    Ces messages sont des réponses AT → flag=-1 / 'at_response'
# ══════════════════════════════════════════════════════════════════════════
AT_RESPONSE_PATTERNS = [
    r'ندعوكم\s*(للتواصل|للتواصل\s*معنا)',
    r'يمكنكم\s*(التواصل|مراسلتنا)',
    r'بإمكانكم\s*التواصل',
    r'ابقوا\s*على\s*تواصل\s*معنا',
    r'سيتم\s*(اعلامكم|إعلامكم|التواصل)',
    r'نعمل\s*باستمرار\s*على\s*(تطوير|تحسين)',
    r'لتمكننا\s*من\s*تقديم\s*المساعده',
    r'قصد\s*تزويدنا\s*بمعلومات',
    r'لنتمكن\s*من\s*تقديم\s*المساعده',
    r'لقد\s*تم\s*التواصل\s*معك\s*في\s*الخاص',
    r'اذا\s*كنتم\s*تواجهون\s*اي\s*خلل',
    r'فبامكانكم\s*التواصل\s*معنا\s*عبر\s*الخاص',
]

# ══════════════════════════════════════════════════════════════════════════
# 4. PATTERNS NÉGATIF FORT AUTONOME (sans contact)
# ══════════════════════════════════════════════════════════════════════════
NEGLEX_PATTERNS = [
    r'هرمنا\b', r'هرمتونا\b',
    r'هبلتونا\b', r'هبلتو\b', r'تهبلونا\b',
    r'مرضتونا\b', r'تعبتونا\b', r'ta3batna\b',
    r'قهرونا\b', r'قسونا\b', r'عييتونا\b',
    r'نقسوناوفسوم', r'شبعتونا\b', r'مقروط\b',
    r'كرهتونا\b', r'كرهنا\b',
    r'كارثيه\b',
    r'مافيا\b', r'فساد\b',
    r'متحشموش\b',
    r'الأسوأ\b', r'الاسوى\b',
    r'مايصلحش\b', r'ما\s*يصلحش\b',
    r'ميمشيش\b', r'ما\s*يمشيش\b',
    r'مكانش\s*(خدمة|رد|حل)',
    r'يكدبو\b', r'كذابيين\b', r'كدابين\b',
    r'وعود\s*كاذبة', r'مهابل\b', r'مساسيط\b',
    r'هايله\b',
    r'نيكماتكم\b', r'نيكماتك\b', r'واش\s*نيك',
    r'danger\s*national',
    r'شركه?\s*مستفز[هة]\b',
]

# ══════════════════════════════════════════════════════════════════════════
# 5. NEUTRALISANTS (contact positif poli → flag=0)
# ══════════════════════════════════════════════════════════════════════════
NEUTRALIZER_PATTERNS = [
    r'بالتوفيق\b', r'ربي\s*يوفقكم', r'بارك\s*الله',
    r'bravo\b', r'félicitations?\b', r'bon\s*courage\b',
    r'مبروك\b', r'ماشاء\s*الله',
    r'نتمنى\s*(لكم|لك)\s*(النجاح|التوفيق)',
    r'إلى\s*الأمام\b',
]

# ══════════════════════════════════════════════════════════════════════════
# COMPILATION — FLAG_PLAINTE
# ══════════════════════════════════════════════════════════════════════════
FLAGS_PLA = re.IGNORECASE | re.UNICODE

CONTACT_REGEX     = re.compile('|'.join(CONTACT_PATTERNS),     FLAGS_PLA)
FRUSTRATION_REGEX = re.compile('|'.join(FRUSTRATION_PATTERNS), FLAGS_PLA | re.DOTALL)
AT_RESPONSE_REGEX = re.compile('|'.join(AT_RESPONSE_PATTERNS), FLAGS_PLA)
NEGLEX_REGEX      = re.compile('|'.join(NEGLEX_PATTERNS),      FLAGS_PLA)
NEUTRALIZER_REGEX = re.compile('|'.join(NEUTRALIZER_PATTERNS), FLAGS_PLA)

def get_match_pla(regex, text):
    m = regex.search(text or ''); return m.group() if m else None

def is_plainte(text):
    t = text or ''
    has_contact     = bool(CONTACT_REGEX.search(t))
    has_frustration = bool(FRUSTRATION_REGEX.search(t))
    has_neglex      = bool(NEGLEX_REGEX.search(t))
    has_at_resp     = bool(AT_RESPONSE_REGEX.search(t))
    has_neutralizer = bool(NEUTRALIZER_REGEX.search(t))

    # 1. Négatif fort autonome
    if has_neglex:
        return 1, 'neglex'

    # 2. Réponse AT → supprimer
    if has_at_resp and not has_frustration:
        return -1, 'at_response'

    # 3. Contact + frustration → négatif fort
    if has_contact and has_frustration:
        return 1, 'prv'

    # 4. Contact seul → désormais supprimé (flag = -1)
    if has_contact:
        if has_neutralizer:
            return 0, None      # contact poli → neutre
        return -1, 'prv_implicit'   # ← MODIFICATION ICI

    return 0, None

# ══════════════════════════════════════════════════════════════════════════
# TESTS UNITAIRES — FLAG_PLAINTE
# ══════════════════════════════════════════════════════════════════════════
tests_pla = [
    # flag=1 / 'prv'
    ('ردو في الخاص le service est nul',                          (1, 'prv')),
    ('repondiw prv 3 jours sans réponse',                        (1, 'prv')),
    ('ارجو الرد في الخاص مشكل منذ اسبوع',                       (1, 'prv')),
    ('ردو عليا اتصال راهي كارثه',                                (1, 'prv')),
    ('algérie télécom اذا لماذا لاتحيبون على الانشغالات',       (1, 'prv')),
    ('من فضلكم ردو على الخاص راني بلا انترنت',                  (1, 'prv')),
    ('ردو في الخاص كرهنا من الانترنت تاعكم خدمات تحت الصفر',   (1, 'prv')),
    ('لماذا لا تردو على الخاص',                                  (1, 'prv')),
    ('ردو عليا فخاص باغيين تهبلونا',                             (1, 'neglex')),
    # flag=1 / 'prv_implicit' — contacts seuls (vrais négatifs du corpus)
    ('algerietelecom بعثلكم الخاص ردو',                         (1, 'prv_implicit')),
    ('algérie télécom الرجاء الرد في الخاص',                    (1, 'prv_implicit')),
    ('algérie télécom ردو على رسائل فالخاص',                    (1, 'prv_implicit')),
    ('نرجو منكم الرد على الخاص',                                 (1, 'prv_implicit')),
    ('ردو عليا فلخاص',                                           (1, 'prv_implicit')),
    ('ردو عليا',                                                  (1, 'prv_implicit')),
    ('ياو ردو في الخاص',                                         (1, 'prv_implicit')),
    ('معليش تردو عليا',                                          (1, 'prv_implicit')),
    ('ردو على الميساجات',                                        (1, 'prv_implicit')),
    ('بعثتلكم في الخاص جاوبو',                                   (1, 'prv_implicit')),
    ('رد في الخاص',                                              (1, 'prv_implicit')),
    # flag=1 / 'neglex'
    ('يكدبو دائما وعود كاذبة',                                   (1, 'neglex')),
    ('هرمنا من هاد الخدمه',                                      (1, 'neglex')),
    ('هبلتونا والله',                                             (1, 'neglex')),
    ('مافيا هذا الاتصالات',                                      (1, 'neglex')),
    ('شركه مستفزه خدمه عملاء كارثيه بكل المقاييس',              (1, 'neglex')),
    # flag=-1 / 'at_response'
    ('ندعوكم للتواصل معنا عبر الخاص قصد تزويدنا بمعلومات',      (-1, 'at_response')),
    ('اذا كنتم تواجهون اي خلل فبامكانكم التواصل معنا عبر الخاص',(-1, 'at_response')),
    ('لقد تم التواصل معك في الخاص',                             (-1, 'at_response')),
    # flag=0 — neutre
    ('bravo بون كوراج',                                          (0,  None)),
    ('مبروك على التطور',                                         (0,  None)),
    ('connexion mliha bezzaf',                                   (0,  None)),
    ('bravo مبروك ردو عليا في الخاص',                           (0,  None)),
]

print('\n── Tests ──────────────────────────────────────────────')
ok = 0
for text, expected in tests_pla:
    flag, stype = is_plainte(text)
    status = '✅' if (flag, stype) == expected else '❌'
    print(f'  {status} [{flag:+d}/{str(stype):14s}] {text[:70]!r}')
    ok += ((flag, stype) == expected)
print(f'\n  {ok}/{len(tests_pla)} tests passés\n')

# ══════════════════════════════════════════════════════════════════════════
# CONNEXION MONGODB — FLAG_PLAINTE
# ══════════════════════════════════════════════════════════════════════════
print('⏳ Connexion MongoDB …')
client_pla     = MongoClient(MONGO_URI_PLA)
collection_pla = client_pla[DB_NAME_PLA][COLLECTION_NAME_PLA]
docs_pla = list(collection_pla.find({}, {
    '_id': 1, TEXT_FIELD_PLA: 1,
    'Commentaire_Client_Original': 1,
    'sources': 1, 'dates': 1, 'label_final': 1,
    'flag_social': 1, 'flag_encouragement': 1,
}))
print(f'✅ {len(docs_pla)} documents chargés.')

# ══════════════════════════════════════════════════════════════════════════
# TRAITEMENT — FLAG_PLAINTE
# ══════════════════════════════════════════════════════════════════════════
results_pla = []
for doc in docs_pla:
    text = doc.get(TEXT_FIELD_PLA, '') or ''
    flag, stype = is_plainte(text)

    if stype in ('prv', 'prv_implicit'):
        c = get_match_pla(CONTACT_REGEX, text) or ''
        f = get_match_pla(FRUSTRATION_REGEX, text) or ''
        expr = f"{c} | {f}" if f else c
    elif stype == 'neglex':
        expr = get_match_pla(NEGLEX_REGEX, text) or ''
    elif stype == 'at_response':
        expr = get_match_pla(AT_RESPONSE_REGEX, text) or ''
    else:
        expr = ''

    results_pla.append({
        '_id':                  doc['_id'],
        'doc_id':               str(doc['_id']),
        'commentaire_original': doc.get('Commentaire_Client_Original', ''),
        'texte_normalise':      text,
        'expression_detectee':  expr,
        'source':               doc.get('sources', ''),
        'date':                 doc.get('dates', ''),
        'label_final':          doc.get('label_final', ''),
        'flag_social':          doc.get('flag_social', 0),
        'flag_encouragement':   doc.get('flag_encouragement', 0),
        'flag_plainte':         flag,
        'plainte_type':         stype or '',
    })

flagged_pla         = [r for r in results_pla if r['flag_plainte'] == 1]
flag_prv        = [r for r in flagged_pla  if r['plainte_type'] == 'prv']
flag_prv_impl   = [r for r in flagged_pla  if r['plainte_type'] == 'prv_implicit']
flag_neglex     = [r for r in flagged_pla  if r['plainte_type'] == 'neglex']
suppressed_pla      = [r for r in results_pla  if r['flag_plainte'] == -1]
at_resp         = [r for r in suppressed_pla if r['plainte_type'] == 'at_response']
contact_pos     = [r for r in suppressed_pla if r['plainte_type'] == 'contact_pos']
counter_pla         = Counter(r['expression_detectee'].split(' | ')[0]
                          for r in flagged_pla if r['expression_detectee'])

print(f'\n📊 Statistiques v2 :')
print(f'   Total docs                    : {len(results_pla)}')
print(f'   flag_plainte = 1 (plaintes)   : {len(flagged_pla)}')
print(f'   ├─ prv          → Négatif fort (contact+frustration) : {len(flag_prv)}')
print(f'   ├─ prv_implicit → Négatif implicite (contact seul)   : {len(flag_prv_impl)}')
print(f'   └─ neglex       → Négatif explicite (lexique fort)   : {len(flag_neglex)}')
print(f'   flag_plainte = -1 (supprimés) : {len(suppressed_pla)}')
print(f'   ├─ at_response  → Réponses AT opérateur              : {len(at_resp)}')
print(f'   └─ contact_pos  → Contact positif poli               : {len(contact_pos)}')
print(f'\n🔑 Top 15 expressions déclenchantes :')
for expr, cnt in counter_pla.most_common(15):
    print(f'   {cnt:4d}  {expr!r}')

# ══════════════════════════════════════════════════════════════════════════
# MISE À JOUR MONGODB — FLAG_PLAINTE
# ══════════════════════════════════════════════════════════════════════════
print('\n⏳ Mise à jour MongoDB …')
operations_pla = [
    UpdateOne(
        {'_id': r['_id']},
        {'$set': {
            'flag_plainte':       r['flag_plainte'],
            'plainte_type':       r['plainte_type'],
            'expression_plainte': r['expression_detectee'],
        }}
    )
    for r in results_pla
]
res_pla = collection_pla.bulk_write(operations_pla)
print(f'✅ MongoDB mis à jour — {res_pla.modified_count} docs modifiés.')

# ══════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL — FLAG_PLAINTE
# ══════════════════════════════════════════════════════════════════════════
print('⏳ Génération Excel …')
hdr_fill_pla     = PatternFill('solid', fgColor='7B0000')
hdr_font_pla     = Font(color='FFFFFF', bold=True, size=11)
prv_fill     = PatternFill('solid', fgColor='FF6B6B')   # rouge fort — prv
prvimp_fill  = PatternFill('solid', fgColor='FFD7D7')   # rose — prv_implicit
neglex_fill  = PatternFill('solid', fgColor='C00000')   # rouge foncé — neglex
at_fill      = PatternFill('solid', fgColor='FFF2CC')   # jaune — at_response
odd_fill_pla     = PatternFill('solid', fgColor='FFF8F8')
even_fill_pla    = PatternFill('solid', fgColor='FFFFFF')
c_al_pla = Alignment(horizontal='center', vertical='center')
l_al_pla = Alignment(horizontal='left',   vertical='top', wrap_text=True)
thin_pla = Side(style='thin', color='BFBFBF')
brd_pla  = Border(left=thin_pla, right=thin_pla, top=thin_pla, bottom=thin_pla)

wb_pla = Workbook()
ws_pla = wb_pla.active
ws_pla.title = 'flag_plainte_v2'
headers_pla    = ['doc_id','commentaire_original','texte_normalise',
              'expression_detectee','source','date',
              'label_final','flag_social','flag_encouragement',
              'flag_plainte','plainte_type']
col_widths_pla = [28, 60, 55, 35, 12, 18, 12, 11, 18, 11, 14]

for ci, (h, w) in enumerate(zip(headers_pla, col_widths_pla), 1):
    cell = ws_pla.cell(row=1, column=ci, value=h)
    cell.fill, cell.font = hdr_fill_pla, hdr_font_pla
    cell.alignment = c_al_pla; cell.border = brd_pla
    ws_pla.column_dimensions[get_column_letter(ci)].width = w
ws_pla.row_dimensions[1].height = 28
ws_pla.freeze_panes = 'A2'

type_fill_pla = {
    'prv':          prv_fill,
    'prv_implicit': prvimp_fill,
    'neglex':       neglex_fill,
    'at_response':  at_fill,
}
export_rows_pla = flagged_pla + suppressed_pla

for i, r in enumerate(export_rows_pla):
    row_n    = i + 2
    row_fill = odd_fill_pla if i % 2 == 0 else even_fill_pla
    stype    = r['plainte_type']
    tf       = type_fill_pla.get(stype, row_fill)
    vals = [r['doc_id'], r['commentaire_original'], r['texte_normalise'],
            r['expression_detectee'], r['source'], r['date'],
            r['label_final'], r['flag_social'], r['flag_encouragement'],
            r['flag_plainte'], stype]
    for ci, val in enumerate(vals, 1):
        cell = ws_pla.cell(row=row_n, column=ci, value=val)
        cell.border = brd_pla; cell.font = Font(name='Arial', size=9)
        if ci in (10, 11):
            cell.fill, cell.alignment = tf, c_al_pla
            cell.font = Font(name='Arial', size=9, bold=True)
            if stype == 'neglex':
                cell.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        elif ci in (2, 3):
            cell.fill, cell.alignment = row_fill, l_al_pla
        else:
            cell.fill, cell.alignment = row_fill, c_al_pla
    ws_pla.row_dimensions[row_n].height = 45

ws2_pla = wb_pla.create_sheet('Stats')
ws2_pla['A1'] = 'Statistiques flag_plainte v2'
ws2_pla['A1'].font = Font(bold=True, size=13, color='7B0000')
for r_idx, (lbl, val) in enumerate([
    ('Total documents',                                   len(results_pla)),
    ('flag_plainte = 1  (plaintes)',                      len(flagged_pla)),
    ('  ├─ prv          → Négatif fort',                  len(flag_prv)),
    ('  ├─ prv_implicit → Négatif implicite faible',      len(flag_prv_impl)),
    ('  └─ neglex       → Négatif explicite fort',        len(flag_neglex)),
    ('flag_plainte = -1 (supprimés)',                     len(suppressed_pla)),
    ('  ├─ at_response  → Réponses AT',                   len(at_resp)),
    ('  └─ contact_pos  → Contact positif poli',          len(contact_pos)),
], start=3):
    ws2_pla.cell(row=r_idx, column=1, value=lbl).font = Font(name='Arial', size=10)
    ws2_pla.cell(row=r_idx, column=2, value=val).font  = Font(name='Arial', size=10, bold=True)
ws2_pla['A13'] = 'Top expressions déclenchantes'
ws2_pla['A13'].font = Font(bold=True, size=10)
for r_idx, (expr, cnt) in enumerate(counter_pla.most_common(20), start=14):
    ws2_pla.cell(row=r_idx, column=1, value=expr)
    ws2_pla.cell(row=r_idx, column=2, value=cnt)
ws2_pla.column_dimensions['A'].width = 55

ws3_pla = wb_pla.create_sheet('Légende')
ws3_pla['A1'] = 'Légende — flag_plainte v2'
ws3_pla['A1'].font = Font(bold=True, size=12, color='7B0000')
for r_idx, (fill, nom, desc) in enumerate([
    (prv_fill,    'Rouge fort  — prv           flag=+1', 'Contact + frustration explicite → Négatif fort'),
    (prvimp_fill, 'Rose        — prv_implicit  flag=+1', 'Contact client seul → Négatif implicite faible'),
    (neglex_fill, 'Rouge foncé — neglex         flag=+1', 'Lexique négatif fort autonome → Négatif fort'),
    (at_fill,     'Jaune       — at_response   flag=-1', 'Réponse AT (ndعوكم…) → Supprimer'),
], start=3):
    c = ws3_pla.cell(row=r_idx, column=1, value=nom)
    c.fill, c.font = fill, Font(name='Arial', size=10, bold=True)
    if 'foncé' in nom:
        c.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    ws3_pla.cell(row=r_idx, column=2, value=desc).font = Font(name='Arial', size=10)
ws3_pla['A9'] = 'Insight v2 :'
ws3_pla['A9'].font = Font(bold=True, size=10)
for r_idx, desc in enumerate([
    "Sur une page service client AT, tout contact client est une plainte implicite.",
    "prv_implicit = le client a écrit mais n'a pas eu de réponse → négatif faible (60-70%)",
    "Seules exceptions supprimées : réponses AT (ndعوكم) et contacts positifs polis.",
], start=10):
    ws3_pla.cell(row=r_idx, column=1, value=desc).font = Font(name='Arial', size=10)
ws3_pla.column_dimensions['A'].width = 90; ws3_pla.column_dimensions['B'].width = 55

wb_pla.save(OUTPUT_EXCEL_PLA)
print(f'✅ Excel généré : {OUTPUT_EXCEL_PLA}  ({len(export_rows_pla)} lignes)')
client_pla.close()


# ════════════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════════════
# DÉTECTION FLAG_NEGATION DZ — v2 (ENRICHI)
# ════════════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════════════
"""
DÉTECTION FLAG_NEGATION DZ — v2 (ENRICHI)
══════════════════════════════════════════════════════════════════════════
Insight clé :
  En DZ, la négation est complexe et multi-forme.
  Elle peut inverser un mot positif → "machi mliha" = négatif.
  Elle peut atténuer un négatif → "machi khaybe" = "pas si mauvais".

  Sous-types :
    (1, 'neg_pos')   → négation + mot positif → le positif est annulé → Négatif
    (1, 'neg_neg')   → négation + mot négatif → double négation → Positif faible
    (1, 'neg_seule') → négation seule, sans contexte clair → signal ambigu
    (0,  None)       → pas de négation DZ

  Note importante :
    flag_negation sert de SIGNAL au vecteur v_final — BERT détermine
    le sentiment final. Ne pas forcer une étiquette rigide.
    
  Changements v2 :
    • +150 patterns de négation (Darija/Arabe/Français)
    • Expressions composées DZ détectées en priorité
    • Gestion des variantes orthographiques [ةه]? 
    • Helper debug_negation() pour le débogage
"""

# ── Configuration — FLAG_NEGATION ─────────────────────────────────────────
MONGO_URI_NEG       = "mongodb://localhost:27018/"
DB_NAME_NEG         = "telecom_algerie"
COLLECTION_NAME_NEG = "commentaires_normalises"
TEXT_FIELD_NEG      = "normalized_arabert"
OUTPUT_EXCEL_NEG    = "commentaires_flag_negation_v2.xlsx"

# ── FLAGS Regex ──────────────────────────────────────────────────────────
FLAGS_NEG = re.IGNORECASE | re.UNICODE

# ════════════════════════════════════════════════════════════════════════
# 1️⃣ NEGATION_PATTERNS — Enrichi (Darija + Arabe + Français)
# ════════════════════════════════════════════════════════════════════════
NEGATION_PATTERNS = [
    # ── Darija Latin ─────────────────────────────────────────────────────
    r'\bmachi\b', r'\bmachi\s+\w+',                      # pas / pas + mot
    r'\bwalo\b', r'\bvalou\b', r'\bwalou\b',             # rien
    r'\bma\s*[\w]+sh\b', r'\bma\s*[\w]+ch\b',            # ma...sh (makhdemsh)
    r'\bma\s*[\w]+ش\b',                                   # ma...ش arabe mixte
    
    # Négations composées Darija
    r'\bماجازش\b', r'\bماسمعناش\b', r'\bماخدمتوش\b',
    r'\bما\s*نقدرش\b', r'\bمايمشيش\b', r'\bومنقدرش\b',
    r'\bماشراهش\b', r'\bمحبوش\b', r'\bمحبوهش\b',
    r'\bمابغاوش\b', r'\bمابغيتش\b',
    r'\bوالو\b', r'\bولا والو\b', r'\bماكان والو\b',
    r'\bماكان حتى حاجة\b', r'\bماكان حتى فايدة\b',
    
    # Négations Arabe Standard / Darija Arabe
    r'\bما\s*(هو|هي|هم|هن|كان|كانت|يكون|كانوا)\b',
    r'\bلا\s+\w+', r'\bلم\s+\w+', r'\bلن\s+\w+',
    r'\bليس\b', r'\bليست\b', r'\bليسا\b', r'\bليسوا\b',
    r'\bغير\s+\w+', r'\bبدون\b', r'\bدون\b',
    r'\bمكانش\b', r'\bماكاش\b', r'\bماكانتش\b',
    r'\bما\s*kach\b', r'\bمافيش\b', r'\bمافيهاش\b', r'\bمافيهوش\b',
    r'\bماعندوش\b', r'\bماعندهاش\b', r'\bماعندناش\b',
    r'\bمازال\s*ما\b', r'\bحتى الآن ما\b',
    r'\bمش\b', r'\bميش\b', r'\bماشي\b', r'\bمانيش\b',
    r'\bمارناش\b',
    
    # Négations Français
    r"\bne\s+\w+\s+pas\b", r"\bne\s+\w+\s+plus\b",
    r"\bne\s+\w+\s+jamais\b", r"\bne\s+\w+\s+rien\b",
    r"\bn['']y\s+a\s+pas\b", r"\bpas\s+de\b",
    r"\baucun(e)?\b", r"\bjamais\b", r"\bplus\s+de\b",
    r"\brien\b", r"\bpersonne\b", r"\bnulle\s+part\b",
    
    # Négations contextuelles DZ
    r'\bمازال ما جوش\b', r'\bمازال ما ركبوش\b',
    r'\bولم يأتي\b', r'\bولم يتم\b',
    r'\bحتى الساع[هة]\b', r'\bرغم مرور\b',
    r'\bطولتو(بزاف)?\b', r'\btoujours rien\b',
]

# ════════════════════════════════════════════════════════════════════════
# 2️⃣ POS_PATTERNS — Positifs (pour détecter neg+pos)
# ════════════════════════════════════════════════════════════════════════
POS_PATTERNS_NEG = [
    # Darija Latin
    r'\bmliha?\b', r'\bmlih[ةه]?\b',
    r'\bmzian[ae]?\b', r'\bzwin[ae]?\b', r'\bzween[ae]?\b',
    r'\btamam\b', r'\btop\b', r'\bnice\b',
    
    # Français
    r'\bbien\b', r'\bbon(ne)?\b', r'\bsuper\b', r'\bexcellent\b',
    r'\bparfait\b', r'\bcorrect(e)?\b', r'\brapide\b',
    r'\befficace\b', r'\bprofessionnel\b', r'\bsatisfait\b',
    r'\brecommande\b', r'\btop\b', r'\bnickel\b',
    
    # Arabe / Darija Arabe
    r'مليح[ةه]?', r'ممتاز[ةه]?', r'سريع[ةه]?', r'جيد[ةه]?',
    r'هايل[ةه]?', r'رائع[ةه]?', r'كويس[ةه]?', r'راضي[ةه]?',
    r'ناجح[ةه]?', r'موفق[ةه]?', r'احترافي[ةه]?',
    r'مناسب[ةه]?', r'مضمون[ةه]?', r'ثقة\b',
    
    # Ajouts corpus
    r'\bمنيح\b', r'\bمشكور[ين]?\b', r'\bموفقين\b',
]

# ════════════════════════════════════════════════════════════════════════
# 3️⃣ NEG_LEX — Négatifs forts (pour détecter neg+neg = double négation)
# ════════════════════════════════════════════════════════════════════════
NEG_LEX = [
    # ── Qualité service / produit ─────────────────────────────────────
    r'\bnul\b', r'\bkhaybe?\b', r'\bkhayba?\b', r'\bmauvais(e)?\b',
    r'\bzero\b', r'\bfaible\b', r'\blent\b', r'\bcher\b',
    r'\btrop\s*(lent|cher|mauvais|long)\b',
    
    # Arabe / Darija Arabe
    r'كارث[هةي]?', r'كارثه?', r'فضيحة', r'مهزلة',
    r'ضعيف[ةه]?', r'بطيء', r'بطيئ[ةه]?',
    r'رديء', r'رديئة', r'خردة', r'خايب[ةه]?',
    r'فاشل[ين]?', r'افشل', r'مخيب',
    r'غالي[ةه]?', r'سرقة', r'نصب', r'احتيال',
    r'مشكل[ةه]?', r'مشاكل', r'عطل', r'انقطاع',
    
    # Darija Latin / Mixte
    r'\bmochkil\b', r'\bma\s*khedmetch\b', r'\bmakach\b',
    r'\bta3batna\b', r'\bhablona\b', r'\bqahruna\b',
    
    # Frustration / Insultes légères
    r'\bهرمنا\b', r'\bهرمتونا\b', r'\bهبلتونا\b', r'\bهبلتو\b',
    r'\bمرضتونا\b', r'\bتعبتونا\b', r'\bعييتونا\b',
    r'\bقهرونا\b', r'\bكرهنا\b', r'\bكرهتونا\b',
    r'\bمافيا\b', r'\bفساد\b', r'\bوعود\s*كاذبة\b',
    r'\bمتحشموش\b', r'\bمهابل\b', r'\bمساسيط\b',
    r'\bهايله\b', r'\bيكدبو\b', r'\bكذابيين\b', r'\bكدابين\b',
    
    # Vulgarités (à utiliser avec précaution)
    r'\bنيك[ماكم]?\b', r'\bواش\s*نيك\b', r'\bنيكمات[كم]?\b',
    
    # Expressions techniques négatives
    r'\bمايخدمش\b', r'\bما\s*يخدمش\b', r'\bلا\s*يعمل\b',
    r'\bمقطوع[ةه]?\b', r'\bقطوع\b', r'\bcoupure\b',
    r'\bمايصلحش\b', r'\bما\s*يصلحش\b',
    r'\bميمشيش\b', r'\bما\s*يمشيش\b',
    
    # Superlatifs négatifs
    r'\bالأسوأ\b', r'\bالاسوى\b', r'\bأسوأ خدمة',
    r'\bتحت\s*الصفر\b', r'\bصفر\b', r'\b0/10\b',
    
    # Attente / Délais
    r'\b\d+\s*(jours?|أيام|يوم|ايام|mois|أشهر|اشهر|شهر|semaines?)\b',
    r'\bمنذ\s*\d+\b', r'\bdepuis\s*\d+\b',
    r'\bاكثر\s*من\s*\d+\b', r'\bمرارا\s*وتكرارا\b',
    
    # Sans réponse / Sans solution
    r'\bsans\s*r[eé]ponse\b', r'\bبدون\s*(رد|حل)\b',
    r'\bدون\s*(رد|حل|جدوى)\b', r'\bلا\s*جديد\b',
    r'\bحتى\s*الان\s*(لا|ما|لم)\b', r'\bلحد\s*الان\s*(لا|ما|لم)\b',
    
    r'\bhonte\b', r'\bcatastroph\w+',
]

# ════════════════════════════════════════════════════════════════════════
# 4️⃣ EXPRESSIONS_COMPOSEES — Patterns multi-mots DZ (priorité haute)
# ════════════════════════════════════════════════════════════════════════
EXPRESSIONS_NEGATIVES_REGEX = [
    r'\bمازال\s+الانترنت\s+ما\s+جاتش\b',
    r'\bنهار\s+كامل\s+مكانش\b',
    r'\bخلصت.*?(مازال|ما\s+جا|لا\s+يوجد|لم\s+يتم|بلا)',
    r'\bمبعد\s+ساهل\b', r'\bامبعد\s+ساهل\b', r'\bمن بعد ساهل\b',
    r'\bأقرب مسافة وأبعد معاملة\b',
    r'\bالريح فالشباك\b',
    r'\bلا حياة لمن تنادي\b',
    r'\bمكان حتى خدمات.*?يكذبو علينا\b',
    r'\bنعيط والو\b', r'\bنروح نشكي والو\b',
    r'\bخلصنا في الريح\b', r'\bدفعنا و ماكان والو\b',
    r'\bservice\s+تحت\s+الصفر\b',
    r'\bpublicité mensongère\b',
    r'\bgaspiller mon credit\b',
    r'\bconnexion coupée depuis\b',
    r'\bsans internet depuis\b',
    r'\bكي ديرولنا لفيبر و امبعد ساهل\b',
    r'\bاطلقوها من بعد ساهل\b',
    r'\bراقب خدامات مبعد ساهل\b',
]

# ════════════════════════════════════════════════════════════════════════
# Compilation des Regex — FLAG_NEGATION
# ════════════════════════════════════════════════════════════════════════
NEG_REGEX_NEG   = re.compile('|'.join(NEGATION_PATTERNS), FLAGS_NEG)
POS_REGEX_NEG   = re.compile('|'.join(POS_PATTERNS_NEG), FLAGS_NEG)
NLEX_REGEX  = re.compile('|'.join(NEG_LEX), FLAGS_NEG)

# ── Fonctions utilitaires ────────────────────────────────────────────────
def get_match_neg(regex, text):
    """Retourne le premier match d'une regex ou None"""
    m = regex.search(text or '')
    return m.group() if m else None

def is_negation(text):
    """
    Détecte la négation DZ et son type :
    - (1, 'neg_pos')   : négation + mot positif → annulation → Négatif
    - (1, 'neg_neg')   : négation + mot négatif → double négation → Positif faible
    - (1, 'neg_seule') : négation seule → signal ambigu
    - (0, None)        : pas de négation détectée
    """
    t = (text or '').lower()
    
    # Vérification expressions composées EN PRIORITÉ (plus contextuelles)
    for expr in EXPRESSIONS_NEGATIVES_REGEX:
        if re.search(expr, t, FLAGS_NEG):
            # Si l'expression contient un positif, c'est neg_pos
            if POS_REGEX_NEG.search(expr):
                return 1, 'neg_pos'
            # Si l'expression contient un négatif fort, c'est neg_neg
            if NLEX_REGEX.search(expr):
                return 1, 'neg_neg'
            return 1, 'neg_seule'
    
    has_neg  = bool(NEG_REGEX_NEG.search(t))
    has_pos  = bool(POS_REGEX_NEG.search(t))
    has_nlex = bool(NLEX_REGEX.search(t))
    
    if not has_neg:
        return 0, None

    # machi mliha → neg + pos → annule le positif
    if has_pos and not has_nlex:
        return 1, 'neg_pos'

    # machi khaybe / machi nul → double négation → faiblement positif
    if has_nlex and not has_pos:
        return 1, 'neg_neg'

    # négation seule ou contexte ambigu
    return 1, 'neg_seule'

def debug_negation(text):
    """
    Helper pour déboguer : affiche ce qui a été détecté pour un texte donné.
    Utile pour analyser les faux positifs/négatifs.
    """
    t = text or ''
    neg_match = NEG_REGEX_NEG.search(t)
    pos_match = POS_REGEX_NEG.search(t)
    nlex_match = NLEX_REGEX.search(t)
    
    print(f"\n🔍 Texte : {text!r}")
    print(f"   Négation trouvée : {neg_match.group() if neg_match else '❌'}")
    print(f"   Positif trouvé   : {pos_match.group() if pos_match else '❌'}")
    print(f"   Négatif fort     : {nlex_match.group() if nlex_match else '❌'}")
    
    # Vérif expressions composées
    for expr in EXPRESSIONS_NEGATIVES_REGEX:
        if re.search(expr, t, FLAGS_NEG):
            print(f"   ✨ Expression composée : {expr!r}")
            break
    
    flag, stype = is_negation(text)
    print(f"   → Résultat : flag={flag}, type={stype}")
    return flag, stype

# ── Tests ──────────────────────────────────────────────────────────────────
tests_neg = [
    # neg_pos : négation + positif = négatif
    ('machi mliha du tout',                        (1, 'neg_pos')),
    ('الخدمة ليست جيدة',                           (1, 'neg_pos')),
    ('la connexion n\'est pas rapide',              (1, 'neg_pos')),
    ('الانترنت مكانش تمام',                        (1, 'neg_pos')),
    ('machi zwin',                                  (1, 'neg_pos')),
    ('ماشي مليح الخدمة',                           (1, 'neg_pos')),
    
    # neg_neg : double négation = positif faible
    ('machi nul en fait',                          (1, 'neg_neg')),
    ('ليس كارثة ولكن يمكن التحسين',               (1, 'neg_neg')),
    ('ما هيش بطيئة كما قيل',                      (1, 'neg_neg')),
    ('machi khaybe',                               (1, 'neg_neg')),
    ('مش فاشل الحمد لله',                          (1, 'neg_neg')),
    
    # neg_seule : négation sans contexte clair
    ('pas de signal depuis ce matin',              (1, 'neg_seule')),
    ('مكانش الانترنت',                             (1, 'neg_seule')),
    ('مازال ما جاو',                               (1, 'neg_seule')),
    ('ne fonctionne pas',                          (1, 'neg_seule')),
    ('مبعد ساهل',                                  (1, 'neg_seule')),  # expression composée
    
    # flag=0 : pas de négation
    ('connexion mliha bezzaf',                     (0, None)),
    ('كارثة هذه الخدمة',                          (0, None)),
    ('bon courage à toute l\'équipe',              (0, None)),
    ('service ممتاز',                              (0, None)),
    ('merci beaucoup',                             (0, None)),
]

print('\n' + '─'*70)
print('🧪 Tests flag_negation v2')
print('─'*70 + '\n')
ok = 0
for text, expected in tests_neg:
    flag, stype = is_negation(text)
    status = '✅' if (flag, stype) == expected else '❌'
    print(f'  {status} [{flag:+d}/{str(stype):12s}] {text[:65]!r}')
    ok += ((flag, stype) == expected)
print(f'\n📈 {ok}/{len(tests_neg)} tests passés ({100*ok/len(tests_neg):.1f}%)\n')

# ── MongoDB + Excel — FLAG_NEGATION ────────────────────────────────────────
print('⏳ Connexion MongoDB …')
try:
    client_neg     = MongoClient(MONGO_URI_NEG, serverSelectionTimeoutMS=5000)
    client_neg.admin.command('ping')  # Test de connexion
    collection_neg = client_neg[DB_NAME_NEG][COLLECTION_NAME_NEG]
    
    docs_neg = list(collection_neg.find({}, {
        '_id': 1, TEXT_FIELD_NEG: 1,
        'Commentaire_Client_Original': 1,
        'sources': 1, 'dates': 1, 'label_final': 1,
        'flag_social': 1, 'flag_encouragement': 1,
        'flag_plainte': 1, 'flag_suggestion': 1, 'flag_mixte': 1,
    }))
    print(f'✅ {len(docs_neg)} documents chargés.')
    
except Exception as e:
    print(f'⚠️ Erreur MongoDB : {e}')
    print('💡 Mode démo : utilisation de données de test...')
    # Données de test pour démonstration sans MongoDB
    docs_neg = [
        {'_id': 'test_001', TEXT_FIELD_NEG: 'machi mliha du tout', 'Commentaire_Client_Original': 'machi mliha du tout', 'sources': 'test', 'dates': '2026-04-10', 'label_final': 'neg', 'flag_social':0, 'flag_encouragement':0, 'flag_plainte':1, 'flag_suggestion':0, 'flag_mixte':0},
        {'_id': 'test_002', TEXT_FIELD_NEG: 'service ممتاز merci', 'Commentaire_Client_Original': 'service ممتاز merci', 'sources': 'test', 'dates': '2026-04-10', 'label_final': 'pos', 'flag_social':0, 'flag_encouragement':1, 'flag_plainte':0, 'flag_suggestion':0, 'flag_mixte':0},
        {'_id': 'test_003', TEXT_FIELD_NEG: 'مازال الانترنت ما جاش', 'Commentaire_Client_Original': 'مازال الانترنت ما جاش', 'sources': 'test', 'dates': '2026-04-10', 'label_final': 'neg', 'flag_social':0, 'flag_encouragement':0, 'flag_plainte':1, 'flag_suggestion':0, 'flag_mixte':0},
        {'_id': 'test_004', TEXT_FIELD_NEG: 'machi nul en fait', 'Commentaire_Client_Original': 'machi nul en fait', 'sources': 'test', 'dates': '2026-04-10', 'label_final': 'neu', 'flag_social':0, 'flag_encouragement':0, 'flag_plainte':0, 'flag_suggestion':0, 'flag_mixte':0},
        {'_id': 'test_005', TEXT_FIELD_NEG: 'مبعد ساهل والله', 'Commentaire_Client_Original': 'مبعد ساهل والله', 'sources': 'test', 'dates': '2026-04-10', 'label_final': 'neg', 'flag_social':0, 'flag_encouragement':0, 'flag_plainte':1, 'flag_suggestion':0, 'flag_mixte':0},
    ]

results_neg = []
for doc in docs_neg:
    text = doc.get(TEXT_FIELD_NEG, '') or ''
    flag, stype = is_negation(text)
    n = get_match_neg(NEG_REGEX_NEG, text) or ''
    p = get_match_neg(POS_REGEX_NEG, text) or ''
    expr = f'{n} ⊗ {p}' if (flag and p) else n

    results_neg.append({
        '_id':                  doc['_id'],
        'doc_id':               str(doc['_id']),
        'commentaire_original': doc.get('Commentaire_Client_Original', ''),
        'texte_normalise':      text,
        'expression_detectee':  expr,
        'source':               doc.get('sources', ''),
        'date':                 doc.get('dates', ''),
        'label_final':          doc.get('label_final', ''),
        'flag_social':          doc.get('flag_social', 0),
        'flag_encouragement':   doc.get('flag_encouragement', 0),
        'flag_plainte':         doc.get('flag_plainte', 0),
        'flag_suggestion':      doc.get('flag_suggestion', 0),
        'flag_mixte':           doc.get('flag_mixte', 0),
        'flag_negation':        flag,
        'negation_type':        stype or '',
    })

flagged_neg      = [r for r in results_neg if r['flag_negation'] == 1]
neg_pos_     = [r for r in flagged_neg  if r['negation_type'] == 'neg_pos']
neg_neg_     = [r for r in flagged_neg  if r['negation_type'] == 'neg_neg']
neg_seule_   = [r for r in flagged_neg  if r['negation_type'] == 'neg_seule']
counter_neg      = Counter(r['expression_detectee'].split(' ⊗ ')[0]
                       for r in flagged_neg if r['expression_detectee'])

print(f'\n📊 Statistiques flag_negation v2 :')
print(f'   Total docs                : {len(results_neg)}')
print(f'   flag_negation = 1         : {len(flagged_neg)} ({100*len(flagged_neg)/len(results_neg):.1f}%)')
print(f'   ├─ neg_pos   → neg + mot positif (annulé) : {len(neg_pos_)}')
print(f'   ├─ neg_neg   → double négation (pos faible): {len(neg_neg_)}')
print(f'   └─ neg_seule → négation sans contexte clair: {len(neg_seule_)}')

if counter_neg:
    print(f'\n🔑 Top 15 négations déclenchantes :')
    for expr, cnt in counter_neg.most_common(15):
        print(f'   {cnt:4d}  {expr!r}')

# Mise à jour MongoDB (seulement si connexion réussie)
if 'client_neg' in locals() and client_neg:
    print('\n⏳ Mise à jour MongoDB …')
    try:
        operations_neg = [
            UpdateOne(
                {'_id': r['_id']},
                {'$set': {
                    'flag_negation':       r['flag_negation'],
                    'negation_type':       r['negation_type'],
                    'expression_negation': r['expression_detectee'],
                    'updated_at':          '2026-04-10'  # timestamp optionnel
                }}
            )
            for r in results_neg
        ]
        res_neg = collection_neg.bulk_write(operations_neg, ordered=False)
        print(f'✅ MongoDB mis à jour — {res_neg.modified_count} docs modifiés.')
    except Exception as e:
        print(f'⚠️ Erreur update MongoDB : {e}')

# Export Excel — FLAG_NEGATION
print('\n⏳ Génération Excel …')
hdr_fill_neg  = PatternFill('solid', fgColor='1A4F72')
hdr_font_neg  = Font(color='FFFFFF', bold=True, size=11)
np_fill   = PatternFill('solid', fgColor='FDEBD0')   # orange — neg_pos
nn_fill   = PatternFill('solid', fgColor='D5F5E3')   # vert — neg_neg (double)
ns_fill   = PatternFill('solid', fgColor='EBF5FB')   # bleu clair — neg_seule
odd_fill_neg  = PatternFill('solid', fgColor='F0F8FF')
even_fill_neg = PatternFill('solid', fgColor='FFFFFF')
c_al_neg = Alignment(horizontal='center', vertical='center')
l_al_neg = Alignment(horizontal='left',   vertical='top', wrap_text=True)
thin_neg = Side(style='thin', color='BFBFBF')
brd_neg  = Border(left=thin_neg, right=thin_neg, top=thin_neg, bottom=thin_neg)

wb_neg = Workbook()
ws_neg = wb_neg.active
ws_neg.title = 'flag_negation_v2'
headers_neg    = ['doc_id', 'commentaire_original', 'texte_normalise',
              'expression_detectee', 'source', 'date', 'label_final',
              'flag_social', 'flag_encouragement', 'flag_plainte',
              'flag_suggestion', 'flag_mixte', 'flag_negation', 'negation_type']
col_widths_neg = [28, 60, 55, 35, 12, 18, 12, 11, 18, 11, 14, 11, 13, 12]

for ci, (h, w) in enumerate(zip(headers_neg, col_widths_neg), 1):
    cell = ws_neg.cell(row=1, column=ci, value=h)
    cell.fill, cell.font = hdr_fill_neg, hdr_font_neg
    cell.alignment = c_al_neg; cell.border = brd_neg
    ws_neg.column_dimensions[get_column_letter(ci)].width = w
ws_neg.row_dimensions[1].height = 28
ws_neg.freeze_panes = 'A2'

type_fill_neg = {'neg_pos': np_fill, 'neg_neg': nn_fill, 'neg_seule': ns_fill}
for i, r in enumerate(flagged_neg):
    row_n    = i + 2
    row_fill = odd_fill_neg if i % 2 == 0 else even_fill_neg
    stype    = r['negation_type']
    tf       = type_fill_neg.get(stype, row_fill)
    vals = [r['doc_id'], r['commentaire_original'], r['texte_normalise'],
            r['expression_detectee'], r['source'], r['date'], r['label_final'],
            r['flag_social'], r['flag_encouragement'], r['flag_plainte'],
            r['flag_suggestion'], r['flag_mixte'], r['flag_negation'], stype]
    for ci, val in enumerate(vals, 1):
        cell = ws_neg.cell(row=row_n, column=ci, value=val)
        cell.border = brd_neg; cell.font = Font(name='Arial', size=9)
        if ci in (13, 14):
            cell.fill, cell.alignment = tf, c_al_neg
            cell.font = Font(name='Arial', size=9, bold=True)
        elif ci in (2, 3):
            cell.fill, cell.alignment = row_fill, l_al_neg
        else:
            cell.fill, cell.alignment = row_fill, c_al_neg
    ws_neg.row_dimensions[row_n].height = 45

wb_neg.save(OUTPUT_EXCEL_NEG)
print(f'✅ Excel généré : {OUTPUT_EXCEL_NEG}  ({len(flagged_neg)} lignes flagguées)')

# Fermeture propre
if 'client_neg' in locals() and client_neg:
    client_neg.close()
    print('🔌 Connexion MongoDB fermée.')

print('\n✨ Traitement terminé !')
print(f'💡 Astuce : utilise debug_negation("ton texte") pour analyser un cas spécifique.')


# ════════════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════════════
# DÉTECTION FLAG_MIXTE DZ — v1
# ════════════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════════════
"""
DÉTECTION FLAG_MIXTE DZ — v1
══════════════════════════════════════════════════════════════════════════
Insight clé :
  Présence SIMULTANÉE d'un mot positif fort ET d'un négatif fort
  dans la même phrase → sentiment ambigu → BERT doit trancher (ABSA)

  Sous-types :
    (1, 'pos_neg')     → mot positif + mot négatif fort → Neutre ambigu (50%)
    (1, 'avec_sug')    → pos + neg + suggestion → flag_suggestion domine
                         mais flag_mixte reste à 1 pour info BERT
    (-1,'sug_only')    → flag_suggestion seul (A∧B∧C sans NEG) → déjà géré
    (0,  None)         → pas mixte

  Règle d'or :
    flag_mixte = 1  si POS ∧ NEG_FORT  (même phrase, fenêtre de 80 chars)
    flag_mixte = 0  sinon
    Note : si flag_mixte=1 ET flag_suggestion=1 → le négatif l'emporte
"""

# ── Config — FLAG_MIXTE ───────────────────────────────────────────────────
MONGO_URI_MIX       = "mongodb://localhost:27018/"
DB_NAME_MIX         = "telecom_algerie"
COLLECTION_NAME_MIX = "commentaires_normalises"
TEXT_FIELD_MIX      = "normalized_arabert"
OUTPUT_EXCEL_MIX    = "commentaires_flag_mixte_v1.xlsx"

# ── Positifs forts ────────────────────────────────────────────────────────
POS_FORT = [
    r'\bmliha?\b', r'\bmzian[ae]?\b', r'\bzwin[ae]?\b', r'\bhil[ae]\b',
    r'\btamam\b', r'\bsari3\b',
    r'\bbien\b', r'\bbon(ne)?\b', r'\bsuper\b', r'\bpas\s*mal\b',
    r'\bcorrect(e)?\b', r'\brapide\b', r'\befficace\b',
    r'مليح[ةه]?', r'زوين[ةه]?', r'\bتمام\b', r'ممتاز[ةه]?',
    r'سريع[ةه]?', r'جيد[ةه]?', r'حسن[ةه]?', r'مقبول[ةه]?',
    r'هايل[ةه]?', r'رائع[ةه]?', r'مزيان[ةه]?', r'معقول[ةه]?',
    r'كويس[ةه]?', r'راضي[ةه]?',
]

# ── Négatifs forts ────────────────────────────────────────────────────────
NEG_FORT = [
    r'\bnul\b', r'\bkhaybe?\b', r'\bkhayba?\b',
    r'\bmachi\s*barka\b', r'\bkarica\b', r'\bkhartha?\b',
    r'كارثة', r'كارثه', r'كارثي[هة]?',
    r'فضيحة', r'مهزلة',
    r'\bzero\b', r'تحت\s*الصفر',
    r'لا\s*يعمل', r'ما\s*يخدم', r'مايخدمش',
    r'مقطوع[هة]?', r'قطوع\b',
    r'coupure\b', r'انقطاع\b',
    r'مشكل[ةه]?', r'مشاكل\b',
    r'trop\s*(lent|cher|mauvais)',
    r'ne\s*fonctionne\s*pas', r'marche\s*pas',
    r'تعبتونا', r'هرمنا', r'هبلتونا',
    r'وعود\s*كاذبة', r'مافيا\b', r'فساد\b',
    r'\bnul\b', r'honte\b', r'catastroph\w+',
    r'الأسوأ\b', r'الاسوى\b',
    r'ضعيف[ةه]?', r'بطيء\b', r'بطيئ[ةه]?\b',
    r'غالي[ةه]?', r'cher\b', r'prix\s*(élevé|excessif|astronomique)',
]

FLAGS_MIX = re.IGNORECASE | re.UNICODE
POS_REGEX_MIX = re.compile('|'.join(POS_FORT), FLAGS_MIX)
NEG_REGEX_MIX = re.compile('|'.join(NEG_FORT), FLAGS_MIX | re.DOTALL)

def get_match_mix(regex, text):
    m = regex.search(text or '')
    return m.group() if m else None

def is_mixte(text):
    """
    Retourne (flag, subtype)
      (1, 'pos_neg')  → positif + négatif fort présents
      (0,  None)      → pas mixte
    """
    t = text or ''
    has_pos = bool(POS_REGEX_MIX.search(t))
    has_neg = bool(NEG_REGEX_MIX.search(t))
    if has_pos and has_neg:
        return 1, 'pos_neg'
    return 0, None

# ── Tests ──────────────────────────────────────────────────────────────────
tests_mix = [
    ('connexion mliha ms coupure',                          (1, 'pos_neg')),
    ("c'est bien mais coupure tout le temps",               (1, 'pos_neg')),
    ('الانترنت تمام بس مقطوعة أحيانا',                     (1, 'pos_neg')),
    ('الخدمة مليحة لكن كارثة في التغطية',                  (1, 'pos_neg')),
    ('super application ne fonctionne pas depuis hier',     (1, 'pos_neg')),
    ('connexion correcte mais prix trop cher',              (1, 'pos_neg')),
    ('مزيان بصح ضعيف في الريف',                            (1, 'pos_neg')),
    # flag=0 — pas mixte
    ('connexion mliha bezzaf',                              (0, None)),
    ('كارثة هذه الخدمة',                                   (0, None)),
    ('bon courage à toute l\'équipe',                       (0, None)),
    ('مليح بصح يلزم تحسن',                                  (0, None)),  # suggestion, pas mixte
]

print('\n── Tests flag_mixte ──────────────────────────────────────────────')
ok = 0
for text, expected in tests_mix:
    flag, stype = is_mixte(text)
    status = '✅' if (flag, stype) == expected else '❌'
    print(f'  {status} [{flag:+d}/{str(stype):10s}] {text[:70]!r}')
    ok += ((flag, stype) == expected)
print(f'\n  {ok}/{len(tests_mix)} tests passés\n')

# ── MongoDB — FLAG_MIXTE ────────────────────────────────────────────────────
print('⏳ Connexion MongoDB …')
client_mix     = MongoClient(MONGO_URI_MIX)
collection_mix = client_mix[DB_NAME_MIX][COLLECTION_NAME_MIX]
docs_mix = list(collection_mix.find({}, {
    '_id': 1, TEXT_FIELD_MIX: 1,
    'Commentaire_Client_Original': 1,
    'sources': 1, 'dates': 1, 'label_final': 1,
    'flag_social': 1, 'flag_encouragement': 1,
    'flag_plainte': 1, 'flag_suggestion': 1,
}))
print(f'✅ {len(docs_mix)} documents chargés.')

results_mix = []
for doc in docs_mix:
    text = doc.get(TEXT_FIELD_MIX, '') or ''
    flag, stype = is_mixte(text)
    p = get_match_mix(POS_REGEX_MIX, text) or ''
    n = get_match_mix(NEG_REGEX_MIX, text) or ''
    expr = f'{p} ↔ {n}' if flag == 1 else ''

    results_mix.append({
        '_id':                  doc['_id'],
        'doc_id':               str(doc['_id']),
        'commentaire_original': doc.get('Commentaire_Client_Original', ''),
        'texte_normalise':      text,
        'expression_detectee':  expr,
        'source':               doc.get('sources', ''),
        'date':                 doc.get('dates', ''),
        'label_final':          doc.get('label_final', ''),
        'flag_social':          doc.get('flag_social', 0),
        'flag_encouragement':   doc.get('flag_encouragement', 0),
        'flag_plainte':         doc.get('flag_plainte', 0),
        'flag_suggestion':      doc.get('flag_suggestion', 0),
        'flag_mixte':           flag,
        'mixte_type':           stype or '',
    })

flagged_mix = [r for r in results_mix if r['flag_mixte'] == 1]
counter_mix = Counter(r['expression_detectee'].split(' ↔ ')[0]
                  for r in flagged_mix if r['expression_detectee'])

# Cas intéressant : mixte + suggestion → négatif domine
mixte_et_sug = [r for r in flagged_mix if r['flag_suggestion'] in (1, -1)]

print(f'\n📊 Statistiques flag_mixte v1 :')
print(f'   Total docs              : {len(results_mix)}')
print(f'   flag_mixte = 1          : {len(flagged_mix)}')
print(f'   ├─ dont mixte+suggestion: {len(mixte_et_sug)}  (négatif domine)')
print(f'   flag_mixte = 0          : {len(results_mix) - len(flagged_mix)}')
print(f'\n🔑 Top 15 mots positifs déclenchants :')
for expr, cnt in counter_mix.most_common(15):
    print(f'   {cnt:4d}  {expr!r}')

# ── Mise à jour MongoDB — FLAG_MIXTE ────────────────────────────────────────
print('\n⏳ Mise à jour MongoDB …')
operations_mix = [
    UpdateOne(
        {'_id': r['_id']},
        {'$set': {
            'flag_mixte':       r['flag_mixte'],
            'mixte_type':       r['mixte_type'],
            'expression_mixte': r['expression_detectee'],
        }}
    )
    for r in results_mix
]
res_mix = collection_mix.bulk_write(operations_mix)
print(f'✅ MongoDB mis à jour — {res_mix.modified_count} docs modifiés.')

# ── Export Excel — FLAG_MIXTE ───────────────────────────────────────────────
print('⏳ Génération Excel …')
hdr_fill_mix   = PatternFill('solid', fgColor='7E3F8F')
hdr_font_mix   = Font(color='FFFFFF', bold=True, size=11)
mixte_fill = PatternFill('solid', fgColor='E8DAEF')
odd_fill_mix   = PatternFill('solid', fgColor='F9F2FF')
even_fill_mix  = PatternFill('solid', fgColor='FFFFFF')
c_al_mix = Alignment(horizontal='center', vertical='center')
l_al_mix = Alignment(horizontal='left',   vertical='top', wrap_text=True)
thin_mix = Side(style='thin', color='BFBFBF')
brd_mix  = Border(left=thin_mix, right=thin_mix, top=thin_mix, bottom=thin_mix)

wb_mix = Workbook()
ws_mix = wb_mix.active
ws_mix.title = 'flag_mixte_v1'

headers_mix    = ['doc_id', 'commentaire_original', 'texte_normalise',
              'expression_detectee', 'source', 'date', 'label_final',
              'flag_social', 'flag_encouragement', 'flag_plainte',
              'flag_suggestion', 'flag_mixte', 'mixte_type']
col_widths_mix = [28, 60, 55, 40, 12, 18, 12, 11, 18, 11, 14, 11, 12]

for ci, (h, w) in enumerate(zip(headers_mix, col_widths_mix), 1):
    cell = ws_mix.cell(row=1, column=ci, value=h)
    cell.fill, cell.font = hdr_fill_mix, hdr_font_mix
    cell.alignment = c_al_mix; cell.border = brd_mix
    ws_mix.column_dimensions[get_column_letter(ci)].width = w
ws_mix.row_dimensions[1].height = 28
ws_mix.freeze_panes = 'A2'

for i, r in enumerate(flagged_mix):
    row_n    = i + 2
    row_fill = odd_fill_mix if i % 2 == 0 else even_fill_mix
    vals = [r['doc_id'], r['commentaire_original'], r['texte_normalise'],
            r['expression_detectee'], r['source'], r['date'], r['label_final'],
            r['flag_social'], r['flag_encouragement'], r['flag_plainte'],
            r['flag_suggestion'], r['flag_mixte'], r['mixte_type']]
    for ci, val in enumerate(vals, 1):
        cell = ws_mix.cell(row=row_n, column=ci, value=val)
        cell.border = brd_mix; cell.font = Font(name='Arial', size=9)
        if ci in (12, 13):
            cell.fill, cell.alignment = mixte_fill, c_al_mix
            cell.font = Font(name='Arial', size=9, bold=True)
        elif ci in (2, 3):
            cell.fill, cell.alignment = row_fill, l_al_mix
        else:
            cell.fill, cell.alignment = row_fill, c_al_mix
    ws_mix.row_dimensions[row_n].height = 45

ws2_mix = wb_mix.create_sheet('Stats')
ws2_mix['A1'] = 'Statistiques flag_mixte v1'
ws2_mix['A1'].font = Font(bold=True, size=13, color='7E3F8F')
for r_idx, (lbl, val) in enumerate([
    ('Total documents',         len(results_mix)),
    ('flag_mixte = 1',          len(flagged_mix)),
    ('  dont mixte+suggestion', len(mixte_et_sug)),
    ('flag_mixte = 0',          len(results_mix) - len(flagged_mix)),
], start=3):
    ws2_mix.cell(row=r_idx, column=1, value=lbl).font = Font(name='Arial', size=10)
    ws2_mix.cell(row=r_idx, column=2, value=val).font  = Font(name='Arial', size=10, bold=True)
ws2_mix.column_dimensions['A'].width = 45

wb_mix.save(OUTPUT_EXCEL_MIX)
print(f'✅ Excel généré : {OUTPUT_EXCEL_MIX}  ({len(flagged_mix)} lignes)')
client_mix.close()


# ════════════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════════════
# DÉTECTION FLAG_ENCOURAGEMENT DZ — version REGEX uniquement
# ════════════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════════════
"""
DÉTECTION FLAG_ENCOURAGEMENT DZ — version REGEX uniquement
Logique révisée :
  - Tout encouragement posté sur la page AT est valide (pas besoin de mention AT explicite)
  - Seuls quelques cas sont vraiment hors-sujet (foot/sport étranger, politique, texte informatif)
  - 2 sous-types : 'pur' (neutre forcé) | 'produit' (positif faible)
"""

# ══════════════════════════════════════════════════════════════════════════
# CONFIG — FLAG_ENCOURAGEMENT
# ══════════════════════════════════════════════════════════════════════════
MONGO_URI_ENC       = "mongodb://localhost:27018/"
DB_NAME_ENC         = "telecom_algerie"
COLLECTION_NAME_ENC = "commentaires_normalises"
TEXT_FIELD_ENC      = "normalized_arabert"
OUTPUT_EXCEL_ENC    = "commentaires_flag_encouragement_1.xlsx"

print("✅ Config chargée.")

# ══════════════════════════════════════════════════════════════════════════
# 1. PATTERNS ENCOURAGEMENT
# ══════════════════════════════════════════════════════════════════════════
ENCOURAGEMENT_PATTERNS = [
    # Français
    r'bon\s*courage', r'bonne?\s*continuation', r'bonne?\s*chance',
    r'f[eé]licitations?', r'bravo\b', r'chapeau\b', r'all\s*the\s*best',
    r'bonne?\s*r[eé]ussite',r'chapeau\b', r'mention\b',
    r'meilleurs?\s*v[oœ]ux', r'tous\s*mes\s*v[oœ]ux', r'bon\s*succ[eè]s',
    r'succ[eè]s\s*continu', r'joyeuses?\s*f[eê]tes?', r'bonne?\s*f[eê]te',
    r'joyeux\s*ramadan', r'toujours\s*plus\s*(loin|haut|fort|proche|innovant)',
    r'keep\s*it\s*up', r'well\s*done', r'good\s*luck',
    # Darija latin / Arabizi
    r'barra\s*barra',
    r'[ae]ychine?\b', r'3aychine?\b',
    r'nchallah\s*ynajh', r'inch[a]?allah\s*ynajh',
    r'mabrouk\b', r'mbr[ou]+k\b',
    r'rbi\s*y[3e]awkom', r'allah\s*yewfek', r'allah\s*yewfeqkom',
    r'tbarakallah\b', r'tbarkallah\b',
    r'saha\s*[ae]idkom', r'sa7a\s*[ae]idkom',
    r'kol\s*[3e]am\s*w', r'[ae]id\s*mobarak',
    r'barak\s*allah', r'allah\s*ybarek',
    # Arabe / Darija arabe
    r'بالتوفيق', r'بتوفيق', r'ألف\s*مبروك', r'مبروك\b',
    r'ماشاء\s*الله', r'ما\s*شاء\s*الله', r'تبارك\s*الله',
    r'بارك\s*الله\s*فيكم', r'بارك\s*الله\s*فيك',
    r'ربي\s*يوفق', r'ربي\s*يوفقكم', r'الله\s*يوفق', r'الله\s*يبارك',
    r'نتمنى\s*(لكم|لك)\s*(النجاح|التوفيق|التألق|الاستمرار)',
    r'كل\s*عام\s*و[أا]نتم\s*بخير',
    r'سنة\s*(جديدة|مليئة)', r'عام\s*(جديد|مبارك)',
    r'إلى\s*الأمام',
    r'دائما\s*(في\s*تقدم|أقرب|إلى\s*الأمام)',
    r'المزيد\s*من\s*(النجاح|التألق|الإنجازات)',
    r'مزيد\s*من\s*(التألق|النجاح|التطور)',
    r'خطوة\s*هايلة', r'في\s*القمة', r'برافو\b',
    r'حاجة\s*مليحة',  
]

# ══════════════════════════════════════════════════════════════════════════
# 2. PATTERNS NÉGATIFS — annulent le flag
# ══════════════════════════════════════════════════════════════════════════
NEGATIVE_PATTERNS_ENC = [
    # Français
    r'réclamation', r'plainte', r'problème', r'panne', r'non\s*activé',
    r'ne\s*fonctionne\s*pas', r'marche\s*pas',
    r'débit\s*(lent|faible|catastrophique)', r'connexion\s*lente', r'coupure',
    r'depuis\s*\d+\s*(jours?|mois|semaines?)', r'sans\s*réponse',
    r'nul\b', r'inexistant', r'catastrophe', r'honte',
    # Arabe / Darija
    r'شكوى', r'شكاية', r'شكاوي', r'مشكل', r'مشكلة',
    r'لا\s*يعمل', r'مقطوع', r'قطوع', r'بطيء', r'ضعيف',
    r'ما\s*خدم', r'مكانش', r'ما\s*ركبوا', r'ما\s*تواصل', r'ما\s*جاو',
    r'ما\s*حلوا', r'ما\s*صلحوا', r'ما\s*ردوا',
    r'\d+\s*(أشهر|يوم|سنة).*(?:انتظار|دون|بدون)',
    r'معاناة', r'اذلال', r'تعبتونا', r'هبلونا',
    r'وعود\s*كاذبة', r'بدون\s*حل', r'دون\s*أي\s*حل',
    r'فساد', r'مافيا', r'سجن', r'ديكتاتور',
    # Darija latin
    r'mochkil', r'ma\s*khedmetch', r'makach',
    r'hchouma', r'ta3batna', r'hablona',
    # zidou dans plainte
    r'tzido.{0,20}(batel|tgt3|mochkil)',
    r'zidou.{0,20}(batel|tgt3|mochkil)',
    # عام جديد + critique
    r'عام\s*(جديد|سعيد).{0,60}(خيوط|انقطاع|نحاس|adsl|مشكل|عيب)',
    r'(خيوط|انقطاع|نحاس|adsl|مشكل|عيب).{0,60}عام\s*(جديد|سعيد)',
]

# ══════════════════════════════════════════════════════════════════════════
# 3. PATTERNS VRAIMENT HORS-SUJET
#    Seuls ces cas sont exclus : sport étranger, politique, texte culturel/info
# ══════════════════════════════════════════════════════════════════════════
OFF_TOPIC_PATTERNS = [
    # Équipes sportives nommées (pas AT)
    r'\busmk\b', r'\busm\s*khenchela\b', r'\bcrb\b', r'\bmca\b', r'\busma\b',
    r'\bmc\s*alger\b', r'فريق\s+(?!اتصالات)',  # فريق + autre chose qu'اتصالات
    r'لفريق\s+\w+(?<!اتصالات)',
    # Politique / hors-sujet total
    r'على\s*حساب\s*اموال\s*الشعب',
    r'ديكتاتور', r'تبون', r'بوتفليقة',
    # Texte informatif/culturel sans lien (long hors-sujet)
    r'يونيسكو', r'بلغاريا', r'تراث.*ثقافي',
    r'عصور\s*وثنية', r'قبل\s*التاريخ',
]

# ══════════════════════════════════════════════════════════════════════════
# 4. PATTERNS PRODUIT/SERVICE
# ══════════════════════════════════════════════════════════════════════════
PRODUCT_PATTERNS = [
    r'appli(cation)?\b', r'service\b', r'réseau\b', r'connexion\b', r'débit\b',
    r'fibre?\b', r'internet\b', r'offre\b', r'abonnement\b', r'prix\b', r'tarif\b',
    r'vitesse\b', r'signal\b', r'4g\b', r'5g\b', r'wifi\b', r'box\b', r'modem\b',
    r"l['\u2019]appli", r'le\s*service', r'le\s*réseau', r'la\s*connexion',
    r'gpon\b', r'xgs.?pon\b', r'fttx\b', r'adsl\b', r'vdsl\b', r'idoom\b',
    r'الانترنت', r'انترنت\b', r'الشبكة', r'شبكة\b',
    r'الفيبر', r'فيبر\b', r'الاشتراك', r'الخدمة',
    r'السعر', r'التدفق', r'الاتصال', r'المودم', r'مودم\b', r'الباقة',
    r'الالياف\s*البصرية', r'الياف\b',
    r'تطبيق', r'الاندرويد', r'التطبيق',
]

# ══════════════════════════════════════════════════════════════════════════
# COMPILATION — FLAG_ENCOURAGEMENT
# ══════════════════════════════════════════════════════════════════════════
FLAGS_ENC = re.IGNORECASE | re.UNICODE
ENC_REGEX       = re.compile('|'.join(ENCOURAGEMENT_PATTERNS), FLAGS_ENC)
NEG_REGEX_ENC   = re.compile('|'.join(NEGATIVE_PATTERNS_ENC),  FLAGS_ENC | re.DOTALL)
OFF_TOPIC_REGEX = re.compile('|'.join(OFF_TOPIC_PATTERNS),     FLAGS_ENC)
PRODUCT_REGEX   = re.compile('|'.join(PRODUCT_PATTERNS),       FLAGS_ENC)

def get_match_enc(text):
    m = ENC_REGEX.search(text or '')
    return m.group() if m else None

def is_encouragement(text):
    """
    Retourne (flag_encouragement, sous_type)
      (1, 'pur')     → encouragement valide, pas d'avis produit  → Neutre forcé
      (1, 'produit') → encouragement + avis produit              → Positif faible
      (0, 'negatif') → pattern négatif détecté                   → annulé
      (0, 'offtopic')→ vraiment hors-sujet (sport/politique)     → supprimé
      (0, None)      → pas un encouragement
    """
    t = text or ''
    if not ENC_REGEX.search(t):    return 0, None
    if NEG_REGEX_ENC.search(t):    return 0, 'negatif'
    if OFF_TOPIC_REGEX.search(t):  return 0, 'offtopic'
    if PRODUCT_REGEX.search(t):    return 1, 'produit'
    return 1, 'pur'

# ══════════════════════════════════════════════════════════════════════════
# TESTS — FLAG_ENCOURAGEMENT
# ══════════════════════════════════════════════════════════════════════════
tests_enc = [
    # ── Doivent être flag=1 ───────────────────────────────────────────────
    ('Bon courage les gars !',                         (1, 'pur')),
    ('Bravo et bonne continuation',                    (1, 'pur')),
    ('بالتوفيق والنجاح الدائم ان شاء الله',            (1, 'pur')),
    ('كل عام وانتم بخير',                              (1, 'pur')),
    ('Bonne continuation inchallah',                   (1, 'pur')),
    ('Félicitations',                                  (1, 'pur')),
    ('Good luck',                                      (1, 'pur')),
    ('assegas ameggaz chapeau',                        (1, 'pur')),
    ('بارك الله فيكم إتقاااان وتفاني في العمل',        (1, 'pur')),
    ('ربي يوفقكم',                                     (1, 'pur')),
    ('برافو مزيد من التألق والنجاح',                    (1, 'pur')),
    # ── Encouragement + produit → flag=1/produit ─────────────────────────
    ('بالتوفيق والنجاح نتمنى تعميم الالياف البصرية',   (1, 'produit')),
    ('على العموم حاجة مليحة التطورات بالتوفيق',        (1, 'produit')),  # تطورات = produit context
    ('بالتوفيق اذا فيها السرعات المطروحة كافية مودم',  (1, 'produit')),
    ('bravo bonne continuation le service internet',   (1, 'produit')),
    # ── Doivent être flag=0 ───────────────────────────────────────────────
    ('على حساب اموال الشعب الف مبروك لفريق خنشلة',    (0, 'offtopic')),
    ('عجيب أمر هذه الدنيا يونيسكو بلغاريا',            (0, 'offtopic')),
    ('قديمة و ما عندها حتى معنى عيب تبداو بيها',       (0, 'negatif')),
    ('bon courage mais le service est nul',            (0, 'negatif')),
    ('عام جديد وما زالنا مع خيوط النحاس adsl',         (0, 'negatif')),
]
print('\n── Tests ──────────────────────────────────────────────')
ok = 0
for text, expected in tests_enc:
    flag, stype = is_encouragement(text)
    status = '✅' if (flag, stype) == expected else '❌'
    print(f'  {status} [{flag}/{str(stype):10s}] {text[:60]!r}')
    ok += ((flag, stype) == expected)
print(f'\n  {ok}/{len(tests_enc)} tests passés\n')

# ══════════════════════════════════════════════════════════════════════════
# CONNEXION MONGODB — FLAG_ENCOURAGEMENT
# ══════════════════════════════════════════════════════════════════════════
print('⏳ Connexion MongoDB …')
client_enc     = MongoClient(MONGO_URI_ENC)
collection_enc = client_enc[DB_NAME_ENC][COLLECTION_NAME_ENC]
docs_enc = list(collection_enc.find({}, {
    '_id': 1, TEXT_FIELD_ENC: 1,
    'Commentaire_Client_Original': 1,
    'sources': 1, 'dates': 1, 'label_final': 1, 'flag_social': 1,
}))
print(f'✅ {len(docs_enc)} documents chargés.')

# ══════════════════════════════════════════════════════════════════════════
# TRAITEMENT — FLAG_ENCOURAGEMENT
# ══════════════════════════════════════════════════════════════════════════
results_enc = []
for doc in docs_enc:
    text = doc.get(TEXT_FIELD_ENC, '') or ''
    flag, stype = is_encouragement(text)
    results_enc.append({
        '_id':                  doc['_id'],
        'doc_id':               str(doc['_id']),
        'commentaire_original': doc.get('Commentaire_Client_Original', ''),
        'texte_normalise':      text,
        'expression_detectee':  get_match_enc(text) or '',
        'source':               doc.get('sources', ''),
        'date':                 doc.get('dates', ''),
        'label_final':          doc.get('label_final', ''),
        'flag_social':          doc.get('flag_social', 0),
        'flag_encouragement':   flag,
        'encouragement_type':   stype or '',
    })

flagged_enc    = [r for r in results_enc if r['flag_encouragement'] == 1]
flag_pur_enc   = [r for r in flagged_enc if r['encouragement_type'] == 'pur']
flag_prod_enc  = [r for r in flagged_enc if r['encouragement_type'] == 'produit']
counter_enc    = Counter(r['expression_detectee'] for r in flagged_enc)

print(f'\n📊 Statistiques :')
print(f'   Total docs            : {len(results_enc)}')
print(f'   flag_encouragement=1  : {len(flagged_enc)}')
print(f'   └─ pur     → Neutre forcé   : {len(flag_pur_enc)}')
print(f'   └─ produit → Positif faible : {len(flag_prod_enc)}')
print(f'\n🔑 Top 10 expressions :')
for expr, cnt in counter_enc.most_common(10):
    print(f'   {cnt:4d}  {expr!r}')

# ══════════════════════════════════════════════════════════════════════════
# MISE À JOUR MONGODB — FLAG_ENCOURAGEMENT
# ══════════════════════════════════════════════════════════════════════════
print('\n⏳ Mise à jour MongoDB …')
operations_enc = [
    UpdateOne(
        {'_id': r['_id']},
        {'$set': {
            'flag_encouragement':       r['flag_encouragement'],
            'encouragement_type':       r['encouragement_type'],
            'expression_encouragement': r['expression_detectee'],
        }}
    )
    for r in results_enc
]
res_enc = collection_enc.bulk_write(operations_enc)
print(f'✅ MongoDB mis à jour — {res_enc.modified_count} docs modifiés.')

# ══════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL — FLAG_ENCOURAGEMENT = 1 uniquement
# ══════════════════════════════════════════════════════════════════════════
print('⏳ Génération Excel …')
hdr_fill_enc  = PatternFill('solid', fgColor='1F4E79')
hdr_font_enc  = Font(color='FFFFFF', bold=True, size=11)
pur_fill_enc  = PatternFill('solid', fgColor='D9E1F2')
prod_fill_enc = PatternFill('solid', fgColor='C6EFCE')
odd_fill_enc  = PatternFill('solid', fgColor='EBF3FB')
even_fill_enc = PatternFill('solid', fgColor='FFFFFF')
c_al_enc = Alignment(horizontal='center', vertical='center')
l_al_enc = Alignment(horizontal='left',   vertical='top', wrap_text=True)
thin_enc = Side(style='thin', color='BFBFBF')
brd_enc  = Border(left=thin_enc, right=thin_enc, top=thin_enc, bottom=thin_enc)

wb_enc = Workbook()
ws_enc = wb_enc.active
ws_enc.title = 'flag_encouragement_1'

headers_enc    = ['doc_id','commentaire_original','texte_normalise',
              'expression_detectee','source','date',
              'label_final','flag_social','flag_encouragement','encouragement_type']
col_widths_enc = [28, 60, 55, 25, 12, 18, 12, 11, 18, 14]

for ci, (h, w) in enumerate(zip(headers_enc, col_widths_enc), 1):
    cell = ws_enc.cell(row=1, column=ci, value=h)
    cell.fill, cell.font = hdr_fill_enc, hdr_font_enc
    cell.alignment = c_al_enc
    cell.border    = brd_enc
    ws_enc.column_dimensions[get_column_letter(ci)].width = w
ws_enc.row_dimensions[1].height = 28
ws_enc.freeze_panes = 'A2'

type_fill_enc = {'pur': pur_fill_enc, 'produit': prod_fill_enc}

for i, r in enumerate(flagged_enc):
    row_n    = i + 2
    row_fill = odd_fill_enc if i % 2 == 0 else even_fill_enc
    stype    = r['encouragement_type']
    tf       = type_fill_enc.get(stype, row_fill)

    vals = [r['doc_id'], r['commentaire_original'], r['texte_normalise'],
            r['expression_detectee'], r['source'], r['date'],
            r['label_final'], r['flag_social'], r['flag_encouragement'], stype]

    for ci, val in enumerate(vals, 1):
        cell = ws_enc.cell(row=row_n, column=ci, value=val)
        cell.border = brd_enc
        cell.font   = Font(name='Arial', size=9)
        if ci in (9, 10):
            cell.fill, cell.alignment = tf, c_al_enc
            cell.font = Font(name='Arial', size=9, bold=True)
        elif ci in (2, 3):
            cell.fill, cell.alignment = row_fill, l_al_enc
        else:
            cell.fill, cell.alignment = row_fill, c_al_enc
    ws_enc.row_dimensions[row_n].height = 45

# Stats
ws2_enc = wb_enc.create_sheet('Stats')
ws2_enc['A1'] = 'Statistiques flag_encouragement'
ws2_enc['A1'].font = Font(bold=True, size=13, color='1F4E79')
for r_idx, (lbl, val) in enumerate([
    ('Total documents', len(results_enc)),
    ('flag_encouragement = 1', len(flagged_enc)),
    ('  └─ pur     → Neutre forcé',   len(flag_pur_enc)),
    ('  └─ produit → Positif faible', len(flag_prod_enc)),
], start=3):
    ws2_enc.cell(row=r_idx, column=1, value=lbl).font = Font(name='Arial', size=10)
    ws2_enc.cell(row=r_idx, column=2, value=val).font  = Font(name='Arial', size=10, bold=True)
ws2_enc['A10'] = 'Top expressions'
ws2_enc['A10'].font = Font(bold=True, size=10)
for r_idx, (expr, cnt) in enumerate(counter_enc.most_common(20), start=11):
    ws2_enc.cell(row=r_idx, column=1, value=expr)
    ws2_enc.cell(row=r_idx, column=2, value=cnt)
ws2_enc.column_dimensions['A'].width = 45

# Légende
ws3_enc = wb_enc.create_sheet('Légende')
ws3_enc['A1'] = 'Légende — flag_encouragement'
ws3_enc['A1'].font = Font(bold=True, size=12, color='1F4E79')
for r_idx, (fill, nom, desc) in enumerate([
    (pur_fill_enc,  'Bleu  — pur     flag=1', 'Encouragement valide, pas d\'avis produit → Neutre forcé'),
    (prod_fill_enc, 'Vert  — produit flag=1', 'Encouragement + avis produit/service → Positif faible'),
], start=3):
    c = ws3_enc.cell(row=r_idx, column=1, value=nom)
    c.fill, c.font = fill, Font(name='Arial', size=10, bold=True)
    ws3_enc.cell(row=r_idx, column=2, value=desc).font = Font(name='Arial', size=10)
ws3_enc['A7'] = 'Exclus (flag=0) :'
ws3_enc['A7'].font = Font(bold=True, size=10)
for r_idx, desc in enumerate([
    "offtopic : félicite équipe sportive (USMK…), texte politique/culturel hors-sujet",
    "negatif  : contient une plainte ou critique → pattern négatif détecté",
], start=8):
    ws3_enc.cell(row=r_idx, column=1, value=desc).font = Font(name='Arial', size=10)
ws3_enc.column_dimensions['A'].width = 70

wb_enc.save(OUTPUT_EXCEL_ENC)
print(f'✅ Excel généré : {OUTPUT_EXCEL_ENC}  ({len(flagged_enc)} lignes)')
client_enc.close()

print('\n' + '=' * 60)
print('  TOUS LES FLAGS TRAITÉS AVEC SUCCÈS ✅')
print('=' * 60)
print(f'  flag_social       → {OUTPUT_EXCEL}')
print(f'  flag_suggestion   → {OUTPUT_EXCEL_SUG}')
print(f'  flag_plainte      → {OUTPUT_EXCEL_PLA}')
print(f'  flag_negation     → {OUTPUT_EXCEL_NEG}')
print(f'  flag_mixte        → {OUTPUT_EXCEL_MIX}')
print(f'  flag_encouragement→ {OUTPUT_EXCEL_ENC}')
print('=' * 60)
# #!/usr/bin/env python3
# # -*- coding: utf-8 -*-
# """
# FLAGS COMPLET - Version avec création de dataset_unifie
# Lit les commentaires avec traite=false dans commentaires_normalises
# Copie TOUS les champs dans dataset_unifie (NOUVELLE base) + ajoute les flags
# Marque traite=true dans commentaires_normalises après traitement
# """

# import re
# from pymongo import MongoClient, UpdateOne
# from datetime import datetime

# # ══════════════════════════════════════════════════════════════════════════
# # CONFIGURATION
# # ══════════════════════════════════════════════════════════════════════════
# MONGO_URI = "mongodb://localhost:27018/"
# DB_NAME = "telecom_algerie"
# COLLECTION_SOURCE = "commentaires_normalises"
# COLLECTION_DATASET = "dataset_unifie"
# TEXT_FIELD = "normalized_arabert"


# # ══════════════════════════════════════════════════════════════════════════
# # FONCTIONS DE GESTION DES FLAGS
# # ══════════════════════════════════════════════════════════════════════════

# def get_nouveaux_commentaires_count():
#     """Compte les commentaires avec traite=false dans la source"""
#     client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)
#     db = client[DB_NAME]
#     collection = db[COLLECTION_SOURCE]
#     count = collection.count_documents({"traite": False})
#     client.close()
#     return count


# def marquer_comme_traite(ids):
#     """Marque les commentaires comme traités (traite=True) dans commentaires_normalises"""
#     if not ids:
#         print("   ⚠️ Aucun ID à marquer")
#         return
    
#     client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)
#     db = client[DB_NAME]
#     collection = db[COLLECTION_SOURCE]
    
#     valid_ids = [str(id_str).strip() for id_str in ids if str(id_str).strip()]
    
#     if valid_ids:
#         resultat = collection.update_many(
#             {"_id": {"$in": valid_ids}},
#             {"$set": {"traite": True, "date_traitement_flags": datetime.now()}}
#         )
#         print(f"   ✅ {resultat.modified_count} commentaires marqués traite=True dans commentaires_normalises")
    
#     client.close()


# def creer_dataset_unifie(documents):
#     """Crée la collection dataset_unifie avec TOUS les champs + flags"""
#     if not documents:
#         print("   ⚠️ Aucun document à insérer")
#         return
    
#     client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)
#     db = client[DB_NAME]
    
#     # Supprimer l'ancienne collection si elle existe
#     if COLLECTION_DATASET in db.list_collection_names():
#         db[COLLECTION_DATASET].drop()
#         print("   🧹 Ancienne collection dataset_unifie supprimée")
    
#     collection_dataset = db[COLLECTION_DATASET]
#     collection_dataset.insert_many(documents)
#     print(f"   ✅ {len(documents)} documents insérés dans dataset_unifie")
    
#     client.close()


# # ══════════════════════════════════════════════════════════════════════════
# # PATTERNS FLAG_SOCIAL
# # ══════════════════════════════════════════════════════════════════════════

# SOCIAL_PATTERNS = [
#     r'بارك\s*الله', r'الله\s*يبارك', r'الله\s*يعاون', r'ربي\s*يعاون',
#     r'ربي\s*يجازي', r'جزاك\s*الله\s*خير', r'جزاكم\s*الله\s*خير',
#     r'الله\s*يوفق', r'ربي\s*يحفظ', r'ربي\s*يوفق', r'بالتوفيق',
#     r'رمضان\s*كريم', r'رمضان\s*مبارك', r'عيد\s*مبارك', r'عيد\s*سعيد',
#     r'كل\s*عام\s*و[أا]نتم', r'كل\s*عام\s*وأنتم', r'مبروك',
#     r'تبارك\s*الله', r'ماشاء\s*الله', r'ما\s*شاء\s*الله', r'الله\s*معاكم',
#     r'ربي\s*يوفقكم', r'ربي\s*يوفقك\b', r'الله\s*يوفقكم',
#     r'شكرا\s*على\s*جهودكم', r'شكرا\s*على\s*المجهودات',
#     r'يسلموا?\b', r'الله\s*يسلمك', r'صح\s*فطورك', r'صح\s*فطوركم',
#     r'صح\s*عيدك', r'صح\s*عيدكم', r'حفظكم\s*الله',
#     r'ربي\s*يبارك\s*فيكم', r'ربي\s*يبارك\s*فيك\b',
#     r'rbi\s*y[3e]awkom', r'rabi\s*y[3e]awkom', r'barak\s*allah',
#     r'baraka\s*llah', r'allah\s*ybarek', r'allah\s*yewfek',
#     r'allah\s*yewfeqkom', r'tbarakallah', r'tbarkallah',
#     r'nchallah\s*ynajh', r'inch[a]?allah\s*ynajh', r'saha\s*ftourkoum',
#     r'saha\s*[ae]idkom', r'sa7a\s*[ae]idkom', r'kol\s*[3e]am\s*w',
#     r'[ae]id\s*mobarak', r'[ae]id\s*sa[3i]id', r'mabrouk', r'mbr[ou]+k',
#     r'bon\s*courage', r'bonne?\s*continuation', r'merci\s*pour\s*vos\s*efforts',
#     r'merci\s*pour\s*le\s*travail', r'courage\s*[àa]\s*tou[st]',
#     r'bonne?\s*ann[eé]e', r'assegwas\s*amegas', r'vive\s*alg[eé]rie',
#     r'allah\s*ma[3e]akom', r'rabi\s*ma[3e]akom', r'barra\s*barra\s*likom',
#     r'[ae]ychine\b', r'3aychine\b', r'rbi\s*yjazik', r'yeslamo\b',
#     r'ya3tikom\s*essa7a', r'joyeux\s*ramadan', r'bon\s*travail',
#     r'bravo\s*(à\s*)?l[a-z]*\s*[eé]quipe', r'bonne?\s*f[eê]te',
#     r'f[eé]licitations?', r'bonne?\s*chance', r'bravo\s*pour',
#     r'chapeau\s*pour', r'joyeuses?\s*f[eê]tes?', r'meilleurs?\s*v[oœ]ux',
#     r'bonne?\s*journ[eé]e', r'bon\s*succ[eè]s', r'bonne?\s*r[eé]ussite',
#     r'tous\s*mes\s*v[oœ]ux',
# ]

# NEGATIVE_PATTERNS = [
#     r'مشكل', r'مشكلة', r'شكاية', r'شكوى', r'شكاوي', r'ما\s*خدم',
#     r'مكانش', r'ما\s*ركبوا', r'ما\s*تواصل', r'ما\s*جاو',
#     r'ما\s*حلوا', r'ما\s*صلحوا', r'ما\s*ردوا', r'انتظار',
#     r'أنتظر', r'نستنى', r'وعود\s*كاذبة', r'بدون\s*حل',
#     r'دون\s*حل', r'هل\s*يعقل', r'نقطة\s*الصفر',
#     r'معاناة', r'مافيا', r'فساد',
#     r'خدمة.*شهر', r'شهر.*ما.*ركب', r'ست\s*اشهر',
#     r'لم\s*يتواصلوا', r'لم\s*يرد',
#     r'تعبتونا', r'هبلونا',
# ]

# SOCIAL_REGEX = re.compile('|'.join(SOCIAL_PATTERNS), flags=re.IGNORECASE | re.UNICODE)
# NEGATIVE_REGEX = re.compile('|'.join(NEGATIVE_PATTERNS), flags=re.IGNORECASE | re.UNICODE)


# def is_social(text):
#     t = text or ''
#     if not SOCIAL_REGEX.search(t):
#         return 0
#     if NEGATIVE_REGEX.search(t):
#         return 0
#     return 1


# # ══════════════════════════════════════════════════════════════════════════
# # PATTERNS FLAG_SUGGESTION
# # ══════════════════════════════════════════════════════════════════════════

# POS_PATTERNS_SUG = [
#     r'\bmliha?\b', r'\bmzian[ae]?\b', r'\bzwin[ae]?\b', r'\btamam\b',
#     r'\bbien\b', r'\bbon(ne)?\b', r'\bsuper\b', r'\bpas\s*mal\b',
#     r'مليح[ةه]?', r'زوين[ةه]?', r'\bتمام\b', r'ممتاز[ةه]?',
#     r'سريع[ةه]?', r'جيد[ةه]?', r'حسن[ةه]?', r'خدمة\s*(مليحة|جيدة)',
# ]

# ADV_PATTERNS = [
#     r'\bms\b', r'\bmais\b', r'\bcependant\b', r'\bبصح\b', r'\bلكن\b',
#     r'\bwalakin\b', r'\bpar\s*contre\b',
# ]

# AMEL_PATTERNS = [
#     r'\btehsen[ou]*\b', r'\btzid[ou]*\b', r'amélior\w+', r'optim[iu]s\w+',
#     r'تحسين', r'يتحسن', r'تحسنو', r'تزيدو', r'نتمنى\s*تحسين',
#     r'يلزم\s*تحسن', r'ينقصه?\s*(فقط|بس)',
# ]

# NEG_FORT_PATTERNS = [
#     r'\bnul\b', r'\bkhaybe?\b', r'كارثة', r'كارثه', r'مشكل', r'مشكلة',
#     r'coupure\b', r'انقطاع\b', r'شكوى', r'شكاية', r'تعبتونا', r'هرمنا',
# ]

# POS_REGEX_SUG = re.compile('|'.join(POS_PATTERNS_SUG), re.IGNORECASE | re.UNICODE)
# ADV_REGEX = re.compile('|'.join(ADV_PATTERNS), re.IGNORECASE | re.UNICODE)
# AMEL_REGEX = re.compile('|'.join(AMEL_PATTERNS), re.IGNORECASE | re.UNICODE)
# NEG_FORT_REGEX = re.compile('|'.join(NEG_FORT_PATTERNS), re.IGNORECASE | re.UNICODE | re.DOTALL)


# def is_suggestion(text):
#     t = text or ''
#     has_pos = bool(POS_REGEX_SUG.search(t))
#     has_adv = bool(ADV_REGEX.search(t))
#     has_amel = bool(AMEL_REGEX.search(t))
#     has_neg = bool(NEG_FORT_REGEX.search(t))

#     if not (has_pos and has_adv):
#         return 0, ''

#     if has_amel and has_neg:
#         return -1, 'masked_neg'
#     if has_amel and not has_neg:
#         return 1, 'pur'
#     if not has_amel and not has_neg:
#         return 1, 'pos_adv'
#     return 0, ''


# # ══════════════════════════════════════════════════════════════════════════
# # PATTERNS FLAG_PLAINTE
# # ══════════════════════════════════════════════════════════════════════════

# CONTACT_PATTERNS = [
#     r'ردو?\s*(في\s*|على\s*|فال|فل|ف)?\s*[ال]?خاص\b', r'الرد\s*(في|على)\s*الخاص',
#     r'نرجو\s*الرد\s*(في|على)?\s*الخاص', r'ارجو\s*(منكم\s*)?الرد\s*(في|على)\s*الخاص',
#     r'يرجى\s*الرد\s*(على|في)\s*الخاص', r'ردو\s*عليا\b', r'ردو\s*علينا\b',
#     r'بعثتلكم?\s*(في|على|فال|فل)?\s*الخاص', r'\bprv\b', r'\binbox\b',
# ]

# FRUSTRATION_PATTERNS = [
#     r'\d+\s*(jours?|أيام|يوم|ايام|mois|أشهر|اشهر|شهر)', r'منذ\s*\d+',
#     r'شكوى', r'شكاية', r'مشكل', r'مشكلة', r'coupure\b', r'انقطاع\b',
#     r'لا\s*يعمل\b', r'ما\s*يخدم\b', r'تعبتونا\b', r'هرمنا\b', r'هبلتونا\b',
# ]

# AT_RESPONSE_PATTERNS = [
#     r'ندعوكم\s*(للتواصل|للتواصل\s*معنا)', r'يمكنكم\s*(التواصل|مراسلتنا)',
#     r'بإمكانكم\s*التواصل', r'لقد\s*تم\s*التواصل\s*معك\s*في\s*الخاص',
# ]

# NEGLEX_PATTERNS = [
#     r'هرمنا\b', r'هبلتونا\b', r'تعبتونا\b', r'كارثيه\b', r'مافيا\b', r'فساد\b',
# ]

# NEUTRALIZER_PATTERNS = [
#     r'بالتوفيق\b', r'ربي\s*يوفقكم', r'bravo\b', r'مبروك\b', r'bon\s*courage\b',
# ]

# CONTACT_REGEX = re.compile('|'.join(CONTACT_PATTERNS), re.IGNORECASE | re.UNICODE)
# FRUSTRATION_REGEX = re.compile('|'.join(FRUSTRATION_PATTERNS), re.IGNORECASE | re.UNICODE | re.DOTALL)
# AT_RESPONSE_REGEX = re.compile('|'.join(AT_RESPONSE_PATTERNS), re.IGNORECASE | re.UNICODE)
# NEGLEX_REGEX = re.compile('|'.join(NEGLEX_PATTERNS), re.IGNORECASE | re.UNICODE)
# NEUTRALIZER_REGEX = re.compile('|'.join(NEUTRALIZER_PATTERNS), re.IGNORECASE | re.UNICODE)


# def is_plainte(text):
#     t = text or ''
#     has_contact = bool(CONTACT_REGEX.search(t))
#     has_frustration = bool(FRUSTRATION_REGEX.search(t))
#     has_neglex = bool(NEGLEX_REGEX.search(t))
#     has_at_resp = bool(AT_RESPONSE_REGEX.search(t))
#     has_neutralizer = bool(NEUTRALIZER_REGEX.search(t))

#     if has_neglex:
#         return 1, 'neglex'
#     if has_at_resp and not has_frustration:
#         return -1, 'at_response'
#     if has_contact and has_frustration:
#         return 1, 'prv'
#     if has_contact:
#         if has_neutralizer:
#             return 0, ''
#         return -1, 'prv_implicit'
#     return 0, ''


# # ══════════════════════════════════════════════════════════════════════════
# # FLAG_NEGATION
# # ══════════════════════════════════════════════════════════════════════════

# NEGATION_PATTERNS = [
#     r'\bmachi\b', r'\bwalou\b', r'\bma\w*sh\b', r'\bلا\b', r'\bلم\b',
#     r'\bليس\b', r'\bمكانش\b', r'\bماكاش\b', r'\bمش\b', r'\bميش\b',
#     r'\bne\s+\w+\s+pas\b', r'\bpas\b', r'\bjamais\b',
# ]

# POS_PATTERNS_NEG = [
#     r'\bmliha?\b', r'\bmzian[ae]?\b', r'\btamam\b', r'\bbien\b', r'\bbon\b',
#     r'مليح[ةه]?', r'تمام\b', r'ممتاز[ةه]?', r'جيد[ةه]?',
# ]

# NEGATION_REGEX = re.compile('|'.join(NEGATION_PATTERNS), re.IGNORECASE | re.UNICODE)
# POS_REGEX_NEG = re.compile('|'.join(POS_PATTERNS_NEG), re.IGNORECASE | re.UNICODE)


# def is_negation(text):
#     t = text or ''
#     has_neg = bool(NEGATION_REGEX.search(t))
#     if not has_neg:
#         return 0, ''
    
#     has_pos = bool(POS_REGEX_NEG.search(t))
#     if has_pos:
#         return 1, 'neg_pos'
#     return 1, 'neg_seule'


# # ══════════════════════════════════════════════════════════════════════════
# # FLAG_MIXTE
# # ══════════════════════════════════════════════════════════════════════════

# POS_FORT = [
#     r'\bmliha?\b', r'\bmzian[ae]?\b', r'\btamam\b', r'\bbien\b', r'\bbon(ne)?\b',
#     r'مليح[ةه]?', r'تمام\b', r'ممتاز[ةه]?', r'جيد[ةه]?',
# ]

# NEG_FORT_MIX = [
#     r'\bnul\b', r'\bkhaybe?\b', r'كارثة', r'كارثه', r'مشكل', r'مشكلة',
#     r'coupure\b', r'انقطاع\b', r'تعبتونا', r'هرمنا',
# ]

# POS_REGEX_MIX = re.compile('|'.join(POS_FORT), re.IGNORECASE | re.UNICODE)
# NEG_REGEX_MIX = re.compile('|'.join(NEG_FORT_MIX), re.IGNORECASE | re.UNICODE)


# def is_mixte(text):
#     t = text or ''
#     has_pos = bool(POS_REGEX_MIX.search(t))
#     has_neg = bool(NEG_REGEX_MIX.search(t))
#     if has_pos and has_neg:
#         return 1, 'pos_neg'
#     return 0, ''


# # ══════════════════════════════════════════════════════════════════════════
# # FLAG_ENCOURAGEMENT
# # ══════════════════════════════════════════════════════════════════════════

# ENCOURAGEMENT_PATTERNS = [
#     r'bon\s*courage', r'bravo\b', r'f[eé]licitations?', r'بالتوفيق',
#     r'ربي\s*يوفقكم', r'مبروك\b', r'bonne?\s*continuation', r'good\s*luck',
# ]

# PRODUCT_PATTERNS = [
#     r'service\b', r'internet\b', r'connexion\b', r'fibre\b', r'wifi\b',
#     r'الانترنت', r'الخدمة', r'شبكة',
# ]

# ENC_REGEX = re.compile('|'.join(ENCOURAGEMENT_PATTERNS), re.IGNORECASE | re.UNICODE)
# PRODUCT_REGEX = re.compile('|'.join(PRODUCT_PATTERNS), re.IGNORECASE | re.UNICODE)


# def is_encouragement(text):
#     t = text or ''
#     if not ENC_REGEX.search(t):
#         return 0, ''
#     if PRODUCT_REGEX.search(t):
#         return 1, 'produit'
#     return 1, 'pur'


# # ══════════════════════════════════════════════════════════════════════════
# # FONCTION PRINCIPALE
# # ══════════════════════════════════════════════════════════════════════════

# def main():
#     print("=" * 70)
#     print("🔍 EXTRACTION DES FLAGS (social, suggestion, plainte, etc.)")
#     print("   🏷️  Filtre: traite=false dans commentaires_normalises")
#     print("   📥 Source: commentaires_normalises")
#     print("   📤 Destination: dataset_unifie (NOUVELLE collection avec TOUS les champs)")
#     print("=" * 70)

#     # Vérifier les nouveaux commentaires
#     nouveaux_count = get_nouveaux_commentaires_count()

#     if nouveaux_count == 0:
#         print("\n✅ Aucun nouveau commentaire à traiter (traite=false)")
#         print("   Le pipeline est à jour.")
#         return

#     print(f"\n📥 {nouveaux_count} nouveaux commentaires à traiter (traite=false)")

#     # Connexion MongoDB
#     client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)
#     db = client[DB_NAME]
#     collection_source = db[COLLECTION_SOURCE]

#     # Lire TOUS les champs des commentaires avec traite=false
#     docs = list(collection_source.find({"traite": False}))
#     print(f"✅ {len(docs)} documents chargés (avec TOUS leurs champs)")

#     # Traitement des flags
#     print("\n⏳ Application des patterns...")
    
#     documents_dataset = []
#     ids_traites = []
    
#     for doc in docs:
#         text = doc.get(TEXT_FIELD, '') or ''
        
#         # Calculer les flags
#         flag_social = is_social(text)
#         flag_suggestion, sug_type = is_suggestion(text)
#         flag_plainte, plainte_type = is_plainte(text)
#         flag_negation, negation_type = is_negation(text)
#         flag_mixte, mixte_type = is_mixte(text)
#         flag_encouragement, enc_type = is_encouragement(text)
        
#         # Ajouter les flags au document existant (copie TOUS les champs)
#         doc['flag_social'] = flag_social
#         doc['flag_suggestion'] = flag_suggestion
#         doc['suggestion_type'] = sug_type
#         doc['flag_plainte'] = flag_plainte
#         doc['plainte_type'] = plainte_type
#         doc['flag_negation'] = flag_negation
#         doc['negation_type'] = negation_type
#         doc['flag_mixte'] = flag_mixte
#         doc['mixte_type'] = mixte_type
#         doc['flag_encouragement'] = flag_encouragement
#         doc['encouragement_type'] = enc_type
#         doc['date_traitement_flags'] = datetime.now()
        
#         documents_dataset.append(doc)
#         ids_traites.append(str(doc['_id']))

#     # Créer dataset_unifie avec TOUS les champs
#     print("\n⏳ Création de dataset_unifie...")
#     creer_dataset_unifie(documents_dataset)

#     # Marquer les commentaires comme traités dans commentaires_normalises
#     if ids_traites:
#         print("\n🏷️  Marquage des commentaires traités dans commentaires_normalises...")
#         marquer_comme_traite(ids_traites)

#     # Statistiques
#     total = len(documents_dataset)
#     n_social = sum(1 for d in documents_dataset if d.get('flag_social', 0) == 1)
#     n_suggestion = sum(1 for d in documents_dataset if d.get('flag_suggestion', 0) == 1)
#     n_plainte = sum(1 for d in documents_dataset if d.get('flag_plainte', 0) == 1)
#     n_negation = sum(1 for d in documents_dataset if d.get('flag_negation', 0) == 1)
#     n_mixte = sum(1 for d in documents_dataset if d.get('flag_mixte', 0) == 1)
#     n_encouragement = sum(1 for d in documents_dataset if d.get('flag_encouragement', 0) == 1)

#     print(f"\n📊 STATISTIQUES FLAGS:")
#     print(f"   flag_social        : {n_social} / {total} ({100*n_social/total:.1f}%)")
#     print(f"   flag_suggestion    : {n_suggestion} / {total} ({100*n_suggestion/total:.1f}%)")
#     print(f"   flag_plainte       : {n_plainte} / {total} ({100*n_plainte/total:.1f}%)")
#     print(f"   flag_negation      : {n_negation} / {total} ({100*n_negation/total:.1f}%)")
#     print(f"   flag_mixte         : {n_mixte} / {total} ({100*n_mixte/total:.1f}%)")
#     print(f"   flag_encouragement : {n_encouragement} / {total} ({100*n_encouragement/total:.1f}%)")

#     # Vérification finale
#     restants = get_nouveaux_commentaires_count()
#     print(f"\n📊 Collection source (commentaires_normalises): {restants} commentaires restants (traite=false)")
#     print(f"📊 Collection dataset_unifie: {total} commentaires (avec TOUS les champs)")

#     # Afficher un exemple des champs
#     if documents_dataset:
#         print("\n📋 Exemple des champs dans dataset_unifie:")
#         print(f"   {list(documents_dataset[0].keys())[:15]}...")

#     client.close()
#     print("\n🎉 Extraction des flags terminée !")


# if __name__ == "__main__":
#     main()