"""
DÉTECTION FLAG_MIXTE DZ — v1
══════════════════════════════════════════════════════════════════════════
Insight clé :
  Présence SIMULTANÉE d'un mot positif fort ET d'un négatif fort
  dans la même phrase → sentiment ambigu → BERT doit trancher (ABSA)

  Sous-types :
    (1, 'pos_neg')     → mot positif + mot négatif fort → Neutre ambigu (50%)
    (1, 'avec_sug')    → pos + neg + suggestion → flag_suggestion domine
                         mais flag_mixte reste à 1 pour info BERT
    (-1,'sug_only')    → flag_suggestion seul (A∧B∧C sans NEG) → déjà géré
    (0,  None)         → pas mixte

  Règle d'or :
    flag_mixte = 1  si POS ∧ NEG_FORT  (même phrase, fenêtre de 80 chars)
    flag_mixte = 0  sinon
    Note : si flag_mixte=1 ET flag_suggestion=1 → le négatif l'emporte
"""

import re
from collections import Counter
from pymongo import MongoClient, UpdateOne
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MONGO_URI       = "mongodb://localhost:27017/"
DB_NAME         = "telecom_algerie"
COLLECTION_NAME = "commentaires_normalises_tfidf"
TEXT_FIELD      = "normalized_arabert"
OUTPUT_EXCEL    = "commentaires_flag_mixte_v1.xlsx"

# ── Positifs forts ────────────────────────────────────────────────────────
POS_FORT = [
    r'\bmliha?\b', r'\bmzian[ae]?\b', r'\bzwin[ae]?\b', r'\bhil[ae]\b',
    r'\btamam\b', r'\bsari3\b',
    r'\bbien\b', r'\bbon(ne)?\b', r'\bsuper\b', r'\bpas\s*mal\b',
    r'\bcorrect(e)?\b', r'\brapide\b', r'\befficace\b',
    r'مليح[ةه]?', r'زوين[ةه]?', r'\bتمام\b', r'ممتاز[ةه]?',
    r'سريع[ةه]?', r'جيد[ةه]?', r'حسن[ةه]?', r'مقبول[ةه]?',
    r'هايل[ةه]?', r'رائع[ةه]?', r'مزيان[ةه]?', r'معقول[ةه]?',
    r'كويس[ةه]?', r'راضي[ةه]?',
]

# ── Négatifs forts ────────────────────────────────────────────────────────
NEG_FORT = [
    r'\bnul\b', r'\bkhaybe?\b', r'\bkhayba?\b',
    r'\bmachi\s*barka\b', r'\bkarica\b', r'\bkhartha?\b',
    r'كارثة', r'كارثه', r'كارثي[هة]?',
    r'فضيحة', r'مهزلة',
    r'\bzero\b', r'تحت\s*الصفر',
    r'لا\s*يعمل', r'ما\s*يخدم', r'مايخدمش',
    r'مقطوع[هة]?', r'قطوع\b',
    r'coupure\b', r'انقطاع\b',
    r'مشكل[ةه]?', r'مشاكل\b',
    r'trop\s*(lent|cher|mauvais)',
    r'ne\s*fonctionne\s*pas', r'marche\s*pas',
    r'تعبتونا', r'هرمنا', r'هبلتونا',
    r'وعود\s*كاذبة', r'مافيا\b', r'فساد\b',
    r'\bnul\b', r'honte\b', r'catastroph\w+',
    r'الأسوأ\b', r'الاسوى\b',
    r'ضعيف[ةه]?', r'بطيء\b', r'بطيئ[ةه]?\b',
    r'غالي[ةه]?', r'cher\b', r'prix\s*(élevé|excessif|astronomique)',
]

FLAGS = re.IGNORECASE | re.UNICODE
POS_REGEX = re.compile('|'.join(POS_FORT), FLAGS)
NEG_REGEX = re.compile('|'.join(NEG_FORT), FLAGS | re.DOTALL)

def get_match(regex, text):
    m = regex.search(text or '')
    return m.group() if m else None

def is_mixte(text):
    """
    Retourne (flag, subtype)
      (1, 'pos_neg')  → positif + négatif fort présents
      (0,  None)      → pas mixte
    """
    t = text or ''
    has_pos = bool(POS_REGEX.search(t))
    has_neg = bool(NEG_REGEX.search(t))
    if has_pos and has_neg:
        return 1, 'pos_neg'
    return 0, None

# ── Tests ──────────────────────────────────────────────────────────────────
tests = [
    ('connexion mliha ms coupure',                          (1, 'pos_neg')),
    ("c'est bien mais coupure tout le temps",               (1, 'pos_neg')),
    ('الانترنت تمام بس مقطوعة أحيانا',                     (1, 'pos_neg')),
    ('الخدمة مليحة لكن كارثة في التغطية',                  (1, 'pos_neg')),
    ('super application ne fonctionne pas depuis hier',     (1, 'pos_neg')),
    ('connexion correcte mais prix trop cher',              (1, 'pos_neg')),
    ('مزيان بصح ضعيف في الريف',                            (1, 'pos_neg')),
    # flag=0 — pas mixte
    ('connexion mliha bezzaf',                              (0, None)),
    ('كارثة هذه الخدمة',                                   (0, None)),
    ('bon courage à toute l\'équipe',                       (0, None)),
    ('مليح بصح يلزم تحسن',                                  (0, None)),  # suggestion, pas mixte
]

print('\n── Tests flag_mixte ──────────────────────────────────────────────')
ok = 0
for text, expected in tests:
    flag, stype = is_mixte(text)
    status = '✅' if (flag, stype) == expected else '❌'
    print(f'  {status} [{flag:+d}/{str(stype):10s}] {text[:70]!r}')
    ok += ((flag, stype) == expected)
print(f'\n  {ok}/{len(tests)} tests passés\n')

# ── MongoDB ────────────────────────────────────────────────────────────────
print('⏳ Connexion MongoDB …')
client     = MongoClient(MONGO_URI)
collection = client[DB_NAME][COLLECTION_NAME]
docs = list(collection.find({}, {
    '_id': 1, TEXT_FIELD: 1,
    'Commentaire_Client_Original': 1,
    'sources': 1, 'dates': 1, 'label_final': 1,
    'flag_social': 1, 'flag_encouragement': 1,
    'flag_plainte': 1, 'flag_suggestion': 1,
}))
print(f'✅ {len(docs)} documents chargés.')

results = []
for doc in docs:
    text = doc.get(TEXT_FIELD, '') or ''
    flag, stype = is_mixte(text)
    p = get_match(POS_REGEX, text) or ''
    n = get_match(NEG_REGEX, text) or ''
    expr = f'{p} ↔ {n}' if flag == 1 else ''

    results.append({
        '_id':                  doc['_id'],
        'doc_id':               str(doc['_id']),
        'commentaire_original': doc.get('Commentaire_Client_Original', ''),
        'texte_normalise':      text,
        'expression_detectee':  expr,
        'source':               doc.get('sources', ''),
        'date':                 doc.get('dates', ''),
        'label_final':          doc.get('label_final', ''),
        'flag_social':          doc.get('flag_social', 0),
        'flag_encouragement':   doc.get('flag_encouragement', 0),
        'flag_plainte':         doc.get('flag_plainte', 0),
        'flag_suggestion':      doc.get('flag_suggestion', 0),
        'flag_mixte':           flag,
        'mixte_type':           stype or '',
    })

flagged = [r for r in results if r['flag_mixte'] == 1]
counter = Counter(r['expression_detectee'].split(' ↔ ')[0]
                  for r in flagged if r['expression_detectee'])

# Cas intéressant : mixte + suggestion → négatif domine
mixte_et_sug = [r for r in flagged if r['flag_suggestion'] in (1, -1)]

print(f'\n📊 Statistiques flag_mixte v1 :')
print(f'   Total docs              : {len(results)}')
print(f'   flag_mixte = 1          : {len(flagged)}')
print(f'   ├─ dont mixte+suggestion: {len(mixte_et_sug)}  (négatif domine)')
print(f'   flag_mixte = 0          : {len(results) - len(flagged)}')
print(f'\n🔑 Top 15 mots positifs déclenchants :')
for expr, cnt in counter.most_common(15):
    print(f'   {cnt:4d}  {expr!r}')

# ── Mise à jour MongoDB ────────────────────────────────────────────────────
print('\n⏳ Mise à jour MongoDB …')
operations = [
    UpdateOne(
        {'_id': r['_id']},
        {'$set': {
            'flag_mixte':       r['flag_mixte'],
            'mixte_type':       r['mixte_type'],
            'expression_mixte': r['expression_detectee'],
        }}
    )
    for r in results
]
res = collection.bulk_write(operations)
print(f'✅ MongoDB mis à jour — {res.modified_count} docs modifiés.')

# ── Export Excel ───────────────────────────────────────────────────────────
print('⏳ Génération Excel …')
hdr_fill   = PatternFill('solid', fgColor='7E3F8F')
hdr_font   = Font(color='FFFFFF', bold=True, size=11)
mixte_fill = PatternFill('solid', fgColor='E8DAEF')
odd_fill   = PatternFill('solid', fgColor='F9F2FF')
even_fill  = PatternFill('solid', fgColor='FFFFFF')
c_al = Alignment(horizontal='center', vertical='center')
l_al = Alignment(horizontal='left',   vertical='top', wrap_text=True)
thin = Side(style='thin', color='BFBFBF')
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = 'flag_mixte_v1'

headers    = ['doc_id', 'commentaire_original', 'texte_normalise',
              'expression_detectee', 'source', 'date', 'label_final',
              'flag_social', 'flag_encouragement', 'flag_plainte',
              'flag_suggestion', 'flag_mixte', 'mixte_type']
col_widths = [28, 60, 55, 40, 12, 18, 12, 11, 18, 11, 14, 11, 12]

for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.fill, cell.font = hdr_fill, hdr_font
    cell.alignment = c_al; cell.border = brd
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 28
ws.freeze_panes = 'A2'

for i, r in enumerate(flagged):
    row_n    = i + 2
    row_fill = odd_fill if i % 2 == 0 else even_fill
    vals = [r['doc_id'], r['commentaire_original'], r['texte_normalise'],
            r['expression_detectee'], r['source'], r['date'], r['label_final'],
            r['flag_social'], r['flag_encouragement'], r['flag_plainte'],
            r['flag_suggestion'], r['flag_mixte'], r['mixte_type']]
    for ci, val in enumerate(vals, 1):
        cell = ws.cell(row=row_n, column=ci, value=val)
        cell.border = brd; cell.font = Font(name='Arial', size=9)
        if ci in (12, 13):
            cell.fill, cell.alignment = mixte_fill, c_al
            cell.font = Font(name='Arial', size=9, bold=True)
        elif ci in (2, 3):
            cell.fill, cell.alignment = row_fill, l_al
        else:
            cell.fill, cell.alignment = row_fill, c_al
    ws.row_dimensions[row_n].height = 45

ws2 = wb.create_sheet('Stats')
ws2['A1'] = 'Statistiques flag_mixte v1'
ws2['A1'].font = Font(bold=True, size=13, color='7E3F8F')
for r_idx, (lbl, val) in enumerate([
    ('Total documents',         len(results)),
    ('flag_mixte = 1',          len(flagged)),
    ('  dont mixte+suggestion', len(mixte_et_sug)),
    ('flag_mixte = 0',          len(results) - len(flagged)),
], start=3):
    ws2.cell(row=r_idx, column=1, value=lbl).font = Font(name='Arial', size=10)
    ws2.cell(row=r_idx, column=2, value=val).font  = Font(name='Arial', size=10, bold=True)
ws2.column_dimensions['A'].width = 45

wb.save(OUTPUT_EXCEL)
print(f'✅ Excel généré : {OUTPUT_EXCEL}  ({len(flagged)} lignes)')
client.close()