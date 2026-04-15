"""
DÉTECTION FLAG_SUGGESTION DZ — v1
══════════════════════════════════════════════════════════════════════════
Insight clé du corpus :
  "Positif Conditionnel DZ" — satisfaction partielle très fréquente dans
  les avis algériens en ligne. L'utilisateur est globalement satisfait
  MAIS signale un manque ou un potentiel d'amélioration.

  Structure obligatoire : [Mot positif DZ] + [Connecteur adversatif DZ]
                          + [Verbe / mot d'amélioration DZ]

  Sous-types :
    (1, 'pur')        → 3 conditions réunies, aucun négatif fort
                        → Positif conditionnel (60–70%) — jamais Négatif
    (1, 'pos_adv')    → positif + adversatif SANS verbe d'amélioration
                        → Positif atténué faible (50–60%)
    (-1,'masked_neg') → les 3 conditions présentes MAIS négatif fort aussi
                        → flag_mixte l'emporte, flag_suggestion supprimé
    (0, None)         → pas un positif conditionnel

  Règle d'or :
    flag_suggestion = 1 (pur)     si A ∧ B ∧ C ∧ ¬NEG_FORT
    flag_suggestion = 1 (pos_adv) si A ∧ B ∧ ¬C ∧ ¬NEG_FORT
    flag_suggestion = -1          si (A ∧ B ∧ C) ∧ NEG_FORT  → flag_mixte domine
    flag_suggestion = 0           sinon
"""

import re
from collections import Counter
from pymongo import MongoClient, UpdateOne
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════════════
MONGO_URI       = "mongodb://localhost:27017/"
DB_NAME         = "telecom_algerie"
COLLECTION_NAME = "commentaires_normalises_tfidf"
TEXT_FIELD      = "normalized_arabert"
OUTPUT_EXCEL    = "commentaires_flag_suggestion_v1.xlsx"

print("✅ Config chargée.")

# ══════════════════════════════════════════════════════════════════════════
# 1. CONDITION A — MOTS POSITIFS DZ
#    Présence d'un terme évaluatif positif (qualité, satisfaction)
# ══════════════════════════════════════════════════════════════════════════
POS_PATTERNS = [
    # ── Darija latin / arabizi ─────────────────────────────────────────────
    r'\bmliha?\b',          # mliha / mlih
    r'\bmzian[ae]?\b',      # mzian / mziana
    r'\bzwin[ae]?\b',       # zwina / zwine
    r'\bhil[ae]\b',         # hila / hile
    r'\btamam\b',           # tamam (trop bien)
    r'\bmazyan[ae]?\b',     # mazyan / mazyana
    r'\bsari3\b',           # rapide (arabizi)
    r'\bb[ae]rak\s*allah\b',# baraka allah (positif pur, pas social ici)
    # ── Français ──────────────────────────────────────────────────────────
    r'\bbien\b',
    r'\bbon(ne)?\b',
    r'\bsuper\b',
    r'\bpas\s*mal\b',
    r'\bcorrect(e)?\b',
    r'\bsatisfais?ant(e)?\b',
    r'\brapide\b',
    r'\befficace\b',
    r"\bc'est\s*bien\b",        
    r'\bça\s*march[e]?\b',
    r'\bça\s*fonctionn\w+\b',
    r'\bbon\s*d[ée]but\b',
    r'\bencourageant(e)?\b',
    r'\bglobalement\s*(positif|bien|correct)\b',
    r'\bbon\s*travail\b',
    r'\bp[ae]s\s*mal\b',
    r'\bbonne\s*(qualité|vitesse|connexion)\b',
    r'\bla\s*connexion\s*(est\s*)?(bonne|correcte|rapide|stable)\b',
    # ── Arabe / Darija arabe ───────────────────────────────────────────────
    r'مليح[ةه]?',
    r'زوين[ةه]?',
    r'\bتمام\b',
    r'ممتاز[ةه]?',
    r'سريع[ةه]?',
    r'جيد[ةه]?',
    r'حسن[ةه]?',
    r'مقبول[ةه]?',
    r'هايل[ةه]?',
    r'رائع[ةه]?',
    r'مزيان[ةه]?',
    r'معقول[ةه]?',
    r'كويس[ةه]?',
    r'نظيف[ةه]?',
    r'مرتاح[ةه]?',
    r'راضي[ةه]?',
    r'خدمة\s*(مليحة|جيدة|حسنة|معقولة)',
    r'الخدمة\s*(مليحة|جيدة|ممتازة)',
    r'اتصال\s*(مليح|جيد|سريع)',
    r'الانترنت\s*(مليح|مليحة|سريع|تمام)',
    r'سرعة\s*(كويسة|مليحة|جيدة)',
]

# ══════════════════════════════════════════════════════════════════════════
# 2. CONDITION B — CONNECTEURS ADVERSATIFS DZ
#    Marqueur d'opposition / de restriction (mais, cependant, seulement)
# ══════════════════════════════════════════════════════════════════════════
ADV_PATTERNS = [
    # ── Darija latin ──────────────────────────────────────────────────────
    r'\bms\b',              # mais (darija lat.)
    r'\bbss\b',             # seulement / mais
    r'\blukan\b',           # si seulement
    r'\blou?kan\b',
    r'\bwalakin\b',         # mais (arabe dial.)
    # ── Français ──────────────────────────────────────────────────────────
    r'\bmais\b',
    r'\bcependant\b',
    r'\btoutefois\b',
    r'\bseulement\b',
    r'\bpar\s*contre\b',
    r'\bn[eé]anmoins\b',
    r'\bsi\s*possible\b',
    r'\bsi\s*seulement\b',
    r'\bsauf\s*que\b',
    r'\bexcept[eé]\s*que\b',
    r'\bà\s*condition\s*que\b',
    # ── Arabe / Darija arabe ───────────────────────────────────────────────
    r'\bبصح\b',             # mais (darija ar.) — PRIORITÉ HAUTE
    r'\bلكن\b',
    r'\bوالكن\b',
    r'\bغير\s*أن\b',
    r'\bإلا\s*أن\b',
    r'\bبس\b',
    r'\bلو\s*كان\b',
    r'لو\s*كانو',
    r'\bفقط\s*لو\b',
    r'\bولكن\b',
    r'\bمع\s*ذلك\b',
    r'\bرغم\s*ذلك\b',
    r'\bبالرغم\s*من\b',
]

# ══════════════════════════════════════════════════════════════════════════
# 3. CONDITION C — VERBES / MOTS D'AMÉLIORATION DZ
#    Demande explicite d'une action correctrice ou d'un ajout
# ══════════════════════════════════════════════════════════════════════════
AMEL_PATTERNS = [
    # ── Darija latin — verbes d'action ────────────────────────────────────
    r'\btehsen[ou]*\b',     # améliorez
    r'\btzid[ou]*\b',       # continuez / ajoutez
    r'\bbeddl[ou]*\b',      # changez
    r'\bcorrigi\b',         # corrigez
    r'\bdebugi\b',          # déboguez
    r'\brajou[ou]*\b',      # revenez sur
    r'\bml[ou]+\b',         # remplissez
    r'\bkamml[ou]*\b',      # complétez
    r'\bhasn[ou]*\b',       # améliorez (var.)
    # ── Français — demandes constructives ─────────────────────────────────
    r'amélior\w+',
    r'optim[iu]s\w+',
    r'augment\w+\s*(la\s*|le\s*)?(vitesse|débit|qualité)',
    r'r[eé]duir\w+\s*(le\s*|les\s*)?prix',
    r'baiss\w+\s*(le\s*|les\s*)?prix',
    r'peaufin\w+',
    r'ajust\w+',
    r'travaill\w+\s*sur',
    r'des?\s*am[eé]liorations?\b',
    r'des?\s*efforts?\b',
    r'un\s*peu\s*plus\b',
    r'encore\s*mieux\b',
    r'plus\s*de\s*(rapidité|vitesse|qualité|stabilité)',
    r'il\s*(faut|faudrait)\s*(encore|plus)',
    r'il\s*(reste|y\s*a)\s*(des\s*)?am[eé]liorations?\s*[àa]\s*faire',
    r'ce\s*serait\s*(bien|mieux|parfait)\s*(de|si)',
    r"j'aimerais?\s*(plus|voir|que)",
    r'on\s*attend\s*(mieux|plus)',
    r'si\s*possible\s*(am[eé]lior|r[eé]duir|augment|opt)',
    # ── Arabe / Darija arabe ───────────────────────────────────────────────
    r'تحسين',
    r'يتحسن',
    r'تحسنو',
    r'تزيدو',
    r'نتمنى\s*(يتحسن|تحسين|أكثر|أفضل)',
    r'يلزم\s*تحسن',
    r'خاصكم\s*تحسنو',
    r'لو\s*(زادو|تحسنو|رفعو)',
    r'نشاء\s*الله\s*يتحسن',
    r'إن\s*شاء\s*الله\s*يتحسن',
    r'المزيد\s*من\s*(الجهود|التحسين|التطوير)',
    r'نأمل\s*(في\s*)?(تحسين|تطوير)',
    r'نتمنى\s*(تطوير|تحسين)',
    r'يجب\s*(تحسين|تطوير|رفع)',
    r'أتمنى\s*(أن\s*)?(يتحسن|تحسين)',
    r'نقص\s*في\s*(السعر|الخدمة|التغطية)',    # manque = demande implicite
    r'لو\s*كان\s*السعر\s*(أقل|منخفض)',
    r'لو\s*تخفضو\s*السعر',
    r'ينقصه?\s*(فقط|بس)',                    # il manque juste
    r'ينقص\s*(شوية|قليل|قليلاً)',
]

# ══════════════════════════════════════════════════════════════════════════
# 4. NÉGATIFS FORTS — annulent flag_suggestion → flag_mixte domine
#    Si présents avec A+B+C → flag = -1 / 'masked_neg'
# ══════════════════════════════════════════════════════════════════════════
NEG_FORT_PATTERNS = [
    # ── Mots négatifs forts DZ ─────────────────────────────────────────────
    r'\bnul\b', r'\bkhaybe?\b', r'\bkhayba?\b',
    r'\bmachi\s*barka\b', r'\bkarica\b', r'\bkhartha?\b',
    r'كارثة', r'كارثه', r'كارثي[هة]?',
    r'فضيحة', r'مهزلة',
    r'\bzero\b', r'تحت\s*الصفر',
    # ── Non-fonctionnement ────────────────────────────────────────────────
    r'لا\s*يعمل', r'ما\s*يخدم', r'مايخدمش',
    r'مقطوع[هة]?', r'قطوع\b',
    r'coupure\b', r'انقطاع\b',
    # ── Plaintes explicites ───────────────────────────────────────────────
    r'شكوى', r'شكاية', r'شكاوي',
    r'مشكل[ةه]?', r'مشاكل\b',
    r'r[eé]clamation\b', r'plainte\b',
    # ── Durée d'attente + frustration ─────────────────────────────────────
    r'\d+\s*(أشهر|mois|شهر|jours?|أيام|ايام)\s*(بدون|sans|دون|ولم)',
    r'شهرين.*ما\s*ركب',
    r'ست\s*اشهر',
    r'مازال\s*(ما|لم)\s*(جا[ءو]?|ركب)',
    # ── Frustration forte ────────────────────────────────────────────────
    r'تعبتونا', r'هرمنا', r'هبلتونا',
    r'وعود\s*كاذبة',
    r'مافيا\b', r'فساد\b',
    r'هل\s*يعقل',
    # ── Sarcasme explicite ────────────────────────────────────────────────
    r'والله.*يا\s*(ذيك|راحه)',
    r'في\s*الأحلام\b', r'في\s*المنام\b',
]

# ══════════════════════════════════════════════════════════════════════════
# COMPILATION
# ══════════════════════════════════════════════════════════════════════════
FLAGS = re.IGNORECASE | re.UNICODE

POS_REGEX      = re.compile('|'.join(POS_PATTERNS),      FLAGS)
ADV_REGEX      = re.compile('|'.join(ADV_PATTERNS),      FLAGS)
AMEL_REGEX     = re.compile('|'.join(AMEL_PATTERNS),     FLAGS)
NEG_FORT_REGEX = re.compile('|'.join(NEG_FORT_PATTERNS), FLAGS | re.DOTALL)

def get_match(regex, text):
    m = regex.search(text or ''); return m.group() if m else None

def is_suggestion(text):
    """
    Retourne (flag, subtype) :
      (1,  'pur')        → A ∧ B ∧ C ∧ ¬NEG  — Positif conditionnel fort
      (1,  'pos_adv')    → A ∧ B ∧ ¬C ∧ ¬NEG — Positif atténué faible
      (-1, 'masked_neg') → A ∧ B ∧ C ∧ NEG   — Masqué par négatif fort
      (0,  None)         → pas de positif conditionnel
    """
    t = text or ''
    has_pos  = bool(POS_REGEX.search(t))
    has_adv  = bool(ADV_REGEX.search(t))
    has_amel = bool(AMEL_REGEX.search(t))
    has_neg  = bool(NEG_FORT_REGEX.search(t))

    # Cas de base : au moins pos + adversatif requis
    if not (has_pos and has_adv):
        return 0, None

    # A ∧ B ∧ C ∧ NEG → masqué — flag_mixte domine
    if has_amel and has_neg:
        return -1, 'masked_neg'

    # A ∧ B ∧ C ∧ ¬NEG → positif conditionnel pur
    if has_amel and not has_neg:
        return 1, 'pur'

    # A ∧ B ∧ ¬C ∧ ¬NEG → positif atténué (adversatif sans verbe d'amélioration)
    if not has_amel and not has_neg:
        return 1, 'pos_adv'

    return 0, None

# ══════════════════════════════════════════════════════════════════════════
# TESTS UNITAIRES
# ══════════════════════════════════════════════════════════════════════════
tests = [
    # ── flag=1 / 'pur' — les 3 conditions ─────────────────────────────────
    ('connexion mliha ms tehseno fiha',                                    (1, 'pur')),
    ("l'appli zwina bss lukan tzidou fiha",                                (1, 'pur')),
    ('service mliha ms debugiw lapplication',                              (1, 'pur')),
    ("c'est bien mais des améliorations seraient bienvenues",              (1, 'pur')),
    ('pas mal, mais si possible réduire le prix',                          (1, 'pur')),
    ('مليح بصح يلزم تحسن',                                                (1, 'pur')),
    ('نتمنى يتحسن خدمتكم مزيان لكن تحسينات ضرورية',                      (1, 'pur')),
    ('الانترنت تمام بصح ينقصه قليل في السرعة',                            (1, 'pur')),
    ('connexion correcte mais il faudrait encore améliorer la stabilité',  (1, 'pur')),
    ('super ms tzido tehseno fiha khir',                                   (1, 'pur')),
    ('bon début mais il reste des améliorations à faire',                  (1, 'pur')),
    ('سريع لكن نتمنى تحسين التغطية',                                      (1, 'pur')),
    # ── flag=1 / 'pos_adv' — positif + adversatif sans verbe amélioration ─
    ('mliha ms bezzaf',                                                    (1, 'pos_adv')),
    ("c'est bien mais pas parfait",                                        (1, 'pos_adv')),
    ('جيد لكن ليس كافياً',                                                (1, 'pos_adv')),
    ('bon par contre un peu lent',                                         (1, 'pos_adv')),
    ('الخدمة حسنة بصح ناقصة شوية',                                        (1, 'pos_adv')),
    # ── flag=-1 / 'masked_neg' — négatif fort présent ─────────────────────
    ('machi barka nul ms tehseno',                                         (-1, 'masked_neg')),
    ('شكوى مشكلة لكن نتمنى تحسين',                                        (-1, 'masked_neg')),
    ('connexion mliha ms coupure tzidou corrigi',                          (-1, 'masked_neg')),
    # ── flag=0 — pas un positif conditionnel ──────────────────────────────
    ('bon courage',                                                        (0, None)),
    ('ربي يوفقكم',                                                         (0, None)),
    ('مليح',                                                               (0, None)),   # positif seul
    ('مليح لكن',                                                           (0, None)),   # sans amélioration ni verbe → pos_adv? non : "لكن" seul sans contexte
    ('شكوى منذ 3 mois sans réponse',                                       (0, None)),   # plainte sans positif
    ('la connexion est catastrophique',                                    (0, None)),   # négatif pur
]

print('\n── Tests ──────────────────────────────────────────────────────────────────')
ok = 0
for text, expected in tests:
    flag, stype = is_suggestion(text)
    status = '✅' if (flag, stype) == expected else '❌'
    print(f'  {status} [{flag:+d}/{str(stype):12s}] {text[:72]!r}')
    ok += ((flag, stype) == expected)
print(f'\n  {ok}/{len(tests)} tests passés\n')

# ══════════════════════════════════════════════════════════════════════════
# CONNEXION MONGODB
# ══════════════════════════════════════════════════════════════════════════
print('⏳ Connexion MongoDB …')
client     = MongoClient(MONGO_URI)
collection = client[DB_NAME][COLLECTION_NAME]
docs = list(collection.find({}, {
    '_id': 1, TEXT_FIELD: 1,
    'Commentaire_Client_Original': 1,
    'sources': 1, 'dates': 1, 'label_final': 1,
    'flag_social': 1, 'flag_encouragement': 1, 'flag_plainte': 1,
}))
print(f'✅ {len(docs)} documents chargés.')

# ══════════════════════════════════════════════════════════════════════════
# TRAITEMENT
# ══════════════════════════════════════════════════════════════════════════
results = []
for doc in docs:
    text = doc.get(TEXT_FIELD, '') or ''
    flag, stype = is_suggestion(text)

    if stype == 'pur':
        p = get_match(POS_REGEX,  text) or ''
        a = get_match(ADV_REGEX,  text) or ''
        c = get_match(AMEL_REGEX, text) or ''
        expr = ' | '.join(filter(None, [p, a, c]))
    elif stype == 'pos_adv':
        p = get_match(POS_REGEX, text) or ''
        a = get_match(ADV_REGEX, text) or ''
        expr = ' | '.join(filter(None, [p, a]))
    elif stype == 'masked_neg':
        n = get_match(NEG_FORT_REGEX, text) or ''
        expr = f'[NEG] {n}'
    else:
        expr = ''

    results.append({
        '_id':                  doc['_id'],
        'doc_id':               str(doc['_id']),
        'commentaire_original': doc.get('Commentaire_Client_Original', ''),
        'texte_normalise':      text,
        'expression_detectee':  expr,
        'source':               doc.get('sources', ''),
        'date':                 doc.get('dates', ''),
        'label_final':          doc.get('label_final', ''),
        'flag_social':          doc.get('flag_social', 0),
        'flag_encouragement':   doc.get('flag_encouragement', 0),
        'flag_plainte':         doc.get('flag_plainte', 0),
        'flag_suggestion':      flag,
        'suggestion_type':      stype or '',
    })

flagged      = [r for r in results if r['flag_suggestion'] == 1]
flag_pur     = [r for r in flagged  if r['suggestion_type'] == 'pur']
flag_posadv  = [r for r in flagged  if r['suggestion_type'] == 'pos_adv']
suppressed   = [r for r in results  if r['flag_suggestion'] == -1]
counter      = Counter(r['expression_detectee'].split(' | ')[0]
                       for r in flagged if r['expression_detectee'])

print(f'\n📊 Statistiques v1 :')
print(f'   Total docs                          : {len(results)}')
print(f'   flag_suggestion = 1  (positif cond.): {len(flagged)}')
print(f'   ├─ pur      → Positif conditionnel (A∧B∧C)     : {len(flag_pur)}')
print(f'   └─ pos_adv  → Positif atténué      (A∧B, sans C): {len(flag_posadv)}')
print(f'   flag_suggestion = -1 (masked_neg)   : {len(suppressed)}')
print(f'   flag_suggestion = 0  (neutre)       : {len(results)-len(flagged)-len(suppressed)}')
print(f'\n🔑 Top 15 expressions déclenchantes :')
for expr, cnt in counter.most_common(15):
    print(f'   {cnt:4d}  {expr!r}')

# ══════════════════════════════════════════════════════════════════════════
# MISE À JOUR MONGODB
# ══════════════════════════════════════════════════════════════════════════
print('\n⏳ Mise à jour MongoDB …')
operations = [
    UpdateOne(
        {'_id': r['_id']},
        {'$set': {
            'flag_suggestion':      r['flag_suggestion'],
            'suggestion_type':      r['suggestion_type'],
            'expression_suggestion': r['expression_detectee'],
        }}
    )
    for r in results
]
res = collection.bulk_write(operations)
print(f'✅ MongoDB mis à jour — {res.modified_count} docs modifiés.')

# ══════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════════════
print('⏳ Génération Excel …')

# ── Couleurs ──────────────────────────────────────────────────────────────
hdr_fill     = PatternFill('solid', fgColor='1A5276')   # bleu foncé — header
hdr_font     = Font(color='FFFFFF', bold=True, size=11)
pur_fill     = PatternFill('solid', fgColor='A9DFBF')   # vert — pur (positif conditionnel)
posadv_fill  = PatternFill('solid', fgColor='D5F5E3')   # vert clair — pos_adv
masked_fill  = PatternFill('solid', fgColor='FAD7A0')   # orange — masked_neg
odd_fill     = PatternFill('solid', fgColor='EBF5FB')
even_fill    = PatternFill('solid', fgColor='FFFFFF')
c_al = Alignment(horizontal='center', vertical='center')
l_al = Alignment(horizontal='left',   vertical='top', wrap_text=True)
thin = Side(style='thin', color='BFBFBF')
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = 'flag_suggestion_v1'

headers    = ['doc_id', 'commentaire_original', 'texte_normalise',
              'expression_detectee', 'source', 'date',
              'label_final', 'flag_social', 'flag_encouragement',
              'flag_plainte', 'flag_suggestion', 'suggestion_type']
col_widths = [28, 60, 55, 40, 12, 18, 12, 11, 18, 11, 14, 13]

for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.fill, cell.font = hdr_fill, hdr_font
    cell.alignment = c_al; cell.border = brd
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 28
ws.freeze_panes = 'A2'

type_fill = {
    'pur':        pur_fill,
    'pos_adv':    posadv_fill,
    'masked_neg': masked_fill,
}
export_rows = flagged + suppressed

for i, r in enumerate(export_rows):
    row_n    = i + 2
    row_fill = odd_fill if i % 2 == 0 else even_fill
    stype    = r['suggestion_type']
    tf       = type_fill.get(stype, row_fill)
    vals = [r['doc_id'], r['commentaire_original'], r['texte_normalise'],
            r['expression_detectee'], r['source'], r['date'],
            r['label_final'], r['flag_social'], r['flag_encouragement'],
            r['flag_plainte'], r['flag_suggestion'], stype]
    for ci, val in enumerate(vals, 1):
        cell = ws.cell(row=row_n, column=ci, value=val)
        cell.border = brd; cell.font = Font(name='Arial', size=9)
        if ci in (11, 12):
            cell.fill, cell.alignment = tf, c_al
            cell.font = Font(name='Arial', size=9, bold=True)
        elif ci in (2, 3):
            cell.fill, cell.alignment = row_fill, l_al
        else:
            cell.fill, cell.alignment = row_fill, c_al
    ws.row_dimensions[row_n].height = 45

# ── Feuille Stats ─────────────────────────────────────────────────────────
ws2 = wb.create_sheet('Stats')
ws2['A1'] = 'Statistiques flag_suggestion v1'
ws2['A1'].font = Font(bold=True, size=13, color='1A5276')
for r_idx, (lbl, val) in enumerate([
    ('Total documents',                                                         len(results)),
    ('flag_suggestion = 1  (positif conditionnel)',                             len(flagged)),
    ('  ├─ pur      → Positif conditionnel (A∧B∧C)',                            len(flag_pur)),
    ('  └─ pos_adv  → Positif atténué (A∧B sans verbe amélioration)',           len(flag_posadv)),
    ('flag_suggestion = -1 (masked_neg → flag_mixte domine)',                   len(suppressed)),
    ('flag_suggestion = 0  (neutre, pas de pattern)',                           len(results)-len(flagged)-len(suppressed)),
], start=3):
    ws2.cell(row=r_idx, column=1, value=lbl).font = Font(name='Arial', size=10)
    ws2.cell(row=r_idx, column=2, value=val).font  = Font(name='Arial', size=10, bold=True)

ws2['A11'] = 'Top expressions déclenchantes'
ws2['A11'].font = Font(bold=True, size=10)
for r_idx, (expr, cnt) in enumerate(counter.most_common(20), start=12):
    ws2.cell(row=r_idx, column=1, value=expr)
    ws2.cell(row=r_idx, column=2, value=cnt)
ws2.column_dimensions['A'].width = 60
ws2.column_dimensions['B'].width = 15

# ── Feuille Légende ───────────────────────────────────────────────────────
ws3 = wb.create_sheet('Légende')
ws3['A1'] = 'Légende — flag_suggestion v1'
ws3['A1'].font = Font(bold=True, size=12, color='1A5276')
for r_idx, (fill, nom, desc) in enumerate([
    (pur_fill,    'Vert       — pur         flag=+1', 'A∧B∧C ∧ ¬NEG → Positif conditionnel (60–70%) — jamais Négatif'),
    (posadv_fill, 'Vert clair — pos_adv     flag=+1', 'A∧B ∧ ¬C ∧ ¬NEG → Positif atténué (50–60%)'),
    (masked_fill, 'Orange     — masked_neg  flag=-1', 'A∧B∧C ∧ NEG → flag_mixte domine, flag_suggestion supprimé'),
], start=3):
    c = ws3.cell(row=r_idx, column=1, value=nom)
    c.fill, c.font = fill, Font(name='Arial', size=10, bold=True)
    ws3.cell(row=r_idx, column=2, value=desc).font = Font(name='Arial', size=10)

ws3['A8'] = 'Règle logique :'
ws3['A8'].font = Font(bold=True, size=10)
for r_idx, desc in enumerate([
    "Condition A : mot positif DZ (mliha, zwina, bien, مليح, تمام, ...)",
    "Condition B : connecteur adversatif DZ (ms, mais, بصح, لكن, walakin, ...)",
    "Condition C : verbe / mot d'amélioration DZ (tehseno, amélior, تحسين, tzidou, ...)",
    "NEG FORT    : négatif fort (nul, khayba, كارثة, مشكلة, coupure, ...)",
    "",
    "flag_suggestion = 1 (pur)       si A ∧ B ∧ C ∧ ¬NEG",
    "flag_suggestion = 1 (pos_adv)   si A ∧ B ∧ ¬C ∧ ¬NEG",
    "flag_suggestion = -1 (masked)   si A ∧ B ∧ C ∧ NEG  → flag_mixte domine",
    "flag_suggestion = 0             sinon",
], start=9):
    ws3.cell(row=r_idx, column=1, value=desc).font = Font(name='Arial', size=10)
ws3.column_dimensions['A'].width = 75
ws3.column_dimensions['B'].width = 70

wb.save(OUTPUT_EXCEL)
print(f'✅ Excel généré : {OUTPUT_EXCEL}  ({len(export_rows)} lignes)')
client.close()
