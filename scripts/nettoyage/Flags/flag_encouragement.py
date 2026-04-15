"""
DÉTECTION FLAG_ENCOURAGEMENT DZ — version REGEX uniquement
Logique révisée :
  - Tout encouragement posté sur la page AT est valide (pas besoin de mention AT explicite)
  - Seuls quelques cas sont vraiment hors-sujet (foot/sport étranger, politique, texte informatif)
  - 2 sous-types : 'pur' (neutre forcé) | 'produit' (positif faible)
"""

import re
from collections import Counter
from pymongo import MongoClient, UpdateOne
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════════════
MONGO_URI       = "mongodb://localhost:27017/"
DB_NAME         = "telecom_algerie"
COLLECTION_NAME = "commentaires_normalises_tfidf"
TEXT_FIELD      = "normalized_arabert"
OUTPUT_EXCEL    = "commentaires_flag_encouragement_1.xlsx"

print("✅ Config chargée.")

# ══════════════════════════════════════════════════════════════════════════
# 1. PATTERNS ENCOURAGEMENT
# ══════════════════════════════════════════════════════════════════════════
ENCOURAGEMENT_PATTERNS = [
    # Français
    r'bon\s*courage', r'bonne?\s*continuation', r'bonne?\s*chance',
    r'f[eé]licitations?', r'bravo\b', r'chapeau\b', r'all\s*the\s*best',
    r'bonne?\s*r[eé]ussite',r'chapeau\b', r'mention\b',
    r'meilleurs?\s*v[oœ]ux', r'tous\s*mes\s*v[oœ]ux', r'bon\s*succ[eè]s',
    r'succ[eè]s\s*continu', r'joyeuses?\s*f[eê]tes?', r'bonne?\s*f[eê]te',
    r'joyeux\s*ramadan', r'toujours\s*plus\s*(loin|haut|fort|proche|innovant)',
    r'keep\s*it\s*up', r'well\s*done', r'good\s*luck',
    # Darija latin / Arabizi
    r'barra\s*barra',
    r'[ae]ychine?\b', r'3aychine?\b',
    r'nchallah\s*ynajh', r'inch[a]?allah\s*ynajh',
    r'mabrouk\b', r'mbr[ou]+k\b',
    r'rbi\s*y[3e]awkom', r'allah\s*yewfek', r'allah\s*yewfeqkom',
    r'tbarakallah\b', r'tbarkallah\b',
    r'saha\s*[ae]idkom', r'sa7a\s*[ae]idkom',
    r'kol\s*[3e]am\s*w', r'[ae]id\s*mobarak',
    r'barak\s*allah', r'allah\s*ybarek',
    # Arabe / Darija arabe
    r'بالتوفيق', r'بتوفيق', r'ألف\s*مبروك', r'مبروك\b',
    r'ماشاء\s*الله', r'ما\s*شاء\s*الله', r'تبارك\s*الله',
    r'بارك\s*الله\s*فيكم', r'بارك\s*الله\s*فيك',
    r'ربي\s*يوفق', r'ربي\s*يوفقكم', r'الله\s*يوفق', r'الله\s*يبارك',
    r'نتمنى\s*(لكم|لك)\s*(النجاح|التوفيق|التألق|الاستمرار)',
    r'كل\s*عام\s*و[أا]نتم\s*بخير',
    r'سنة\s*(جديدة|مليئة)', r'عام\s*(جديد|مبارك)',
    r'إلى\s*الأمام',
    r'دائما\s*(في\s*تقدم|أقرب|إلى\s*الأمام)',
    r'المزيد\s*من\s*(النجاح|التألق|الإنجازات)',
    r'مزيد\s*من\s*(التألق|النجاح|التطور)',
    r'خطوة\s*هايلة', r'في\s*القمة', r'برافو\b',
    r'حاجة\s*مليحة',  
]

# ══════════════════════════════════════════════════════════════════════════
# 2. PATTERNS NÉGATIFS — annulent le flag
# ══════════════════════════════════════════════════════════════════════════
NEGATIVE_PATTERNS = [
    # Français
    r'réclamation', r'plainte', r'problème', r'panne', r'non\s*activé',
    r'ne\s*fonctionne\s*pas', r'marche\s*pas',
    r'débit\s*(lent|faible|catastrophique)', r'connexion\s*lente', r'coupure',
    r'depuis\s*\d+\s*(jours?|mois|semaines?)', r'sans\s*réponse',
    r'nul\b', r'inexistant', r'catastrophe', r'honte',
    # Arabe / Darija
    r'شكوى', r'شكاية', r'شكاوي', r'مشكل', r'مشكلة',
    r'لا\s*يعمل', r'مقطوع', r'قطوع', r'بطيء', r'ضعيف',
    r'ما\s*خدم', r'مكانش', r'ما\s*ركبوا', r'ما\s*تواصل', r'ما\s*جاو',
    r'ما\s*حلوا', r'ما\s*صلحوا', r'ما\s*ردوا',
    r'\d+\s*(أشهر|يوم|سنة).*(?:انتظار|دون|بدون)',
    r'معاناة', r'اذلال', r'تعبتونا', r'هبلونا',
    r'وعود\s*كاذبة', r'بدون\s*حل', r'دون\s*أي\s*حل',
    r'فساد', r'مافيا', r'سجن', r'ديكتاتور',
    # Darija latin
    r'mochkil', r'ma\s*khedmetch', r'makach',
    r'hchouma', r'ta3batna', r'hablona',
    # zidou dans plainte
    r'tzido.{0,20}(batel|tgt3|mochkil)',
    r'zidou.{0,20}(batel|tgt3|mochkil)',
    # عام جديد + critique
    r'عام\s*(جديد|سعيد).{0,60}(خيوط|انقطاع|نحاس|adsl|مشكل|عيب)',
    r'(خيوط|انقطاع|نحاس|adsl|مشكل|عيب).{0,60}عام\s*(جديد|سعيد)',
]

# ══════════════════════════════════════════════════════════════════════════
# 3. PATTERNS VRAIMENT HORS-SUJET
#    Seuls ces cas sont exclus : sport étranger, politique, texte culturel/info
# ══════════════════════════════════════════════════════════════════════════
OFF_TOPIC_PATTERNS = [
    # Équipes sportives nommées (pas AT)
    r'\busmk\b', r'\busm\s*khenchela\b', r'\bcrb\b', r'\bmca\b', r'\busma\b',
    r'\bmc\s*alger\b', r'فريق\s+(?!اتصالات)',  # فريق + autre chose qu'اتصالات
    r'لفريق\s+\w+(?<!اتصالات)',
    # Politique / hors-sujet total
    r'على\s*حساب\s*اموال\s*الشعب',
    r'ديكتاتور', r'تبون', r'بوتفليقة',
    # Texte informatif/culturel sans lien (long hors-sujet)
    r'يونيسكو', r'بلغاريا', r'تراث.*ثقافي',
    r'عصور\s*وثنية', r'قبل\s*التاريخ',
]

# ══════════════════════════════════════════════════════════════════════════
# 4. PATTERNS PRODUIT/SERVICE
# ══════════════════════════════════════════════════════════════════════════
PRODUCT_PATTERNS = [
    r'appli(cation)?\b', r'service\b', r'réseau\b', r'connexion\b', r'débit\b',
    r'fibre?\b', r'internet\b', r'offre\b', r'abonnement\b', r'prix\b', r'tarif\b',
    r'vitesse\b', r'signal\b', r'4g\b', r'5g\b', r'wifi\b', r'box\b', r'modem\b',
    r"l['\u2019]appli", r'le\s*service', r'le\s*réseau', r'la\s*connexion',
    r'gpon\b', r'xgs.?pon\b', r'fttx\b', r'adsl\b', r'vdsl\b', r'idoom\b',
    r'الانترنت', r'انترنت\b', r'الشبكة', r'شبكة\b',
    r'الفيبر', r'فيبر\b', r'الاشتراك', r'الخدمة',
    r'السعر', r'التدفق', r'الاتصال', r'المودم', r'مودم\b', r'الباقة',
    r'الالياف\s*البصرية', r'الياف\b',
    r'تطبيق', r'الاندرويد', r'التطبيق',
]

# ══════════════════════════════════════════════════════════════════════════
# COMPILATION
# ══════════════════════════════════════════════════════════════════════════
FLAGS = re.IGNORECASE | re.UNICODE
ENC_REGEX       = re.compile('|'.join(ENCOURAGEMENT_PATTERNS), FLAGS)
NEG_REGEX       = re.compile('|'.join(NEGATIVE_PATTERNS),      FLAGS | re.DOTALL)
OFF_TOPIC_REGEX = re.compile('|'.join(OFF_TOPIC_PATTERNS),     FLAGS)
PRODUCT_REGEX   = re.compile('|'.join(PRODUCT_PATTERNS),       FLAGS)

def get_match(text):
    m = ENC_REGEX.search(text or '')
    return m.group() if m else None

def is_encouragement(text):
    """
    Retourne (flag_encouragement, sous_type)
      (1, 'pur')     → encouragement valide, pas d'avis produit  → Neutre forcé
      (1, 'produit') → encouragement + avis produit              → Positif faible
      (0, 'negatif') → pattern négatif détecté                   → annulé
      (0, 'offtopic')→ vraiment hors-sujet (sport/politique)     → supprimé
      (0, None)      → pas un encouragement
    """
    t = text or ''
    if not ENC_REGEX.search(t):    return 0, None
    if NEG_REGEX.search(t):        return 0, 'negatif'
    if OFF_TOPIC_REGEX.search(t):  return 0, 'offtopic'
    if PRODUCT_REGEX.search(t):    return 1, 'produit'
    return 1, 'pur'

# ══════════════════════════════════════════════════════════════════════════
# TESTS
# ══════════════════════════════════════════════════════════════════════════
tests = [
    # ── Doivent être flag=1 ───────────────────────────────────────────────
    ('Bon courage les gars !',                         (1, 'pur')),
    ('Bravo et bonne continuation',                    (1, 'pur')),
    ('بالتوفيق والنجاح الدائم ان شاء الله',            (1, 'pur')),
    ('كل عام وانتم بخير',                              (1, 'pur')),
    ('Bonne continuation inchallah',                   (1, 'pur')),
    ('Félicitations',                                  (1, 'pur')),
    ('Good luck',                                      (1, 'pur')),
    ('assegas ameggaz chapeau',                        (1, 'pur')),
    ('بارك الله فيكم إتقاااان وتفاني في العمل',        (1, 'pur')),
    ('ربي يوفقكم',                                     (1, 'pur')),
    ('برافو مزيد من التألق والنجاح',                    (1, 'pur')),
    # ── Encouragement + produit → flag=1/produit ─────────────────────────
    ('بالتوفيق والنجاح نتمنى تعميم الالياف البصرية',   (1, 'produit')),
    ('على العموم حاجة مليحة التطورات بالتوفيق',        (1, 'produit')),  # تطورات = produit context
    ('بالتوفيق اذا فيها السرعات المطروحة كافية مودم',  (1, 'produit')),
    ('bravo bonne continuation le service internet',   (1, 'produit')),
    # ── Doivent être flag=0 ───────────────────────────────────────────────
    ('على حساب اموال الشعب الف مبروك لفريق خنشلة',    (0, 'offtopic')),
    ('عجيب أمر هذه الدنيا يونيسكو بلغاريا',            (0, 'offtopic')),
    ('قديمة و ما عندها حتى معنى عيب تبداو بيها',       (0, 'negatif')),
    ('bon courage mais le service est nul',            (0, 'negatif')),
    ('عام جديد وما زالنا مع خيوط النحاس adsl',         (0, 'negatif')),
]
print('\n── Tests ──────────────────────────────────────────────')
ok = 0
for text, expected in tests:
    flag, stype = is_encouragement(text)
    status = '✅' if (flag, stype) == expected else '❌'
    print(f'  {status} [{flag}/{str(stype):10s}] {text[:60]!r}')
    ok += ((flag, stype) == expected)
print(f'\n  {ok}/{len(tests)} tests passés\n')

# ══════════════════════════════════════════════════════════════════════════
# CONNEXION MONGODB
# ══════════════════════════════════════════════════════════════════════════
print('⏳ Connexion MongoDB …')
client     = MongoClient(MONGO_URI)
collection = client[DB_NAME][COLLECTION_NAME]
docs = list(collection.find({}, {
    '_id': 1, TEXT_FIELD: 1,
    'Commentaire_Client_Original': 1,
    'sources': 1, 'dates': 1, 'label_final': 1, 'flag_social': 1,
}))
print(f'✅ {len(docs)} documents chargés.')

# ══════════════════════════════════════════════════════════════════════════
# TRAITEMENT
# ══════════════════════════════════════════════════════════════════════════
results = []
for doc in docs:
    text = doc.get(TEXT_FIELD, '') or ''
    flag, stype = is_encouragement(text)
    results.append({
        '_id':                  doc['_id'],
        'doc_id':               str(doc['_id']),
        'commentaire_original': doc.get('Commentaire_Client_Original', ''),
        'texte_normalise':      text,
        'expression_detectee':  get_match(text) or '',
        'source':               doc.get('sources', ''),
        'date':                 doc.get('dates', ''),
        'label_final':          doc.get('label_final', ''),
        'flag_social':          doc.get('flag_social', 0),
        'flag_encouragement':   flag,
        'encouragement_type':   stype or '',
    })

flagged    = [r for r in results if r['flag_encouragement'] == 1]
flag_pur   = [r for r in flagged if r['encouragement_type'] == 'pur']
flag_prod  = [r for r in flagged if r['encouragement_type'] == 'produit']
counter    = Counter(r['expression_detectee'] for r in flagged)

print(f'\n📊 Statistiques :')
print(f'   Total docs            : {len(results)}')
print(f'   flag_encouragement=1  : {len(flagged)}')
print(f'   └─ pur     → Neutre forcé   : {len(flag_pur)}')
print(f'   └─ produit → Positif faible : {len(flag_prod)}')
print(f'\n🔑 Top 10 expressions :')
for expr, cnt in counter.most_common(10):
    print(f'   {cnt:4d}  {expr!r}')

# ══════════════════════════════════════════════════════════════════════════
# MISE À JOUR MONGODB
# ══════════════════════════════════════════════════════════════════════════
print('\n⏳ Mise à jour MongoDB …')
operations = [
    UpdateOne(
        {'_id': r['_id']},
        {'$set': {
            'flag_encouragement':       r['flag_encouragement'],
            'encouragement_type':       r['encouragement_type'],
            'expression_encouragement': r['expression_detectee'],
        }}
    )
    for r in results
]
res = collection.bulk_write(operations)
print(f'✅ MongoDB mis à jour — {res.modified_count} docs modifiés.')

# ══════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL — flag_encouragement = 1 uniquement
# ══════════════════════════════════════════════════════════════════════════
print('⏳ Génération Excel …')
hdr_fill  = PatternFill('solid', fgColor='1F4E79')
hdr_font  = Font(color='FFFFFF', bold=True, size=11)
pur_fill  = PatternFill('solid', fgColor='D9E1F2')
prod_fill = PatternFill('solid', fgColor='C6EFCE')
odd_fill  = PatternFill('solid', fgColor='EBF3FB')
even_fill = PatternFill('solid', fgColor='FFFFFF')
c_al = Alignment(horizontal='center', vertical='center')
l_al = Alignment(horizontal='left',   vertical='top', wrap_text=True)
thin = Side(style='thin', color='BFBFBF')
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = 'flag_encouragement_1'

headers    = ['doc_id','commentaire_original','texte_normalise',
              'expression_detectee','source','date',
              'label_final','flag_social','flag_encouragement','encouragement_type']
col_widths = [28, 60, 55, 25, 12, 18, 12, 11, 18, 14]

for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.fill, cell.font = hdr_fill, hdr_font
    cell.alignment = c_al
    cell.border    = brd
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 28
ws.freeze_panes = 'A2'

type_fill = {'pur': pur_fill, 'produit': prod_fill}

for i, r in enumerate(flagged):
    row_n    = i + 2
    row_fill = odd_fill if i % 2 == 0 else even_fill
    stype    = r['encouragement_type']
    tf       = type_fill.get(stype, row_fill)

    vals = [r['doc_id'], r['commentaire_original'], r['texte_normalise'],
            r['expression_detectee'], r['source'], r['date'],
            r['label_final'], r['flag_social'], r['flag_encouragement'], stype]

    for ci, val in enumerate(vals, 1):
        cell = ws.cell(row=row_n, column=ci, value=val)
        cell.border = brd
        cell.font   = Font(name='Arial', size=9)
        if ci in (9, 10):
            cell.fill, cell.alignment = tf, c_al
            cell.font = Font(name='Arial', size=9, bold=True)
        elif ci in (2, 3):
            cell.fill, cell.alignment = row_fill, l_al
        else:
            cell.fill, cell.alignment = row_fill, c_al
    ws.row_dimensions[row_n].height = 45

# Stats
ws2 = wb.create_sheet('Stats')
ws2['A1'] = 'Statistiques flag_encouragement'
ws2['A1'].font = Font(bold=True, size=13, color='1F4E79')
for r_idx, (lbl, val) in enumerate([
    ('Total documents', len(results)),
    ('flag_encouragement = 1', len(flagged)),
    ('  └─ pur     → Neutre forcé',   len(flag_pur)),
    ('  └─ produit → Positif faible', len(flag_prod)),
], start=3):
    ws2.cell(row=r_idx, column=1, value=lbl).font = Font(name='Arial', size=10)
    ws2.cell(row=r_idx, column=2, value=val).font  = Font(name='Arial', size=10, bold=True)
ws2['A10'] = 'Top expressions'
ws2['A10'].font = Font(bold=True, size=10)
for r_idx, (expr, cnt) in enumerate(counter.most_common(20), start=11):
    ws2.cell(row=r_idx, column=1, value=expr)
    ws2.cell(row=r_idx, column=2, value=cnt)
ws2.column_dimensions['A'].width = 45

# Légende
ws3 = wb.create_sheet('Légende')
ws3['A1'] = 'Légende — flag_encouragement'
ws3['A1'].font = Font(bold=True, size=12, color='1F4E79')
for r_idx, (fill, nom, desc) in enumerate([
    (pur_fill,  'Bleu  — pur     flag=1', 'Encouragement valide, pas d\'avis produit → Neutre forcé'),
    (prod_fill, 'Vert  — produit flag=1', 'Encouragement + avis produit/service → Positif faible'),
], start=3):
    c = ws3.cell(row=r_idx, column=1, value=nom)
    c.fill, c.font = fill, Font(name='Arial', size=10, bold=True)
    ws3.cell(row=r_idx, column=2, value=desc).font = Font(name='Arial', size=10)
ws3['A7'] = 'Exclus (flag=0) :'
ws3['A7'].font = Font(bold=True, size=10)
for r_idx, desc in enumerate([
    "offtopic : félicite équipe sportive (USMK…), texte politique/culturel hors-sujet",
    "negatif  : contient une plainte ou critique → pattern négatif détecté",
], start=8):
    ws3.cell(row=r_idx, column=1, value=desc).font = Font(name='Arial', size=10)
ws3.column_dimensions['A'].width = 70

wb.save(OUTPUT_EXCEL)
print(f'✅ Excel généré : {OUTPUT_EXCEL}  ({len(flagged)} lignes)')
client.close()