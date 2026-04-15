"""
DÉTECTION FLAG_SOCIAL DZ — version REGEX uniquement
Détecte les commentaires de vœux/encouragements sociaux
Export Excel : uniquement flag_social = 1
"""

import re
from collections import Counter
from pymongo import MongoClient, UpdateOne
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════════════

# "mongodb+srv://yousrahadjabderrahmane_db_user:C8wjIvWqOBUjK66u@ac-1ksfahb-shard-00-00.gejzu4a.mongodb.net/?retryWrites=true&w=majority&appName=Cluster0"
MONGO_URI       = "mongodb://localhost:27017/"
DB_NAME         = "telecom_algerie"
COLLECTION_NAME = "commentaires_normalises"
TEXT_FIELD      = "normalized_arabert"
OUTPUT_EXCEL    = "commentaires_flag_social_1.xlsx"

print("✅ Config chargée.")

# ══════════════════════════════════════════════════════════════════════════
# PATTERNS POSITIFS — vœux / encouragements sociaux
# ══════════════════════════════════════════════════════════════════════════
SOCIAL_PATTERNS = [

    # ── Arabe / darija arabe ──────────────────────────────────────────────
    r'بارك\s*الله',
    r'الله\s*يبارك',
    r'الله\s*يعاون',
    r'ربي\s*يعاون',
    r'ربي\s*يجازي',
    r'جزاك\s*الله\s*خير',
    r'جزاكم\s*الله\s*خير',
    r'الله\s*يوفق',
    r'ربي\s*يحفظ',
    r'ربي\s*يوفق',
    r'بالتوفيق',
    r'رمضان\s*كريم',
    r'رمضان\s*مبارك',
    r'عيد\s*مبارك',
    r'عيد\s*سعيد',
    r'كل\s*عام\s*و[أا]نتم',
    r'كل\s*عام\s*وأنتم',
    r'مبروك',
    r'تبارك\s*الله',
    r'ماشاء\s*الله',
    r'ما\s*شاء\s*الله',
    r'الله\s*معاكم',
    r'الله\s*معاك\b',
    r'ربي\s*يوفقكم',
    r'ربي\s*يوفقك\b',
    r'الله\s*يوفقكم',
    r'شكرا\s*على\s*جهودكم',
    r'شكرا\s*على\s*المجهودات',
    r'ربي\s*يجازيكم',
    r'ربي\s*يعاونكم',
    r'إن\s*شاء\s*الله\s*ينجح',
    r'نشالله\s*ينجح',
    r'يسلموا?\b',
    r'الله\s*يسلمك',
    r'صح\s*فطورك',
    r'صح\s*فطوركم',
    r'صح\s*عيدك',
    r'صح\s*عيدكم',
    r'حفظكم\s*الله',
    r'وفقكم\s*الله',
    r'الله\s*يعطيكم\s*الصحة',
    r'ربي\s*يبارك\s*فيكم',
    r'ربي\s*يبارك\s*فيك\b',

    # ── Darija latin / arabizi ────────────────────────────────────────────
    r'rbi\s*y[3e]awkom',
    r'rabi\s*y[3e]awkom',
    r'barak\s*allah',
    r'baraka\s*llah',
    r'allah\s*ybarek',
    r'allah\s*yewfek',
    r'allah\s*yewfeqkom',
    r'tbarakallah',
    r'tbarkallah',
    r'nchallah\s*ynajh',
    r'inch[a]?allah\s*ynajh',
    r'saha\s*ftourkoum',
    r'saha\s*[ae]idkom',
    r'sa7a\s*[ae]idkom',
    r'kol\s*[3e]am\s*w',
    r'[ae]id\s*mobarak',
    r'[ae]id\s*sa[3i]id',
    r'mabrouk',
    r'mbr[ou]+k',
    r'bon\s*courage',
    r'bonne?\s*continuation',
    r'merci\s*pour\s*vos\s*efforts',
    r'merci\s*pour\s*le\s*travail',
    r'courage\s*[àa]\s*tou[st]',
    r'bonne?\s*ann[eé]e',
    r'assegwas\s*amegas',
    r'vive\s*alg[eé]rie',
    r'allah\s*ma[3e]akom',
    r'rabi\s*ma[3e]akom',
    r'barra\s*barra\s*likom',
    r'[ae]ychine\b',
    r'3aychine\b',
    r'rbi\s*yjazik',
    r'yeslamo\b',
    r'ya3tikom\s*essa7a',

    # ── Français ─────────────────────────────────────────────────────────
    r'joyeux\s*ramadan',
    r'bon\s*travail',
    r'bravo\s*(à\s*)?l[a-z]*\s*[eé]quipe',
    r'bonne?\s*f[eê]te',
    r'f[eé]licitations?',
    r'bonne?\s*ann[eé]e',
    r'bonne?\s*continuation',
    r'merci\s*pour\s*vos\s*efforts',
    r'bonne?\s*chance',
    r'bravo\s*pour',
    r'chapeau\s*pour',
    r'joyeuses?\s*f[eê]tes?',
    r'meilleurs?\s*v[oœ]eux',
    r'bonne?\s*journ[eé]e',
    r'bon\s*succ[eè]s',
    r'bonne?\s*r[eé]ussite',
    r'tous\s*mes\s*v[oœ]eux',
]

# ══════════════════════════════════════════════════════════════════════════
# PATTERNS NÉGATIFS — mots-clés qui ANNULENT le flag
# (pour éviter les faux positifs du type "جزاكم الله" dans une longue plainte)
# ══════════════════════════════════════════════════════════════════════════
NEGATIVE_PATTERNS = [
    r'مشكل',
    r'مشكلة',
    r'شكاية',
    r'شكوى', r'شكاوي',
    r'ما\s*خدم',
    r'مكانش',
    r'ما\s*عندي',
    r'ما\s*ركبوا',
    r'ما\s*تواصل',
    r'ما\s*جاو',
    r'ما\s*حلوا',
    r'ما\s*صلحوا',
    r'ما\s*ردوا',
    r'مشكل.*جزاك',
    r'انتظار',
    r'أنتظر',
    r'نستنى',
    r'وعود\s*كاذبة',
    r'سوء\s*الاستقبال',
    r'بدون\s*حل',
    r'دون\s*[أا]ي\s*حل',
    r'هل\s*يعقل',
    r'هل\s*منطق',
    r'نقطة\s*الصفر',
    r'معاناة',
    r'اضطهاد',
    r'ظلم',
    r'إرهاب',
    r'مافيا',
    r'فساد',
    r'سجن',
    r'ديكتاتور',
    r'تبون',
    r'بوتفليقة',
    r'فاضح',
    r'راكم\s*غير\s*تزيدوا\s*تبعدوا',
    r'خدمة.*شهر',
    r'شهر.*ما.*ركب',
    r'شهرين.*ولحد',
    r'ست\s*اشهر',
    r'شهر\s*ونص',
    r'يحرم\s*من\s*خدمة',
    r'لم\s*يتواصلوا',
    r'لم\s*نستطع',
    r'لم\s*يرد',
    r'خط\s*أخضر.*لا\s*يرد',
    r'تعبتونا',
    r'هبلونا',
    r'حيكم',
    r'اذلال',
]

SOCIAL_REGEX   = re.compile('|'.join(SOCIAL_PATTERNS),   flags=re.IGNORECASE | re.UNICODE)
NEGATIVE_REGEX = re.compile('|'.join(NEGATIVE_PATTERNS), flags=re.IGNORECASE | re.UNICODE)

def get_match(text):
    m = SOCIAL_REGEX.search(text or '')
    return m.group() if m else None

def is_social(text):
    t = text or ''
    # Doit matcher un pattern positif ET ne pas matcher un pattern négatif
    if not SOCIAL_REGEX.search(t):
        return 0
    if NEGATIVE_REGEX.search(t):
        return 0
    return 1

# ── Test rapide ───────────────────────────────────────────────────────────
tests = [
    # Vrais sociaux → 1
    ('ربي يوفقكم',                                              1),
    ('barak allah fik',                                         1),
    ('bon courage',                                             1),
    ('mabrouk',                                                 1),
    ('بالتوفيق والسداد',                                        1),
    # Faux positifs corrigés → 0
    ('والله يا ذيك راحه',                                       0),
    ('تعبتونا والله',                                           0),
    ('المودم مكانش',                                            0),
    # cas réels signalés → 0
    ('واش من اقرب راكم غير تزيدوا تبعدوا علينا بلاصيت فيبر',   0),
    ('قدمت عدت شكاوي جزاكم الله خيرا',                         0),  # plainte avec جزاكم
    ('ست اشهر يستنى انترنت هل يعقل',                           0),
    ('الله يوفق كل من يسعى لإصلاح هذا',                        1),  # court, positif pur
]
print('Test des patterns :')
ok = 0
for text, expected in tests:
    result  = is_social(text)
    match   = get_match(text)
    status  = '✅' if result == expected else '❌'
    print(f'  {status} [{result}] {text[:50]!r:55s} → match={match!r}')
    ok += (result == expected)
print(f'\n{ok}/{len(tests)} tests passés.')

# ══════════════════════════════════════════════════════════════════════════
# CHARGEMENT MONGODB
# ══════════════════════════════════════════════════════════════════════════
print('⏳ Connexion MongoDB …')
client     = MongoClient(MONGO_URI)
collection = client[DB_NAME][COLLECTION_NAME]

docs = list(collection.find({}, {
    '_id': 1,
    TEXT_FIELD: 1,
    'Commentaire_Client_Original': 1,
    'sources': 1,
    'dates': 1,
    'label_final': 1,
}))
print(f'✅ {len(docs)} documents chargés.')

# ══════════════════════════════════════════════════════════════════════════
# APPLICATION DES REGEX
# ══════════════════════════════════════════════════════════════════════════
print('⏳ Application des patterns …')

results = []
for doc in docs:
    text  = doc.get(TEXT_FIELD) or ''
    flag  = is_social(text)
    match = get_match(text) if flag else ''
    results.append({
        '_id':                  doc['_id'],
        'doc_id':               str(doc['_id']),
        'texte_normalise':      text,
        'commentaire_original': doc.get('Commentaire_Client_Original', ''),
        'source':               doc.get('sources', ''),
        'date':                 doc.get('dates', ''),
        'label_final':          doc.get('label_final', ''),
        'flag_social':          flag,
        'expression_detectee':  match or '',
    })

n_social = sum(1 for r in results if r['flag_social'] == 1)
total    = len(results)
print(f'✅ flag_social=1 : {n_social} / {total} ({100*n_social/total:.1f}%)')

# ══════════════════════════════════════════════════════════════════════════
# MISE À JOUR MONGODB
# ══════════════════════════════════════════════════════════════════════════
print('⏳ Mise à jour MongoDB …')

operations = [
    UpdateOne(
        {'_id': r['_id']},
        {'$set': {
            'flag_social':       r['flag_social'],
            'expression_sociale': r['expression_detectee'],
        }}
    )
    for r in results
]
res = collection.bulk_write(operations)
print(f'✅ MongoDB mis à jour — {res.modified_count} docs modifiés.')

# ══════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL — uniquement flag_social = 1
# ══════════════════════════════════════════════════════════════════════════
print('⏳ Génération du fichier Excel …')

social_results = [r for r in results if r['flag_social'] == 1]

# ── Styles ────────────────────────────────────────────────────────────────
hdr_fill    = PatternFill('solid', fgColor='1F4E79')
hdr_font    = Font(color='FFFFFF', bold=True, size=11)
green_fill  = PatternFill('solid', fgColor='E2EFDA')
yellow_fill = PatternFill('solid', fgColor='FFF2CC')
blue_fill   = PatternFill('solid', fgColor='D9E1F2')
sec_fill    = PatternFill('solid', fgColor='D9E1F2')
c_al        = Alignment(horizontal='center', vertical='center')
l_al        = Alignment(horizontal='left',   vertical='center', wrap_text=True)
border      = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

def write_header(ws, cols, widths):
    ws.row_dimensions[1].height = 28
    for ci, (h, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = c_al
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w

wb  = Workbook()

# ── Feuille 1 : flag_social_1 ─────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'flag_social_1'
write_header(ws1,
    ['doc_id', 'commentaire_original', 'texte_normalise',
     'expression_detectee', 'source', 'date', 'label_final', 'flag_social'],
    [28, 60, 60, 30, 14, 18, 14, 12]
)
for ri, r in enumerate(social_results, 2):
    ws1.row_dimensions[ri].height = 22
    vals   = [r['doc_id'], r['commentaire_original'], r['texte_normalise'],
              r['expression_detectee'], r['source'], r['date'],
              r['label_final'], r['flag_social']]
    fills  = [None, green_fill, green_fill, yellow_fill, None, None, None, blue_fill]
    aligns = [c_al, l_al, l_al, c_al, c_al, c_al, c_al, c_al]
    for ci, (val, fill, align) in enumerate(zip(vals, fills, aligns), 1):
        cell = ws1.cell(row=ri, column=ci, value=val)
        cell.border = border
        cell.alignment = align
        if fill:
            cell.fill = fill
        if ci == 8:
            cell.font = Font(bold=True, color='276221')

# ── Feuille 2 : Stats ─────────────────────────────────────────────────────
ws2 = wb.create_sheet('Stats')
ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 20

source_counts = Counter(r['source'] for r in social_results)
match_counts  = Counter(r['expression_detectee'].lower()
                        for r in social_results if r['expression_detectee'])

stats_rows = [
    ('RÉSULTATS — DÉTECTION flag_social = 1', ''),
    ('', ''),
    ('Total commentaires traités',   total),
    ('flag_social = 1',              n_social),
    ('flag_social = 0',              total - n_social),
    ('% commentaires sociaux',       f'{100*n_social/total:.2f}%'),
    ('Méthode',                      'Regex lexicale + filtre négatif'),
    ('', ''),
    ('RÉPARTITION PAR SOURCE', ''),
]
for src, cnt in source_counts.most_common():
    stats_rows.append((f'  {src}', cnt))

stats_rows += [('', ''), ('TOP 10 EXPRESSIONS DÉTECTÉES', '')]
for expr, cnt in match_counts.most_common(10):
    stats_rows.append((f'  {expr}', cnt))

for ri, (label, val) in enumerate(stats_rows, 1):
    ws2.row_dimensions[ri].height = 22
    ca = ws2.cell(row=ri, column=1, value=label)
    cb = ws2.cell(row=ri, column=2, value=val)
    is_sec = any(x in str(label) for x in ['RÉSULTATS', 'RÉPARTITION', 'TOP 10'])
    if is_sec:
        ca.font = Font(bold=True, size=11, color='1F4E79')
        ca.fill = sec_fill
        cb.fill = sec_fill
    ca.alignment = l_al
    cb.alignment = c_al
    if isinstance(val, int):
        cb.font = Font(bold=True)

wb.save(OUTPUT_EXCEL)
print(f'✅ Excel généré : {OUTPUT_EXCEL}')
print(f'   → {n_social} commentaires sociaux exportés')
print()
print('=' * 55)
print('  RÉSUMÉ FINAL')
print('=' * 55)
print(f'  Total docs traités    : {total}')
print(f'  flag_social = 1       : {n_social} ({100*n_social/total:.1f}%)')
print(f'  flag_social = 0       : {total - n_social} ({100*(total-n_social)/total:.1f}%)')
print(f'  Méthode               : Regex + filtre négatif')
print(f'  MongoDB mis à jour    : ✅')
print(f'  Excel généré          : {OUTPUT_EXCEL} ✅')
print('=' * 55)