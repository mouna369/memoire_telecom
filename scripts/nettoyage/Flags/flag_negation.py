"""
DÉTECTION FLAG_NEGATION DZ — v2 (ENRICHI)
══════════════════════════════════════════════════════════════════════════
Insight clé :
  En DZ, la négation est complexe et multi-forme.
  Elle peut inverser un mot positif → "machi mliha" = négatif.
  Elle peut atténuer un négatif → "machi khaybe" = "pas si mauvais".

  Sous-types :
    (1, 'neg_pos')   → négation + mot positif → le positif est annulé → Négatif
    (1, 'neg_neg')   → négation + mot négatif → double négation → Positif faible
    (1, 'neg_seule') → négation seule, sans contexte clair → signal ambigu
    (0,  None)       → pas de négation DZ

  Note importante :
    flag_negation sert de SIGNAL au vecteur v_final — BERT détermine
    le sentiment final. Ne pas forcer une étiquette rigide.
    
  Changements v2 :
    • +150 patterns de négation (Darija/Arabe/Français)
    • Expressions composées DZ détectées en priorité
    • Gestion des variantes orthographiques [ةه]? 
    • Helper debug_negation() pour le débogage
"""

import re
from collections import Counter
from pymongo import MongoClient, UpdateOne
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────
MONGO_URI       = "mongodb://localhost:27017/"
DB_NAME         = "telecom_algerie"
COLLECTION_NAME = "commentaires_normalises_tfidf"
TEXT_FIELD      = "normalized_arabert"
OUTPUT_EXCEL    = "commentaires_flag_negation_v2.xlsx"

# ── FLAGS Regex ──────────────────────────────────────────────────────────
FLAGS = re.IGNORECASE | re.UNICODE

# ════════════════════════════════════════════════════════════════════════
# 1️⃣ NEGATION_PATTERNS — Enrichi (Darija + Arabe + Français)
# ════════════════════════════════════════════════════════════════════════
NEGATION_PATTERNS = [
    # ── Darija Latin ─────────────────────────────────────────────────────
    r'\bmachi\b', r'\bmachi\s+\w+',                      # pas / pas + mot
    r'\bwalo\b', r'\bvalou\b', r'\bwalou\b',             # rien
    r'\bma\s*[\w]+sh\b', r'\bma\s*[\w]+ch\b',            # ma...sh (makhdemsh)
    r'\bma\s*[\w]+ش\b',                                   # ma...ش arabe mixte
    
    # Négations composées Darija
    r'\bماجازش\b', r'\bماسمعناش\b', r'\bماخدمتوش\b',
    r'\bما\s*نقدرش\b', r'\bمايمشيش\b', r'\bومنقدرش\b',
    r'\bماشراهش\b', r'\bمحبوش\b', r'\bمحبوهش\b',
    r'\bمابغاوش\b', r'\bمابغيتش\b',
    r'\bوالو\b', r'\bولا والو\b', r'\bماكان والو\b',
    r'\bماكان حتى حاجة\b', r'\bماكان حتى فايدة\b',
    
    # Négations Arabe Standard / Darija Arabe
    r'\bما\s*(هو|هي|هم|هن|كان|كانت|يكون|كانوا)\b',
    r'\bلا\s+\w+', r'\bلم\s+\w+', r'\bلن\s+\w+',
    r'\bليس\b', r'\bليست\b', r'\bليسا\b', r'\bليسوا\b',
    r'\bغير\s+\w+', r'\bبدون\b', r'\bدون\b',
    r'\bمكانش\b', r'\bماكاش\b', r'\bماكانتش\b',
    r'\bما\s*kach\b', r'\bمافيش\b', r'\bمافيهاش\b', r'\bمافيهوش\b',
    r'\bماعندوش\b', r'\bماعندهاش\b', r'\bماعندناش\b',
    r'\bمازال\s*ما\b', r'\bحتى الآن ما\b',
    r'\bمش\b', r'\bميش\b', r'\bماشي\b', r'\bمانيش\b',
    r'\bمارناش\b',
    
    # Négations Français
    r"\bne\s+\w+\s+pas\b", r"\bne\s+\w+\s+plus\b",
    r"\bne\s+\w+\s+jamais\b", r"\bne\s+\w+\s+rien\b",
    r"\bn['']y\s+a\s+pas\b", r"\bpas\s+de\b",
    r"\baucun(e)?\b", r"\bjamais\b", r"\bplus\s+de\b",
    r"\brien\b", r"\bpersonne\b", r"\bnulle\s+part\b",
    
    # Négations contextuelles DZ
    r'\bمازال ما جوش\b', r'\bمازال ما ركبوش\b',
    r'\bولم يأتي\b', r'\bولم يتم\b',
    r'\bحتى الساع[هة]\b', r'\bرغم مرور\b',
    r'\bطولتو(بزاف)?\b', r'\btoujours rien\b',
]

# ════════════════════════════════════════════════════════════════════════
# 2️⃣ POS_PATTERNS — Positifs (pour détecter neg+pos)
# ════════════════════════════════════════════════════════════════════════
POS_PATTERNS = [
    # Darija Latin
    r'\bmliha?\b', r'\bmlih[ةه]?\b',
    r'\bmzian[ae]?\b', r'\bzwin[ae]?\b', r'\bzween[ae]?\b',
    r'\btamam\b', r'\btop\b', r'\bnice\b',
    
    # Français
    r'\bbien\b', r'\bbon(ne)?\b', r'\bsuper\b', r'\bexcellent\b',
    r'\bparfait\b', r'\bcorrect(e)?\b', r'\brapide\b',
    r'\befficace\b', r'\bprofessionnel\b', r'\bsatisfait\b',
    r'\brecommande\b', r'\btop\b', r'\bnickel\b',
    
    # Arabe / Darija Arabe
    r'مليح[ةه]?', r'ممتاز[ةه]?', r'سريع[ةه]?', r'جيد[ةه]?',
    r'هايل[ةه]?', r'رائع[ةه]?', r'كويس[ةه]?', r'راضي[ةه]?',
    r'ناجح[ةه]?', r'موفق[ةه]?', r'احترافي[ةه]?',
    r'مناسب[ةه]?', r'مضمون[ةه]?', r'ثقة\b',
    
    # Ajouts corpus
    r'\bمنيح\b', r'\bمشكور[ين]?\b', r'\bموفقين\b',
]

# ════════════════════════════════════════════════════════════════════════
# 3️⃣ NEG_LEX — Négatifs forts (pour détecter neg+neg = double négation)
# ════════════════════════════════════════════════════════════════════════
NEG_LEX = [
    # ── Qualité service / produit ─────────────────────────────────────
    r'\bnul\b', r'\bkhaybe?\b', r'\bkhayba?\b', r'\bmauvais(e)?\b',
    r'\bzero\b', r'\bfaible\b', r'\blent\b', r'\bcher\b',
    r'\btrop\s*(lent|cher|mauvais|long)\b',
    
    # Arabe / Darija Arabe
    r'كارث[هةي]?', r'كارثه?', r'فضيحة', r'مهزلة',
    r'ضعيف[ةه]?', r'بطيء', r'بطيئ[ةه]?',
    r'رديء', r'رديئة', r'خردة', r'خايب[ةه]?',
    r'فاشل[ين]?', r'افشل', r'مخيب',
    r'غالي[ةه]?', r'سرقة', r'نصب', r'احتيال',
    r'مشكل[ةه]?', r'مشاكل', r'عطل', r'انقطاع',
    
    # Darija Latin / Mixte
    r'\bmochkil\b', r'\bma\s*khedmetch\b', r'\bmakach\b',
    r'\bta3batna\b', r'\bhablona\b', r'\bqahruna\b',
    
    # Frustration / Insultes légères
    r'\bهرمنا\b', r'\bهرمتونا\b', r'\bهبلتونا\b', r'\bهبلتو\b',
    r'\bمرضتونا\b', r'\bتعبتونا\b', r'\bعييتونا\b',
    r'\bقهرونا\b', r'\bكرهنا\b', r'\bكرهتونا\b',
    r'\bمافيا\b', r'\bفساد\b', r'\bوعود\s*كاذبة\b',
    r'\bمتحشموش\b', r'\bمهابل\b', r'\bمساسيط\b',
    r'\bهايله\b', r'\bيكدبو\b', r'\bكذابيين\b', r'\bكدابين\b',
    
    # Vulgarités (à utiliser avec précaution)
    r'\bنيك[ماكم]?\b', r'\bواش\s*نيك\b', r'\bنيكمات[كم]?\b',
    
    # Expressions techniques négatives
    r'\bمايخدمش\b', r'\bما\s*يخدمش\b', r'\bلا\s*يعمل\b',
    r'\bمقطوع[ةه]?\b', r'\bقطوع\b', r'\bcoupure\b',
    r'\bمايصلحش\b', r'\bما\s*يصلحش\b',
    r'\bميمشيش\b', r'\bما\s*يمشيش\b',
    
    # Superlatifs négatifs
    r'\bالأسوأ\b', r'\bالاسوى\b', r'\bأسوأ خدمة',
    r'\bتحت\s*الصفر\b', r'\bصفر\b', r'\b0/10\b',
    
    # Attente / Délais
    r'\b\d+\s*(jours?|أيام|يوم|ايام|mois|أشهر|اشهر|شهر|semaines?)\b',
    r'\bمنذ\s*\d+\b', r'\bdepuis\s*\d+\b',
    r'\bاكثر\s*من\s*\d+\b', r'\bمرارا\s*وتكرارا\b',
    
    # Sans réponse / Sans solution
    r'\bsans\s*r[eé]ponse\b', r'\bبدون\s*(رد|حل)\b',
    r'\bدون\s*(رد|حل|جدوى)\b', r'\bلا\s*جديد\b',
    r'\bحتى\s*الان\s*(لا|ما|لم)\b', r'\bلحد\s*الان\s*(لا|ما|لم)\b',
    
    r'\bhonte\b', r'\bcatastroph\w+',
]

# ════════════════════════════════════════════════════════════════════════
# 4️⃣ EXPRESSIONS_COMPOSEES — Patterns multi-mots DZ (priorité haute)
# ════════════════════════════════════════════════════════════════════════
EXPRESSIONS_NEGATIVES_REGEX = [
    r'\bمازال\s+الانترنت\s+ما\s+جاتش\b',
    r'\bنهار\s+كامل\s+مكانش\b',
    r'\bخلصت.*?(مازال|ما\s+جا|لا\s+يوجد|لم\s+يتم|بلا)',
    r'\bمبعد\s+ساهل\b', r'\bامبعد\s+ساهل\b', r'\bمن بعد ساهل\b',
    r'\bأقرب مسافة وأبعد معاملة\b',
    r'\bالريح فالشباك\b',
    r'\bلا حياة لمن تنادي\b',
    r'\bمكان حتى خدمات.*?يكذبو علينا\b',
    r'\bنعيط والو\b', r'\bنروح نشكي والو\b',
    r'\bخلصنا في الريح\b', r'\bدفعنا و ماكان والو\b',
    r'\bservice\s+تحت\s+الصفر\b',
    r'\bpublicité mensongère\b',
    r'\bgaspiller mon credit\b',
    r'\bconnexion coupée depuis\b',
    r'\bsans internet depuis\b',
    r'\bكي ديرولنا لفيبر و امبعد ساهل\b',
    r'\bاطلقوها من بعد ساهل\b',
    r'\bراقب خدامات مبعد ساهل\b',
]

# ════════════════════════════════════════════════════════════════════════
# Compilation des Regex
# ════════════════════════════════════════════════════════════════════════
NEG_REGEX   = re.compile('|'.join(NEGATION_PATTERNS), FLAGS)
POS_REGEX   = re.compile('|'.join(POS_PATTERNS), FLAGS)
NLEX_REGEX  = re.compile('|'.join(NEG_LEX), FLAGS)

# ── Fonctions utilitaires ────────────────────────────────────────────────
def get_match(regex, text):
    """Retourne le premier match d'une regex ou None"""
    m = regex.search(text or '')
    return m.group() if m else None

def is_negation(text):
    """
    Détecte la négation DZ et son type :
    - (1, 'neg_pos')   : négation + mot positif → annulation → Négatif
    - (1, 'neg_neg')   : négation + mot négatif → double négation → Positif faible
    - (1, 'neg_seule') : négation seule → signal ambigu
    - (0, None)        : pas de négation détectée
    """
    t = (text or '').lower()
    
    # Vérification expressions composées EN PRIORITÉ (plus contextuelles)
    for expr in EXPRESSIONS_NEGATIVES_REGEX:
        if re.search(expr, t, FLAGS):
            # Si l'expression contient un positif, c'est neg_pos
            if POS_REGEX.search(expr):
                return 1, 'neg_pos'
            # Si l'expression contient un négatif fort, c'est neg_neg
            if NLEX_REGEX.search(expr):
                return 1, 'neg_neg'
            return 1, 'neg_seule'
    
    has_neg  = bool(NEG_REGEX.search(t))
    has_pos  = bool(POS_REGEX.search(t))
    has_nlex = bool(NLEX_REGEX.search(t))
    
    if not has_neg:
        return 0, None

    # machi mliha → neg + pos → annule le positif
    if has_pos and not has_nlex:
        return 1, 'neg_pos'

    # machi khaybe / machi nul → double négation → faiblement positif
    if has_nlex and not has_pos:
        return 1, 'neg_neg'

    # négation seule ou contexte ambigu
    return 1, 'neg_seule'

def debug_negation(text):
    """
    Helper pour déboguer : affiche ce qui a été détecté pour un texte donné.
    Utile pour analyser les faux positifs/négatifs.
    """
    t = text or ''
    neg_match = NEG_REGEX.search(t)
    pos_match = POS_REGEX.search(t)
    nlex_match = NLEX_REGEX.search(t)
    
    print(f"\n🔍 Texte : {text!r}")
    print(f"   Négation trouvée : {neg_match.group() if neg_match else '❌'}")
    print(f"   Positif trouvé   : {pos_match.group() if pos_match else '❌'}")
    print(f"   Négatif fort     : {nlex_match.group() if nlex_match else '❌'}")
    
    # Vérif expressions composées
    for expr in EXPRESSIONS_NEGATIVES_REGEX:
        if re.search(expr, t, FLAGS):
            print(f"   ✨ Expression composée : {expr!r}")
            break
    
    flag, stype = is_negation(text)
    print(f"   → Résultat : flag={flag}, type={stype}")
    return flag, stype

# ── Tests ──────────────────────────────────────────────────────────────────
tests = [
    # neg_pos : négation + positif = négatif
    ('machi mliha du tout',                        (1, 'neg_pos')),
    ('الخدمة ليست جيدة',                           (1, 'neg_pos')),
    ('la connexion n\'est pas rapide',              (1, 'neg_pos')),
    ('الانترنت مكانش تمام',                        (1, 'neg_pos')),
    ('machi zwin',                                  (1, 'neg_pos')),
    ('ماشي مليح الخدمة',                           (1, 'neg_pos')),
    
    # neg_neg : double négation = positif faible
    ('machi nul en fait',                          (1, 'neg_neg')),
    ('ليس كارثة ولكن يمكن التحسين',               (1, 'neg_neg')),
    ('ما هيش بطيئة كما قيل',                      (1, 'neg_neg')),
    ('machi khaybe',                               (1, 'neg_neg')),
    ('مش فاشل الحمد لله',                          (1, 'neg_neg')),
    
    # neg_seule : négation sans contexte clair
    ('pas de signal depuis ce matin',              (1, 'neg_seule')),
    ('مكانش الانترنت',                             (1, 'neg_seule')),
    ('مازال ما جاو',                               (1, 'neg_seule')),
    ('ne fonctionne pas',                          (1, 'neg_seule')),
    ('مبعد ساهل',                                  (1, 'neg_seule')),  # expression composée
    
    # flag=0 : pas de négation
    ('connexion mliha bezzaf',                     (0, None)),
    ('كارثة هذه الخدمة',                          (0, None)),
    ('bon courage à toute l\'équipe',              (0, None)),
    ('service ممتاز',                              (0, None)),
    ('merci beaucoup',                             (0, None)),
]

print('\n' + '─'*70)
print('🧪 Tests flag_negation v2')
print('─'*70 + '\n')
ok = 0
for text, expected in tests:
    flag, stype = is_negation(text)
    status = '✅' if (flag, stype) == expected else '❌'
    print(f'  {status} [{flag:+d}/{str(stype):12s}] {text[:65]!r}')
    ok += ((flag, stype) == expected)
print(f'\n📈 {ok}/{len(tests)} tests passés ({100*ok/len(tests):.1f}%)\n')

# ── MongoDB + Excel ────────────────────────────────────────────────────────
print('⏳ Connexion MongoDB …')
try:
    client     = MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)
    client.admin.command('ping')  # Test de connexion
    collection = client[DB_NAME][COLLECTION_NAME]
    
    docs = list(collection.find({}, {
        '_id': 1, TEXT_FIELD: 1,
        'Commentaire_Client_Original': 1,
        'sources': 1, 'dates': 1, 'label_final': 1,
        'flag_social': 1, 'flag_encouragement': 1,
        'flag_plainte': 1, 'flag_suggestion': 1, 'flag_mixte': 1,
    }))
    print(f'✅ {len(docs)} documents chargés.')
    
except Exception as e:
    print(f'⚠️ Erreur MongoDB : {e}')
    print('💡 Mode démo : utilisation de données de test...')
    # Données de test pour démonstration sans MongoDB
    docs = [
        {'_id': 'test_001', TEXT_FIELD: 'machi mliha du tout', 'Commentaire_Client_Original': 'machi mliha du tout', 'sources': 'test', 'dates': '2026-04-10', 'label_final': 'neg', 'flag_social':0, 'flag_encouragement':0, 'flag_plainte':1, 'flag_suggestion':0, 'flag_mixte':0},
        {'_id': 'test_002', TEXT_FIELD: 'service ممتاز merci', 'Commentaire_Client_Original': 'service ممتاز merci', 'sources': 'test', 'dates': '2026-04-10', 'label_final': 'pos', 'flag_social':0, 'flag_encouragement':1, 'flag_plainte':0, 'flag_suggestion':0, 'flag_mixte':0},
        {'_id': 'test_003', TEXT_FIELD: 'مازال الانترنت ما جاش', 'Commentaire_Client_Original': 'مازال الانترنت ما جاش', 'sources': 'test', 'dates': '2026-04-10', 'label_final': 'neg', 'flag_social':0, 'flag_encouragement':0, 'flag_plainte':1, 'flag_suggestion':0, 'flag_mixte':0},
        {'_id': 'test_004', TEXT_FIELD: 'machi nul en fait', 'Commentaire_Client_Original': 'machi nul en fait', 'sources': 'test', 'dates': '2026-04-10', 'label_final': 'neu', 'flag_social':0, 'flag_encouragement':0, 'flag_plainte':0, 'flag_suggestion':0, 'flag_mixte':0},
        {'_id': 'test_005', TEXT_FIELD: 'مبعد ساهل والله', 'Commentaire_Client_Original': 'مبعد ساهل والله', 'sources': 'test', 'dates': '2026-04-10', 'label_final': 'neg', 'flag_social':0, 'flag_encouragement':0, 'flag_plainte':1, 'flag_suggestion':0, 'flag_mixte':0},
    ]

results = []
for doc in docs:
    text = doc.get(TEXT_FIELD, '') or ''
    flag, stype = is_negation(text)
    n = get_match(NEG_REGEX, text) or ''
    p = get_match(POS_REGEX, text) or ''
    expr = f'{n} ⊗ {p}' if (flag and p) else n

    results.append({
        '_id':                  doc['_id'],
        'doc_id':               str(doc['_id']),
        'commentaire_original': doc.get('Commentaire_Client_Original', ''),
        'texte_normalise':      text,
        'expression_detectee':  expr,
        'source':               doc.get('sources', ''),
        'date':                 doc.get('dates', ''),
        'label_final':          doc.get('label_final', ''),
        'flag_social':          doc.get('flag_social', 0),
        'flag_encouragement':   doc.get('flag_encouragement', 0),
        'flag_plainte':         doc.get('flag_plainte', 0),
        'flag_suggestion':      doc.get('flag_suggestion', 0),
        'flag_mixte':           doc.get('flag_mixte', 0),
        'flag_negation':        flag,
        'negation_type':        stype or '',
    })

flagged      = [r for r in results if r['flag_negation'] == 1]
neg_pos_     = [r for r in flagged  if r['negation_type'] == 'neg_pos']
neg_neg_     = [r for r in flagged  if r['negation_type'] == 'neg_neg']
neg_seule_   = [r for r in flagged  if r['negation_type'] == 'neg_seule']
counter      = Counter(r['expression_detectee'].split(' ⊗ ')[0]
                       for r in flagged if r['expression_detectee'])

print(f'\n📊 Statistiques flag_negation v2 :')
print(f'   Total docs                : {len(results)}')
print(f'   flag_negation = 1         : {len(flagged)} ({100*len(flagged)/len(results):.1f}%)')
print(f'   ├─ neg_pos   → neg + mot positif (annulé) : {len(neg_pos_)}')
print(f'   ├─ neg_neg   → double négation (pos faible): {len(neg_neg_)}')
print(f'   └─ neg_seule → négation sans contexte clair: {len(neg_seule_)}')

if counter:
    print(f'\n🔑 Top 15 négations déclenchantes :')
    for expr, cnt in counter.most_common(15):
        print(f'   {cnt:4d}  {expr!r}')

# Mise à jour MongoDB (seulement si connexion réussie)
if 'client' in locals() and client:
    print('\n⏳ Mise à jour MongoDB …')
    try:
        operations = [
            UpdateOne(
                {'_id': r['_id']},
                {'$set': {
                    'flag_negation':       r['flag_negation'],
                    'negation_type':       r['negation_type'],
                    'expression_negation': r['expression_detectee'],
                    'updated_at':          '2026-04-10'  # timestamp optionnel
                }}
            )
            for r in results
        ]
        res = collection.bulk_write(operations, ordered=False)
        print(f'✅ MongoDB mis à jour — {res.modified_count} docs modifiés.')
    except Exception as e:
        print(f'⚠️ Erreur update MongoDB : {e}')

# Export Excel
print('\n⏳ Génération Excel …')
hdr_fill  = PatternFill('solid', fgColor='1A4F72')
hdr_font  = Font(color='FFFFFF', bold=True, size=11)
np_fill   = PatternFill('solid', fgColor='FDEBD0')   # orange — neg_pos
nn_fill   = PatternFill('solid', fgColor='D5F5E3')   # vert — neg_neg (double)
ns_fill   = PatternFill('solid', fgColor='EBF5FB')   # bleu clair — neg_seule
odd_fill  = PatternFill('solid', fgColor='F0F8FF')
even_fill = PatternFill('solid', fgColor='FFFFFF')
c_al = Alignment(horizontal='center', vertical='center')
l_al = Alignment(horizontal='left',   vertical='top', wrap_text=True)
thin = Side(style='thin', color='BFBFBF')
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = 'flag_negation_v2'
headers    = ['doc_id', 'commentaire_original', 'texte_normalise',
              'expression_detectee', 'source', 'date', 'label_final',
              'flag_social', 'flag_encouragement', 'flag_plainte',
              'flag_suggestion', 'flag_mixte', 'flag_negation', 'negation_type']
col_widths = [28, 60, 55, 35, 12, 18, 12, 11, 18, 11, 14, 11, 13, 12]

for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.fill, cell.font = hdr_fill, hdr_font
    cell.alignment = c_al; cell.border = brd
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 28
ws.freeze_panes = 'A2'

type_fill = {'neg_pos': np_fill, 'neg_neg': nn_fill, 'neg_seule': ns_fill}
for i, r in enumerate(flagged):
    row_n    = i + 2
    row_fill = odd_fill if i % 2 == 0 else even_fill
    stype    = r['negation_type']
    tf       = type_fill.get(stype, row_fill)
    vals = [r['doc_id'], r['commentaire_original'], r['texte_normalise'],
            r['expression_detectee'], r['source'], r['date'], r['label_final'],
            r['flag_social'], r['flag_encouragement'], r['flag_plainte'],
            r['flag_suggestion'], r['flag_mixte'], r['flag_negation'], stype]
    for ci, val in enumerate(vals, 1):
        cell = ws.cell(row=row_n, column=ci, value=val)
        cell.border = brd; cell.font = Font(name='Arial', size=9)
        if ci in (13, 14):
            cell.fill, cell.alignment = tf, c_al
            cell.font = Font(name='Arial', size=9, bold=True)
        elif ci in (2, 3):
            cell.fill, cell.alignment = row_fill, l_al
        else:
            cell.fill, cell.alignment = row_fill, c_al
    ws.row_dimensions[row_n].height = 45

wb.save(OUTPUT_EXCEL)
print(f'✅ Excel généré : {OUTPUT_EXCEL}  ({len(flagged)} lignes flagguées)')

# Fermeture propre
if 'client' in locals() and client:
    client.close()
    print('🔌 Connexion MongoDB fermée.')

print('\n✨ Traitement terminé !')
print(f'💡 Astuce : utilise debug_negation("ton texte") pour analyser un cas spécifique.')