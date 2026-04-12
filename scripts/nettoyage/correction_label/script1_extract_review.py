"""
╔══════════════════════════════════════════════════════════════════════╗
║  SCRIPT 1 — Extraction Neutre + Conflit pour vérification manuelle  ║
║  Génère deux fichiers Excel identiques (toi + camarade)              ║
║  Source : MongoDB  telecom_algerie_new > dataset_unifie_sans_doublons║
╚══════════════════════════════════════════════════════════════════════╝

USAGE:
    python script1_extract_review.py
    
    → Génère: review_MOI.xlsx et review_CAMARADE.xlsx
"""

import pandas as pd
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
import math

# ══════════════════════════════════════════════
# ⚙️  CONFIGURATION — adapte ici
# ══════════════════════════════════════════════
MONGO_URI  = "mongodb://yousrahadjabderrahmane_db_user:C8wjIvWqOBUjK66u@ac-1ksfahb-shard-00-00.gejzu4a.mongodb.net:27017,ac-1ksfahb-shard-00-01.gejzu4a.mongodb.net:27017,ac-1ksfahb-shard-00-02.gejzu4a.mongodb.net:27017/?ssl=true&replicaSet=atlas-mdnqx7-shard-0&authSource=admin&appName=Cluster0"
DB_NAME    = "telecom_algerie_new"
COLLECTION = "dataset_unifie_label_mzl"

# Si tu travailles depuis un CSV/JSON exporté (sans MongoDB), 
# mets le chemin ici et commente le bloc MongoClient plus bas
CSV_INPUT  = None 

# Noms des champs dans ton document MongoDB (d'après la capture)
CHAMPS = {
    "text_original": "Commentaire_Client_Original",
    "text_client":   "Commentaire_Client",
    "label":         "label_final",
    "conflit":       "conflit",
    "score":         "score",
    "confidence":    "confidence",
    "annote":        "annoté",
    "reason":        "reason",
    "source":        "sources",
    "flag_social":          "flag_social",
    "flag_encouragement":   "flag_encouragement",
    "flag_plainte":         "flag_plainte",
    "flag_suggestion":      "flag_suggestion",
    "flag_negation":        "flag_negation",
    "flag_mixte":           "flag_mxte",   # nom exact de ta capture
    "normalised":           "normalised_full",
}

OUTPUT_MOI      = "review_MOI_mzl.xlsx"
OUTPUT_CAMARADE = "review_CAMARADE_mzl.xlsx"
# ══════════════════════════════════════════════

# ── Couleurs ──
C_HEADER   = "1F4E79"   # bleu foncé
C_NEUTRE   = "D6E4F0"   # bleu clair
C_CONFLIT  = "FFF2CC"   # jaune
C_CORRIGE  = "E2EFDA"   # vert clair (colonne à remplir)
C_LEGEND   = "F2F2F2"
C_WHITE    = "FFFFFF"

thin = Side(style="thin", color="CCCCCC")
border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

def load_data():
    """Charge les données depuis MongoDB ou CSV."""
    if CSV_INPUT:
        print(f"📂 Chargement depuis CSV : {CSV_INPUT}")
        try:
            df = pd.read_csv(CSV_INPUT, encoding="utf-8")
        except:
            df = pd.read_csv(CSV_INPUT, encoding="latin1")
        return df

    print(f"🔌 Connexion MongoDB : {MONGO_URI}")
    client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)
    db     = client[DB_NAME]
    col    = db[COLLECTION]

    # Récupère uniquement neutre + conflit
    query  = {"$or": [
        {CHAMPS["label"]: "neutre"},
        {CHAMPS["conflit"]: True},
        {CHAMPS["conflit"]: "true"},
    ]}
    docs   = list(col.find(query))
    print(f"   ✅ {len(docs)} documents trouvés (neutre + conflit)")
    df = pd.DataFrame(docs)
    if "_id" in df.columns:
        df["_id"] = df["_id"].astype(str)
    return df

def detect_cas(row):
    """Détermine si c'est un conflit ou juste neutre."""
    conflit_field = row.get(CHAMPS["conflit"], False)
    if conflit_field in [True, "true", 1, "1"]:
        return "⚡ CONFLIT"
    label = str(row.get(CHAMPS["label"], "")).lower()
    if label == "neutre":
        return "🔵 NEUTRE"
    return "autre"

def get_val(row, key, default=0):
    """Récupère une valeur de flag (0/1)."""
    v = row.get(CHAMPS.get(key, key), default)
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return 0
    try:
        return int(float(v))
    except:
        return 0

def build_excel(df, output_path, reviewer_name):
    """Construit le fichier Excel de review."""
    wb = Workbook()

    # ── Feuille 1 : Données à corriger ──────────────────────────────────
    ws = wb.active
    ws.title = "📝 A corriger"
    ws.sheet_view.rightToLeft = False

    # Titre
    ws.merge_cells("A1:P1")
    ws["A1"] = f"📋 VÉRIFICATION MANUELLE — {reviewer_name}  |  Fichier : {CHAMPS['label']}"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    ws["A1"].fill      = PatternFill("solid", start_color=C_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Légende
    ws.merge_cells("A2:P2")
    ws["A2"] = ("🟦 NEUTRE à vérifier   |   🟨 CONFLIT flags positif+négatif   |   "
                "✏️  Remplis UNIQUEMENT la colonne 'label_corrige'   |   "
                "Valeurs : positif  /  negatif  /  neutre  /  supprimer")
    ws["A2"].font      = Font(italic=True, size=9, color="595959", name="Arial")
    ws["A2"].fill      = PatternFill("solid", start_color=C_LEGEND)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # En-têtes
    headers = [
        ("#",            5),
        ("ID / _id",    24),
        ("Texte original",  45),
        ("Texte normalisé", 40),
        ("label_final",  12),
        ("conflit",       9),
        ("score",         8),
        ("confidence",    10),
        ("reason",        20),
        ("🚩 flag_social",  10),
        ("🚩 flag_enc",     9),
        ("🚩 flag_plnt",    9),
        ("🚩 flag_sug",     9),
        ("🚩 flag_neg",     9),
        ("🚩 flag_mix",     9),
        ("⭐ label_corrige — REMPLIR ICI", 22),
    ]

    ws.row_dimensions[3].height = 32
    for col_i, (h, w) in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_i, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        cell.fill      = PatternFill("solid", start_color=C_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border_thin
        ws.column_dimensions[get_column_letter(col_i)].width = w

    # Données
    flag_keys = ["flag_social","flag_encouragement","flag_plainte",
                 "flag_suggestion","flag_negation","flag_mixte"]

    for i, (_, row) in enumerate(df.iterrows(), start=1):
        r   = i + 3
        cas = detect_cas(row)
        row_fill = PatternFill("solid", start_color=C_CONFLIT if "CONFLIT" in cas else C_NEUTRE)

        text_orig  = str(row.get(CHAMPS["text_original"], row.get(CHAMPS["text_client"], "")))
        text_norm  = str(row.get(CHAMPS["normalised"], ""))
        doc_id     = str(row.get("_id", i))
        label_cur  = str(row.get(CHAMPS["label"], ""))
        conflit_v  = str(row.get(CHAMPS["conflit"], ""))
        score_v    = row.get(CHAMPS["score"], "")
        conf_v     = row.get(CHAMPS["confidence"], "")
        reason_v   = str(row.get(CHAMPS["reason"], ""))

        vals = ([i, doc_id, text_orig, text_norm, label_cur, conflit_v,
                 score_v, conf_v, reason_v]
                + [get_val(row, k) for k in flag_keys]
                + [""])

        for col_i, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col_i, value=val)
            cell.border = border_thin
            is_last = col_i == len(headers)
            cell.fill = PatternFill("solid", start_color=C_CORRIGE) if is_last else row_fill
            if is_last:
                cell.font = Font(bold=True, color="1F4E79", name="Arial")
            else:
                cell.font = Font(size=9, name="Arial")
            wrap = col_i in (3, 4, 9)
            cell.alignment = Alignment(wrap_text=wrap, vertical="top")

        h_row = 45 if len(text_orig) > 80 else (32 if len(text_orig) > 40 else 20)
        ws.row_dimensions[r].height = h_row

    # Figer la ligne d'en-tête
    ws.freeze_panes = "A4"

    # Filtre auto sur toute la plage
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{3 + len(df)}"

    # ── Feuille 2 : Guide de correction ─────────────────────────────────
    ws2 = wb.create_sheet("📖 Guide")
    guide_data = [
        ("CHAMP",               "VALEUR",     "SIGNIFICATION"),
        ("label_corrige",       "positif",    "Le commentaire exprime une satisfaction / opinion positive"),
        ("label_corrige",       "negatif",    "Plainte, frustration, critique négative"),
        ("label_corrige",       "neutre",     "Formule sociale, hors-sujet, aucune opinion réelle"),
        ("label_corrige",       "supprimer",  "Spam, incompréhensible, demande prv sans contexte"),
        ("","",""),
        ("CAS CONFLIT",         "🟨",         "flag_positif=1 ET flag_négatif=1 en même temps → choisir dominant"),
        ("CAS NEUTRE",          "🟦",         "label=neutre → confirmer ou corriger si tu vois une opinion"),
        ("","",""),
        ("EXEMPLES","",""),
        ("'connexion mliha ms le prix ghali'", "→ negatif", "flag_mixte — négatif domine selon rapport"),
        ("'rbi y3awkom'",                      "→ neutre",  "Bénédiction sociale → neutre forcé"),
        ("'courage zidou'",                    "→ neutre",  "Encouragement pur sans avis produit"),
        ("'zidou hih l appli zwina'",          "→ positif", "Encouragement + avis positif produit"),
        ("'repondiw prv, 7 jours sans rep'",   "→ negatif", "Plainte implicite DZ"),
        ("'connexion mliha ms tehseno'",       "→ positif", "Positif conditionnel DZ (55-65%)"),
    ]
    ws2["A1"] = "📖 GUIDE DE CORRECTION — Valeurs autorisées et exemples DZ"
    ws2["A1"].font = Font(bold=True, size=12, color="FFFFFF", name="Arial")
    ws2["A1"].fill = PatternFill("solid", start_color=C_HEADER)
    ws2.merge_cells("A1:C1")
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 25

    for row_i, (a, b, c) in enumerate(guide_data, start=2):
        ws2.cell(row=row_i, column=1, value=a).font = Font(bold=(row_i in (2,3,4,5,7,8,9)), name="Arial", size=10)
        ws2.cell(row=row_i, column=2, value=b).font = Font(color="1F4E79", bold=True, name="Arial", size=10)
        ws2.cell(row=row_i, column=3, value=c).font = Font(name="Arial", size=10, color="595959")
        ws2.row_dimensions[row_i].height = 18

    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 50

    # ── Feuille 3 : Statistiques rapides ────────────────────────────────
    ws3 = wb.create_sheet("📊 Stats")
    ws3["A1"] = "📊 STATISTIQUES DU FICHIER D'EXTRACTION"
    ws3["A1"].font = Font(bold=True, size=12, color="FFFFFF", name="Arial")
    ws3["A1"].fill = PatternFill("solid", start_color=C_HEADER)
    ws3.merge_cells("A1:B1")
    ws3["A1"].alignment = Alignment(horizontal="center")

    total   = len(df)
    neutres = len(df[df.apply(lambda r: detect_cas(r) == "🔵 NEUTRE", axis=1)])
    conflits= total - neutres

    stats = [
        ("Total commentaires extraits", total),
        ("Dont NEUTRE",                 neutres),
        ("Dont CONFLIT",                conflits),
        ("", ""),
        ("Part NEUTRE",                 f"=B3/B2*100"),
        ("Part CONFLIT",                f"=B4/B2*100"),
    ]
    for row_i, (k, v) in enumerate(stats, start=2):
        ws3.cell(row=row_i, column=1, value=k).font = Font(name="Arial", size=10)
        ws3.cell(row=row_i, column=2, value=v).font  = Font(bold=True, name="Arial", size=10, color="1F4E79")
        ws3.row_dimensions[row_i].height = 18

    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 16

    wb.save(output_path)
    print(f"   ✅ Fichier créé : {output_path}  ({total} lignes)")


def main():
    print("=" * 60)
    print("  SCRIPT 1 — Extraction pour vérification manuelle")
    print("=" * 60)

    df = load_data()
    if df is None or len(df) == 0:
        print("❌ Aucune donnée chargée. Vérifie la connexion ou le CSV.")
        return

    # Filtrer uniquement neutre + conflit (si MongoDB retourne autre chose)
    mask = (
        (df.get(CHAMPS["label"], pd.Series(dtype=str)).astype(str).str.lower() == "neutre") |
        (df.get(CHAMPS["conflit"], pd.Series(dtype=bool)).astype(str).str.lower().isin(["true","1"]))
    )
    df_review = df[mask].copy().reset_index(drop=True)
    print(f"\n📊 Total à réviser : {len(df_review)}")

    # Partage en deux moitiés
    mid = len(df_review) // 2
    df_moi      = df_review.iloc[:mid].reset_index(drop=True)
    df_camarade = df_review.iloc[mid:].reset_index(drop=True)

    print(f"   → Toi      : {len(df_moi)} commentaires")
    print(f"   → Camarade : {len(df_camarade)} commentaires")
    print()

    build_excel(df_moi,      OUTPUT_MOI,      "MOI")
    build_excel(df_camarade, OUTPUT_CAMARADE, "CAMARADE")

    print("\n✅ Terminé ! Ouvre les fichiers Excel, remplis la colonne")
    print("   'label_corrige' puis lance script2_reimport.py")

if __name__ == "__main__":
    main()
