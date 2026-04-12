# """
# clean_telecom_FINAL_v3.py — Nettoyage corpus Télécom DZ
# ═══════════════════════════════════════════════════════════════════════════════
# Nouveautés v3 :
#   - Fix bug : return None mal placé dans _is_suppressed
#   - Nouveaux patterns suppression : bruit rires (hhh), URLs, contenu vide
#     (ok/oui/يخي/كيفاش seul), technique seul avec unité (60 Mbps, adsl)
#   - Corrections arabes : معاكم→معكم, االتزام→التزام, بيس→شريحة, alef répété
#   - Ajout dict : bonus→بونوس, puce→شريحة, البونوس, بيس
#   - Nouvelles couleurs Excel pour les nouveaux types de suppression
#   - Écriture des résultats dans MongoDB
# ═══════════════════════════════════════════════════════════════════════════════
# """

# import re
# import pandas as pd
# from openpyxl.styles import Font, PatternFill, Alignment
# from openpyxl.utils import get_column_letter
# from pymongo import MongoClient

# # ── CONFIG ───────────────────────────────────────────────────────────────────


# LIMIT   = None
# OUTPUT  = "nettoyage_resultats_v3.xlsx"


# MONGO_URI = "mongodb://localhost:27018/"
# DB_NAME = "telecom_algerie"

# from dict_data import DICT_NORM, DICT_STOP


# # ══════════════════════════════════════════════════════════════════════════════
# #  SUPPRESSION PATTERNS
# # ══════════════════════════════════════════════════════════════════════════════

# PRIX_RE = re.compile(
#     r'\b\d[\d\s]*(?:da|dz|dzd|دج|دينار|dinars?|euros?|€)\b'
#     r'|\b(?:da|dz|dzd|دج|دينار)\s*\d[\d\s]*\b'
#     r'|\b\d+\s*(?:da|dz)\b',
#     re.IGNORECASE,
# )
# TEL_RE = re.compile(
#     r'\b(?:0|\+213|00213)[\s\-]?[5-7]\d[\s\-]?\d{2}[\s\-]?\d{2}[\s\-]?\d{2}\b'
#     r'|\b0[5-7]\d{8}\b'
# )
# TECH_WORDS = re.compile(
#     r'\b(4g|5g|3g|sbox|inbox|lte|apn|sms|mms|ussd|forfait|recharge|solde)\b',
#     re.IGNORECASE,
# )
# CHARABIA_LAT = re.compile(r'\b[b-df-hj-np-tv-z]{6,}\b', re.IGNORECASE)
# CHARABIA_AR  = re.compile(r'[\u0621-\u064A]{1,2}([\u0621-\u064A])\1{3,}')
# SUBS         = re.compile(r'\b\w{3,}\b')

# # ── NOUVEAUX patterns de suppression ─────────────────────────────────────────

# # Rires / bruit clavier : hhh, hhhh, lol, mdr, hjijhghh, hgfmlkjhh...
# BRUIT_LAT = re.compile(
#     r'^[\s]*(?:h{3,}|(?:[a-z]{1,3})?([a-z])\1{3,}|lol+|mdr+|xd+|haha+|hihi+|héhé+)[\s!.]*$',
#     re.IGNORECASE,
# )

# # URL seule
# URL_ONLY = re.compile(
#     r'^[\s]*(?:https?[\s:/\\]+\S*|www\.\S+)[\s]*$',
#     re.IGNORECASE,
# )

# # Contenu vide : réponses courtes sans information ni sentiment
# # - Questions darija seules : كيفاش, وقتاش, شحال, واش, علاش, فين, كيف
# # - Réponses courtes : ok, oui, non, no, تم, يخي, ايه, لا, ها, واه
# # - Interjections arabes pures : يخي, خي, ههه, بلا
# # - Prix/chiffres seuls sans contexte : "60", "100"
# VIDE_CONTENT = re.compile(
#     r'^[\s]*(?:'
#     # questions darija seules (avec ou sans ponctuation)
#     r'كيفاش|كيف|وقتاش|شحال|واش|علاش|فين|قداش|'
#     # réponses oui/non courtes
#     r'ok|oui|non|no|yes|yep|nope|'
#     # darija oui/non
#     r'تم|يخي|ايه|إيه|آه|واه|لا|ها|هه|'
#     # interjections/ponctuations pures
#     r'خي|يا خي|آ خي|بلا|هاك|هاكا|هيا|يلا|'
#     # bruit arabe répété : اااا, تتت
#     r'[اى]{3,}|ات+ال+ات+|[أإآا]{4,}'
#     r')[\s?!.،؟…]*$',
#     re.IGNORECASE | re.UNICODE,
# )

# # Technique seul avec unité : "60 Mbps", "adsl", "4G", "100 mb"
# # (chiffre optionnel + unité, sans contexte)
# TECH_UNIT_ONLY = re.compile(
#     r'^[\s]*\d*[\s]*(mbps|mb/s|kbps|gb/s|mbits?|mb|gb|go|mo|ko|ghz|mhz|'
#     r'adsl|vdsl|4g|5g|3g|2g|lte|ftth|fttb|gpon|ping|dns)[\s]*$',
#     re.IGNORECASE,
# )

# # Interjections / expressions vides sans info télécom
# # let's go, let's goo, lies, ni, intéressé, ok cool, yalla, etc.
# INTERJECTION_VIDE = re.compile(
#     r'^[\s]*(?:'
#     r"let'?s\s*goo*|"                       # let's go, let's goo
#     r'yall?a+|yala+|'                        # yalla, yala
#     r'cool+|super+|bravo+|bien+|'
#     r'ni|lies?|'                             # ni / lies (hors contexte)
#     r'intéressé[e]?|interested|'
#     r'photo(?:\s+\w{1,10})?|'               # photo / photo fortnite
#     r'ok\s+cool|cool\s+ok'
#     r')[\s?!.،؟…]*$',
#     re.IGNORECASE | re.UNICODE,
# )

# # Prix mentionné seul (mot "prix" / "سعر" / "ثمن" sans contexte)
# PRIX_MOT_SEUL = re.compile(
#     r'^[\s]*(?:'
#     r'(?:le\s+)?prix|سعر|ثمن|تمن'
#     r')[\s?!.،؟…]*$',
#     re.IGNORECASE | re.UNICODE,
# )

# # Chiffre + mot numérique vague (1 ooo, 3 millions, million, 3 three, 3 com supprimé…)
# # → nombre suivi de mot non-technique, ou mot numérique isolé
# CHIFFRE_MOT_VAGUE = re.compile(
#     r'^[\s]*(?:'
#     r'\d[\d\s]*(?:ooo+|millions?|billions?|three|two|one|zero|com\b.*)'
#     r'|millions?|billions?'                  # mot numérique seul
#     r')[\s?!.،؟…]*$',
#     re.IGNORECASE | re.UNICODE,
# )

# # Noms de villes algériennes seuls (sans contexte)
# VILLE_SEULE = re.compile(
#     r'^[\s]*(?:'
#     r'sidi\s+bel\s+abb[eè]s|oran|alger(?:s)?|constantine|annaba|'
#     r'tizi\s+ouzou|bejaia|béjaïa|blida|setif|sétif|batna|biskra|'
#     r'tlemcen|chlef|médéa|mostaganem|mascara|djelfa|m.?sila|'
#     r'skikda|jijel|guelma|souk\s+ahras|tiaret|ghardaia|bechar'
#     r')[\s?!.,،؟…]*$',
#     re.IGNORECASE | re.UNICODE,
# )

# # Caractères spéciaux / diacritiques non-arabes sans sens (zźźź1 aàà, zin øü)
# CHARABIA_DIACRITIC = re.compile(
#     r'^[\s]*[a-zA-Zàáâãäåæçèéêëìíîïðñòóôõöùúûüýþÿżźśøü\d\s]{1,20}[\s]*$'
#     r'(?=.*[żźśøüðþ])',   # doit contenir au moins un caractère hors ASCII standard
# )
# # Variante plus simple : texte court avec diacritiques exotiques
# CHARABIA_DIACRITIC = re.compile(
#     r'^[\s\w]{0,5}[żźśøüðþæœ][\s\w]{0,10}$',
#     re.IGNORECASE | re.UNICODE,
# )

# # Interjections arabes manquantes dans VIDE_CONTENT
# # ييا يياا, يهاه, همم, ههخخ, هذا 4g, امم, واي فاي 6
# INTERJECTION_AR_VIDE = re.compile(
#     r'^[\s]*(?:'
#     r'[يه]{2,}[اوى]+[اه]*|'                # ييا, يياا, يهاه
#     r'هم+|هه+خ+|ام+|اهه+|'                 # همم, ههخخ, امم
#     r'هذا\s*(?:4g|5g|3g|lte|adsl|wifi)|'   # هذا 4g / هذا wifi
#     r'واي\s*فاي\s*\d+|'                    # واي فاي 6
#     r'سيدي\s+\S+(?:\s+\S+)?'              # سيدي بلعباس (ville arabe)
#     r')[\s?!.،؟…]*$',
#     re.IGNORECASE | re.UNICODE,
# )

# # Chiffre grand + interjection arabe (185000 امم, 1000 ping adsl)
# CHIFFRE_INTERJECTION = re.compile(
#     r'^[\s]*\d[\d\s]*\s*(?:'
#     r'ام+|اهه+|هم+|همم+|اممم+|'            # interjections
#     r'ping\s+adsl|adsl\s+ping'              # technique combiné
#     r')[\s?!.،؟…]*$',
#     re.IGNORECASE | re.UNICODE,
# )

# # Terme télécom seul précédé d'article (la fibre, la fibre 2, le réseau...)
# TECH_ARTICLE_SEUL = re.compile(
#     r'^[\s]*(?:la\s+|le\s+|les\s+|une?\s+|du\s+|de\s+)?'
#     r'(?:fibre|fiber|adsl|vdsl|4g|5g|3g|lte|ftth|fttb|wifi|connexion|réseau|network)'
#     r'(?:\s+\d+)?[\s?!.،؟…]*$',
#     re.IGNORECASE | re.UNICODE,
# )

# # ping + chiffre + terme tech seul (ping 1000 adsl, ping 50ms)
# PING_TECH_SEUL = re.compile(
#     r'^[\s]*(?:ping\s*\d*\s*(?:adsl|vdsl|4g|5g|lte|ms)?'
#     r'|\d+\s*(?:ping|ms)\s*(?:adsl|vdsl|4g|5g)?)[\s?!.]*$',
#     re.IGNORECASE,
# )

# # Séquences arabes de lettres répétées sans sens (ييا يياا)
# SEQUENCE_AR_VIDE = re.compile(
#     r'^[\s]*[يا]{3,}[\s]*[يا]{0,5}[\s?!.،؟…]*$',
#     re.UNICODE,
# )

# # Charabia latin à très faible ratio de voyelles (hjijhghh, hjjgh…)
# def _is_low_vowel_noise(text: str) -> bool:
#     """Mot latin court (< 12 chars) avec < 15% de voyelles → bruit clavier."""
#     t = text.strip().lower()
#     if not t or len(t) > 12 or not t.isalpha():
#         return False
#     vowels = sum(1 for c in t if c in 'aeiouéèêëàâùûü')
#     return (vowels / len(t)) < 0.15


# def _is_suppressed(t: str):
#     """
#     Retourne la raison de suppression, ou None si le doc est à conserver.
#     IMPORTANT : chaque return None doit être à la FIN de la fonction,
#     pas imbriqué dans un if.
#     """
#     # ── 1. Bruit rires / clavier latin ───────────────────────────────────
#     if BRUIT_LAT.match(t):
#         return "bruit_rire_clavier"

#     # ── 2. URL seule ──────────────────────────────────────────────────────
#     if URL_ONLY.match(t):
#         return "url_seule"

#     # ── 3. Contenu vide (ok, يخي, كيفاش seul…) ───────────────────────────
#     if VIDE_CONTENT.match(t):
#         return "contenu_vide"

#     # ── 4. Technique seul avec unité (60 Mbps, adsl…) ────────────────────
#     if TECH_UNIT_ONLY.match(t):
#         return "technique_seul"

#     # ── 4b. Prix mot seul (prix, le prix, سعر, ثمن…) ─────────────────────
#     if PRIX_MOT_SEUL.match(t):
#         return "prix_detecte"

#     # ── 4c. Interjection vide (let's go, photo, lies, ni, intéressé…) ────
#     if INTERJECTION_VIDE.match(t):
#         return "contenu_vide"

#     # ── 4d. Chiffre + mot numérique vague (1 ooo, 3 millions, million…) ──
#     if CHIFFRE_MOT_VAGUE.match(t):
#         return "contenu_vide"

#     # ── 4e. Nom de ville seul ─────────────────────────────────────────────
#     if VILLE_SEULE.match(t):
#         return "contenu_vide"

#     # ── 4f. Interjections arabes manquantes (ييا, همم, هذا 4g, واي فاي 6)
#     if INTERJECTION_AR_VIDE.match(t):
#         return "contenu_vide"

#     # ── 4g. Chiffre + interjection arabe (185000 امم, ping 1000 adsl) ────
#     if CHIFFRE_INTERJECTION.match(t):
#         return "contenu_vide"

#     # ── 4h. Charabia diacritiques exotiques (zźźź1, zin øü…) ─────────────
#     if CHARABIA_DIACRITIC.match(t):
#         return "charabia_clavier"

#     # ── 4i. Terme télécom + article seul (la fibre, la fibre 2…) ─────────
#     if TECH_ARTICLE_SEUL.match(t):
#         return "technique_seul"

#     # ── 4j. Ping + chiffre + tech seul (ping 1000 adsl) ──────────────────
#     if PING_TECH_SEUL.match(t):
#         return "technique_seul"

#     # ── 4k. Séquence arabe vide (ييا يياا) ───────────────────────────────
#     if SEQUENCE_AR_VIDE.match(t):
#         return "contenu_vide"

#     # ── 4l. Charabia latin faible voyelles (hjijhghh) ─────────────────────
#     if _is_low_vowel_noise(t):
#         return "charabia_clavier"

#     # ── 5. Prix seul ──────────────────────────────────────────────────────
#     if PRIX_RE.search(t):
#         rest = PRIX_RE.sub("", t).strip()
#         if len(SUBS.findall(rest)) <= 1:
#             return "prix_detecte"

#     # ── 6. Téléphone seul ─────────────────────────────────────────────────
#     if TEL_RE.search(t):
#         rest = TEL_RE.sub("", t).strip()
#         if len(SUBS.findall(rest)) <= 2:
#             return "telephone_detecte"

#     # ── 7. Technique seul (mots-clés sans contexte) ───────────────────────
#     tech_rest = re.sub(r'\d+', '', TECH_WORDS.sub("", t)).strip()
#     if len(SUBS.findall(tech_rest)) == 0 and len(t.split()) <= 6:
#         return "technique_seul"

#     # ── 8. Charabia latin ─────────────────────────────────────────────────
#     lat = CHARABIA_LAT.findall(t)
#     if lat and sum(len(m) for m in lat) > len(t) * 0.35:
#         return "charabia_clavier"

#     # ── 9. Charabia arabe ─────────────────────────────────────────────────
#     if CHARABIA_AR.search(t):
#         ca = CHARABIA_AR.sub("", t).strip()
#         if len(ca) < len(t) * 0.45:
#             return "charabia_arabe"

#     return None   # ← ICI, jamais dans un if


# # ══════════════════════════════════════════════════════════════════════════════
# #  CORRECTIONS ARABES (nouvelles règles v3)
# # ══════════════════════════════════════════════════════════════════════════════
# CORRECTIONS_AR = [
#     # Fautes orthographiques courantes
#     (re.compile(r'\bمعاكم\b', re.UNICODE),    'معكم'),
#     (re.compile(r'\bاالتزام\b', re.UNICODE),   'التزام'),   # double alef
#     (re.compile(r'\bبيس\b', re.UNICODE),       'شريحة'),
#     (re.compile(r'\bالبونوس\b', re.UNICODE),   'بونوس'),
#     # Alef répété : اااا → ا  (garder un seul)
#     (re.compile(r'[اأإآ]{2,}', re.UNICODE),    'ا'),
#     # Interjection vide isolée
#     (re.compile(r'\bيخي\b', re.UNICODE),        ''),
#     # Tirets multiples utilisés comme séparateur décoratif
#     (re.compile(r'-{2,}'),                      ' '),
#     # Flèches textuelles décoratives
#     (re.compile(r'[—–]{2,}'),                   ' '),
# ]


# # ══════════════════════════════════════════════════════════════════════════════
# #  NORMALISATION PATTERNS (compilés une seule fois)
# # ══════════════════════════════════════════════════════════════════════════════
# _UNICODE_TABLE = str.maketrans(DICT_NORM["unicode_arabic"])
# _EMOJIS = sorted(DICT_NORM["emojis"].items(), key=lambda x: len(x[0]), reverse=True)

# _ARABIZI_UP = [(re.compile(rf'\b{re.escape(k)}\b'), v)
#                for k, v in sorted(DICT_NORM["arabizi_upper"].items(),
#                                    key=lambda x: len(x[0]), reverse=True)]
# _ARABIZI = [(re.compile(rf'\b{re.escape(k)}\b', re.IGNORECASE), v)
#             for k, v in sorted(DICT_NORM["arabizi_words"].items(),
#                                 key=lambda x: len(x[0]), reverse=True)]

# _MIXED = []
# for _p, _r in DICT_NORM["mixed_ar_fr_regex"].items():
#     try:
#         _MIXED.append((re.compile(_p, re.IGNORECASE | re.UNICODE), _r))
#     except re.error:
#         pass

# _nv = DICT_NORM["network_variants"]
# _NV = re.compile(
#     r'\b(?:' + '|'.join(re.escape(w) for w in _nv["latin"]) + r')\b'
#     + r'|(?:' + '|'.join(re.escape(w) for w in _nv["arabic"]) + r')',
#     re.IGNORECASE,
# )
# _NV_REPL = _nv["normalized_form"]

# _ABBREV = [(re.compile(rf'\b{re.escape(k)}\b', re.IGNORECASE), v)
#            for k, v in sorted(DICT_NORM["abbreviations"].items(),
#                                key=lambda x: len(x[0]), reverse=True)]
# _TELECOM = [(re.compile(rf'\b{re.escape(k)}\b', re.IGNORECASE), v)
#             for k, v in sorted(DICT_NORM["telecom_tech"].items(),
#                                 key=lambda x: len(x[0]), reverse=True)]
# _FR = []
# for _p, _r in DICT_NORM["french_corrections_regex"].items():
#     try:
#         _FR.append((re.compile(_p, re.IGNORECASE), _r))
#     except re.error:
#         pass

# _HASHTAG = re.compile(r'#\w+')
# _EXTRA   = re.compile(r'\s{2,}')
# _GLUE1   = re.compile(r'([a-zA-Z\u0621-\u064A])(\d)', re.UNICODE)
# _GLUE2   = re.compile(r'(\d)([a-zA-Z\u0621-\u064A])', re.UNICODE)


# # ══════════════════════════════════════════════════════════════════════════════
# #  FONCTION PRINCIPALE
# # ══════════════════════════════════════════════════════════════════════════════
# def clean(text: str):
#     """
#     Retourne (texte_nettoyé | None, raison_suppression | None).
#     Si supprimé → texte_nettoyé = None et raison_suppression est renseignée.
#     """
#     if not isinstance(text, str) or not text.strip():
#         return None, "vide"

#     t = text.strip()

#     # ── Étape 1 : Suppression ─────────────────────────────────────────────
#     reason = _is_suppressed(t)
#     if reason:
#         return None, reason

#     # ── Étape 2 : Unicode arabe ───────────────────────────────────────────
#     t = t.translate(_UNICODE_TABLE)

#     # ── Étape 2bis : Corrections orthographiques arabes (v3) ──────────────
#     for pat, repl in CORRECTIONS_AR:
#         t = pat.sub(repl, t)

#     # ── Étape 3 : Emojis → étiquettes arabes ─────────────────────────────
#     for em, label in _EMOJIS:
#         t = t.replace(em, f" {label} ")

#     # ── Étape 4 : Arabizi UPPER ───────────────────────────────────────────
#     for pat, r in _ARABIZI_UP:
#         t = pat.sub(r, t)

#     # ── Étape 5 : Arabizi words ───────────────────────────────────────────
#     for pat, r in _ARABIZI:
#         t = pat.sub(r, t)

#     # ── Étape 6 : Mixed AR-FR regex ───────────────────────────────────────
#     for pat, r in _MIXED:
#         t = pat.sub(r, t)

#     # ── Étape 7 : Network variants ────────────────────────────────────────
#     t = _NV.sub(_NV_REPL, t)

#     # ── Étape 8 : Abréviations ────────────────────────────────────────────
#     for pat, r in _ABBREV:
#         t = pat.sub(r, t)

#     # ── Étape 9 : Termes télécom ──────────────────────────────────────────
#     for pat, r in _TELECOM:
#         t = pat.sub(r, t)

#     # ── Étape 10 : French corrections ────────────────────────────────────
#     for pat, r in _FR:
#         t = pat.sub(r, t)

#     # ── Étape 11 : Corrections générales ─────────────────────────────────
#     t = t.lower()
#     t = _HASHTAG.sub("", t)
#     t = _GLUE1.sub(r'\1 \2', t)
#     t = _GLUE2.sub(r'\1 \2', t)
#     t = _EXTRA.sub(" ", t).strip()

#     # Si après tout le nettoyage il ne reste rien d'utile
#     if not t or len(t.strip()) < 2:
#         return None, "vide_apres_nettoyage"

#     return t, None


# # ══════════════════════════════════════════════════════════════════════════════
# #  MONGODB + EXPORT
# # ══════════════════════════════════════════════════════════════════════════════
# SUPP_COLORS = {
#     "prix_detecte":          "FFD6D6",
#     "telephone_detecte":     "FFE4C4",
#     "technique_seul":        "E0E0FF",
#     "charabia_clavier":      "FFD6FF",
#     "charabia_arabe":        "D6F0FF",
#     "vide":                  "E8E8E8",
#     # Nouveaux v3
#     "bruit_rire_clavier":    "FFF0B3",   # jaune
#     "url_seule":             "FFDAB9",   # pêche
#     "contenu_vide":          "E0FFE0",   # vert clair
#     "vide_apres_nettoyage":  "D3D3D3",   # gris
# }

# from pymongo import UpdateOne
# from tqdm import tqdm  # pip install tqdm

# def process_collection(col_name: str, limit=None) -> pd.DataFrame:
#     mongo = MongoClient(MONGO_URI)
#     col = mongo[DB_NAME][col_name]

#     cursor = col.find({}, {"_id": 1, "normalized_arabert": 1, "normalized_full": 1})
#     if limit is not None:
#         cursor = cursor.limit(limit)

#     total = col.count_documents({}) if limit is None else limit

#     rows = []
#     updated = 0
#     deleted = 0
#     bulk_ops = []
#     delete_ids = []
#     BATCH_SIZE = 500

#     for doc in tqdm(cursor, total=total, desc="Nettoyage"):
#         ca, ra = clean(doc.get("normalized_arabert", ""))
#         cf, rf = clean(doc.get("normalized_full", ""))

#         # Si les deux champs sont à supprimer → suppression du document
#         if ca is None and cf is None:
#             delete_ids.append(doc["_id"])
#             deleted += 1
#             rows.append({
#                 "_id": str(doc["_id"]),
#                 "normalized_arabert_original": doc.get("normalized_arabert", ""),
#                 "normalized_arabert_nettoyé": "⛔ SUPPRIMÉ",
#                 "normalized_arabert_suppression": ra or rf,
#                 "normalized_full_original": doc.get("normalized_full", ""),
#                 "normalized_full_nettoyé": "⛔ SUPPRIMÉ",
#                 "normalized_full_suppression": rf or ra,
#             })
#             continue

#         # Sinon, mise à jour (un ou deux champs conservés)
#         update_fields = {}
#         if ca is not None:
#             update_fields["normalized_arabert"] = ca
#         else:
#             update_fields["normalized_arabert_supprime"] = ra

#         if cf is not None:
#             update_fields["normalized_full"] = cf
#         else:
#             update_fields["normalized_full_supprime"] = rf

#         update_fields["nettoyage_v3_applique"] = True
#         bulk_ops.append(UpdateOne({"_id": doc["_id"]}, {"$set": update_fields}))
#         updated += 1

#         rows.append({
#             "_id": str(doc["_id"]),
#             "normalized_arabert_original": doc.get("normalized_arabert", ""),
#             "normalized_arabert_nettoyé": ca if ca else "⛔ SUPPRIMÉ",
#             "normalized_arabert_suppression": ra or "✓ conservé",
#             "normalized_full_original": doc.get("normalized_full", ""),
#             "normalized_full_nettoyé": cf if cf else "⛔ SUPPRIMÉ",
#             "normalized_full_suppression": rf or "✓ conservé",
#         })

#         # Envoi par lots
#         if len(bulk_ops) >= BATCH_SIZE:
#             if bulk_ops:
#                 col.bulk_write(bulk_ops, ordered=False)
#                 bulk_ops = []
#         if len(delete_ids) >= BATCH_SIZE:
#             col.delete_many({"_id": {"$in": delete_ids}})
#             delete_ids = []

#     # Derniers lots
#     if bulk_ops:
#         col.bulk_write(bulk_ops, ordered=False)
#     if delete_ids:
#         col.delete_many({"_id": {"$in": delete_ids}})

#     mongo.close()
#     print(f"\n✅ MongoDB mis à jour : {updated} docs nettoyés, {deleted} docs supprimés")
#     return pd.DataFrame(rows)

# def export_excel(df: pd.DataFrame, path: str):
#     HDR = PatternFill("solid", start_color="1B3A5C", end_color="1B3A5C")

#     with pd.ExcelWriter(path, engine="openpyxl") as writer:

#         # ── Feuille 1 : Résultats ─────────────────────────────────────────
#         df.to_excel(writer, sheet_name="Résultats_nettoyage", index=False)
#         ws = writer.sheets["Résultats_nettoyage"]

#         for cell in ws[1]:
#             cell.font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
#             cell.fill = HDR
#             cell.alignment = Alignment(horizontal="center", wrap_text=True)

#         for row in ws.iter_rows(min_row=2):
#             sup_arabert = str(row[3].value or "")
#             fill = None
#             for reason, color in SUPP_COLORS.items():
#                 if reason in sup_arabert:
#                     fill = PatternFill("solid", start_color=color, end_color=color)
#                     break
#             if "conservé" in sup_arabert:
#                 fill = PatternFill("solid", start_color="D4EDDA", end_color="D4EDDA")
#             if fill:
#                 for cell in row:
#                     cell.fill = fill
#             for cell in row:
#                 cell.alignment = Alignment(wrap_text=True, vertical="top")

#         for col in ws.columns:
#             ws.column_dimensions[get_column_letter(col[0].column)].width = 45

#         # ── Feuille 2 : Statistiques ──────────────────────────────────────
#         total = len(df)
#         stats = []
#         for field in ["normalized_arabert_suppression", "normalized_full_suppression"]:
#             for reason, cnt in df[field].value_counts().items():
#                 stats.append({
#                     "champ":    field.replace("_suppression", ""),
#                     "raison":   reason,
#                     "nb":       int(cnt),
#                     "% total":  f"{100*cnt/total:.1f}%",
#                 })
#         pd.DataFrame(stats).to_excel(writer, sheet_name="Statistiques", index=False)
#         ws2 = writer.sheets["Statistiques"]
#         for cell in ws2[1]:
#             cell.font = Font(bold=True, color="FFFFFF", name="Arial")
#             cell.fill = HDR
#         for col in ws2.columns:
#             ws2.column_dimensions[get_column_letter(col[0].column)].width = 30

#         # ── Feuille 3 : Exemples supprimés ───────────────────────────────
#         mask = df["normalized_arabert_suppression"] != "✓ conservé"
#         if mask.any():
#             samples = (df[mask]
#                        .groupby("normalized_arabert_suppression")
#                        .head(5)
#                        [["_id", "normalized_arabert_original",
#                          "normalized_arabert_suppression"]]
#                        .rename(columns={
#                            "normalized_arabert_original": "texte_original",
#                            "normalized_arabert_suppression": "raison",
#                        }))
#             samples.to_excel(writer, sheet_name="Exemples_supprimés", index=False)
#             ws3 = writer.sheets["Exemples_supprimés"]
#             for cell in ws3[1]:
#                 cell.font = Font(bold=True, color="FFFFFF", name="Arial")
#                 cell.fill = HDR
#             for col in ws3.columns:
#                 ws3.column_dimensions[get_column_letter(col[0].column)].width = 42

#         # ── Feuille 4 : Légende couleurs (v3) ────────────────────────────
#         legende_data = [
#             {"couleur_hex": c, "raison": r, "description": {
#                 "prix_detecte":         "Texte contenant uniquement un prix (DA, DZD…)",
#                 "telephone_detecte":    "Numéro de téléphone seul",
#                 "technique_seul":       "Terme technique ou unité seul (60 Mbps, adsl…)",
#                 "charabia_clavier":     "Suite de consonnes sans sens (latin)",
#                 "charabia_arabe":       "Répétition de lettres arabes sans sens",
#                 "vide":                 "Texte vide ou None",
#                 "bruit_rire_clavier":   "Rires (hhh, lol, mdr…) ou bruit clavier",
#                 "url_seule":            "URL seule sans commentaire",
#                 "contenu_vide":         "Réponse sans information (ok, يخي, كيفاش seul…)",
#                 "vide_apres_nettoyage": "Texte vide après application du pipeline",
#             }.get(r, "")}
#             for r, c in SUPP_COLORS.items()
#         ]
#         pd.DataFrame(legende_data).to_excel(
#             writer, sheet_name="Légende_couleurs", index=False
#         )
#         ws4 = writer.sheets["Légende_couleurs"]
#         for cell in ws4[1]:
#             cell.font = Font(bold=True, color="FFFFFF", name="Arial")
#             cell.fill = HDR
#         for col in ws4.columns:
#             ws4.column_dimensions[get_column_letter(col[0].column)].width = 35

#     print(f"✅ Export → {path}")


# # ══════════════════════════════════════════════════════════════════════════════
# if __name__ == "__main__":
#     mongo = MongoClient(MONGO_URI)
#     db = mongo[DB_NAME]
#     collections = db.list_collection_names()
#     mongo.close()
#     print(f"Collections disponibles : {collections}")


#     target = "dataset_unifie"
#     print(f"Traitement : '{target}'  (limite {LIMIT} docs)\n")

#     # ── Test rapide sur 20 docs avant le vrai lancement ──────────────────
#     print("── Test sur 20 docs ──")
#     df_test = process_collection(target, limit=20)
#     print(df_test[["normalized_arabert_original",
#                    "normalized_arabert_nettoyé",
#                    "normalized_arabert_suppression"]].to_string())

#     rep = input("\n✅ Le test semble bon ? Lancer sur TOUTE la collection ? (oui/non) : ")
#     if rep.strip().lower() in ("oui", "o", "yes", "y"):
#         print(f"\n🚀 Lancement complet sur '{target}'...")
#         df = process_collection(target, limit=LIMIT)
#         print(f"\nDocs traités : {len(df)}")
#         for field in ["normalized_arabert_suppression", "normalized_full_suppression"]:
#             print(f"\n[{field}]")
#             print(df[field].value_counts(dropna=False).to_string())
#         export_excel(df, OUTPUT)
#     else:
#         print("❌ Annulé. Modifiez les paramètres et relancez.")

"""
clean_telecom_FINAL_v4.py — Nettoyage corpus Télécom DZ
═══════════════════════════════════════════════════════════════════════════════
Nouveautés v4 (par rapport à v3) :
  - PONCTUATION_SEULE      : virgule seule, lettre seule, chiffre seul
  - TECH_UNIT_ONLY corrigé : accepte ponctuation finale (2g ?, 5g !)
  - MOTS_VIDES_AR          : سعر، شكرا، رائع، غالي، والحراش، وراكم…
  - MOTS_VIDES_LAT         : bravo, lies, tysm, cv, f 2026, gif, photo…
  - MOTS_VIDES_NUM         : million, milliard seuls
  - SCORE_SEUL             : 0/10, 25, 00⁰⁰00, 1ooo, ١ يي
  - QUESTION_TECH_VAGUE    : هل هو 4g, هل يوجد adsl
  - FIX_ROUTING_CAMPAGNE   : hashtag campagne copier-coller
  - TIFINAGH               : script berbère hors scope
═══════════════════════════════════════════════════════════════════════════════
"""

import re
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pymongo import MongoClient

# ── CONFIG ───────────────────────────────────────────────────────────────────

LIMIT   = None
OUTPUT  = "nettoyage_resultats_v4.xlsx"

MONGO_URI = "mongodb://localhost:27018/"
DB_NAME   = "telecom_algerie"

from dict_data import DICT_NORM, DICT_STOP


# ══════════════════════════════════════════════════════════════════════════════
#  SUPPRESSION PATTERNS
# ══════════════════════════════════════════════════════════════════════════════

PRIX_RE = re.compile(
    r'\b\d[\d\s]*(?:da|dz|dzd|دج|دينار|dinars?|euros?|€)\b'
    r'|\b(?:da|dz|dzd|دج|دينار)\s*\d[\d\s]*\b'
    r'|\b\d+\s*(?:da|dz)\b',
    re.IGNORECASE,
)
TEL_RE = re.compile(
    r'\b(?:0|\+213|00213)[\s\-]?[5-7]\d[\s\-]?\d{2}[\s\-]?\d{2}[\s\-]?\d{2}\b'
    r'|\b0[5-7]\d{8}\b'
)
TECH_WORDS = re.compile(
    r'\b(4g|5g|3g|sbox|inbox|lte|apn|sms|mms|ussd|forfait|recharge|solde)\b',
    re.IGNORECASE,
)
CHARABIA_LAT = re.compile(r'\b[b-df-hj-np-tv-z]{6,}\b', re.IGNORECASE)
CHARABIA_AR  = re.compile(r'[\u0621-\u064A]{1,2}([\u0621-\u064A])\1{3,}')
SUBS         = re.compile(r'\b\w{3,}\b')

# ── Bruit rires / clavier latin ──────────────────────────────────────────────
BRUIT_LAT = re.compile(
    r'^[\s]*(?:h{3,}|(?:[a-z]{1,3})?([a-z])\1{3,}|lol+|mdr+|xd+|haha+|hihi+|héhé+)[\s!.]*$',
    re.IGNORECASE,
)

# ── URL seule ─────────────────────────────────────────────────────────────────
URL_ONLY = re.compile(
    r'^[\s]*(?:https?[\s:/\\]+\S*|www\.\S+)[\s]*$',
    re.IGNORECASE,
)

# ── Contenu vide court (ok, يخي, كيفاش seul…) ────────────────────────────────
VIDE_CONTENT = re.compile(
    r'^[\s]*(?:'
    r'كيفاش|كيف|وقتاش|شحال|واش|علاش|فين|قداش|'
    r'ok|oui|non|no|yes|yep|nope|'
    r'تم|يخي|ايه|إيه|آه|واه|لا|ها|هه|'
    r'خي|يا خي|آ خي|بلا|هاك|هاكا|هيا|يلا|'
    r'[اى]{3,}|ات+ال+ات+|[أإآا]{4,}'
    r')[\s?!.،؟…]*$',
    re.IGNORECASE | re.UNICODE,
)

# ── Technique seul avec unité — VERSION CORRIGÉE v4 (accepte ponctuation) ────
TECH_UNIT_ONLY = re.compile(
    r'^[\s]*\d*[\s]*(mbps|mb/s|kbps|gb/s|mbits?|mb|gb|go|mo|ko|ghz|mhz|'
    r'adsl|vdsl|4g|5g|3g|2g|lte|ftth|fttb|gpon|ping|dns)[\s?!.،؟…]*$',
    re.IGNORECASE,
)

# ── Interjection vide (let's go, photo, lies, ni…) ───────────────────────────
INTERJECTION_VIDE = re.compile(
    r'^[\s]*(?:'
    r"let'?s\s*goo*|"
    r'yall?a+|yala+|'
    r'cool+|super+|bravo+|bien+|'
    r'ni|lies?|'
    r'intéressé[e]?|interested|'
    r'photo(?:\s+\w{1,10})?|'
    r'ok\s+cool|cool\s+ok'
    r')[\s?!.،؟…]*$',
    re.IGNORECASE | re.UNICODE,
)

# ── Prix mot seul ─────────────────────────────────────────────────────────────
PRIX_MOT_SEUL = re.compile(
    r'^[\s]*(?:'
    r'(?:le\s+)?prix|سعر|ثمن|تمن|لسعر|وسعر'
    r')[\s?!.،؟…]*$',
    re.IGNORECASE | re.UNICODE,
)

# ── Chiffre + mot numérique vague ─────────────────────────────────────────────
CHIFFRE_MOT_VAGUE = re.compile(
    r'^[\s]*(?:'
    r'\d[\d\s]*(?:ooo+|millions?|billions?|three|two|one|zero|com\b.*)'
    r'|millions?|billions?'
    r')[\s?!.،؟…]*$',
    re.IGNORECASE | re.UNICODE,
)

# ── Villes algériennes seules ─────────────────────────────────────────────────
VILLE_SEULE = re.compile(
    r'^[\s]*(?:'
    r'sidi\s+bel\s+abb[eè]s|oran|alger(?:s)?|constantine|annaba|'
    r'tizi\s+ouzou|bejaia|béjaïa|blida|setif|sétif|batna|biskra|'
    r'tlemcen|chlef|médéa|mostaganem|mascara|djelfa|m.?sila|'
    r'skikda|jijel|guelma|souk\s+ahras|tiaret|ghardaia|bechar'
    r')[\s?!.,،؟…]*$',
    re.IGNORECASE | re.UNICODE,
)

# ── Charabia diacritiques exotiques ──────────────────────────────────────────
CHARABIA_DIACRITIC = re.compile(
    r'^[\s\w]{0,5}[żźśøüðþæœ][\s\w]{0,10}$',
    re.IGNORECASE | re.UNICODE,
)

# ── Interjections arabes (ييا, همم, هذا 4g, واي فاي 6) ───────────────────────
INTERJECTION_AR_VIDE = re.compile(
    r'^[\s]*(?:'
    r'[يه]{2,}[اوى]+[اه]*|'
    r'هم+|هه+خ+|ام+|اهه+|'
    r'هذا\s*(?:4g|5g|3g|lte|adsl|wifi)|'
    r'واي\s*فاي\s*\d+|'
    r'سيدي\s+\S+(?:\s+\S+)?'
    r')[\s?!.،؟…]*$',
    re.IGNORECASE | re.UNICODE,
)

# ── Chiffre + interjection arabe (185000 امم, ping 1000 adsl) ────────────────
CHIFFRE_INTERJECTION = re.compile(
    r'^[\s]*\d[\d\s]*\s*(?:'
    r'ام+|اهه+|هم+|همم+|اممم+|'
    r'ping\s+adsl|adsl\s+ping'
    r')[\s?!.،؟…]*$',
    re.IGNORECASE | re.UNICODE,
)

# ── Terme télécom + article seul (la fibre, la fibre 2…) ─────────────────────
TECH_ARTICLE_SEUL = re.compile(
    r'^[\s]*(?:la\s+|le\s+|les\s+|une?\s+|du\s+|de\s+)?'
    r'(?:fibre|fiber|adsl|vdsl|4g|5g|3g|lte|ftth|fttb|wifi|connexion|réseau|network)'
    r'(?:\s+\d+)?[\s?!.،؟…]*$',
    re.IGNORECASE | re.UNICODE,
)

# ── Ping + tech seul (ping 1000 adsl, ping!…) ────────────────────────────────
PING_TECH_SEUL = re.compile(
    r'^[\s]*(?:ping\s*\d*\s*(?:adsl|vdsl|4g|5g|lte|ms)?'
    r'|\d+\s*(?:ping|ms)\s*(?:adsl|vdsl|4g|5g)?)[\s?!.]*$',
    re.IGNORECASE,
)

# ── Séquences arabes vides (ييا يياا) ────────────────────────────────────────
SEQUENCE_AR_VIDE = re.compile(
    r'^[\s]*[يا]{3,}[\s]*[يا]{0,5}[\s?!.،؟…]*$',
    re.UNICODE,
)

# ══════════════════════════════════════════════════════════════════════════════
#  NOUVEAUX PATTERNS v4
# ══════════════════════════════════════════════════════════════════════════════

# Ponctuation / caractère unique / chiffre seul (, . 0 y ر م…)
PONCTUATION_SEULE = re.compile(r'^[\s\W\d]{1,5}$', re.UNICODE)

# Mots vides arabes isolés (sans info télécom utile)
MOTS_VIDES_AR = re.compile(
    r'^[\s]*(?:'
    # Prix/valeur seuls
    r'سعر|ثمن|تمن|الثمن|السعر|لسعر|وسعر|بشحال|'
    # Réactions positives seules
    r'رائع|روعه|جيد|هايله|منيح|ممتاز|'
    # Réactions négatives seules
    r'غالي|غاليه|ثقيل[ه]?|ضعيف[ه]?|فاشل[ه]?|زباله|زيرو|ازبل|اطز|'
    r'طزطز+|طووز|مافيا|خداع|خيانه|سارق|'
    # Salutations/politesse seules
    r'سلام|شكرا|بالتوفيق|باتوفيق|بتوفيق|موفقين|مشكورين|'
    # Réponses courtes sans info
    r'نعم|همم|والوو+|'
    # Demandes sans contexte
    r'ردوا|ردووا|ريبوندو+|روودووا|وجاوبو|راسلوني|'
    # Mécontentement seul sans description du problème
    r'تعبتونا|تعبتوني|انهكتنا|هرمنا|شبعتونا|مرضتونا|'
    # Villes/lieux arabes seuls
    r'والحراش|قسنطينه|وراكم|'
    # Questions vagues sans contexte
    r'وش\s*هي\s*[؟?]?|وش\s*هو\s*[؟?]?|'
    # Mots courts sans sens télécom
    r'راد|جهال|حب|خاص|ناقص|مزال|جيبولنا|نول|'
    # Lettres arabes isolées (1-2 lettres)
    r'[ءأإآابتثجحخدذرزسشصضطظعغفقكلمنهوي]{1,2}'
    r')[\s?!.،؟…،؛]*$',
    re.IGNORECASE | re.UNICODE,
)

# Mots vides latins/français isolés
MOTS_VIDES_LAT = re.compile(
    r'^[\s]*(?:'
    r'bravo+|cool+|super+|bien+|nul+|nuull+|'
    r'mensonge|lies?|impossible|gratuit|'
    r'interested|intéressé[e]?|'
    r'tysm|tm|cv|ps\d*|xc\d+[-\w]*|'
    r'y\b|w\b|ni\b|nn\s*hh|'
    r'gif|photo(?:\s+\w{1,15})?|'
    r'f\s+\d{4}|'                          # F 2026
    r'www(?:\.\S+)?|'
    r"let'?s\s*goo+|let'?s\s*go|"
    r'yall?a+|'
    r'no\s+comment|no\s+nut\s+\w+|nuh\s+uh|'
    r'nobody\s+is\s+\w+|'
    r'ok\s*(?:cool)?|'
    r'jamais\s+de\s+la\s+vie|pas\s+encore'
    r')[\s?!.،؟…]*$',
    re.IGNORECASE | re.UNICODE,
)

# Mots numériques vagues seuls
MOTS_VIDES_NUM = re.compile(
    r'^[\s]*(?:million+|billion+|milliard+)[\s?!.،؟…]*$',
    re.IGNORECASE,
)

# Score / note / chiffre seul sans contexte (0/10, 25, 1ooo, ١ يي…)
SCORE_SEUL = re.compile(
    r'^[\s]*(?:'
    r'\d+/\d+|'                         # 0/10, 5/10
    r'0+[⁰°₀]*0*|'                     # 00, 00⁰⁰00
    r'\d{1,5}(?:\s+ooo+)?|'            # 25, 1 ooo, 185000
    r'\d+\s*[mo]o?\b|'                 # 2 Mo, 3m
    r'[١٢٣٤٥٦٧٨٩٠]\s*[يا]{1,3}'      # ١ يي
    r')[\s?!.]*$',
    re.IGNORECASE | re.UNICODE,
)

# Question technique vague sans info réelle
QUESTION_TECH_VAGUE = re.compile(
    r'^[\s]*(?:هل\s+هو|هل\s+هي|هل\s+يوجد|هل\s+كاين)\s*'
    r'(?:4g|5g|3g|adsl|lte|wifi|fibre)?[\s?!.،؟…]*$',
    re.IGNORECASE | re.UNICODE,
)

# Campagne FixRouting copier-coller (sans contenu personnel)
FIX_ROUTING_CAMPAGNE = re.compile(
    r'FixRoutingAT|FixRoutingInAlgeria|FixeDSLAlgeria',
    re.IGNORECASE,
)

# Script Tifinagh (berbère) — hors scope corpus
TIFINAGH = re.compile(r'[\u2D30-\u2D7F]{3,}')


# ── Charabia latin faible voyelles ───────────────────────────────────────────
def _is_low_vowel_noise(text: str) -> bool:
    """Mot latin court (< 12 chars) avec < 15% de voyelles → bruit clavier."""
    t = text.strip().lower()
    if not t or len(t) > 12 or not t.isalpha():
        return False
    vowels = sum(1 for c in t if c in 'aeiouéèêëàâùûü')
    return (vowels / len(t)) < 0.15


# ══════════════════════════════════════════════════════════════════════════════
#  FONCTION _is_suppressed — VERSION COMPLÈTE v4
# ══════════════════════════════════════════════════════════════════════════════
def _is_suppressed(t: str):
    """
    Retourne la raison de suppression, ou None si le doc est à conserver.
    """
    # ── 0. Ponctuation / caractère unique / chiffre seul ─────────────────
    if PONCTUATION_SEULE.match(t):
        return "contenu_vide"

    # ── 1. Bruit rires / clavier latin ───────────────────────────────────
    if BRUIT_LAT.match(t):
        return "bruit_rire_clavier"

    # ── 2. URL seule ──────────────────────────────────────────────────────
    if URL_ONLY.match(t):
        return "url_seule"

    # ── 3. Contenu vide (ok, يخي, كيفاش seul…) ───────────────────────────
    if VIDE_CONTENT.match(t):
        return "contenu_vide"

    # ── 4. Technique seul avec unité (60 Mbps, adsl, 2g ?…) ──────────────
    if TECH_UNIT_ONLY.match(t):
        return "technique_seul"

    # ── 4b. Prix mot seul (prix, le prix, سعر, ثمن…) ─────────────────────
    if PRIX_MOT_SEUL.match(t):
        return "contenu_vide"

    # ── 4c. Interjection vide (let's go, photo, lies, ni…) ───────────────
    if INTERJECTION_VIDE.match(t):
        return "contenu_vide"

    # ── 4d. Chiffre + mot numérique vague (1 ooo, 3 millions…) ───────────
    if CHIFFRE_MOT_VAGUE.match(t):
        return "contenu_vide"

    # ── 4e. Nom de ville seul ─────────────────────────────────────────────
    if VILLE_SEULE.match(t):
        return "contenu_vide"

    # ── 4f. Interjections arabes (ييا, همم, هذا 4g, واي فاي 6) ──────────
    if INTERJECTION_AR_VIDE.match(t):
        return "contenu_vide"

    # ── 4g. Chiffre + interjection arabe (185000 امم, ping 1000 adsl) ────
    if CHIFFRE_INTERJECTION.match(t):
        return "contenu_vide"

    # ── 4h. Charabia diacritiques exotiques (zźźź1, zin øü…) ─────────────
    if CHARABIA_DIACRITIC.match(t):
        return "charabia_clavier"

    # ── 4i. Terme télécom + article seul (la fibre, la fibre 2…) ─────────
    if TECH_ARTICLE_SEUL.match(t):
        return "technique_seul"

    # ── 4j. Ping + chiffre + tech seul (ping 1000 adsl) ──────────────────
    if PING_TECH_SEUL.match(t):
        return "technique_seul"

    # ── 4k. Séquence arabe vide (ييا يياا) ───────────────────────────────
    if SEQUENCE_AR_VIDE.match(t):
        return "contenu_vide"

    # ── 4l. Charabia latin faible voyelles (hjijhghh) ─────────────────────
    if _is_low_vowel_noise(t):
        return "charabia_clavier"

    # ══ NOUVEAUX v4 ════════════════════════════════════════════════════════

    # ── N1. Mots vides arabes isolés ──────────────────────────────────────
    if MOTS_VIDES_AR.match(t):
        return "contenu_vide"

    # ── N2. Mots vides latins isolés ──────────────────────────────────────
    if MOTS_VIDES_LAT.match(t):
        return "contenu_vide"

    # ── N3. Mots numériques vagues (million, milliard…) ───────────────────
    if MOTS_VIDES_NUM.match(t):
        return "contenu_vide"

    # ── N4. Score / note / chiffre seul (0/10, 25, 1ooo, ١ يي) ──────────
    if SCORE_SEUL.match(t):
        return "contenu_vide"

    # ── N5. Question technique vague (هل هو 4g) ───────────────────────────
    if QUESTION_TECH_VAGUE.match(t):
        return "contenu_vide"

    # ── N6. Script Tifinagh hors scope ────────────────────────────────────
    if TIFINAGH.search(t) and len(t) < 30:
        return "charabia_clavier"

    # ── N7. Campagne FixRouting copier-coller ─────────────────────────────
    if FIX_ROUTING_CAMPAGNE.search(t) and len(t.split()) <= 20:
        return "contenu_vide"

    # ══ PATTERNS EXISTANTS ═════════════════════════════════════════════════

    # ── 5. Prix seul ──────────────────────────────────────────────────────
    if PRIX_RE.search(t):
        rest = PRIX_RE.sub("", t).strip()
        if len(SUBS.findall(rest)) <= 1:
            return "prix_detecte"

    # ── 6. Téléphone seul ─────────────────────────────────────────────────
    if TEL_RE.search(t):
        rest = TEL_RE.sub("", t).strip()
        if len(SUBS.findall(rest)) <= 2:
            return "telephone_detecte"

    # ── 7. Technique seul (mots-clés sans contexte) ───────────────────────
    tech_rest = re.sub(r'\d+', '', TECH_WORDS.sub("", t)).strip()
    if len(SUBS.findall(tech_rest)) == 0 and len(t.split()) <= 6:
        return "technique_seul"

    # ── 8. Charabia latin ─────────────────────────────────────────────────
    lat = CHARABIA_LAT.findall(t)
    if lat and sum(len(m) for m in lat) > len(t) * 0.35:
        return "charabia_clavier"

    # ── 9. Charabia arabe ─────────────────────────────────────────────────
    if CHARABIA_AR.search(t):
        ca = CHARABIA_AR.sub("", t).strip()
        if len(ca) < len(t) * 0.45:
            return "charabia_arabe"

    return None


# ══════════════════════════════════════════════════════════════════════════════
#  CORRECTIONS ARABES
# ══════════════════════════════════════════════════════════════════════════════
CORRECTIONS_AR = [
    (re.compile(r'\bمعاكم\b', re.UNICODE),    'معكم'),
    (re.compile(r'\bاالتزام\b', re.UNICODE),   'التزام'),
    (re.compile(r'\bبيس\b', re.UNICODE),       'شريحة'),
    (re.compile(r'\bالبونوس\b', re.UNICODE),   'بونوس'),
    (re.compile(r'[اأإآ]{2,}', re.UNICODE),    'ا'),
    (re.compile(r'\bيخي\b', re.UNICODE),        ''),
    (re.compile(r'-{2,}'),                      ' '),
    (re.compile(r'[—–]{2,}'),                   ' '),
]


# ══════════════════════════════════════════════════════════════════════════════
#  NORMALISATION PATTERNS
# ══════════════════════════════════════════════════════════════════════════════
_UNICODE_TABLE = str.maketrans(DICT_NORM["unicode_arabic"])
_EMOJIS = sorted(DICT_NORM["emojis"].items(), key=lambda x: len(x[0]), reverse=True)

_ARABIZI_UP = [(re.compile(rf'\b{re.escape(k)}\b'), v)
               for k, v in sorted(DICT_NORM["arabizi_upper"].items(),
                                   key=lambda x: len(x[0]), reverse=True)]
_ARABIZI = [(re.compile(rf'\b{re.escape(k)}\b', re.IGNORECASE), v)
            for k, v in sorted(DICT_NORM["arabizi_words"].items(),
                                key=lambda x: len(x[0]), reverse=True)]

_MIXED = []
for _p, _r in DICT_NORM["mixed_ar_fr_regex"].items():
    try:
        _MIXED.append((re.compile(_p, re.IGNORECASE | re.UNICODE), _r))
    except re.error:
        pass

_nv = DICT_NORM["network_variants"]
_NV = re.compile(
    r'\b(?:' + '|'.join(re.escape(w) for w in _nv["latin"]) + r')\b'
    + r'|(?:' + '|'.join(re.escape(w) for w in _nv["arabic"]) + r')',
    re.IGNORECASE,
)
_NV_REPL = _nv["normalized_form"]

_ABBREV = [(re.compile(rf'\b{re.escape(k)}\b', re.IGNORECASE), v)
           for k, v in sorted(DICT_NORM["abbreviations"].items(),
                               key=lambda x: len(x[0]), reverse=True)]
_TELECOM = [(re.compile(rf'\b{re.escape(k)}\b', re.IGNORECASE), v)
            for k, v in sorted(DICT_NORM["telecom_tech"].items(),
                                key=lambda x: len(x[0]), reverse=True)]
_FR = []
for _p, _r in DICT_NORM["french_corrections_regex"].items():
    try:
        _FR.append((re.compile(_p, re.IGNORECASE), _r))
    except re.error:
        pass

_HASHTAG = re.compile(r'#\w+')
_EXTRA   = re.compile(r'\s{2,}')
_GLUE1   = re.compile(r'([a-zA-Z\u0621-\u064A])(\d)', re.UNICODE)
_GLUE2   = re.compile(r'(\d)([a-zA-Z\u0621-\u064A])', re.UNICODE)


# ══════════════════════════════════════════════════════════════════════════════
#  FONCTION PRINCIPALE clean()
# ══════════════════════════════════════════════════════════════════════════════
def clean(text: str):
    """
    Retourne (texte_nettoyé | None, raison_suppression | None).
    Si supprimé → texte_nettoyé = None et raison_suppression est renseignée.
    """
    if not isinstance(text, str) or not text.strip():
        return None, "vide"

    t = text.strip()

    # ── Étape 1 : Suppression ─────────────────────────────────────────────
    reason = _is_suppressed(t)
    if reason:
        return None, reason

    # ── Étape 2 : Unicode arabe ───────────────────────────────────────────
    t = t.translate(_UNICODE_TABLE)

    # ── Étape 2bis : Corrections orthographiques arabes ───────────────────
    for pat, repl in CORRECTIONS_AR:
        t = pat.sub(repl, t)

    # ── Étape 3 : Emojis → étiquettes arabes ─────────────────────────────
    for em, label in _EMOJIS:
        t = t.replace(em, f" {label} ")

    # ── Étape 4 : Arabizi UPPER ───────────────────────────────────────────
    for pat, r in _ARABIZI_UP:
        t = pat.sub(r, t)

    # ── Étape 5 : Arabizi words ───────────────────────────────────────────
    for pat, r in _ARABIZI:
        t = pat.sub(r, t)

    # ── Étape 6 : Mixed AR-FR regex ───────────────────────────────────────
    for pat, r in _MIXED:
        t = pat.sub(r, t)

    # ── Étape 7 : Network variants ────────────────────────────────────────
    t = _NV.sub(_NV_REPL, t)

    # ── Étape 8 : Abréviations ────────────────────────────────────────────
    for pat, r in _ABBREV:
        t = pat.sub(r, t)

    # ── Étape 9 : Termes télécom ──────────────────────────────────────────
    for pat, r in _TELECOM:
        t = pat.sub(r, t)

    # ── Étape 10 : French corrections ─────────────────────────────────────
    for pat, r in _FR:
        t = pat.sub(r, t)

    # ── Étape 11 : Corrections générales ──────────────────────────────────
    t = t.lower()
    t = _HASHTAG.sub("", t)
    t = _GLUE1.sub(r'\1 \2', t)
    t = _GLUE2.sub(r'\1 \2', t)
    t = _EXTRA.sub(" ", t).strip()

    if not t or len(t.strip()) < 2:
        return None, "vide_apres_nettoyage"

    return t, None


# ══════════════════════════════════════════════════════════════════════════════
#  COULEURS EXCEL
# ══════════════════════════════════════════════════════════════════════════════
SUPP_COLORS = {
    "prix_detecte":          "FFD6D6",
    "telephone_detecte":     "FFE4C4",
    "technique_seul":        "E0E0FF",
    "charabia_clavier":      "FFD6FF",
    "charabia_arabe":        "D6F0FF",
    "vide":                  "E8E8E8",
    "bruit_rire_clavier":    "FFF0B3",
    "url_seule":             "FFDAB9",
    "contenu_vide":          "E0FFE0",
    "vide_apres_nettoyage":  "D3D3D3",
}


# ══════════════════════════════════════════════════════════════════════════════
#  MONGODB + TRAITEMENT
# ══════════════════════════════════════════════════════════════════════════════
from pymongo import UpdateOne
from tqdm import tqdm

def process_collection(col_name: str, limit=None) -> pd.DataFrame:
    mongo = MongoClient(MONGO_URI)
    col = mongo[DB_NAME][col_name]

    cursor = col.find({}, {"_id": 1, "normalized_arabert": 1, "normalized_full": 1})
    if limit is not None:
        cursor = cursor.limit(limit)

    total = col.count_documents({}) if limit is None else limit

    rows = []
    updated = 0
    deleted = 0
    bulk_ops = []
    delete_ids = []
    BATCH_SIZE = 500

    for doc in tqdm(cursor, total=total, desc="Nettoyage v4"):
        ca, ra = clean(doc.get("normalized_arabert", ""))
        cf, rf = clean(doc.get("normalized_full", ""))

        if ca is None and cf is None:
            delete_ids.append(doc["_id"])
            deleted += 1
            rows.append({
                "_id": str(doc["_id"]),
                "normalized_arabert_original": doc.get("normalized_arabert", ""),
                "normalized_arabert_nettoyé": "⛔ SUPPRIMÉ",
                "normalized_arabert_suppression": ra or rf,
                "normalized_full_original": doc.get("normalized_full", ""),
                "normalized_full_nettoyé": "⛔ SUPPRIMÉ",
                "normalized_full_suppression": rf or ra,
            })
            continue

        update_fields = {}
        if ca is not None:
            update_fields["normalized_arabert"] = ca
        else:
            update_fields["normalized_arabert_supprime"] = ra

        if cf is not None:
            update_fields["normalized_full"] = cf
        else:
            update_fields["normalized_full_supprime"] = rf

        update_fields["nettoyage_v4_applique"] = True
        bulk_ops.append(UpdateOne({"_id": doc["_id"]}, {"$set": update_fields}))
        updated += 1

        rows.append({
            "_id": str(doc["_id"]),
            "normalized_arabert_original": doc.get("normalized_arabert", ""),
            "normalized_arabert_nettoyé": ca if ca else "⛔ SUPPRIMÉ",
            "normalized_arabert_suppression": ra or "✓ conservé",
            "normalized_full_original": doc.get("normalized_full", ""),
            "normalized_full_nettoyé": cf if cf else "⛔ SUPPRIMÉ",
            "normalized_full_suppression": rf or "✓ conservé",
        })

        if len(bulk_ops) >= BATCH_SIZE:
            if bulk_ops:
                col.bulk_write(bulk_ops, ordered=False)
                bulk_ops = []
        if len(delete_ids) >= BATCH_SIZE:
            col.delete_many({"_id": {"$in": delete_ids}})
            delete_ids = []

    if bulk_ops:
        col.bulk_write(bulk_ops, ordered=False)
    if delete_ids:
        col.delete_many({"_id": {"$in": delete_ids}})

    mongo.close()
    print(f"\n✅ MongoDB mis à jour : {updated} docs nettoyés, {deleted} docs supprimés")
    return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════════════
#  EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def export_excel(df: pd.DataFrame, path: str):
    HDR = PatternFill("solid", start_color="1B3A5C", end_color="1B3A5C")

    with pd.ExcelWriter(path, engine="openpyxl") as writer:

        # Feuille 1 : Résultats
        df.to_excel(writer, sheet_name="Résultats_nettoyage", index=False)
        ws = writer.sheets["Résultats_nettoyage"]

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
            cell.fill = HDR
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for row in ws.iter_rows(min_row=2):
            sup_arabert = str(row[3].value or "")
            fill = None
            for reason, color in SUPP_COLORS.items():
                if reason in sup_arabert:
                    fill = PatternFill("solid", start_color=color, end_color=color)
                    break
            if "conservé" in sup_arabert:
                fill = PatternFill("solid", start_color="D4EDDA", end_color="D4EDDA")
            if fill:
                for cell in row:
                    cell.fill = fill
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 45

        # Feuille 2 : Statistiques
        total = len(df)
        stats = []
        for field in ["normalized_arabert_suppression", "normalized_full_suppression"]:
            for reason, cnt in df[field].value_counts().items():
                stats.append({
                    "champ":    field.replace("_suppression", ""),
                    "raison":   reason,
                    "nb":       int(cnt),
                    "% total":  f"{100*cnt/total:.1f}%",
                })
        pd.DataFrame(stats).to_excel(writer, sheet_name="Statistiques", index=False)
        ws2 = writer.sheets["Statistiques"]
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF", name="Arial")
            cell.fill = HDR
        for col in ws2.columns:
            ws2.column_dimensions[get_column_letter(col[0].column)].width = 30

        # Feuille 3 : Exemples supprimés
        mask = df["normalized_arabert_suppression"] != "✓ conservé"
        if mask.any():
            samples = (df[mask]
                       .groupby("normalized_arabert_suppression")
                       .head(5)
                       [["_id", "normalized_arabert_original",
                         "normalized_arabert_suppression"]]
                       .rename(columns={
                           "normalized_arabert_original": "texte_original",
                           "normalized_arabert_suppression": "raison",
                       }))
            samples.to_excel(writer, sheet_name="Exemples_supprimés", index=False)
            ws3 = writer.sheets["Exemples_supprimés"]
            for cell in ws3[1]:
                cell.font = Font(bold=True, color="FFFFFF", name="Arial")
                cell.fill = HDR
            for col in ws3.columns:
                ws3.column_dimensions[get_column_letter(col[0].column)].width = 42

        # Feuille 4 : Légende couleurs
        legende_data = [
            {"couleur_hex": c, "raison": r, "description": {
                "prix_detecte":         "Texte contenant uniquement un prix (DA, DZD…)",
                "telephone_detecte":    "Numéro de téléphone seul",
                "technique_seul":       "Terme technique ou unité seul (60 Mbps, adsl, 2g ?…)",
                "charabia_clavier":     "Suite de consonnes sans sens (latin/tifinagh)",
                "charabia_arabe":       "Répétition de lettres arabes sans sens",
                "vide":                 "Texte vide ou None",
                "bruit_rire_clavier":   "Rires (hhh, lol, mdr…) ou bruit clavier",
                "url_seule":            "URL seule sans commentaire",
                "contenu_vide":         "Sans info utile (ok, يخي, سعر, bravo, 25, ييا…)",
                "vide_apres_nettoyage": "Texte vide après application du pipeline",
            }.get(r, "")}
            for r, c in SUPP_COLORS.items()
        ]
        pd.DataFrame(legende_data).to_excel(
            writer, sheet_name="Légende_couleurs", index=False
        )
        ws4 = writer.sheets["Légende_couleurs"]
        for cell in ws4[1]:
            cell.font = Font(bold=True, color="FFFFFF", name="Arial")
            cell.fill = HDR
        for col in ws4.columns:
            ws4.column_dimensions[get_column_letter(col[0].column)].width = 35

    print(f"✅ Export → {path}")


# ══════════════════════════════════════════════════════════════════════════════
#  POINT D'ENTRÉE
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    mongo = MongoClient(MONGO_URI)
    db = mongo[DB_NAME]
    collections = db.list_collection_names()
    mongo.close()
    print(f"Collections disponibles : {collections}")

    target = "dataset_unifie"
    print(f"Traitement : '{target}'  (limite {LIMIT} docs)\n")

    # Test rapide sur 20 docs
    print("── Test sur 20 docs ──")
    df_test = process_collection(target, limit=20)
    print(df_test[["normalized_arabert_original",
                   "normalized_arabert_nettoyé",
                   "normalized_arabert_suppression"]].to_string())

    rep = input("\n✅ Le test semble bon ? Lancer sur TOUTE la collection ? (oui/non) : ")
    if rep.strip().lower() in ("oui", "o", "yes", "y"):
        print(f"\n🚀 Lancement complet sur '{target}'...")
        df = process_collection(target, limit=LIMIT)
        print(f"\nDocs traités : {len(df)}")
        for field in ["normalized_arabert_suppression", "normalized_full_suppression"]:
            print(f"\n[{field}]")
            print(df[field].value_counts(dropna=False).to_string())
        export_excel(df, OUTPUT)
    else:
        print("❌ Annulé. Modifiez les paramètres et relancez.")